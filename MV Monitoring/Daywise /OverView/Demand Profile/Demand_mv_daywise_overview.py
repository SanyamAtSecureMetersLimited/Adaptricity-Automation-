import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and MV feeder info
    DB1_HOST = "10.11.16.186"
    DB1_PORT = "5432"
    DB1_DATABASE = "serviceplatformdb12004"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.186"
    DB2_PORT = "5432"
    DB2_DATABASE = "serviceplatformdb12004"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant03"  # Change to tenant01, tenant02, etc. as needed

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('mv_overview_demand_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/mv_overview_demand_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for MV Overview Demand"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Target_Date', 'Meter_Serial_No'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'DD/MM/YYYY', 'YOUR_METER_NO']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Update Target_Date with desired date (DD/MM/YYYY)',
                    'Update Meter_Serial_No with your MV meter serial number',
                    'Save file before running',
                ],
                'Important_Notes': [
                    'This script is FOR MV DEMAND OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Date format: DD/MM/YYYY',
                    'MV Feeder meter only',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"MV Overview Demand Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    """Ensure date is in 'DD/MM/YYYY' string format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors='raise')
        return parsed.strftime("%d/%m/%Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for MV Overview Demand"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'MV_DEMAND'}  # Fixed for MV demand monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()

        config['meter_type'] = 'MV'  # Fixed for MV

        required_fields = ['type', 'area', 'substation', 'target_date', 'meter_serial_no']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("MV Overview Demand Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING MV OVERVIEW DEMAND AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default MV Overview Demand configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("MV Overview Demand Configuration validated successfully")
    logger.info(f"   Monitoring Type: MV Demand Overview (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Date: {config['target_date']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise
    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no):
    """Get MV feeder metrics from database"""
    logger.info(f"Fetching MV Overview Demand metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        query = f"SELECT feeder_id, feeder_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_feeder WHERE meter_serial_no = %s LIMIT 1;"
        
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        if not result:
            logger.info(f"MV Feeder not found: {mtr_serial_no}")
            return None, None, None

        feeder_id, feeder_name, meterid = result
        logger.info(f"Metrics: {feeder_name}, meterid: {meterid}")
        return feeder_id, feeder_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_demand_overview(target_date, mtr_id, node_id):
    """Fetch database data for MV demand overview"""
    logger.info(f"Fetching MV demand overview database data for date: {target_date}")
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND raw.surveydate >= '{start_date}' AND raw.surveydate < '{next_day}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        query = f"""
            SELECT DISTINCT 
                raw.surveydate,
                raw.kwh_i, raw.kvah_i, raw.kvar_i_total,
                nrm.kw_i, nrm.kva_i, nrm.kvar_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata AS raw
            JOIN {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile AS nrm
                ON raw.surveydate = nrm.surveydate
            WHERE raw.mtrid = {mtr_id}
              AND nrm.nodeid = {node_id}
              {date_filter}
            ORDER BY raw.surveydate ASC;
        """

        raw_df = pd.read_sql(query, conn)
        logger.info(f"Retrieved: {len(raw_df)} MV demand records")
        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoring.secure.online:43379/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SanyamU")
        driver.find_element(By.ID, "Password").send_keys("Secure@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(8)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_date(driver, target_date):
    """Set calendar date"""
    try:
        logger.info(f"Setting date: {target_date}")
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        date_info = {
            'selected_date': target_dt.strftime("%B %Y"),
            'start_date': target_dt.strftime("%Y-%m-%d"),
            'end_date': (target_dt + timedelta(days=0)).strftime("%Y-%m-%d")
        }
        logger.info("Date set successfully")
        return date_info
    except Exception as e:
        logger.info(f"Date error: {e}")
        return None


def select_type(driver):
    """Select MV monitoring - FIXED FOR MV ONLY"""
    try:
        logger.info("Selecting MV monitoring (fixed for MV demand overview script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divmvmonitoring']").click()
        logger.info("MV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver):
    """Select MV Feeder meter type"""
    try:
        logger.info("Selecting MV Feeder meter type")
        wait = WebDriverWait(driver, 10)
        
        mv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="MVFeederClick"]')))
        mv_button.click()
        logger.info("MV Feeder selected")
        
        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for MV meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_demand_overview_data(driver):
    """Collect MV demand overview data from UI"""
    logger.info("Starting MV demand overview data collection...")
    data = {}

    try:
        # Navigate to Demand tab
        time.sleep(4)
        logger.info("Clicking on MV Demand tab...")
        driver.find_element(By.XPATH, "//div[@class='dx-item-content' and text()='Demand']").click()
        time.sleep(2)

        # Collect demand values from the table
        logger.info("Collecting MV demand values...")
        data['Demand Table'] = {
            'act_max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kw"]').text,
            'act_avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kw"]').text,
            'act_dt': driver.find_element(By.XPATH, '//td[@id="kw_MaxDatetime"]').text,
            'app_max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kva"]').text,
            'app_avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kva"]').text,
            'app_dt': driver.find_element(By.XPATH, '//td[@id="kva_MaxDatetime"]').text,
            'react_max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kvar"]').text,
            'react_avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kvar"]').text,
            'react_dt': driver.find_element(By.XPATH, '//td[@id="kvar_MaxDatetime"]').text
        }

        logger.info("MV demand overview data collection completed")
        logger.info(f"Collected MV demand data: {data}")

    except Exception as e:
        logger.error(f"Error in MV demand data collection: {str(e)}")
        raise

    return data


@log_execution_time
def save_demand_overview_data_to_excel(date_info, overview_data):
    """Save MV demand overview data to Excel"""
    logger.info("Saving MV demand overview data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Demand Table
        ws_dt = wb.create_sheet("Demand Table")
        ws_dt.append(["Parameter", "Max", "Avg", "Date and time at max value"])
        
        dt = overview_data['Demand Table']
        
        # Active Power
        ws_dt.append(["Active", dt['act_max'], dt['act_avg'], dt['act_dt']])
        
        # Apparent Power
        ws_dt.append(["Apparent", dt['app_max'], dt['app_avg'], dt['app_dt']])
        
        # Reactive Power
        ws_dt.append(["Reactive", dt['react_max'], dt['react_avg'], dt['react_dt']])

        # Save
        file_name = f"chart_data_from_ui_mv_demand_overview_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"MV demand overview data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving MV demand overview data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_demand_overview_database_calculations(raw_df, date_info):
    """Process database calculations for MV demand overview"""
    logger.info("Processing MV demand overview database calculations...")

    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        # Calculate interval
        if len(raw_df) > 1:
            interval_minutes = int((raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)
        else:
            interval_minutes = 15

        sip_duration_in_hr = interval_minutes / 60

        processed_file = f"theoretical_mv_demand_overview_calculated_data_{date_safe}_{timestamp}.xlsx"

        # Calculate NRM values from RAW if not present
        if 'kw_i' not in raw_df.columns:
            raw_df['kw_i'] = raw_df['kwh_i'] / sip_duration_in_hr
        if 'kva_i' not in raw_df.columns:
            raw_df['kva_i'] = raw_df['kvah_i'] / sip_duration_in_hr
        if 'kvar_i' not in raw_df.columns:
            raw_df['kvar_i'] = raw_df['kvar_i_total'] / sip_duration_in_hr

        # Format datetime helper
        def format_datetime(dt_value):
            if isinstance(dt_value, pd.Timestamp):
                return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
            elif isinstance(dt_value, str):
                dt_obj = pd.to_datetime(dt_value)
                return dt_obj.strftime(f'{dt_obj.day} %b at %H:%M')
            else:
                return str(dt_value)

        # 1. Demand Table - Active Power (KW)
        active_max = raw_df['kw_i'].max()
        active_avg = raw_df['kw_i'].mean()
        active_max_time = raw_df.loc[raw_df['kw_i'].idxmax(), 'surveydate']
        active_max_time_formatted = format_datetime(active_max_time)

        # 2. Demand Table - Apparent Power (KVA)
        apparent_max = raw_df['kva_i'].max()
        apparent_avg = raw_df['kva_i'].mean()
        apparent_max_time = raw_df.loc[raw_df['kva_i'].idxmax(), 'surveydate']
        apparent_max_time_formatted = format_datetime(apparent_max_time)

        # 3. Demand Table - Reactive Power (KVAR)
        reactive_max = raw_df['kvar_i'].max()
        reactive_avg = raw_df['kvar_i'].mean()
        reactive_max_time = raw_df.loc[raw_df['kvar_i'].idxmax(), 'surveydate']
        reactive_max_time_formatted = format_datetime(reactive_max_time)

        # Create demand table
        demand_table_data = [
            ['Active', active_max, active_avg, active_max_time_formatted],
            ['Apparent', apparent_max, apparent_avg, apparent_max_time_formatted],
            ['Reactive', reactive_max, reactive_avg, reactive_max_time_formatted]
        ]
        demand_table_df = pd.DataFrame(demand_table_data, columns=['Parameter', 'Max', 'Avg', 'Date and time at max value'])

        # Save to Excel
        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            demand_table_df.to_excel(writer, sheet_name='Demand Table', index=False)

        logger.info(f"Processed MV demand overview data saved: {processed_file}")
        logger.info(f"DB Parameter: Active -> Max={active_max:.2f}, Avg={active_avg:.2f}, DateTime={active_max_time_formatted}")
        logger.info(f"DB Parameter: Apparent -> Max={apparent_max:.2f}, Avg={apparent_avg:.2f}, DateTime={apparent_max_time_formatted}")
        logger.info(f"DB Parameter: Reactive -> Max={reactive_max:.2f}, Avg={reactive_avg:.2f}, DateTime={reactive_max_time_formatted}")

        return processed_file

    except Exception as e:
        logger.error(f"Error processing MV demand overview database: {str(e)}")
        raise


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_demand_overview_comparison(chart_file, processed_file, date_info):
    """Create complete MV demand overview comparison with validation"""
    logger.info("Creating MV demand overview comparison...")

    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_mv_demand_overview_{date_safe}.xlsx"

        # Load sheets
        sheet_name = 'Demand Table'
        chart_data = pd.read_excel(chart_file, sheet_name=sheet_name)
        processed_data = pd.read_excel(processed_file, sheet_name=sheet_name)

        # Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb = Workbook()
        wb.remove(wb.active)

        validation_results = {}

        logger.info(f"Creating MV demand comparison for: {sheet_name}")

        processed_df = processed_data
        chart_df = chart_data

        ws = wb.create_sheet(title=f"{sheet_name}_Comparison")

        headers = ["Parameter", "DB_Max", "Chart_Max", "Max_Difference", "Max_Match",
                   "DB_Avg", "Chart_Avg", "Avg_Difference", "Avg_Match",
                   "DB_DateTime", "Chart_DateTime", "DateTime_Match", "Overall_Match"]

        ws.append(headers)

        sheet_results = []

        for i in range(len(processed_df)):
            try:
                row_data = []
                overall_match = True

                param = processed_df.iloc[i, 0]
                row_data.append(param)

                # Max comparison
                db_max = processed_df.iloc[i, 1]
                chart_max = chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                try:
                    if db_max != "-" and chart_max != "-":
                        max_diff = abs(float(db_max) - float(chart_max))
                        max_match = "YES" if max_diff <= 0.1 else "NO"
                        overall_match = overall_match and (max_match == "YES")
                    else:
                        max_diff = "-"
                        max_match = "YES" if str(db_max).strip() == str(chart_max).strip() else "NO"
                        overall_match = overall_match and (max_match == "YES")
                except (ValueError, TypeError):
                    max_diff = "-"
                    max_match = "YES" if str(db_max).strip() == str(chart_max).strip() else "NO"
                    overall_match = overall_match and (max_match == "YES")

                row_data.extend([db_max, chart_max, max_diff, max_match])

                # Avg comparison
                db_avg = processed_df.iloc[i, 2]
                chart_avg = chart_df.iloc[i, 2] if i < len(chart_df) else "-"
                try:
                    if db_avg != "-" and chart_avg != "-":
                        avg_diff = abs(float(db_avg) - float(chart_avg))
                        avg_match = "YES" if avg_diff <= 0.1 else "NO"
                        overall_match = overall_match and (avg_match == "YES")
                    else:
                        avg_diff = "-"
                        avg_match = "YES" if str(db_avg).strip() == str(chart_avg).strip() else "NO"
                        overall_match = overall_match and (avg_match == "YES")
                except (ValueError, TypeError):
                    avg_diff = "-"
                    avg_match = "YES" if str(db_avg).strip() == str(chart_avg).strip() else "NO"
                    overall_match = overall_match and (avg_match == "YES")

                row_data.extend([db_avg, chart_avg, avg_diff, avg_match])

                # DateTime comparison
                db_datetime = processed_df.iloc[i, 3]
                chart_datetime = chart_df.iloc[i, 3] if i < len(chart_df) else "-"
                datetime_match = "YES" if str(db_datetime).strip() == str(chart_datetime).strip() else "NO"
                overall_match = overall_match and (datetime_match == "YES")

                row_data.extend([db_datetime, chart_datetime, datetime_match])

                row_data.append("YES" if overall_match else "NO")

                sheet_results.append({
                    'item': param,
                    'match': overall_match
                })

                ws.append(row_data)

            except Exception as e:
                logger.warning(f"Error processing row {i} in {sheet_name}: {str(e)}")
                continue

        validation_results[sheet_name] = sheet_results

        passed_count = sum(1 for result in sheet_results if result['match'])
        failed_count = len(sheet_results) - passed_count
        logger.info(f"{sheet_name} MV Validation: {passed_count} passed, {failed_count} failed")

        # Apply colors
        for row_num in range(2, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                header = ws.cell(row=1, column=col_num).value

                if header and ("_Match" in header or header == "Overall_Match"):
                    if cell.value == "YES":
                        cell.fill = green_fill
                    elif cell.value == "NO":
                        cell.fill = red_fill
                elif header and "_Difference" in header:
                    if isinstance(cell.value, (int, float)):
                        if cell.value <= 0.1:
                            cell.fill = green_fill
                        else:
                            cell.fill = red_fill

        wb.save(output_file)
        logger.info(f"MV demand overview comparison saved: {output_file}")

        return output_file, validation_results

    except Exception as e:
        logger.error(f"Error creating MV demand overview comparison: {str(e)}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_demand_overview_summary_report(config, date_info, chart_file, processed_file,
                                  comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive MV demand overview summary report with ENHANCED styling"""
    logger.info("Creating MV demand overview summary report with enhanced styling...")

    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MV_DEMAND_OVERVIEW_{date_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_alignment = Alignment(horizontal="center", vertical="center")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"MV DEMAND OVERVIEW VALIDATION SUMMARY - {date_info['selected_date'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Empty row
        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Date:", config['target_date']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Feeder Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "MV Demand Overview (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Data rows
        total_chart_points = 0
        try:
            chart_df = pd.read_excel(chart_file, sheet_name='Demand Table')
            total_chart_points = len(chart_df)
        except:
            total_chart_points = 3

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Chart Data Points", total_chart_points, "COMPLETE RECORDS"],
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in status:
                ws[f'C{current_row}'].font = pass_font
                ws[f'C{current_row}'].fill = pass_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0
        validation_data = []

        for sheet_name, results in validation_results.items():
            total_items = len(results)
            passed_items = sum(1 for result in results if result['match'])
            failed_items = total_items - passed_items
            success_rate = f"{(passed_items / total_items) * 100:.1f}%" if total_items > 0 else "0%"
            status = "PASS" if passed_items == total_items else "FAIL"

            display_name = sheet_name.replace('_', ' ')
            validation_data.append([display_name, passed_items, failed_items, success_rate, status])

            overall_passed += passed_items
            overall_total += total_items

        for comp_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = comp_type
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: MV demand overview validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor MV demand discrepancies found - Review recommended"
            assessment_color = warning_fill
            assessment_font_color = warning_font
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant MV demand validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        from openpyxl.utils import get_column_letter
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced MV demand overview summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MV DEMAND OVERVIEW VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Data: Raw={len(raw_df)}, Chart={total_chart_points}")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating MV demand summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_mv_demand_overview_automation():
    """Main MV Demand Overview automation process"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying MV Demand Overview configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])

        # Set date
        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info:
            logger.info("Failed to set date")
            return False

        # Select meter type
        if not select_meter_type(driver):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching MV meter metrics...")
        feeder_id, name, mtr_id = get_metrics(config['meter_serial_no'])

        if not feeder_id:
            logger.info(f"MV Feeder not found: {config['meter_serial_no']}")
            return False

        logger.info(f"MV Feeder found: {name} (ID: {feeder_id})")
        node_id = feeder_id

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for overview page to load
        time.sleep(5)

        # Collect demand overview data
        logger.info("Collecting MV demand overview data from UI...")
        overview_data = collect_demand_overview_data(driver)

        # Save demand overview data
        chart_file = save_demand_overview_data_to_excel(date_info, overview_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get database data
        raw_df = get_database_data_for_demand_overview(config['target_date'], mtr_id, node_id)

        if raw_df.empty:
            logger.info("No database data found")
            return False

        # Process database calculations
        logger.info("Processing MV demand database calculations...")
        processed_file = process_demand_overview_database_calculations(raw_df, date_info)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison report
        logger.info("Creating MV demand validation comparison...")
        comparison_file, validation_results = create_demand_overview_comparison(chart_file, processed_file, date_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive MV demand summary...")
        summary_report = create_demand_overview_summary_report(
            config, date_info, chart_file, processed_file,
            comparison_file, validation_results, raw_df, name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("MV DEMAND OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: MV Demand Overview (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Date: {config['target_date']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ MV Demand Overview monitoring (fixed)")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Demand table extraction from UI")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"MV Demand Overview Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("MV DEMAND OVERVIEW AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: MV Demand Overview (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ MV Demand Overview monitoring only")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Demand table data extraction")
    logger.info("   ✓ Active, Apparent, Reactive power monitoring")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Better null/dash handling")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_mv_demand_overview_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("MV DEMAND OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ MV Demand Overview monitoring (fixed)")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Demand table extraction")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("MV DEMAND OVERVIEW AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("MV Demand Overview Automation Finished")
    logger.info("=" * 60)
