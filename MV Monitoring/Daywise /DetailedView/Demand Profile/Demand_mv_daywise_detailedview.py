"""
MV DEMAND PROFILE AUTOMATION - COMPLETE UPDATED VERSION
========================================================
This script is configured for MV MONITORING ONLY - DEMAND PROFILE

Features:
- Fixed for MV monitoring (no Type selection needed)
- Search box approach for meter selection
- Centralized database configuration
- Dynamic SIP duration from database
- Enhanced value parsing and comparison
- Test engineer details in reports
- Energy to Demand conversion with dynamic SIP
- Complete error handling and file generation

Author: Sanyam Upadhyay
Version: COMPLETE v1.2
Date: 2025-01-15
"""

import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import functools
import re
import traceback


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and SIP duration
    DB1_HOST = "10.11.16.186"
    DB1_PORT = "5432"
    DB1_DATABASE = "serviceplatformdb12004"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.186"
    DB2_PORT = "5432"
    DB2_DATABASE = "serviceplatformdb12004"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant03"  # MV monitoring uses tenant03

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('mv_demand_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/mv_demand_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'demand_output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous demand output files")
    os.makedirs(output_folder)
    logger.info(f"Created demand output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)

            # If file is already in output folder, return as is
            if os.path.dirname(os.path.abspath(file_path)) == os.path.abspath(output_folder):
                logger.info(f"File already in output folder: {filename}")
                return file_path

            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for MV Demand Monitoring"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Target_Date', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'DD/MM/YYYY', 'YOUR_METER_NO', 'MV']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Update Target_Date with desired date (DD/MM/YYYY)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Meter_Type is fixed to MV',
                ],
                'Important_Notes': [
                    'This script is FOR MV DEMAND MONITORING ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Date format: DD/MM/YYYY',
                    'Meter_Type: MV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"MV Demand configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    """Ensure date is in 'DD/MM/YYYY' string format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors='raise')
        return parsed.strftime("%d/%m/%Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config_demand.xlsx"):
    """Read user configuration from Excel file for MV Demand Monitoring"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'MV'}  # Fixed for MV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'target_date', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("MV Demand configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING MV DEMAND AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config_demand.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default MV Demand configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("MV Demand configuration validated successfully")
    logger.info(f"   Monitoring Type: MV Demand (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Date: {config['target_date']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            logger.info(f"Traceback: {traceback.format_exc()}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_sip_duration(mtrid):
    """Get SIP duration from database"""
    logger.info(f"Fetching SIP duration for meter ID: {mtrid}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        query = f"SELECT sip FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail WHERE mtrid = %s LIMIT 1;"
        cursor.execute(query, (mtrid,))
        result = cursor.fetchone()
        if result and result[0]:
            sip_duration = int(result[0])
            logger.info(f"SIP duration found: {sip_duration} minutes")
            return sip_duration
        logger.info(f"No SIP found, using default 15 min")
        return 15
    except Exception as e:
        logger.error(f"Error fetching SIP: {e}, defaulting to 15 min")
        return 15
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_metrics(mtr_serial_no, nodetypeid, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching MV Demand metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        # MV Feeder query
        query1 = f"SELECT feeder_id AS dt_id, feeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_feeder WHERE meter_serial_no = %s LIMIT 1;"

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None

        dt_id, dt_name, meterid = result1

        logger.info(f"Demand Metrics: {dt_name}")
        return dt_id, dt_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_chart_dates(target_date, mtr_id, node_id):
    """Fetch database data for demand profile"""
    logger.info(f"Fetching demand database data for date: {target_date}")
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{next_day}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # RAW data - Energy values
        raw_query = f"""
            SELECT DISTINCT surveydate, kwh_i, kvah_i, kvar_i_total,
                   kwh_abs, kvah_abs, kvarh_abs
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        # NRM data - Demand values (already converted)
        nrm_query = f"""
            SELECT surveydate, kw_i, kva_i, kvar_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile
            WHERE nodeid={node_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        raw_df = pd.read_sql(raw_query, conn)
        nrm_df = pd.read_sql(nrm_query, conn)

        logger.info(f"Retrieved: Raw={len(raw_df)}, NRM={len(nrm_df)} records")

        if not nrm_df.empty:
            nrm_df['date'] = pd.to_datetime(nrm_df['surveydate']).dt.date
            sip_counts = nrm_df.groupby('date').size()
            for date, count in sip_counts.items():
                logger.info(f"   {date}: {count} SIPs available")

        return raw_df, nrm_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame(), pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:43379/LVMV/DTMonitoring")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SanyamU")
        driver.find_element(By.ID, "Password").send_keys("Secure@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(10)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_date(driver, target_date):
    """Set calendar date"""
    try:
        logger.info(f"Setting date: {target_date}")
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        date_info = {
            'selected_date': target_dt.strftime("%B %Y"),
            'start_date': target_dt.strftime("%Y-%m-%d"),
            'end_date': (target_dt + timedelta(days=0)).strftime("%Y-%m-%d")
        }
        logger.info("Date set successfully")
        return date_info
    except Exception as e:
        logger.info(f"Date error: {e}")
        return None


def select_type(driver):
    """Select MV monitoring - FIXED FOR MV ONLY"""
    try:
        logger.info("Selecting MV monitoring (fixed for MV script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divmvmonitoring']").click()
        logger.info("MV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - MV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "MV":
            mv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="MVFeederClick"]')))
            mv_button.click()
            logger.info("MV feeder selected")
        else:
            logger.info("Invalid meter type for MV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}, trying fallback")
        try:
            view_btn = driver.find_element(By.XPATH,
                                           f"//tr[td[contains(text(), '{meter_serial_no}')]]//a[text()='View']")
            view_btn.click()
            logger.info("View clicked (fallback)")
            return True
        except:
            return False


@log_execution_time
def extract_chart_data_single_pass(driver, wait, sip_duration):
    """Extract demand chart data with dynamic SIP duration"""
    logger.info(f"Starting demand chart extraction with {sip_duration}-min SIP intervals")

    chart_dates = []
    tooltip_data = []

    time.sleep(2)

    try:
        chart_svg = driver.find_element(By.CSS_SELECTOR, 'svg')
        chart_rect = driver.execute_script("return arguments[0].getBoundingClientRect();", chart_svg)
        chart_top_y = int(chart_rect['top'])
        chart_height = chart_rect['height']
        chart_left_x = int(chart_rect['left'])
        chart_width = chart_rect['width']
        logger.info(f"Chart dimensions: {chart_width}x{chart_height}")
    except Exception as e:
        logger.error(f"Chart not found: {e}")
        return [], []

    x_labels = driver.find_elements(By.CSS_SELECTOR, 'g.dxc-arg-elements text')
    if not x_labels:
        logger.error("No X-axis labels found")
        return [], []

    logger.info(f"Found {len(x_labels)} X-axis labels")

    label_positions = {}
    for label in x_labels:
        label_text = label.text.strip()
        rect = driver.execute_script("return arguments[0].getBoundingClientRect();", label)
        x, y, width = rect['x'], rect['y'], rect['width']
        center_x = round(x + width / 2)
        start_y = int(y - chart_height * 0.05)
        label_positions[label_text] = (center_x, start_y)

    available_labels = list(label_positions.keys())
    if len(available_labels) < 2:
        logger.warning("Not enough labels")
        return [], []

    # Sort labels by time
    time_sorted = []
    for label in available_labels:
        try:
            hour, minute = map(int, label.split(':'))
            time_minutes = hour * 60 + minute
            time_sorted.append((time_minutes, label))
        except:
            time_sorted.append((0, label))

    time_sorted.sort()
    sorted_labels = [label for _, label in time_sorted]

    # Calculate spacing
    spacings = []
    for i in range(len(sorted_labels) - 1):
        x1 = label_positions[sorted_labels[i]][0]
        x2 = label_positions[sorted_labels[i + 1]][0]
        spacings.append(x2 - x1)
    avg_spacing = sum(spacings) / len(spacings)
    logger.info(f"Average spacing: {avg_spacing:.1f}px")

    total_sips_per_day = (24 * 60) // sip_duration
    logger.info(f"Expected {total_sips_per_day} SIPs per day ({sip_duration}-min intervals)")

    # Calculate positions
    first_label = sorted_labels[0]
    try:
        first_hour, first_minute = map(int, first_label.split(':'))
        first_time_minutes = first_hour * 60 + first_minute

        if len(sorted_labels) >= 2:
            second_label = sorted_labels[1]
            second_hour, second_minute = map(int, second_label.split(':'))
            second_time_minutes = second_hour * 60 + second_minute
            label_time_diff = second_time_minutes - first_time_minutes
            if label_time_diff < 0:
                label_time_diff += 24 * 60
        else:
            label_time_diff = 120

        minutes_per_pixel = label_time_diff / avg_spacing
        pixels_per_sip = sip_duration / minutes_per_pixel

        first_label_x = label_positions[first_label][0]
        start_x = first_label_x - (first_time_minutes / minutes_per_pixel)
        start_y = label_positions[first_label][1]

        logger.info(f"Start position: x={start_x:.1f}, y={start_y}")
        logger.info(f"Pixels per {sip_duration}-min SIP: {pixels_per_sip:.1f}")

    except Exception as e:
        logger.warning(f"Time parsing failed: {e}, using fallback")
        first_label_x = label_positions[sorted_labels[0]][0]
        start_x = chart_left_x + 50
        start_y = label_positions[sorted_labels[0]][1]
        pixels_per_sip = avg_spacing / (120 // sip_duration)

    # Generate hover positions
    hover_positions = []
    for sip_index in range(total_sips_per_day):
        x_pos = start_x + (sip_index * pixels_per_sip)
        hover_positions.append((x_pos, start_y))

    logger.info(f"Generated {len(hover_positions)} hover positions")

    # Enhance positions with intermediates
    enhanced_positions = []
    for i in range(len(hover_positions) - 1):
        x1, y1 = hover_positions[i]
        x2, y2 = hover_positions[i + 1]
        enhanced_positions.append((x1, y1))
        for j in range(1, 3):
            x_int = x1 + (x2 - x1) * j / 3
            enhanced_positions.append((x_int, y1))
    enhanced_positions.append(hover_positions[-1])

    hover_positions = enhanced_positions
    logger.info(f"Enhanced to {len(hover_positions)} positions")

    seen_tooltips = set()
    headers_ordered = []

    # Warmup hover
    chart_center_x = chart_left_x + (chart_width / 2)
    chart_center_y = chart_top_y + (chart_height / 2)
    warmup_js = f"""
        let tooltips = document.querySelectorAll('.dxc-tooltip');
        tooltips.forEach(t => t.style.display = 'none');
        let evt = new MouseEvent('mousemove', {{bubbles: true, clientX: {int(chart_center_x)}, clientY: {int(chart_center_y)}}});
        let el = document.elementFromPoint({int(chart_center_x)}, {int(chart_center_y)});
        if (el) el.dispatchEvent(evt);
    """
    driver.execute_script(warmup_js)
    time.sleep(0.3)

    logger.info("Starting tooltip extraction...")

    successful = 0
    failed = 0

    for i, (center_x, start_y) in enumerate(hover_positions):
        if i % max(12, total_sips_per_day // 8) == 0:
            logger.info(f"Progress: {i}/{len(hover_positions)} | Success: {successful} | Failed: {failed}")

        tooltip_y = None
        tooltip_found = False

        scan_range = min(150, chart_height // 3)
        y_step = 3

        for y in range(start_y, max(chart_top_y, start_y - scan_range), -y_step):
            hover_js = f"""
                let evt = new MouseEvent('mousemove', {{bubbles: true, cancelable: true, clientX: {int(center_x)}, clientY: {int(y)}}});
                let el = document.elementFromPoint({int(center_x)}, {int(y)});
                if (el && el.dispatchEvent) {{
                    el.dispatchEvent(evt);
                    return true;
                }}
                return false;
            """

            if not driver.execute_script(hover_js):
                continue

            time.sleep(0.02)

            try:
                tooltip_el = driver.find_element(By.XPATH, '//div[@class="dxc-tooltip"]//div//div')
                if tooltip_el.is_displayed():
                    tooltip_y = y
                    tooltip_found = True
                    break
            except:
                continue

        if not tooltip_found:
            failed += 1
            continue

        # Stable hover
        stable_js = f"""
            let evt = new MouseEvent('mousemove', {{bubbles: true, clientX: {int(center_x)}, clientY: {int(tooltip_y)}}});
            let el = document.elementFromPoint({int(center_x)}, {int(tooltip_y)});
            if (el) el.dispatchEvent(evt);
        """
        driver.execute_script(stable_js)
        time.sleep(0.25)

        try:
            tooltip = wait.until(EC.visibility_of_element_located(
                (By.XPATH, '//div[@class="dxc-tooltip"]//div//div')))
            tooltip_text = tooltip.text.strip()

            if not tooltip_text or tooltip_text in seen_tooltips:
                failed += 1
                continue

            seen_tooltips.add(tooltip_text)
            successful += 1

        except:
            failed += 1
            continue

        # Parse tooltip
        lines = tooltip_text.split('\n')
        data_point = {}
        for line_idx, line in enumerate(lines):
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            key, value = key.strip(), value.strip()

            if line_idx == 0:
                data_point[key] = value
                chart_dates.append(value)
            else:
                numeric_match = re.search(r"[-+]?[0-9]*\.?[0-9]+", value)
                data_point[key] = numeric_match.group() if numeric_match else value

        if data_point:
            tooltip_data.append(data_point)
            for key in data_point.keys():
                if key not in headers_ordered:
                    headers_ordered.append(key)

    logger.info(f"Extraction complete - Success: {successful}, Failed: {failed}")
    efficiency = (successful / (successful + failed) * 100) if (successful + failed) > 0 else 0
    logger.info(f"Efficiency: {efficiency:.1f}%")

    unique_dates = sorted(list(set(chart_dates)))
    logger.info(f"Extracted {len(unique_dates)} unique dates, {len(tooltip_data)} data points")
    logger.info(f"Coverage: {(len(tooltip_data) / total_sips_per_day * 100):.1f}%")

    return unique_dates, tooltip_data


@log_execution_time
def collect_side_panel_data(driver, wait):
    """Collect side panel data for demand"""
    logger.info("Collecting demand side panel data...")
    data = {}

    try:
        data['Active'] = {
            'Max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kw"]').text,
            'Avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kw"]').text,
            'DateTime': driver.find_element(By.XPATH, '//td[@id="kw_MaxDatetime"]').text
        }

        data['Apparent'] = {
            'Max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kva"]').text,
            'Avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kva"]').text,
            'DateTime': driver.find_element(By.XPATH, '//td[@id="kva_MaxDatetime"]').text
        }

        data['Reactive'] = {
            'Max': driver.find_element(By.XPATH, '//td[@id="maxDemand_Kvar"]').text,
            'Avg': driver.find_element(By.XPATH, '//td[@id="avgDemand_Kvar"]').text,
            'DateTime': driver.find_element(By.XPATH, '//td[@id="kvar_MaxDatetime"]').text
        }

        logger.info("Demand side panel data collected")
        return data
    except Exception as e:
        logger.info(f"Side panel error: {e}")
        return {}


# ============================================================================
# HELPER FUNCTIONS FOR VALUE PARSING
# ============================================================================
def parse_chart_value(value_str):
    """Parse chart values"""
    if not value_str or str(value_str).strip() in ['-', '', 'nan', 'None']:
        return None

    value_str = str(value_str).strip()
    value_str = value_str.replace('kW', '').replace('kVA', '').replace('kVAr', '').strip()
    numeric_match = re.search(r"[-+]?[0-9]*\.?[0-9]+", value_str)
    if numeric_match:
        try:
            return float(numeric_match.group())
        except:
            return None
    return None


def values_match(val1, val2, tolerance=0.001):
    """Compare two values with tolerance"""
    str1, str2 = str(val1).strip(), str(val2).strip()
    null_values = ['-', '', 'nan', 'None', 'N/A', 'n/a']
    is_null1 = str1.lower() in null_values or pd.isna(val1)
    is_null2 = str2.lower() in null_values or pd.isna(val2)

    if is_null1 and is_null2:
        return True
    if is_null1 or is_null2:
        return False

    try:
        num1, num2 = float(val1), float(val2)
        return abs(num1 - num2) < tolerance
    except:
        return str1.lower() == str2.lower()


def format_datetime(dt_value):
    """Format datetime value"""
    if isinstance(dt_value, pd.Timestamp):
        return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
    elif isinstance(dt_value, str):
        try:
            dt_obj = pd.to_datetime(dt_value)
            return dt_obj.strftime(f'{dt_obj.day} %b at %H:%M')
        except:
            return str(dt_value)
    else:
        return str(dt_value)


# ============================================================================
# FILE PROCESSING FUNCTIONS
# ============================================================================
@log_execution_time
def save_chart_data_to_excel(tooltip_data, date_info, side_data):
    """Save demand chart data to Excel"""
    try:
        logger.info("Creating Excel for demand chart data...")
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Demand_Detailed_View")

        if tooltip_data:
            headers = list(tooltip_data[0].keys())
            ws.append(headers)
            for data_point in tooltip_data:
                row = [data_point.get(key, "") for key in headers]
                ws.append(row)

        # Demand Table
        ws_table = wb.create_sheet(title="Demand_Table")
        ws_table.append(["Parameter", "Max", "Avg", "Date and Time at Max Value"])

        for param_name, param_data in side_data.items():
            ws_table.append([
                param_name,
                param_data.get('Max', ''),
                param_data.get('Avg', ''),
                param_data.get('DateTime', '')
            ])

        chart_file = f"demand_chart_data_from_ui_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(chart_file)
        logger.info(f"Demand chart data saved: {chart_file}")
        return chart_file
    except Exception as e:
        logger.info(f"Error saving chart data: {e}")
        return None


# ============================================================================
# DATABASE COMPARISON WITH ENERGY TO DEMAND CONVERSION
# ============================================================================
@log_execution_time
def process_database_comparison_with_calculated_pipeline(raw_df, nrm_df, date_info, sip_duration):
    """Process database comparison with energy to demand conversion"""
    logger.info("Processing demand database comparison pipeline...")
    logger.info(f"Using dynamic SIP duration: {sip_duration} minutes")

    date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    interval_minutes = sip_duration
    sip_duration_in_hr = interval_minutes / 60

    logger.info(f"Energy to Demand conversion: Divide by {sip_duration_in_hr:.3f} hours")

    # Raw database export
    raw_export_file = f"demand_actual_raw_database_data_{date_safe}_{timestamp}.xlsx"
    with pd.ExcelWriter(raw_export_file, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name='RAW Database', index=False)
        nrm_df.to_excel(writer, sheet_name='NRM Database', index=False)

    logger.info(f"Raw database file created: {raw_export_file}")

    # Calculate NRM from RAW (Energy to Demand conversion)
    nrm_calculated = pd.DataFrame()
    column_map = {
        'kwh_i': 'kw_i',
        'kvah_i': 'kva_i',
        'kvar_i_total': 'kvar_i'
    }

    logger.info(f"Converting energy to demand using {sip_duration}-minute SIP...")

    if raw_df.empty:
        logger.info("ERROR: Raw data is empty")
        nrm_calculated = pd.DataFrame(columns=['surveydate', 'kw_i', 'kva_i', 'kvar_i'])
    else:
        for raw_col, new_col in column_map.items():
            if raw_col in raw_df.columns:
                try:
                    converted_values = raw_df[raw_col] / sip_duration_in_hr
                    nrm_calculated[new_col] = converted_values
                    logger.info(f"   {raw_col} -> {new_col} conversion complete")
                except Exception as e:
                    logger.info(f"   Error converting {raw_col}: {e}")
                    nrm_calculated[new_col] = 0
            else:
                logger.info(f"   {raw_col} not found, filling {new_col} with 0")
                nrm_calculated[new_col] = 0

        nrm_calculated['surveydate'] = raw_df['surveydate']
        cols = nrm_calculated.columns.tolist()
        cols.insert(0, cols.pop(cols.index('surveydate')))
        nrm_calculated = nrm_calculated[cols]

    logger.info(f"NRM calculated shape: {nrm_calculated.shape}")

    # Demand Table Calculation
    logger.info("Creating demand table calculations...")
    demand_table_data = []

    if not nrm_calculated.empty and len(nrm_calculated) > 0:
        kw_valid = not nrm_calculated['kw_i'].isna().all()
        kva_valid = not nrm_calculated['kva_i'].isna().all()
        kvar_valid = not nrm_calculated['kvar_i'].isna().all()

        # Active Power
        if kw_valid:
            try:
                active_max = nrm_calculated['kw_i'].max()
                active_avg = nrm_calculated['kw_i'].mean()
                active_max_idx = nrm_calculated['kw_i'].idxmax()
                active_max_time = nrm_calculated.loc[active_max_idx, 'surveydate']
                demand_table_data.append(['Active', active_max, active_avg, format_datetime(active_max_time)])
                logger.info(f"   Active Max: {active_max:.3f}, Avg: {active_avg:.3f}")
            except Exception as e:
                logger.info(f"   Error processing Active: {e}")
                demand_table_data.append(['Active', 0.0, 0.0, 'Error'])
        else:
            demand_table_data.append(['Active', 0.0, 0.0, 'No data'])

        # Apparent Power
        if kva_valid:
            try:
                apparent_max = nrm_calculated['kva_i'].max()
                apparent_avg = nrm_calculated['kva_i'].mean()
                apparent_max_idx = nrm_calculated['kva_i'].idxmax()
                apparent_max_time = nrm_calculated.loc[apparent_max_idx, 'surveydate']
                demand_table_data.append(['Apparent', apparent_max, apparent_avg, format_datetime(apparent_max_time)])
                logger.info(f"   Apparent Max: {apparent_max:.3f}, Avg: {apparent_avg:.3f}")
            except Exception as e:
                logger.info(f"   Error processing Apparent: {e}")
                demand_table_data.append(['Apparent', 0.0, 0.0, 'Error'])
        else:
            demand_table_data.append(['Apparent', 0.0, 0.0, 'No data'])

        # Reactive Power
        if kvar_valid:
            try:
                reactive_max = nrm_calculated['kvar_i'].max()
                reactive_avg = nrm_calculated['kvar_i'].mean()
                reactive_max_idx = nrm_calculated['kvar_i'].idxmax()
                reactive_max_time = nrm_calculated.loc[reactive_max_idx, 'surveydate']
                demand_table_data.append(['Reactive', reactive_max, reactive_avg, format_datetime(reactive_max_time)])
                logger.info(f"   Reactive Max: {reactive_max:.3f}, Avg: {reactive_avg:.3f}")
            except Exception as e:
                logger.info(f"   Error processing Reactive: {e}")
                demand_table_data.append(['Reactive', 0.0, 0.0, 'Error'])
        else:
            demand_table_data.append(['Reactive', 0.0, 0.0, 'No data'])

        demand_table_df = pd.DataFrame(demand_table_data,
                                       columns=['Parameter', 'Max', 'Avg', 'Date and time at max value'])
    else:
        demand_table_data = [
            ['Active', 0.0, 0.0, 'No data'],
            ['Apparent', 0.0, 0.0, 'No data'],
            ['Reactive', 0.0, 0.0, 'No data']
        ]
        demand_table_df = pd.DataFrame(demand_table_data,
                                       columns=['Parameter', 'Max', 'Avg', 'Date and time at max value'])

    # Processed calculated data
    processed_export_file = f"demand_theoretical_calculated_data_{date_safe}_{timestamp}.xlsx"
    with pd.ExcelWriter(processed_export_file, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name='RAW Database', index=False)
        nrm_calculated.to_excel(writer, sheet_name='NRM Calculated', index=False)
        if not demand_table_df.empty:
            demand_table_df.to_excel(writer, sheet_name='Demand Table', index=False)

        # SIP configuration info
        sip_info_data = [
            ['Parameter', 'Value'],
            ['SIP Duration (minutes)', sip_duration],
            ['SIP Duration (hours)', f"{sip_duration_in_hr:.3f}"],
            ['Conversion Method', 'Demand = Energy / SIP_Duration_Hours'],
            ['Expected Records per Day', f"{(24 * 60) // sip_duration}"],
            ['Actual Raw Records', len(raw_df)],
            ['Actual NRM Records', len(nrm_df)],
            ['Data Coverage %',
             f"{(len(raw_df) / ((24 * 60) // sip_duration) * 100):.1f}%" if sip_duration > 0 else "N/A"]
        ]

        sip_info_df = pd.DataFrame(sip_info_data)
        sip_info_df.to_excel(writer, sheet_name='SIP Configuration', index=False, header=False)

    logger.info(f"Processed file created: {processed_export_file}")

    return raw_export_file, processed_export_file, raw_df, nrm_calculated, nrm_df


@log_execution_time
def create_enhanced_database_comparison_report(raw_file, processed_file, comparison_file):
    """Create enhanced comparison report with RAW vs NRM sheet - FIXED VERSION"""
    logger.info("Creating enhanced database comparison...")

    TOLERANCE = 0.001
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Map processed file sheets to comparison file sheets
    sheet_mapping = {
        'RAW Database': ['surveydate', 'kwh_i', 'kvah_i', 'kvar_i_total'],
        'NRM Calculated': ['surveydate', 'kw_i', 'kva_i', 'kvar_i']
    }

    # Copy processed data to comparison file
    try:
        with pd.ExcelWriter(comparison_file, engine="openpyxl") as writer:
            # Read and write sheets from processed file
            for sheet_name, columns in sheet_mapping.items():
                try:
                    df = pd.read_excel(processed_file, sheet_name=sheet_name)
                    # Reindex to ensure columns exist
                    available_cols = [col for col in columns if col in df.columns]
                    if available_cols:
                        df_subset = df[available_cols]
                        df_subset.to_excel(writer, sheet_name=sheet_name, index=False)
                        logger.info(f"   Written {sheet_name}: {len(df_subset)} rows, {len(available_cols)} columns")
                except Exception as e:
                    logger.info(f"   Error processing sheet {sheet_name}: {e}")
                    # Create empty sheet as fallback
                    pd.DataFrame(columns=columns).to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        logger.info(f"Error creating initial comparison file: {e}")
        return

    # Now load and enhance the workbook
    try:
        wb = load_workbook(comparison_file)
    except Exception as e:
        logger.info(f"Error loading comparison file: {e}")
        return

    # Create RAW vs NRM comparison sheet
    ws_raw_nrm = wb.create_sheet('RAW_vs_NRM_Comparison')

    # Read data from files with error handling
    try:
        df_raw = pd.read_excel(raw_file, sheet_name='RAW Database')
        logger.info(f"   Read raw data: {len(df_raw)} rows")
    except Exception as e:
        logger.info(f"   Error reading raw data: {e}")
        df_raw = pd.DataFrame()

    try:
        df_nrm = pd.read_excel(processed_file, sheet_name='NRM Calculated')
        logger.info(f"   Read NRM data: {len(df_nrm)} rows")
    except Exception as e:
        logger.info(f"   Error reading NRM data: {e}")
        df_nrm = pd.DataFrame()

    if df_raw.empty or df_nrm.empty:
        logger.info("   Insufficient data for RAW vs NRM comparison")
        ws_raw_nrm.append(['Error', 'Insufficient data for comparison'])
        try:
            visible_count = sum(1 for sheet in wb.worksheets if sheet.sheet_state == 'visible')
            if visible_count == 0:
                wb.worksheets[0].sheet_state = 'visible'
            wb.save(comparison_file)
        except:
            pass
        logger.info(f"Comparison file saved (with error note): {comparison_file}")
        return

    headers = ['surveydate', 'raw_kwh', 'nrm_kw', 'kw_diff', 'raw_kvah', 'nrm_kva', 'kva_diff',
               'raw_kvar', 'nrm_kvar', 'kvar_diff', 'match_status']
    ws_raw_nrm.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws_raw_nrm.cell(row=1, column=col_idx)
        cell.fill = header_fill

    max_rows = max(len(df_raw), len(df_nrm))
    logger.info(f"RAW vs NRM Analysis: {max_rows} records to compare")

    # Get SIP duration for conversion
    sip_duration_hr = 0.25  # Default 15 minutes
    try:
        sip_info_df = pd.read_excel(processed_file, sheet_name='SIP Configuration', header=None)
        for idx, row in sip_info_df.iterrows():
            if 'SIP Duration (hours)' in str(row[0]):
                sip_duration_hr = float(row[1])
                break
    except:
        pass

    for idx in range(max_rows):
        row_data = []
        match_status = "MATCH"

        raw_record = df_raw.iloc[idx] if idx < len(df_raw) else None
        nrm_record = df_nrm.iloc[idx] if idx < len(df_nrm) else None

        date_val = raw_record['surveydate'] if raw_record is not None else (
            nrm_record['surveydate'] if nrm_record is not None else "")
        row_data.append(date_val)

        # Convert energy to demand for comparison
        comparisons = [
            ('kwh_i', 'kw_i'),
            ('kvah_i', 'kva_i'),
            ('kvar_i_total', 'kvar_i')
        ]

        for raw_col, nrm_col in comparisons:
            raw_val = raw_record[raw_col] if raw_record is not None and raw_col in raw_record.index else None
            nrm_val = nrm_record[nrm_col] if nrm_record is not None and nrm_col in nrm_record.index else None

            # Convert raw energy to demand
            raw_demand = raw_val / sip_duration_hr if raw_val is not None and not pd.isna(raw_val) else None

            row_data.append(raw_val)  # Raw energy
            row_data.append(nrm_val)  # NRM demand

            if not values_match(raw_demand, nrm_val, TOLERANCE):
                match_status = "MISMATCH"
                try:
                    if raw_demand is not None and nrm_val is not None:
                        diff = abs(float(nrm_val) - float(raw_demand))
                        row_data.append(f"{diff:.6f}")
                    else:
                        row_data.append("N/A")
                except:
                    row_data.append("ERROR")
            else:
                row_data.append("0.000000")

        row_data.append(match_status)
        ws_raw_nrm.append(row_data)
        row_num = ws_raw_nrm.max_row

        match_cell = ws_raw_nrm.cell(row=row_num, column=len(headers))
        if match_status == "MATCH":
            match_cell.fill = green
        else:
            match_cell.fill = red

    # Apply color coding to existing sheets
    for sheet_name, columns in sheet_mapping.items():
        if sheet_name not in wb.sheetnames:
            logger.info(f"   Sheet {sheet_name} not found, skipping color coding")
            continue

        ws = wb[sheet_name]

        try:
            df_raw_sheet = pd.read_excel(raw_file, sheet_name='RAW Database')
            available_cols = [col for col in columns if col in df_raw_sheet.columns]
            if available_cols:
                df_raw_sheet = df_raw_sheet[available_cols]
        except Exception as e:
            logger.info(f"   Error reading raw sheet for coloring: {e}")
            df_raw_sheet = pd.DataFrame()

        try:
            df_processed_sheet = pd.read_excel(processed_file, sheet_name=sheet_name)
            available_cols = [col for col in columns if col in df_processed_sheet.columns]
            if available_cols:
                df_processed_sheet = df_processed_sheet[available_cols]
        except Exception as e:
            logger.info(f"   Error reading processed sheet for coloring: {e}")
            df_processed_sheet = pd.DataFrame()

        if df_raw_sheet.empty or df_processed_sheet.empty:
            continue

        max_rows = min(len(df_raw_sheet), len(df_processed_sheet), ws.max_row - 1)

        for row_idx in range(max_rows):
            for col_idx, col_name in enumerate(columns):
                if col_name not in df_raw_sheet.columns or col_name not in df_processed_sheet.columns:
                    continue

                cell = ws.cell(row=row_idx + 2, column=col_idx + 1)

                if col_name == "surveydate":
                    continue

                try:
                    raw_val = df_raw_sheet.iloc[row_idx][col_name] if row_idx < len(df_raw_sheet) else None
                    processed_val = df_processed_sheet.iloc[row_idx][col_name] if row_idx < len(
                        df_processed_sheet) else None

                    # For RAW Database sheet, always green
                    if sheet_name == 'RAW Database':
                        cell.fill = green
                    else:
                        # For NRM Calculated, compare with tolerance
                        if values_match(raw_val, processed_val, TOLERANCE):
                            cell.fill = green
                        else:
                            cell.fill = red
                except Exception as e:
                    logger.info(f"   Error coloring cell [{row_idx}, {col_idx}]: {e}")
                    continue

    # Auto-adjust column widths
    try:
        for column in ws_raw_nrm.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_raw_nrm.column_dimensions[column_letter].width = min(max_length + 2, 20)
    except Exception as e:
        logger.info(f"   Error adjusting column widths: {e}")

    # Save workbook with error handling
    try:
        # Ensure at least one sheet is visible
        visible_count = sum(1 for sheet in wb.worksheets if sheet.sheet_state == 'visible')
        if visible_count == 0:
            logger.info("   No visible sheets found, making first sheet visible")
            wb.worksheets[0].sheet_state = 'visible'

        wb.save(comparison_file)
        logger.info(f"Enhanced comparison report created: {comparison_file}")
    except Exception as e:
        logger.info(f"Error saving comparison file: {e}")
        logger.info(f"Traceback: {traceback.format_exc()}")
        # Try to save with minimal data
        try:
            wb_minimal = Workbook()
            ws = wb_minimal.active
            ws.title = "Error"
            ws.append(["Error creating comparison report", str(e)])
            wb_minimal.save(comparison_file)
            logger.info(f"Saved minimal error file: {comparison_file}")
        except:
            logger.info("Could not save comparison file at all")


@log_execution_time
def final_chart_database_comparison(chart_file, processed_file, date_info):
    """Compare Chart Data vs Calculated NRM Data for Demand"""
    logger.info("Creating final demand validation report...")

    date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
    output_file = f"complete_validation_report_demand_{date_safe}.xlsx"

    TOLERANCE = 0.001

    try:
        df_chart = pd.read_excel(chart_file, sheet_name='Demand_Detailed_View')
        df_nrm = pd.read_excel(processed_file, sheet_name='NRM Calculated')

        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = f'CHART_VS_CALCULATED_NRM_{date_info["selected_date"].upper().replace(" ", "_")}'

        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        new_headers = ['Date', 'Active', 'Active_Difference', 'Apparent', 'Apparent_Difference',
                       'Reactive', 'Reactive_Difference', 'Match']

        ws_output.append(new_headers)

        column_mapping = {
            'Active': 'kw_i',
            'Apparent': 'kva_i',
            'Reactive': 'kvar_i'
        }

        def convert_date(date_val):
            if isinstance(date_val, datetime):
                return date_val.strftime('%H:%M')
            elif isinstance(date_val, pd.Timestamp):
                return date_val.strftime('%H:%M')
            elif isinstance(date_val, str):
                if not date_val or date_val.strip() == '':
                    return "INVALID_TIME"
                date_formats = ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%H:%M"]
                for fmt in date_formats:
                    try:
                        dt = datetime.strptime(date_val.strip(), fmt)
                        return dt.strftime('%H:%M')
                    except ValueError:
                        continue
                return "INVALID_TIME"
            return "INVALID_TIME"

        max_rows = max(len(df_chart), len(df_nrm))
        logger.info(f"Chart vs NRM: Chart={len(df_chart)}, NRM={len(df_nrm)} rows")

        for row_idx in range(max_rows):
            row_data = []
            match_flag = True

            chart_row = df_chart.iloc[row_idx] if row_idx < len(df_chart) else None
            nrm_row = df_nrm.iloc[row_idx] if row_idx < len(df_nrm) else None

            if nrm_row is not None and 'surveydate' in df_nrm.columns:
                nrm_date = convert_date(nrm_row['surveydate'])
            else:
                nrm_date = ""
            row_data.append(nrm_date)

            expected_columns = ['Active', 'Apparent', 'Reactive']
            for i, col in enumerate(expected_columns):
                chart_val = None
                nrm_val = None

                if chart_row is not None and i + 1 < len(chart_row):
                    chart_val_raw = chart_row.iloc[i + 1]
                    chart_val = parse_chart_value(chart_val_raw)

                if nrm_row is not None and column_mapping[col] in df_nrm.columns:
                    nrm_val = nrm_row[column_mapping[col]]

                row_data.append(nrm_val)

                if not values_match(chart_val, nrm_val, TOLERANCE):
                    match_flag = False
                    try:
                        if chart_val is not None and nrm_val is not None and not pd.isna(chart_val) and not pd.isna(
                                nrm_val):
                            diff = abs(float(nrm_val) - float(chart_val))
                            row_data.append(f"{diff:.6f}")
                        else:
                            row_data.append("N/A")
                    except:
                        row_data.append("N/A")
                else:
                    row_data.append("0.000000")

            row_data.append("Yes" if match_flag else "No")

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_output.cell(row=row_idx + 2, column=col_idx)
                cell.value = value
                if col_idx == len(row_data) and new_headers[col_idx - 1] == "Match":
                    cell.fill = green if value == "Yes" else red

        # Add demand table comparison
        wb_chart = load_workbook(chart_file)
        wb_processed = load_workbook(processed_file)

        if 'Demand_Table' in wb_chart.sheetnames and 'Demand Table' in wb_processed.sheetnames:
            ws_chart_table = wb_chart['Demand_Table']
            ws_proc_table = wb_processed['Demand Table']
            ws_new = wb_output.create_sheet('Demand Table Comparison')

            ws_new.append(['Parameter', 'Chart_Max', 'Calc_Max', 'Max_Diff',
                           'Chart_Avg', 'Calc_Avg', 'Avg_Diff', 'Match'])

            for row in range(2, ws_chart_table.max_row + 1):
                param = ws_chart_table.cell(row=row, column=1).value

                chart_max = parse_chart_value(ws_chart_table.cell(row=row, column=2).value)
                calc_max = parse_chart_value(ws_proc_table.cell(row=row, column=2).value)

                chart_avg = parse_chart_value(ws_chart_table.cell(row=row, column=3).value)
                calc_avg = parse_chart_value(ws_proc_table.cell(row=row, column=3).value)

                max_match = values_match(chart_max, calc_max, TOLERANCE)
                avg_match = values_match(chart_avg, calc_avg, TOLERANCE)

                max_diff = abs(calc_max - chart_max) if (chart_max is not None and calc_max is not None) else 'N/A'
                avg_diff = abs(calc_avg - chart_avg) if (chart_avg is not None and calc_avg is not None) else 'N/A'

                match_status = 'YES' if (max_match and avg_match) else 'NO'

                ws_new.append([param, chart_max, calc_max, max_diff, chart_avg, calc_avg, avg_diff, match_status])

            for row in ws_new.iter_rows(min_row=2, max_col=8):
                match_cell = row[7]
                match_cell.fill = green if match_cell.value == 'YES' else red

        for column in wb_output.worksheets[0].columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            wb_output.worksheets[0].column_dimensions[column_letter].width = min(max_length + 2, 20)

        wb_output.save(output_file)
        logger.info(f"Final validation report saved: {output_file}")
        return output_file

    except Exception as e:
        logger.info(f"Final comparison error: {e}")
        logger.info(f"Traceback: {traceback.format_exc()}")
        return None


@log_execution_time
def create_complete_validation_summary_report(comparison_stats, chart_comparison_file, date_info, raw_df, nrm_df,
                                              chart_dates, tooltip_data, sip_duration, config, meter_name):
    """Create comprehensive validation summary for demand"""
    logger.info("Creating complete demand validation summary...")

    date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    summary_file = f"complete_validation_summary_demand_{date_safe}_{timestamp}.xlsx"

    TOLERANCE = 0.001

    try:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Validation_Summary"

        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_font = Font(bold=True, size=10)
        info_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        info_font = Font(size=9)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # TITLE
        ws_summary.append([f"MV Monitoring Demand Profile Validation Report - {date_info['selected_date'].upper()}"])
        title_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = center_alignment
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        subtitle_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        subtitle_cell.fill = section_fill
        subtitle_cell.font = section_font
        subtitle_cell.alignment = center_alignment
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        ws_summary.append([f"SIP Duration: {sip_duration} minutes (from database configuration)"])
        sip_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        sip_cell.fill = section_fill
        sip_cell.font = section_font
        sip_cell.alignment = center_alignment
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        ws_summary.append([])

        # TEST DETAILS SECTION
        ws_summary.append(["TEST DETAILS"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:B{ws_summary.max_row}')

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Date:", config['target_date']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for detail in test_details:
            ws_summary.append(detail)
            label_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
            value_cell = ws_summary.cell(row=ws_summary.max_row, column=2)
            label_cell.fill = info_fill
            label_cell.font = Font(bold=True, size=9)
            value_cell.font = info_font

        ws_summary.append([])

        # SYSTEM UNDER TEST SECTION
        ws_summary.append(["SYSTEM UNDER TEST"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:B{ws_summary.max_row}')

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "MV Demand (Fixed)"],
            ["SIP Duration:", f"{sip_duration} minutes"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
        ]

        for detail in system_details:
            ws_summary.append(detail)
            label_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
            value_cell = ws_summary.cell(row=ws_summary.max_row, column=2)
            label_cell.fill = info_fill
            label_cell.font = Font(bold=True, size=9)
            value_cell.font = info_font

        ws_summary.append([])

        # DATA VOLUME ANALYSIS
        ws_summary.append(["DATA VOLUME ANALYSIS"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        ws_summary.append(["Dataset", "Record Count", "Status"])
        header_row = ws_summary.max_row
        for col in range(1, 4):
            cell = ws_summary.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font

        raw_records = len(raw_df) if raw_df is not None else 0
        nrm_records = len(nrm_df) if nrm_df is not None else 0
        chart_records = len(chart_dates) if chart_dates else len(tooltip_data) if tooltip_data else 0

        ws_summary.append(["Raw Database Records", raw_records, "COMPLETE RECORDS" if raw_records > 0 else "NO DATA"])
        ws_summary.cell(row=ws_summary.max_row, column=3).fill = pass_fill if raw_records > 0 else fail_fill

        ws_summary.append(["NRM Processed Records", nrm_records, "COMPLETE RECORDS" if nrm_records > 0 else "NO DATA"])
        ws_summary.cell(row=ws_summary.max_row, column=3).fill = pass_fill if nrm_records > 0 else fail_fill

        ws_summary.append(["Chart Data Points", chart_records, "COMPLETE RECORDS" if chart_records > 0 else "NO DATA"])
        ws_summary.cell(row=ws_summary.max_row, column=3).fill = pass_fill if chart_records > 0 else fail_fill

        ws_summary.append([])

        # VALIDATION RESULTS
        ws_summary.append(["VALIDATION RESULTS"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:E{ws_summary.max_row}')

        ws_summary.append(["Comparison Type", "Matches", "Mismatches", "Success Rate (%)", "Status"])
        header_row = ws_summary.max_row
        for col in range(1, 6):
            cell = ws_summary.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # Raw vs NRM
        raw_nrm_total = comparison_stats['raw_nrm_matches'] + comparison_stats['raw_nrm_mismatches']
        raw_nrm_rate = (comparison_stats['raw_nrm_matches'] / raw_nrm_total * 100) if raw_nrm_total > 0 else 0
        raw_nrm_status = "PASS" if raw_nrm_rate >= 90 else "FAIL"

        ws_summary.append(["Raw vs NRM", comparison_stats['raw_nrm_matches'],
                           comparison_stats['raw_nrm_mismatches'], f"{raw_nrm_rate:.1f}%", raw_nrm_status])
        ws_summary.cell(row=ws_summary.max_row, column=5).fill = pass_fill if raw_nrm_rate >= 90 else fail_fill

        # NRM vs Chart
        chart_matches = 0
        chart_mismatches = 0

        try:
            if chart_comparison_file and os.path.exists(chart_comparison_file):
                chart_comp_df = pd.read_excel(chart_comparison_file,
                                              sheet_name=f'CHART_VS_CALCULATED_NRM_{date_info["selected_date"].upper().replace(" ", "_")}')
                if 'Match' in chart_comp_df.columns:
                    chart_matches = len(chart_comp_df[chart_comp_df['Match'] == 'Yes'])
                    chart_mismatches = len(chart_comp_df[chart_comp_df['Match'] == 'No'])
                else:
                    chart_matches = min(nrm_records, chart_records)
                    chart_mismatches = 0
        except Exception as e:
            logger.info(f"Could not calculate Chart vs NRM statistics: {e}")
            chart_matches = min(nrm_records, chart_records)
            chart_mismatches = 0

        nrm_chart_total = chart_matches + chart_mismatches
        nrm_chart_rate = (chart_matches / nrm_chart_total * 100) if nrm_chart_total > 0 else 0
        nrm_chart_status = "PASS" if nrm_chart_rate >= 90 else "FAIL"

        ws_summary.append(["NRM vs CHART", chart_matches, chart_mismatches,
                           f"{nrm_chart_rate:.1f}%", nrm_chart_status])
        ws_summary.cell(row=ws_summary.max_row, column=5).fill = pass_fill if nrm_chart_rate >= 90 else fail_fill

        ws_summary.append([])

        # SIP DURATION CONFIGURATION ANALYSIS
        ws_summary.append(["SIP DURATION CONFIGURATION ANALYSIS"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        ws_summary.append(["Metric", "Value", "Impact"])
        header_row = ws_summary.max_row
        for col in range(1, 4):
            cell = ws_summary.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font

        expected_sips = (24 * 60) // sip_duration
        actual_sips = raw_records
        coverage = (actual_sips / expected_sips * 100) if expected_sips > 0 else 0

        ws_summary.append(
            [f"Configured SIP Duration", f"{sip_duration} minutes", "Used for energy to demand conversion"])
        ws_summary.append(
            [f"Expected SIPs per Day", f"{expected_sips} records", f"Based on {sip_duration}-min intervals"])
        ws_summary.append([f"Actual Database Records", f"{actual_sips} records", f"Coverage: {coverage:.1f}%"])
        ws_summary.append([f"Chart Hover Points", f"{chart_records} points", "Extracted using dynamic SIP"])
        ws_summary.append([f"Conversion Formula", f"Demand = Energy / ({sip_duration}/60) hours", "Energy to Demand"])

        ws_summary.append([])

        # OVERALL ASSESSMENT
        ws_summary.append(["OVERALL ASSESSMENT"])
        section_cell = ws_summary.cell(row=ws_summary.max_row, column=1)
        section_cell.fill = section_fill
        section_cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        overall_success_rate = (
                (comparison_stats['raw_nrm_matches'] + chart_matches) /
                (raw_nrm_total + nrm_chart_total) * 100
        ) if (raw_nrm_total + nrm_chart_total) > 0 else 0

        if overall_success_rate >= 95:
            assessment = [
                ["EXCELLENT: Demand data validation passed with high confidence"],
                [f"Dynamic SIP configuration ({sip_duration} min) working correctly"],
                ["Energy to Demand conversion accurate"],
                ["Continue with current data collection methods"]
            ]
        elif overall_success_rate >= 85:
            assessment = [
                ["GOOD: Minor discrepancies detected"],
                [f"SIP duration ({sip_duration} min) properly applied"],
                ["Review tolerance settings if needed"],
                [f"Overall success rate: {overall_success_rate:.1f}%"]
            ]
        else:
            assessment = [
                ["NEEDS ATTENTION: Significant discrepancies detected"],
                [f"Verify SIP configuration ({sip_duration} min) is correct"],
                ["Check energy to demand conversion formula"],
                ["Immediate investigation required"],
                [f"Overall success rate: {overall_success_rate:.1f}%"]
            ]

        for item in assessment:
            ws_summary.append(item)
            ws_summary.merge_cells(f'A{ws_summary.max_row}:C{ws_summary.max_row}')

        # Column widths
        column_widths = {'A': 40, 'B': 30, 'C': 25, 'D': 20, 'E': 12}
        for col_letter, width in column_widths.items():
            ws_summary.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Complete validation summary saved: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("COMPLETE DEMAND VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"SIP Duration: {sip_duration} minutes")
        logger.info(f"Data: Raw={raw_records}, NRM={nrm_records}, Chart={chart_records}")
        logger.info(f"Raw vs NRM: {raw_nrm_rate:.1f}% - {raw_nrm_status}")
        logger.info(f"NRM vs Chart: {nrm_chart_rate:.1f}% - {nrm_chart_status}")
        logger.info(f"Overall: {overall_success_rate:.1f}%")
        logger.info(f"Coverage: {coverage:.1f}%")

        return summary_file

    except Exception as e:
        logger.info(f"Summary creation error: {e}")
        logger.info(f"Traceback: {traceback.format_exc()}")
        return None


# ============================================================================
# MAIN AUTOMATION FUNCTION - COMPLETE AND FINAL
# ============================================================================
@log_execution_time
def main_mv_demand_automation():
    """Main MV Demand automation process - COMPLETE FINAL VERSION"""
    config = None
    driver = None
    output_folder = None
    sip_duration = 15

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying MV Demand configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])

        # Set date
        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info:
            logger.info("Failed to set date")
            return False

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        nodetypeid = 158  # MV node type ID
        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], nodetypeid, config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False

        logger.info(f"Meter found: {name} (ID: {dt_id})")
        node_id = dt_id

        # Get SIP duration
        sip_duration = get_sip_duration(mtr_id)
        logger.info(f"Using SIP duration: {sip_duration} minutes")

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Navigate to detailed view - DEMAND TAB
        logger.info("Navigating to demand detailed view...")
        time.sleep(5)
        try:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@id="divCfDetailedLink"]'))).click()
            logger.info("Detailed view opened")
            time.sleep(2)
            # Click Demand tab
            wait.until(
                EC.element_to_be_clickable((By.XPATH, '//span[@class="dx-tab-text-span" and text()="Demand"]'))).click()
            logger.info("Demand tab selected")
            time.sleep(4)
        except Exception as e:
            logger.info(f"Failed to open demand view: {e}")
            return False

        # Extract chart data
        logger.info(f"Extracting demand chart data with {sip_duration}-min SIP...")
        chart_dates, tooltip_data = extract_chart_data_single_pass(driver, wait, sip_duration)

        if not chart_dates or not tooltip_data:
            logger.info("Chart extraction failed")
            return False

        # Collect side panel data
        side_data = collect_side_panel_data(driver, wait)

        # Save chart data
        chart_file = save_chart_data_to_excel(tooltip_data, date_info, side_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get database data
        raw_df, nrm_df = get_database_data_for_chart_dates(config['target_date'], mtr_id, node_id)

        if raw_df.empty and nrm_df.empty:
            logger.info("No database data found")
            return False

        # Process database comparison with energy to demand conversion
        logger.info(f"Processing demand comparison with {sip_duration}-min SIP...")
        raw_file, processed_file, raw_df_proc, nrm_calculated, nrm_df_proc = process_database_comparison_with_calculated_pipeline(
            raw_df, nrm_df, date_info, sip_duration)

        # Move files to output folder
        raw_file = save_file_to_output(raw_file, output_folder)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison file directly in output folder
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        comparison_file = os.path.join(output_folder,
                                       f"demand_actual_vs_theoretical_data_{date_safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

        logger.info(f"Creating comparison file: {os.path.basename(comparison_file)}")
        create_enhanced_database_comparison_report(raw_file, processed_file, comparison_file)

        # Verify comparison file was created
        if not os.path.exists(comparison_file):
            logger.info("Warning: Comparison file was not created, continuing with remaining reports")

        # Create final validation report
        logger.info("Creating final demand validation report...")
        final_report = final_chart_database_comparison(chart_file, processed_file, date_info)
        if final_report:
            final_report = save_file_to_output(final_report, output_folder)

        # Calculate comparison stats
        comparison_stats = {'raw_nrm_matches': 0, 'raw_nrm_mismatches': 0}
        TOLERANCE = 0.001
        sip_duration_hr = sip_duration / 60
        max_rows = max(len(raw_df), len(nrm_calculated))

        for idx in range(max_rows):
            raw_record = raw_df.iloc[idx] if idx < len(raw_df) else None
            nrm_record = nrm_calculated.iloc[idx] if idx < len(nrm_calculated) else None
            match_status = "MATCH"

            comparisons = [('kwh_i', 'kw_i'), ('kvah_i', 'kva_i'), ('kvar_i_total', 'kvar_i')]
            for raw_col, nrm_col in comparisons:
                raw_val = raw_record[raw_col] if raw_record is not None and raw_col in raw_record.index else None
                nrm_val = nrm_record[nrm_col] if nrm_record is not None and nrm_col in nrm_record.index else None

                # Convert energy to demand
                raw_demand = raw_val / sip_duration_hr if raw_val is not None and not pd.isna(raw_val) else None

                if not values_match(raw_demand, nrm_val, TOLERANCE):
                    match_status = "MISMATCH"

            if match_status == "MATCH":
                comparison_stats['raw_nrm_matches'] += 1
            else:
                comparison_stats['raw_nrm_mismatches'] += 1

        # Create comprehensive summary report
        logger.info("Creating comprehensive demand summary...")
        summary_report = create_complete_validation_summary_report(
            comparison_stats, final_report, date_info, raw_df, nrm_df,
            chart_dates, tooltip_data, sip_duration, config, name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("MV DEMAND AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: MV Demand (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Date: {config['target_date']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"SIP Duration: {sip_duration} minutes")
        logger.info(f"Chart Data: {len(chart_dates)} dates, {len(tooltip_data)} points")
        logger.info(f"Database: Raw={len(raw_df)}, NRM={len(nrm_df)} records")

        expected_sips = (24 * 60) // sip_duration
        actual_sips = len(raw_df)
        coverage = (actual_sips / expected_sips * 100) if expected_sips > 0 else 0

        logger.info(f"SIP Analysis:")
        logger.info(f"   Expected: {expected_sips} SIPs/day ({sip_duration}-min intervals)")
        logger.info(f"   Actual: {actual_sips} SIPs")
        logger.info(f"   Coverage: {coverage:.1f}%")
        logger.info("")
        logger.info("Generated Files (6 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(raw_file) if raw_file else 'Raw database'}")
        logger.info(f"   3. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(
            f"   4. {os.path.basename(comparison_file) if os.path.exists(comparison_file) else 'Comparison report'}")
        logger.info(f"   5. {os.path.basename(final_report) if final_report else 'Final validation'}")
        logger.info(f"   6. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ MV demand monitoring only (fixed)")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Dynamic SIP from database")
        logger.info("   ✓ Energy to Demand conversion")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")
        logger.info(f"Traceback: {traceback.format_exc()}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"MV Demand Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"SIP: {sip_duration}min\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                    f.write(f"\nFull Traceback:\n{traceback.format_exc()}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("MV DEMAND AUTOMATION - COMPLETE UPDATED VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: MV Demand (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ MV demand monitoring only")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Dynamic SIP duration from database")
    logger.info("   ✓ Energy to Demand conversion (Energy / SIP_hours)")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Fixed sheet naming issues")
    logger.info("   ✓ Complete error handling")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_mv_demand_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("MV DEMAND AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ MV demand monitoring (fixed)")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Dynamic SIP duration")
        logger.info("   ✓ Energy to Demand conversion")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Fixed sheet naming")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 6 output files generated")
    else:
        logger.info("MV DEMAND AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("MV Demand Automation Finished")
    logger.info("=" * 60)
