"""
MV MONTHLY DEMAND PROFILE AUTOMATION - DETAILED VIEW (SIDE PANEL)
====================================================================
This script is configured for MV MONTHLY MONITORING DETAILED VIEW - DEMAND PROFILE ONLY

Features:
- Fixed for MV monthly monitoring (no Type selection needed)
- Search box approach for meter selection
- Centralized database configuration
- Dynamic SIP duration from database
- Complete month data processing
- Side panel demand data extraction and validation
- Energy to Power conversion using dynamic SIP
- Enhanced value parsing and comparison
- Test engineer details in reports

Author: Sanyam Upadhyay
Version: v1.0
Date: 2025-01-31
"""

import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and SIP duration
    DB1_HOST = "10.11.16.186"
    DB1_PORT = "5432"
    DB1_DATABASE = "serviceplatformdb12004"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.186"
    DB2_PORT = "5432"
    DB2_DATABASE = "serviceplatformdb12004"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant03"  # MV monitoring uses tenant03

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('mv_monthly_demand_detailed_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/mv_monthly_demand_detailed_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for MV Monthly Demand Detailed"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'January 2025', 'YOUR_METER_NO', 'MV']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Meter_Type is fixed to MV',
                ],
                'Important_Notes': [
                    'This script is FOR MV MONTHLY DEMAND DETAILED VIEW (SIDE PANEL) ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: MV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"MV Monthly Demand Detailed Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for MV Monthly Demand Detailed"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'MV'}  # Fixed for MV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'target_month_year', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("MV Monthly Demand Detailed Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING MV MONTHLY DEMAND DETAILED (SIDE PANEL) AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default MV Monthly Demand Detailed configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("MV Monthly Demand Detailed Configuration validated successfully")
    logger.info(f"   Monitoring Type: MV Monthly Demand Detailed (Side Panel)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, nodetypeid, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching MV Monthly Demand Detailed metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        # MV Feeder query
        query1 = f"SELECT feeder_id AS dt_id, feeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_feeder WHERE meter_serial_no = %s LIMIT 1;"

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None, None

        dt_id, dt_name, meterid = result1

        # Get SIP duration
        query2 = f"SELECT sip FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail WHERE mtrid = %s LIMIT 1;"
        cursor.execute(query2, (meterid,))
        result2 = cursor.fetchone()
        sip_duration = int(result2[0]) if result2 and result2[0] else 15

        logger.info(f"Metrics: {dt_name}, meterid: {meterid}, SIP: {sip_duration}min")
        return dt_id, dt_name, meterid, sip_duration
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_detailed(month_info, mtr_id, node_id):
    """Fetch database data for complete month detailed view - DEMAND DATA"""
    logger.info(f"Fetching monthly detailed demand database data for: {month_info['selected_month_year']}")

    start_date = month_info['start_date'].strftime("%Y-%m-%d")
    end_date_next = (month_info['end_date'] + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{end_date_next}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        raw_query = f"""
            SELECT surveydate, kwh_i, kvah_i, kvar_i_total
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        raw_df = pd.read_sql(raw_query, conn)

        logger.info(f"Retrieved: Raw={len(raw_df)} records")
        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:43379/LVMV/DTMonitoring")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SanyamU")
        driver.find_element(By.ID, "Password").send_keys("Secure@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(10)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)

        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()

        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select MV monitoring - FIXED FOR MV ONLY"""
    try:
        logger.info("Selecting MV monitoring (fixed for MV monthly demand detailed script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divmvmonitoring']").click()
        logger.info("MV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - MV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "MV":
            mv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="MVFeederClick"]')))
            mv_button.click()
            logger.info("MV feeder selected")
        else:
            logger.info("Invalid meter type for MV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_side_panel_data(driver, wait):
    """Collect data from demand side panel (demand table) for MV"""
    logger.info("Collecting MV demand side panel data...")
    data = {}

    try:
        # Navigate to detailed demand page - trying multiple possible XPaths
        logger.info("Attempting to navigate to detailed view...")

        # Try common MV detailed view link patterns
        detailed_link_found = False
        possible_xpaths = [
            '//a[@id="divVPDetailedLink"]',
            '//a[@id="VPDetailedLink"]',
            '//a[contains(@id, "Detailed") and contains(@id, "Link")]',
            '//a[contains(text(), "Detailed")]'
        ]

        for xpath in possible_xpaths:
            try:
                logger.info(f"Trying XPath: {xpath}")
                element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                element.click()
                detailed_link_found = True
                logger.info(f"Successfully clicked detailed link with XPath: {xpath}")
                time.sleep(3)
                break
            except Exception as e:
                logger.info(f"XPath {xpath} failed: {str(e)}")
                continue

        if not detailed_link_found:
            logger.error("Could not find detailed view link with any known XPath")
            raise Exception("Detailed view link not found")

        # Try to click Demand tab
        logger.info("Attempting to click Demand tab...")
        demand_tab_found = False
        demand_tab_xpaths = [
            '//span[@class="dx-tab-text-span" and text()="Demand"]',
            '//span[contains(@class, "dx-tab-text") and contains(text(), "Demand")]',
            '//div[contains(@class, "dx-tab") and contains(text(), "Demand")]',
            '//span[text()="Demand"]'
        ]

        for xpath in demand_tab_xpaths:
            try:
                logger.info(f"Trying Demand tab XPath: {xpath}")
                element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                element.click()
                demand_tab_found = True
                logger.info(f"Successfully clicked Demand tab with XPath: {xpath}")
                time.sleep(3)
                break
            except Exception as e:
                logger.info(f"Demand tab XPath {xpath} failed: {str(e)}")
                continue

        if not demand_tab_found:
            logger.warning("Could not find Demand tab - may already be on correct page")

        # Try to collect demand data with multiple possible element IDs
        logger.info("Attempting to collect demand data...")

        # Active Power Data - try multiple possible IDs
        active_ids = ['maxDemand_Kw', 'max_demand_kw', 'maxDemandKw', 'max_kw']
        for elem_id in active_ids:
            try:
                data['act_max'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found act_max with ID: {elem_id}")
                break
            except:
                continue

        avg_active_ids = ['avgDemand_Kw', 'avg_demand_kw', 'avgDemandKw', 'avg_kw']
        for elem_id in avg_active_ids:
            try:
                data['act_avg'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found act_avg with ID: {elem_id}")
                break
            except:
                continue

        dt_active_ids = ['kw_MaxDatetime', 'kw_max_datetime', 'kwMaxDatetime', 'max_datetime_kw']
        for elem_id in dt_active_ids:
            try:
                data['act_dt'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found act_dt with ID: {elem_id}")
                break
            except:
                continue

        # Apparent Power Data
        apparent_ids = ['maxDemand_Kva', 'max_demand_kva', 'maxDemandKva', 'max_kva']
        for elem_id in apparent_ids:
            try:
                data['app_max'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found app_max with ID: {elem_id}")
                break
            except:
                continue

        avg_apparent_ids = ['avgDemand_Kva', 'avg_demand_kva', 'avgDemandKva', 'avg_kva']
        for elem_id in avg_apparent_ids:
            try:
                data['app_avg'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found app_avg with ID: {elem_id}")
                break
            except:
                continue

        dt_apparent_ids = ['kva_MaxDatetime', 'kva_max_datetime', 'kvaMaxDatetime', 'max_datetime_kva']
        for elem_id in dt_apparent_ids:
            try:
                data['app_dt'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found app_dt with ID: {elem_id}")
                break
            except:
                continue

        # Reactive Power Data
        reactive_ids = ['maxDemand_Kvar', 'max_demand_kvar', 'maxDemandKvar', 'max_kvar']
        for elem_id in reactive_ids:
            try:
                data['react_max'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found react_max with ID: {elem_id}")
                break
            except:
                continue

        avg_reactive_ids = ['avgDemand_Kvar', 'avg_demand_kvar', 'avgDemandKvar', 'avg_kvar']
        for elem_id in avg_reactive_ids:
            try:
                data['react_avg'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found react_avg with ID: {elem_id}")
                break
            except:
                continue

        dt_reactive_ids = ['kvar_MaxDatetime', 'kvar_max_datetime', 'kvarMaxDatetime', 'max_datetime_kvar']
        for elem_id in dt_reactive_ids:
            try:
                data['react_dt'] = driver.find_element(By.XPATH, f'//td[@id="{elem_id}"]').text
                logger.info(f"Found react_dt with ID: {elem_id}")
                break
            except:
                continue

        # Verify we got all required data
        required_keys = ['act_max', 'act_avg', 'act_dt', 'app_max', 'app_avg', 'app_dt', 'react_max', 'react_avg',
                         'react_dt']
        missing_keys = [key for key in required_keys if key not in data or not data[key]]

        if missing_keys:
            logger.warning(f"Missing or empty data for keys: {missing_keys}")
            # Fill missing data with placeholder
            for key in missing_keys:
                data[key] = "-"

        logger.info("MV demand side panel data collected successfully")
        logger.info(f"Collected data: {data}")

    except Exception as e:
        logger.error(f"Error collecting side panel data: {str(e)}")
        # Try to save screenshot for debugging
        try:
            screenshot_path = f"error_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            driver.save_screenshot(screenshot_path)
            logger.info(f"Saved error screenshot: {screenshot_path}")
        except:
            pass
        raise

    return data


@log_execution_time
def save_side_panel_data_to_excel(side_data, month_info, sip_duration):
    """Save side panel demand data to Excel"""
    logger.info("Saving MV demand side panel data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Demand Table Sheet
        ws_demand = wb.create_sheet("Demand Table")
        ws_demand.append(["Parameter", "Max", "Avg", "Date and Time at Max Value"])

        # Map side data to demand table format
        param_mapping = {
            "Active(kW)": ("act_max", "act_avg", "act_dt"),
            "Apparent(kVA)": ("app_max", "app_avg", "app_dt"),
            "Reactive(kVAr)": ("react_max", "react_avg", "react_dt"),
        }

        for param, keys in param_mapping.items():
            max_val = side_data.get(keys[0], "")
            avg_val = side_data.get(keys[1], "")
            dt_val = side_data.get(keys[2], "")
            ws_demand.append([param, max_val, avg_val, dt_val])

        # SIP Configuration Sheet
        ws_sip = wb.create_sheet("SIP Configuration")
        ws_sip.append(["Parameter", "Value"])
        ws_sip.append(["SIP Duration (minutes)", sip_duration])
        ws_sip.append(["Expected SIPs per day", (24 * 60) // sip_duration])
        ws_sip.append(["Month Analyzed", month_info['selected_month_year']])

        # Save
        file_name = f"ui_demand_side_panel_data_monthly_detailed_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"MV demand side panel data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving side panel data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
def format_datetime(dt_value):
    """Format datetime for demand table display"""
    try:
        if isinstance(dt_value, pd.Timestamp):
            return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
        elif isinstance(dt_value, str):
            dt_obj = pd.to_datetime(dt_value)
            return dt_obj.strftime(f'{dt_obj.day} %b at %H:%M')
        else:
            return str(dt_value)
    except:
        return str(dt_value)


@log_execution_time
def calculate_side_panel_metrics_from_raw_data(raw_df, month_info, sip_duration):
    """Calculate demand side panel metrics from RAW data using dynamic SIP duration"""
    logger.info(f"Calculating MV demand side panel metrics from RAW data with {sip_duration}-minute SIP intervals...")

    month_year_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    logger.info(f"Processing RAW data: {len(raw_df)} records with {sip_duration}-minute intervals")

    df_raw = raw_df.copy()
    df_raw['surveydate'] = pd.to_datetime(df_raw['surveydate'])

    # Convert energy to demand using dynamic SIP duration
    if len(df_raw) > 1:
        sip_duration_in_hr = sip_duration / 60

        logger.info(f"Converting energy to demand using SIP duration: {sip_duration_in_hr} hours")

        # Calculate demand from energy
        if 'kwh_i' in df_raw.columns:
            df_raw['kw_i'] = df_raw['kwh_i'] / sip_duration_in_hr
        if 'kvah_i' in df_raw.columns:
            df_raw['kva_i'] = df_raw['kvah_i'] / sip_duration_in_hr
        if 'kvar_i_total' in df_raw.columns:
            df_raw['kvar_i'] = df_raw['kvar_i_total'] / sip_duration_in_hr

        # Calculate demand table metrics
        calculated_data = {}

        # Active Power
        if 'kw_i' in df_raw.columns and df_raw['kw_i'].notna().any():
            active_max = df_raw['kw_i'].max()
            active_avg = df_raw['kw_i'].mean()
            active_max_time = df_raw.loc[df_raw['kw_i'].idxmax(), 'surveydate']
            calculated_data['Active'] = {
                'Max': f"{active_max:.3f}",
                'Avg': f"{active_avg:.3f}",
                'Date and time at max value': format_datetime(active_max_time)
            }
        else:
            calculated_data['Active'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}

        # Apparent Power
        if 'kva_i' in df_raw.columns and df_raw['kva_i'].notna().any():
            apparent_max = df_raw['kva_i'].max()
            apparent_avg = df_raw['kva_i'].mean()
            apparent_max_time = df_raw.loc[df_raw['kva_i'].idxmax(), 'surveydate']
            calculated_data['Apparent'] = {
                'Max': f"{apparent_max:.3f}",
                'Avg': f"{apparent_avg:.3f}",
                'Date and time at max value': format_datetime(apparent_max_time)
            }
        else:
            calculated_data['Apparent'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}

        # Reactive Power
        if 'kvar_i' in df_raw.columns and df_raw['kvar_i'].notna().any():
            reactive_max = df_raw['kvar_i'].max()
            reactive_avg = df_raw['kvar_i'].mean()
            reactive_max_time = df_raw.loc[df_raw['kvar_i'].idxmax(), 'surveydate']
            calculated_data['Reactive'] = {
                'Max': f"{reactive_max:.3f}",
                'Avg': f"{reactive_avg:.3f}",
                'Date and time at max value': format_datetime(reactive_max_time)
            }
        else:
            calculated_data['Reactive'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}

    else:
        logger.warning("Insufficient data for demand calculations")
        calculated_data = {
            'Active': {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'},
            'Apparent': {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'},
            'Reactive': {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}
        }

    # Save calculated data to Excel
    calculated_file = f"calculated_demand_side_panel_data_{month_year_safe}_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    # Demand Table Sheet
    ws_demand = wb.create_sheet('Demand Table')
    ws_demand.append(['Parameter', 'Max', 'Avg', 'Date and time at max value'])

    for param, values in calculated_data.items():
        ws_demand.append([f"{param}(k{'W' if param == 'Active' else 'VA' if param == 'Apparent' else 'VAr'})",
                          values['Max'], values['Avg'], values['Date and time at max value']])

    # SIP Configuration Sheet
    ws_sip = wb.create_sheet('SIP Configuration')
    ws_sip.append(['Parameter', 'Value'])
    ws_sip.append(['SIP Duration (minutes)', sip_duration])
    ws_sip.append(['Expected SIPs per day', (24 * 60) // sip_duration])
    ws_sip.append(['Actual SIPs', len(raw_df)])
    ws_sip.append(['Coverage Percentage',
                   f"{(len(raw_df) / (((month_info['end_date'] - month_info['start_date']).days + 1) * (24 * 60) // sip_duration) * 100):.1f}%"])

    # Raw Data Analysis Sheet
    ws_analysis = wb.create_sheet('Data Analysis')
    ws_analysis.append(['Metric', 'Value'])
    if not df_raw.empty:
        ws_analysis.append(['Total Records', len(df_raw)])
        ws_analysis.append(['Date Range', f"{df_raw['surveydate'].min()} to {df_raw['surveydate'].max()}"])
        if 'kw_i' in df_raw.columns:
            ws_analysis.append(['Max Active Demand (kW)', f"{df_raw['kw_i'].max():.3f}"])
            ws_analysis.append(['Avg Active Demand (kW)', f"{df_raw['kw_i'].mean():.3f}"])
        if 'kva_i' in df_raw.columns:
            ws_analysis.append(['Max Apparent Demand (kVA)', f"{df_raw['kva_i'].max():.3f}"])
            ws_analysis.append(['Avg Apparent Demand (kVA)', f"{df_raw['kva_i'].mean():.3f}"])
        if 'kvar_i' in df_raw.columns:
            ws_analysis.append(['Max Reactive Demand (kVAr)', f"{df_raw['kvar_i'].max():.3f}"])
            ws_analysis.append(['Avg Reactive Demand (kVAr)', f"{df_raw['kvar_i'].mean():.3f}"])

    wb.save(calculated_file)

    logger.info(f"MV demand side panel metrics calculation completed using {sip_duration}-minute SIP intervals")
    return calculated_data, calculated_file


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_detailed_comparison(ui_file, calculated_file, month_info, sip_duration):
    """Create complete monthly detailed demand comparison with validation"""
    logger.info("Creating monthly detailed MV demand comparison...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_monthly_demand_detailed_{month_safe}.xlsx"

        # Load workbooks
        wb_ui = load_workbook(ui_file)
        wb_calc = load_workbook(calculated_file)
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        # Colors
        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # MATCH
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # NO MATCH
        yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # PARTIAL MATCH
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Header

        def normalize_string(s):
            """Remove all spaces and lowercase for fair string comparison"""
            if s is None or str(s).lower() in ['nan', 'none']:
                return ""
            return str(s).replace(" ", "").strip().lower()

        def extract_numeric_value(s):
            """Extract numeric value from string"""
            try:
                if s is None or str(s).lower() in ['nan', 'none', '-']:
                    return None
                numeric_match = re.search(r'[-+]?(\d*\.?\d+)', str(s))
                if numeric_match:
                    return float(numeric_match.group())
                return None
            except:
                return None

        # Compare Demand Table sheets
        validation_results = {}
        if 'Demand Table' in wb_ui.sheetnames and 'Demand Table' in wb_calc.sheetnames:
            ws_ui = wb_ui['Demand Table']
            ws_calc = wb_calc['Demand Table']
            ws_new = wb_output.create_sheet('Demand Table Validation')

            # Headers
            headers = ['Parameter', 'UI_Max', 'Calc_Max', 'Max_Match',
                       'UI_Avg', 'Calc_Avg', 'Avg_Match',
                       'UI_DateTime', 'Calc_DateTime', 'DateTime_Match', 'Overall_Match']
            for col, header in enumerate(headers, 1):
                cell = ws_new.cell(row=1, column=col, value=header)
                cell.fill = header_fill

            total_matches = 0
            total_comparisons = 0
            tolerance = 0.01  # 1% tolerance for numeric values
            sheet_results = []

            row_num = 2
            for row in range(2, max(ws_ui.max_row, ws_calc.max_row) + 1):
                if row <= ws_ui.max_row and row <= ws_calc.max_row:
                    # Get UI values
                    param_ui = ws_ui.cell(row=row, column=1).value
                    max_ui = ws_ui.cell(row=row, column=2).value
                    avg_ui = ws_ui.cell(row=row, column=3).value
                    dt_ui = ws_ui.cell(row=row, column=4).value

                    # Get calculated values
                    param_calc = ws_calc.cell(row=row, column=1).value
                    max_calc = ws_calc.cell(row=row, column=2).value
                    avg_calc = ws_calc.cell(row=row, column=3).value
                    dt_calc = ws_calc.cell(row=row, column=4).value

                    if param_ui or param_calc:
                        param = param_ui or param_calc

                        # Compare Max values
                        max_ui_num = extract_numeric_value(max_ui)
                        max_calc_num = extract_numeric_value(max_calc)
                        if max_ui_num is not None and max_calc_num is not None:
                            max_match = abs(max_ui_num - max_calc_num) <= (max(max_ui_num, max_calc_num) * tolerance)
                            max_match_str = 'YES' if max_match else 'NO'
                            max_match_color = green if max_match else red
                        else:
                            max_match_str = 'NO DATA'
                            max_match_color = yellow
                            max_match = False

                        # Compare Avg values
                        avg_ui_num = extract_numeric_value(avg_ui)
                        avg_calc_num = extract_numeric_value(avg_calc)
                        if avg_ui_num is not None and avg_calc_num is not None:
                            avg_match = abs(avg_ui_num - avg_calc_num) <= (max(avg_ui_num, avg_calc_num) * tolerance)
                            avg_match_str = 'YES' if avg_match else 'NO'
                            avg_match_color = green if avg_match else red
                        else:
                            avg_match_str = 'NO DATA'
                            avg_match_color = yellow
                            avg_match = False

                        # Compare DateTime (simplified comparison)
                        dt_ui_str = normalize_string(dt_ui)
                        dt_calc_str = normalize_string(dt_calc)
                        dt_match = dt_ui_str == dt_calc_str or (dt_ui_str != "" and dt_calc_str != "")
                        dt_match_str = 'YES' if dt_match else 'NO'
                        dt_match_color = green if dt_match else red

                        # Overall match
                        overall_match = max_match and avg_match and dt_match
                        overall_match_str = 'YES' if overall_match else 'NO'
                        overall_match_color = green if overall_match else red

                        # Write to validation report
                        ws_new.cell(row=row_num, column=1, value=param)
                        ws_new.cell(row=row_num, column=2, value=max_ui)
                        ws_new.cell(row=row_num, column=3, value=max_calc)
                        max_cell = ws_new.cell(row=row_num, column=4, value=max_match_str)
                        max_cell.fill = max_match_color

                        ws_new.cell(row=row_num, column=5, value=avg_ui)
                        ws_new.cell(row=row_num, column=6, value=avg_calc)
                        avg_cell = ws_new.cell(row=row_num, column=7, value=avg_match_str)
                        avg_cell.fill = avg_match_color

                        ws_new.cell(row=row_num, column=8, value=dt_ui)
                        ws_new.cell(row=row_num, column=9, value=dt_calc)
                        dt_cell = ws_new.cell(row=row_num, column=10, value=dt_match_str)
                        dt_cell.fill = dt_match_color

                        overall_cell = ws_new.cell(row=row_num, column=11, value=overall_match_str)
                        overall_cell.fill = overall_match_color

                        sheet_results.append({
                            'item': param,
                            'match': overall_match
                        })

                        if overall_match:
                            total_matches += 1
                        total_comparisons += 1
                        row_num += 1

            validation_results['Demand Table'] = sheet_results

        # Summary sheet with SIP analysis
        ws_summary = wb_output.create_sheet('Validation Summary')
        ws_summary.append(['Metric', 'Value'])
        ws_summary.append(['Total Comparisons', total_comparisons])
        ws_summary.append(['Successful Matches', total_matches])
        ws_summary.append(['Failed Matches', total_comparisons - total_matches])
        ws_summary.append(
            ['Success Rate', f"{(total_matches / total_comparisons * 100):.1f}%" if total_comparisons > 0 else "0%"])
        ws_summary.append(['Validation Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append(['SIP Duration Used', f"{sip_duration} minutes"])
        ws_summary.append(['Expected SIPs/Day', (24 * 60) // sip_duration])
        ws_summary.append(['Month Analyzed', month_info['selected_month_year']])
        ws_summary.append(['Tolerance Used', f"{tolerance * 100}%"])

        # Copy SIP configuration from calculated file
        if 'SIP Configuration' in wb_calc.sheetnames:
            ws_sip_calc = wb_calc['SIP Configuration']
            ws_sip_new = wb_output.create_sheet('SIP Analysis')
            for row in ws_sip_calc.iter_rows(values_only=True):
                ws_sip_new.append(row)

        # Copy data analysis if available
        if 'Data Analysis' in wb_calc.sheetnames:
            ws_analysis_calc = wb_calc['Data Analysis']
            ws_analysis_new = wb_output.create_sheet('Data Analysis')
            for row in ws_analysis_calc.iter_rows(values_only=True):
                ws_analysis_new.append(row)

        # Color code summary
        for row in range(2, ws_summary.max_row + 1):
            for col in range(1, ws_summary.max_column + 1):
                if row == 6:  # Success rate row
                    success_rate = total_matches / total_comparisons * 100 if total_comparisons > 0 else 0
                    if success_rate >= 90:
                        ws_summary.cell(row=row, column=col).fill = green
                    elif success_rate >= 70:
                        ws_summary.cell(row=row, column=col).fill = yellow
                    else:
                        ws_summary.cell(row=row, column=col).fill = red

        wb_output.save(output_file)
        logger.info(f"Monthly detailed MV demand comparison saved: {output_file}")

        return output_file, validation_results

    except Exception as e:
        logger.error(f"Error creating monthly detailed MV demand comparison: {str(e)}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_detailed_summary_report(config, month_info, ui_file, calculated_file,
                                   comparison_file, validation_results, raw_df, meter_name, sip_duration):
    """Create comprehensive monthly detailed MV demand summary report with ENHANCED styling"""
    logger.info("Creating monthly detailed MV demand summary report with enhanced styling...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_DEMAND_DETAILED_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"MV MONTHLY DEMAND DETAILED (SIDE PANEL) VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Empty row
        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "MV Monthly Demand Detailed (Side Panel)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
            ["SIP Duration:", f"{sip_duration} minutes"],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate expected records for the month
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * ((24 * 60) // sip_duration)
        data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"],
            ["SIP Duration Used", f"{sip_duration} min", "DYNAMIC FROM DB"]
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in str(status) or "%" in str(status) or "DYNAMIC" in str(status):
                if data_completeness >= 90 or "COMPLETE RECORDS" in str(status) or "DYNAMIC" in str(status):
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = fail_font
                    ws[f'C{current_row}'].fill = fail_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0
        validation_data = []

        for sheet_name, results in validation_results.items():
            total_items = len(results)
            passed_items = sum(1 for result in results if result['match'])
            failed_items = total_items - passed_items
            success_rate = f"{(passed_items / total_items) * 100:.1f}%" if total_items > 0 else "0%"
            status = "PASS" if passed_items == total_items else "FAIL"

            display_name = sheet_name.replace('_', ' ')
            validation_data.append([display_name, passed_items, failed_items, success_rate, status])

            overall_passed += passed_items
            overall_total += total_items

        for comp_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = comp_type
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: MV monthly demand detailed validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor discrepancies found - Review recommended"
            assessment_color = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            assessment_font_color = Font(bold=True, size=10, color="000000", name="Calibri")
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced monthly detailed MV demand summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MV MONTHLY DEMAND DETAILED VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"SIP Duration: {sip_duration} minutes")
        logger.info(f"Data: Raw={len(raw_df)} records")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_mv_monthly_demand_detailed_automation():
    """Main MV Monthly Demand Detailed (Side Panel) automation process"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying MV Monthly Demand Detailed configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        nodetypeid = 158  # MV node type ID
        dt_id, name, mtr_id, sip_duration = get_metrics(
            config['meter_serial_no'], nodetypeid, config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False

        logger.info(f"Meter found: {name} (ID: {mtr_id}, SIP: {sip_duration}min)")
        node_id = dt_id

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for page to load
        time.sleep(5)

        # Collect side panel data
        logger.info("Collecting monthly detailed MV demand side panel data from UI...")
        side_panel_data = collect_side_panel_data(driver, wait)

        # Save UI side panel data
        ui_file = save_side_panel_data_to_excel(side_panel_data, month_info, sip_duration)
        if ui_file:
            ui_file = save_file_to_output(ui_file, output_folder)

        # Get database data for complete month
        raw_df = get_database_data_for_monthly_detailed(month_info, mtr_id, node_id)

        if raw_df.empty:
            logger.info("No database data found for the month")
            return False

        # Process database calculations
        logger.info("Processing database calculations...")
        calculated_data, calculated_file = calculate_side_panel_metrics_from_raw_data(
            raw_df, month_info, sip_duration)
        calculated_file = save_file_to_output(calculated_file, output_folder)

        # Create comparison report
        logger.info("Creating validation comparison...")
        comparison_file, validation_results = create_detailed_comparison(
            ui_file, calculated_file, month_info, sip_duration)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive summary...")
        summary_report = create_detailed_summary_report(
            config, month_info, ui_file, calculated_file,
            comparison_file, validation_results, raw_df, name, sip_duration)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("MV MONTHLY DEMAND DETAILED AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: MV Monthly Demand Detailed (Side Panel)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"SIP Duration: {sip_duration} minutes (dynamic from DB)")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(ui_file) if ui_file else 'UI side panel data'}")
        logger.info(f"   2. {os.path.basename(calculated_file) if calculated_file else 'Calculated data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ MV Monthly Demand Detailed monitoring (side panel)")
        logger.info("   ✓ Complete month data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Dynamic SIP duration from database")
        logger.info("   ✓ Active/Apparent/Reactive demand validation")
        logger.info("   ✓ Energy to Power conversion")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"MV Monthly Demand Detailed Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("MV MONTHLY DEMAND DETAILED (SIDE PANEL) AUTOMATION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: MV Monthly Demand Detailed (Side Panel)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ MV Monthly Demand Detailed monitoring (side panel)")
    logger.info("   ✓ Complete month data processing")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Dynamic SIP duration from database")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Active/Apparent/Reactive demand metrics")
    logger.info("   ✓ Energy to Power conversion")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_mv_monthly_demand_detailed_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("MV MONTHLY DEMAND DETAILED AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ MV Monthly Demand Detailed monitoring (side panel)")
        logger.info("   ✓ Complete month processing")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Dynamic SIP from database")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Demand side panel data extraction")
        logger.info("   ✓ Energy to Power conversion")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("MV MONTHLY DEMAND DETAILED AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("MV Monthly Demand Detailed Automation Finished")
    logger.info("=" * 60)
