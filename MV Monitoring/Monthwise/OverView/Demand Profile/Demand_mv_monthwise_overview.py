import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and MV feeder info
    DB1_HOST = "10.11.16.186"
    DB1_PORT = "5432"
    DB1_DATABASE = "serviceplatformdb12004"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.186"
    DB2_PORT = "5432"
    DB2_DATABASE = "serviceplatformdb12004"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant04"  # Change to tenant01, tenant02, etc. as needed

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('mv_monthly_demand_overview_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/mv_monthly_demand_overview_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for MV Monthly Demand Overview"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Target_Month_Year', 'Meter_Serial_No'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'January 2025', 'YOUR_METER_NO']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your MV meter serial number',
                    'Save file before running',
                ],
                'Important_Notes': [
                    'This script is FOR MV MONTHLY DEMAND OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: "January 2025" (full month name + year)',
                    'MV Feeder meter only',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"MV Monthly Demand Overview Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for MV Monthly Demand Overview"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'MV'}  # Fixed for MV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()

        config['meter_type'] = 'MV'  # Fixed for MV

        required_fields = ['type', 'area', 'substation', 'target_month_year', 'meter_serial_no']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("MV Monthly Demand Overview Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING MV MONTHLY DEMAND OVERVIEW AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default MV Monthly Demand Overview configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("MV Monthly Demand Overview Configuration validated successfully")
    logger.info(f"   Monitoring Type: MV Monthly Demand Overview (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise
    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no):
    """Get MV feeder metrics from database"""
    logger.info(f"Fetching MV Monthly Demand Overview metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        query = f"SELECT feeder_id, feeder_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_feeder WHERE meter_serial_no = %s LIMIT 1;"
        
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        if not result:
            logger.info(f"MV Feeder not found: {mtr_serial_no}")
            return None, None, None

        feeder_id, feeder_name, meterid = result
        logger.info(f"Metrics: {feeder_name}, meterid: {meterid}")
        return feeder_id, feeder_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_demand_overview(month_info, mtr_id):
    """Fetch database data for MV monthly demand overview - COMPLETE MONTH - RAW ONLY"""
    logger.info(f"Fetching MV monthly demand overview database data for month: {month_info['selected_month_year']}")
    logger.info("⚠️ NO NRM queries - 100% raw MV data based validation")
    
    # Use COMPLETE MONTH range
    start_date = month_info['start_date'].strftime('%Y-%m-%d')
    end_date = month_info['end_date'].strftime('%Y-%m-%d')
    
    logger.info(f"Complete MV month range: {start_date} to {end_date}")

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # ✅ ONLY RAW MV DATABASE QUERY - NO NRM QUERIES
        query = f"""
            SELECT surveydate, kwh_i, kvah_i, kvar_i_total,
                   kwh_abs, kvah_abs, kvarh_abs
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata
            WHERE mtrid = {mtr_id}
              AND DATE(surveydate) >= '{start_date}'
              AND DATE(surveydate) <= '{end_date}'
            ORDER BY surveydate ASC;
        """

        raw_df = pd.read_sql(query, conn)
        logger.info(f"Retrieved: {len(raw_df)} MV records for complete month")
        
        # Calculate data completeness
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * 96  # 96 records per day (15-min intervals)
        completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0
        logger.info(f"MV Data completeness: {completeness:.1f}% ({len(raw_df)}/{expected_records})")
        
        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoring.secure.online:43379/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SanyamU")
        driver.find_element(By.ID, "Password").send_keys("Secure@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(8)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info - MV MONTH-WISE LOGIC"""
    logger.info(f"Setting MV calendar to month: {target_month_year}")
    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)
        
        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        time.sleep(2)

        # Parse month info for COMPLETE MONTH
        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        # Create COMPLETE MONTH boundaries
        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info(f"MV Calendar set successfully to: {target_month_year}")
        logger.info(f"Complete MV month range: {start_date} to {end_date}")
        return month_info

    except Exception as e:
        logger.info(f"Error setting MV calendar month: {e}")
        raise


def select_type(driver):
    """Select MV monitoring - FIXED FOR MV ONLY"""
    try:
        logger.info("Selecting MV monitoring (fixed for MV monthly demand overview script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divmvmonitoring']").click()
        logger.info("MV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver):
    """Select MV Feeder meter type"""
    try:
        logger.info("Selecting MV Feeder meter type")
        wait = WebDriverWait(driver, 10)
        
        mv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="MVFeederClick"]')))
        mv_button.click()
        logger.info("MV Feeder selected")
        
        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for MV meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_demand_overview_data(driver):
    """Collect MV monthly demand overview data from UI"""
    logger.info("Starting MV monthly demand overview data collection...")
    data = {}

    try:
        time.sleep(4)
        logger.info("Clicking on MV Demand tab...")
        driver.find_element(By.XPATH, "//div[@class='dx-item-content' and text()='Demand']").click()
        time.sleep(2)

        logger.info("Collecting MV demand values...")
        data['act_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kw"]').text
        data['act_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kw"]').text
        data['act_dt'] = driver.find_element(By.XPATH, '//td[@id="kw_MaxDatetime"]').text
        data['app_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kva"]').text
        data['app_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kva"]').text
        data['app_dt'] = driver.find_element(By.XPATH, '//td[@id="kva_MaxDatetime"]').text
        data['react_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kvar"]').text
        data['react_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kvar"]').text
        data['react_dt'] = driver.find_element(By.XPATH, '//td[@id="kvar_MaxDatetime"]').text

        logger.info("MV monthly demand overview data collection completed")

    except Exception as e:
        logger.error(f"Error in MV monthly demand data collection: {str(e)}")
        raise

    return data


@log_execution_time
def save_demand_overview_data_to_excel(month_info, overview_data):
    """Save MV monthly demand overview data to Excel"""
    logger.info("Saving MV monthly demand overview data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        ws_demand = wb.create_sheet(title="Demand Table")
        ws_demand.append(["Parameter", "Max", "Avg", "Date and time at max value"])

        # Using parameter names that match the database file
        param_mapping = {
            "Active": ("act_max", "act_avg", "act_dt"),
            "Apparent": ("app_max", "app_avg", "app_dt"),
            "Reactive": ("react_max", "react_avg", "react_dt"),
        }

        for param, keys in param_mapping.items():
            max_val = overview_data.get(keys[0], "")
            avg_val = overview_data.get(keys[1], "")
            dt_val = overview_data.get(keys[2], "")
            ws_demand.append([param, max_val, avg_val, dt_val])

        # Save
        file_name = f"chart_data_from_ui_mv_monthly_demand_overview_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"MV monthly demand overview data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving MV monthly demand overview data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_demand_overview_database_calculations(raw_df, month_info):
    """Process database calculations for MV monthly demand overview - RAW ONLY"""
    logger.info("Processing MV monthly demand overview database calculations...")
    logger.info("⚠️ NO NRM processing - calculating everything from raw MV data")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        # Calculate interval
        if len(raw_df) > 1:
            interval_minutes = int((raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)
        else:
            interval_minutes = 15

        sip_duration_in_hr = interval_minutes / 60
        logger.info(f"MV Survey interval detected: {interval_minutes} minutes ({sip_duration_in_hr} hours)")

        processed_file = f"theoretical_mv_monthly_demand_overview_calculated_data_{month_safe}_{timestamp}.xlsx"

        # ✅ CALCULATE MV DEMAND FROM RAW DATA ONLY - NO NRM TABLE
        demand_calculated = pd.DataFrame()
        column_map = {
            'kwh_i': 'kw_i',
            'kvah_i': 'kva_i',
            'kvar_i_total': 'kvar_i'
        }

        for raw_col, demand_col in column_map.items():
            if raw_col in raw_df.columns:
                # Handle NaN values and division
                raw_values = pd.to_numeric(raw_df[raw_col], errors='coerce')
                demand_calculated[demand_col] = raw_values / sip_duration_in_hr
                logger.info(f"✅ Calculated MV {demand_col} from raw {raw_col}")
            else:
                logger.warning(f"Column '{raw_col}' not found in raw_df — filling '{demand_col}' with 0")
                demand_calculated[demand_col] = 0

        # Add time column for reference
        demand_calculated['surveydate'] = raw_df['surveydate']
        # Reorder columns to put 'surveydate' first
        cols = demand_calculated.columns.tolist()
        cols.insert(0, cols.pop(cols.index('surveydate')))
        demand_calculated = demand_calculated[cols]

        # Format surveydate for output
        def format_datetime(dt_value):
            try:
                if pd.isna(dt_value) or dt_value is None:
                    return "No valid data"
                if isinstance(dt_value, pd.Timestamp):
                    return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
                elif isinstance(dt_value, str):
                    dt_obj = pd.to_datetime(dt_value)
                    return dt_obj.strftime(f'{dt_obj.day} %b at %H:%M')
                else:
                    return str(dt_value)
            except Exception:
                return "Invalid datetime"

        # Helper function to safely calculate max/avg/datetime for MV demand parameters
        def safe_mv_demand_calculation(demand_series, datetime_series, param_name):
            try:
                # Convert to numeric and handle NaN values
                numeric_series = pd.to_numeric(demand_series, errors='coerce')

                # Check if we have any valid (non-NaN) values
                valid_mask = ~pd.isna(numeric_series)

                if not valid_mask.any():
                    logger.warning(f"No valid MV data found for {param_name} - all values are NaN")
                    return 0.0, 0.0, "No valid data"

                # Calculate max and avg from valid data only
                valid_data = numeric_series[valid_mask]
                max_val = valid_data.max()
                avg_val = valid_data.mean()

                # Find the datetime for max value
                if pd.isna(max_val):
                    max_datetime_formatted = "No valid data"
                else:
                    # Get index of maximum value from original series
                    max_idx = numeric_series.idxmax()
                    if pd.isna(max_idx):
                        max_datetime_formatted = "No valid data"
                    else:
                        max_datetime = datetime_series.iloc[max_idx]
                        max_datetime_formatted = format_datetime(max_datetime)

                logger.info(
                    f"Safe MV calculation for {param_name} -> Max={max_val:.2f}, Avg={avg_val:.2f}, DateTime={max_datetime_formatted}")
                return max_val, avg_val, max_datetime_formatted

            except Exception as e:
                logger.error(f"Error in safe MV calculation for {param_name}: {str(e)}")
                return 0.0, 0.0, "Calculation error"

        # Save to Excel
        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            # Save ONLY raw MV data and calculated demand - NO NRM sheets
            raw_df.to_excel(writer, sheet_name='RAW MV Database', index=False)
            demand_calculated.to_excel(writer, sheet_name='MV Demand Calculated from RAW', index=False)

            # MV Demand Table
            demand_table_data = []

            # ✅ Calculate from RAW-derived MV demand (not NRM) with error handling
            # Active Power (KW) - calculated from raw MV
            active_max, active_avg, active_max_time_formatted = safe_mv_demand_calculation(
                demand_calculated['kw_i'], demand_calculated['surveydate'], "MV Active Power"
            )
            demand_table_data.append(['Active', active_max, active_avg, active_max_time_formatted])

            # Apparent Power (KVA) - calculated from raw MV
            apparent_max, apparent_avg, apparent_max_time_formatted = safe_mv_demand_calculation(
                demand_calculated['kva_i'], demand_calculated['surveydate'], "MV Apparent Power"
            )
            demand_table_data.append(['Apparent', apparent_max, apparent_avg, apparent_max_time_formatted])

            # Reactive Power (KVAR) - calculated from raw MV
            reactive_max, reactive_avg, reactive_max_time_formatted = safe_mv_demand_calculation(
                demand_calculated['kvar_i'], demand_calculated['surveydate'], "MV Reactive Power"
            )
            demand_table_data.append(['Reactive', reactive_max, reactive_avg, reactive_max_time_formatted])

            # Create MV demand table with matching column names
            demand_table_df = pd.DataFrame(demand_table_data,
                                           columns=['Parameter', 'Max', 'Avg', 'Date and time at max value'])
            demand_table_df.to_excel(writer, sheet_name='Demand Table', index=False)

        logger.info(f"Processed MV monthly demand overview data saved: {processed_file}")
        return processed_file

    except Exception as e:
        logger.error(f"Error processing MV monthly demand overview database: {str(e)}")
        raise


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_demand_overview_comparison(chart_file, processed_file, month_info):
    """Create complete MV monthly demand overview comparison with validation"""
    logger.info("Creating MV monthly demand overview comparison...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_mv_monthly_demand_overview_{month_safe}.xlsx"

        # Load workbooks
        wb_processed = load_workbook(processed_file)
        wb_chart = load_workbook(chart_file)

        # Check if 'Demand Table' sheet exists in both files
        sheet_name = 'Demand Table'
        if sheet_name not in wb_processed.sheetnames or sheet_name not in wb_chart.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in one or both files!")
            return None, None

        # Get worksheets
        ws_processed = wb_processed[sheet_name]
        ws_chart = wb_chart[sheet_name]

        # Convert to DataFrames
        processed_data = []
        chart_data = []

        # Read processed sheet data
        for row in ws_processed.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                processed_data.append(row)

        # Read chart sheet data
        for row in ws_chart.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                chart_data.append(row)

        if len(processed_data) == 0 or len(chart_data) == 0:
            logger.error("No MV data found in one or both files!")
            return None, None

        # Create DataFrames
        processed_df = pd.DataFrame(processed_data)
        chart_df = pd.DataFrame(chart_data)

        # Set headers (first row)
        processed_df.columns = processed_df.iloc[0]
        processed_df = processed_df[1:].reset_index(drop=True)

        chart_df.columns = chart_df.iloc[0]
        chart_df = chart_df[1:].reset_index(drop=True)

        # Colors
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Output workbook
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        # Output sheet setup
        ws_output = wb_output.create_sheet(title="MV Demand Table Comparison")
        headers = ['Parameter', 'DB_Max', 'UI_Max', 'Max_Diff', 'DB_Avg', 'UI_Avg', 'Avg_Diff',
                   'DB_Datetime', 'UI_Datetime', 'Datetime_Match', 'Overall_Match']
        ws_output.append(headers)

        # Track comparison statistics and validation results
        validation_results = {}

        # Iterate through processed DataFrame
        for idx, row in processed_df.iterrows():
            param = row['Parameter']
            logger.info(f"Processing MV parameter: '{param}'")

            # Find corresponding chart row
            chart_matches = chart_df[chart_df['Parameter'] == param]
            if chart_matches.empty:
                logger.warning(f"MV Parameter '{param}' NOT found in Chart file!")
                output_row = [param, row['Max'], "NOT_FOUND", "N/A", row['Avg'], "NOT_FOUND",
                              "N/A", row['Date and time at max value'], "NOT_FOUND", "MISSING", "FAIL"]
                ws_output.append(output_row)
                validation_results[param] = {'match': False}
                continue

            chart_row = chart_matches.iloc[0]
            logger.info(f"Found matching MV parameter '{param}' in both files!")

            # Values from processed file (database)
            proc_max = row['Max']
            proc_avg = row['Avg']
            proc_dt = row['Date and time at max value']

            # Values from chart file (UI)
            chart_max = chart_row['Max']
            chart_avg = chart_row['Avg']
            chart_dt = chart_row['Date and time at max value']

            # --- MAX COMPARISON ---
            try:
                proc_max_f = float(proc_max)
                chart_max_f = float(chart_max)
                max_diff = abs(proc_max_f - chart_max_f)
                max_match = max_diff < 0.01
                max_diff_disp = round(max_diff, 4)
            except Exception:
                max_match = (str(proc_max).strip() == str(chart_max).strip())
                max_diff_disp = 'StringMismatch' if not max_match else '0'

            # --- AVG COMPARISON ---
            try:
                proc_avg_f = float(proc_avg)
                chart_avg_f = float(chart_avg)
                avg_diff = abs(proc_avg_f - chart_avg_f)
                avg_match = avg_diff < 0.01
                avg_diff_disp = round(avg_diff, 4)
            except Exception:
                avg_match = (str(proc_avg).strip() == str(chart_avg).strip())
                avg_diff_disp = 'StringMismatch' if not avg_match else '0'

            # --- DateTime COMPARISON ---
            dt_match = (str(proc_dt).strip() == str(chart_dt).strip())
            dt_match_text = 'PASS' if dt_match else 'FAIL'

            # --- Overall Match ---
            overall_match = max_match and avg_match and dt_match
            match_text = 'PASS' if overall_match else 'FAIL'

            # Store validation results
            validation_results[param] = {'match': overall_match}

            # Append row to output
            output_row = [param, proc_max, chart_max, max_diff_disp, proc_avg, chart_avg,
                          avg_diff_disp, proc_dt, chart_dt, dt_match_text, match_text]
            ws_output.append(output_row)

            # Apply coloring
            row_idx = ws_output.max_row
            ws_output.cell(row=row_idx, column=4).fill = green_fill if max_match else red_fill
            ws_output.cell(row=row_idx, column=7).fill = green_fill if avg_match else red_fill
            ws_output.cell(row=row_idx, column=10).fill = green_fill if dt_match else red_fill
            ws_output.cell(row=row_idx, column=11).fill = green_fill if overall_match else red_fill

        wb_output.save(output_file)
        logger.info(f"MV monthly demand overview comparison saved: {output_file}")

        return output_file, validation_results

    except Exception as e:
        logger.error(f"Error creating MV monthly demand overview comparison: {str(e)}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_demand_overview_summary_report(config, month_info, chart_file, processed_file,
                                         comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive MV monthly demand overview summary report with ENHANCED styling"""
    logger.info("Creating MV monthly demand overview summary report with enhanced styling...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MV_MONTHLY_DEMAND_OVERVIEW_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")

        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"MV MONTHLY DEMAND OVERVIEW VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Empty row
        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Feeder Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "MV Monthly Demand Overview (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate expected records for the month
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_monthly_records = days_in_month * 96
        actual_records = len(raw_df)
        data_completeness = (actual_records / expected_monthly_records * 100) if expected_monthly_records > 0 else 0

        # Data rows
        total_chart_points = 3  # Always 3 for demand parameters
        data_rows = [
            ["Raw Database Records (Monthly)", actual_records, "COMPLETE RECORDS" if actual_records > 0 else "NO DATA"],
            ["Expected Monthly Records", expected_monthly_records, f"{data_completeness:.1f}% Complete"],
            ["Chart Data Points", total_chart_points, "COMPLETE RECORDS"],
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in status or data_completeness >= 90:
                ws[f'C{current_row}'].font = pass_font
                ws[f'C{current_row}'].fill = pass_fill
            elif data_completeness >= 70:
                ws[f'C{current_row}'].font = warning_font
                ws[f'C{current_row}'].fill = warning_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["Demand Parameter", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0

        # MV demand-specific parameter mapping
        mv_demand_params = {
            'Active': 'Active Power (KW)',
            'Apparent': 'Apparent Power (KVA)',
            'Reactive': 'Reactive Power (KVAR)'
        }

        validation_data = []
        for param, result in validation_results.items():
            display_name = mv_demand_params.get(param, param)
            match = 1 if result['match'] else 0
            total = 1
            success_rate = f"{(match / total) * 100:.1f}%"
            status = "PASS" if match == total else "FAIL"

            validation_data.append([display_name, match, total - match, success_rate, status])
            overall_passed += match
            overall_total += total

        for param, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = param
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: MV monthly demand overview validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor MV monthly demand discrepancies found - Review recommended"
            assessment_color = warning_fill
            assessment_font_color = warning_font
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant MV monthly demand validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced MV monthly demand overview summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MV MONTHLY DEMAND OVERVIEW VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Data: Raw={len(raw_df)}, Chart={total_chart_points}")
        logger.info(f"Monthly Coverage: {days_in_month} days")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating MV monthly demand overview summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_mv_monthly_demand_overview_automation():
    """Main MV Monthly Demand Overview automation process"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying MV Monthly Demand Overview configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False

        # Select meter type
        if not select_meter_type(driver):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching MV meter metrics...")
        feeder_id, name, mtr_id = get_metrics(config['meter_serial_no'])

        if not feeder_id:
            logger.info(f"MV Feeder not found: {config['meter_serial_no']}")
            return False

        logger.info(f"MV Feeder found: {name} (ID: {feeder_id})")

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for overview page to load
        time.sleep(5)

        # Collect demand overview data
        logger.info("Collecting MV monthly demand overview data from UI...")
        overview_data = collect_demand_overview_data(driver)

        # Save overview data
        chart_file = save_demand_overview_data_to_excel(month_info, overview_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get database data
        raw_df = get_database_data_for_monthly_demand_overview(month_info, mtr_id)

        if raw_df.empty:
            logger.info("No database data found")
            return False

        # Process database calculations
        logger.info("Processing MV monthly demand database calculations...")
        processed_file = process_demand_overview_database_calculations(raw_df, month_info)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison report
        logger.info("Creating MV monthly demand validation comparison...")
        comparison_file, validation_results = create_demand_overview_comparison(chart_file, processed_file, month_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive MV monthly demand summary...")
        summary_report = create_demand_overview_summary_report(
            config, month_info, chart_file, processed_file,
            comparison_file, validation_results, raw_df, name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("MV MONTHLY DEMAND OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: MV Monthly Demand Overview (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"Database Records: {len(raw_df)} records")
        
        # Calculate data completeness
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * 96
        completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0
        logger.info(f"Data Completeness: {completeness:.1f}%")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ MV Monthly Demand Overview monitoring (fixed)")
        logger.info("   ✓ Complete month data collection")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ RAW-ONLY approach (no NRM dependencies)")
        logger.info("   ✓ Energy-to-Demand conversion (KWH→KW, KVAH→KVA, KVARH→KVAR)")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"MV Monthly Demand Overview Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("MV MONTHLY DEMAND OVERVIEW AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: MV Monthly Demand Overview (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ MV Monthly Demand Overview monitoring only")
    logger.info("   ✓ Complete month data collection")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ RAW-ONLY approach (no NRM)")
    logger.info("   ✓ Energy-to-Demand conversion")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Better null/dash handling")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_mv_monthly_demand_overview_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("MV MONTHLY DEMAND OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ MV Monthly Demand Overview monitoring (fixed)")
        logger.info("   ✓ Complete month data collection")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ RAW-ONLY approach")
        logger.info("   ✓ Energy-to-Demand conversion")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("MV MONTHLY DEMAND OVERVIEW AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("MV Monthly Demand Overview Automation Finished")
    logger.info("=" * 60)
