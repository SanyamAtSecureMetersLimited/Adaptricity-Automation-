import os
import re
import time
import shutil
import logging
import functools
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Change to tenant01, tenant02, etc. as needed

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('lv_load_piechart_monthwise_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/lv_load_piechart_monthwise_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for LV Load/Pie Chart Monthwise"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'January 2025', 'YOUR_METER_NO',
                      'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'This script is FOR LV LOAD/PIE CHART MONTHWISE ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: DT or LV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"LV Load/Pie Chart Monthwise Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for LV Load/Pie Chart Monthwise"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}  # Fixed for LV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'feeder', 'target_month_year', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("LV Load/Pie Chart Monthwise Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING LV LOAD/PIE CHART MONTHWISE AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default LV Load/Pie Chart Monthwise configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("LV Load/Pie Chart Monthwise Configuration validated successfully")
    logger.info(f"   Monitoring Type: LV Load/Pie Chart Monthwise (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Feeder: {config['feeder']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching LV Load/Pie Chart metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        if meter_type.upper() == 'DT':
            query1 = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        elif meter_type.upper() == 'LV':
            query1 = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        else:
            logger.info(f"Invalid meter type: {meter_type}")
            return None, None, None

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None

        dt_id, dt_name, meterid = result1
        logger.info(f"Metrics: {dt_name}, meterid: {meterid}")
        return dt_id, dt_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_load_monthwise(month_info, dt_name_value, meter_serial_no_value, node_id, meter_type):
    """Get ALL raw database data for COMPLETE MONTH - Load monthwise analysis"""
    logger.info(f"Fetching ALL load database data for COMPLETE MONTH: {month_info['selected_month_year']}")

    start_date = month_info['start_date'].strftime('%Y-%m-%d')
    end_date = month_info['end_date'].strftime('%Y-%m-%d')

    date_filter = f"AND DATE(surveydate) >= '{start_date}' AND DATE(surveydate) <= '{end_date}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        nrm_query = f"""
            SELECT surveydate, kva_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile
            WHERE nodeid={node_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        if meter_type.upper() == "DT":
            rating_query = f"""
                SELECT kva_rating
                FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt
                WHERE dt_name = '{dt_name_value}' AND meter_serial_no = '{meter_serial_no_value}'
                LIMIT 1;
            """
        elif meter_type.upper() == "LV":
            rating_query = f"""
                SELECT conductor_ampacity
                FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder
                WHERE lvfeeder_name = '{dt_name_value}' AND meter_serial_no = '{meter_serial_no_value}'
                LIMIT 1;
            """
        else:
            conn.close()
            raise ValueError("meter_type must be either 'DT' or 'LV'")

        nrm_df = pd.read_sql(nrm_query, conn)

        rating_cursor = conn.cursor()
        rating_cursor.execute(rating_query)
        rating_result = rating_cursor.fetchone()
        rating_cursor.close()
        conn.close()

        rating_value = rating_result[0] if rating_result else None

        logger.info(f"Load MONTHWISE records retrieved: {len(nrm_df)} records")
        if rating_value:
            logger.info(f"Rating fetched: {rating_value}")

        return nrm_df, rating_value
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame(), None
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SANYAM")
        driver.find_element(By.ID, "Password").send_keys("Sanyam@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(8)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)

        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()

        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select LV monitoring - FIXED FOR LV ONLY with better wait handling"""
    try:
        logger.info("Selecting LV monitoring (fixed for LV load/pie chart script)")
        time.sleep(3)  # Wait for page to fully load after login

        # Wait for home button to be present and clickable
        home_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//A[@id='divHome']"))
        )
        time.sleep(1)
        home_button.click()
        time.sleep(5)

        # Wait for LV monitoring button
        lv_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//A[@id='divlvmonitoring']"))
        )
        lv_button.click()
        logger.info("LV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")
        raise


def select_meter_type(driver, meter_type):
    """Select meter type - DT or LV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "DT":
            dt_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]')))
            dt_button.click()
            logger.info("DT selected")
        elif meter_type == "LV":
            lv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]')))
            lv_button.click()
            logger.info("LV feeder selected")
        else:
            logger.info("Invalid meter type for LV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_over_data(driver, metertype):
    """Enhanced load data collection with stale element handling"""
    logger.info("Starting load data collection from overview section...")
    try:
        data = {}
        action = ActionChains(driver)
        wait = WebDriverWait(driver, 10)

        logger.info("Deactivating all legends first...")
        time.sleep(2)  # Wait for chart to load

        # Deactivate all legends with retry logic
        legend_selectors = [
            '#dvLoadingTrend g.dxl-marker rect[fill="#257E94"]',
            '#dvLoadingTrend g.dxl-marker rect[fill="#86B8A5"]',
            '#dvLoadingTrend g.dxl-marker rect[fill="#DEAE2A"]',
            '#dvLoadingTrend g.dxl-marker rect[fill="#E38430"]'
        ]

        for selector in legend_selectors:
            try:
                legend = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                legend.click()
                time.sleep(0.5)
            except:
                pass

        logger.info("Extracting load duration data with enhanced tooltip extraction...")

        # Helper function to extract tooltip with retry
        def extract_tooltip_value(rect_selector, path_selector, max_retries=3):
            for attempt in range(max_retries):
                try:
                    # Activate legend
                    rect = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, rect_selector)))
                    rect.click()
                    time.sleep(1)

                    # Move to path and get tooltip
                    path = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, path_selector)))
                    action.move_to_element(path).perform()
                    time.sleep(1.5)

                    # Get tooltip text
                    tooltip = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.dxc-tooltip svg text')))
                    tooltip_text = tooltip.text

                    # Deactivate legend
                    rect_new = driver.find_element(By.CSS_SELECTOR, rect_selector)
                    rect_new.click()
                    time.sleep(0.5)

                    return tooltip_text

                except Exception as e:
                    logger.warning(f"Attempt {attempt + 1} failed: {str(e)}")
                    if attempt < max_retries - 1:
                        time.sleep(1)
                        continue
                    else:
                        return "-"
            return "-"

        # Duration Load <30%
        p1_text = extract_tooltip_value(
            '#dvLoadingTrend g.dxl-marker rect[fill="#257E94"]',
            '#dvLoadingTrend g.dxc-markers path[fill="#257E94"]'
        )

        # Duration Load 30%-60%
        p2_text = extract_tooltip_value(
            '#dvLoadingTrend g.dxl-marker rect[fill="#86B8A5"]',
            '#dvLoadingTrend g.dxc-markers path[fill="#86B8A5"]'
        )

        # Duration Load 60%-80%
        p3_text = extract_tooltip_value(
            '#dvLoadingTrend g.dxl-marker rect[fill="#DEAE2A"]',
            '#dvLoadingTrend g.dxc-markers path[fill="#DEAE2A"]'
        )

        # Duration Load >80%
        p4_text = extract_tooltip_value(
            '#dvLoadingTrend g.dxl-marker rect[fill="#E38430"]',
            '#dvLoadingTrend g.dxc-markers path[fill="#E38430"]'
        )

        logger.info("Extracting rating and load factor data...")

        if metertype.upper() == "DT":
            rating_label = "KVA Rating"
            raw_text = driver.find_element(By.XPATH, '//label[@id="lblKvaRating"]').text
        else:
            rating_label = "Conductor Ampacity"
            raw_text = driver.find_element(By.XPATH, '//label[@id="lblConductorAmpacity"]').text

        load_fac = driver.find_element(By.XPATH, '//label[@id="lblLoadFactor"]').text
        rating_value = re.findall(r"[\d.]+", raw_text)[0] if raw_text else "-"

        data['Load Table'] = {
            rating_label: rating_value,
            'Load Factor': load_fac,
            'Duration Load < 30%': p1_text,
            'Duration Load 30% - 60%': p2_text,
            'Duration Load 60% - 80%': p3_text,
            'Duration Load > 80%': p4_text
        }

        logger.info("Load data collection completed successfully")
        return data

    except Exception as e:
        logger.info(f"Error in load data collection: {e}")
        raise


@log_execution_time
def save_chart_data_to_excel(month_info, side_data):
    """Enhanced chart data saving with better formatting"""
    logger.info("Saving UI load data to Excel...")
    try:
        wb = Workbook()
        wb.remove(wb.active)

        ws_pft = wb.create_sheet("Load Table")
        ws_pft.append(['Parameter', 'Value'])
        for key, value in side_data['Load Table'].items():
            ws_pft.append([key, value])

        chart_file = f"chart_data_from_ui_piechart_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(chart_file)

        logger.info(f"UI load data saved successfully: {chart_file}")
        return chart_file

    except Exception as e:
        logger.info(f"Error saving chart data to Excel: {e}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_load_database_values_monthwise(nrm_df, rating_value, month_info):
    """Enhanced load processing for COMPLETE MONTH data - monthwise analysis"""
    logger.info("Processing COMPLETE MONTH load database data...")
    logger.info("MONTHWISE ANALYSIS: Using all available data for the complete month")

    try:
        if nrm_df.empty:
            logger.info("Load DataFrame is empty - cannot process monthwise load calculations")
            raise ValueError("No monthwise load data available for processing")

        required_columns = ['kva_i']
        missing_columns = [col for col in required_columns if col not in nrm_df.columns]

        if missing_columns:
            logger.info(f"Missing required load columns: {missing_columns}")
            raise ValueError(f"Missing required load columns: {missing_columns}")

        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')

        # Calculate interval
        if len(nrm_df) > 1:
            interval_minutes = int((nrm_df['surveydate'].iloc[1] - nrm_df['surveydate'].iloc[0]).total_seconds() / 60)
        else:
            interval_minutes = 15

        logger.info(f"Load survey interval detected: {interval_minutes} minutes")
        logger.info(f"MONTHWISE: Processing {len(nrm_df):,} records for complete month analysis")

        rating_float = float(rating_value) if rating_value is not None else 1.0

        if rating_float == 0:
            logger.warning("Rating value is 0, using 1.0 to avoid division by zero")
            rating_float = 1.0

        logger.info("Calculating monthwise load statistics...")

        # Calculate Load Factor
        avg_load = nrm_df['kva_i'].mean()
        max_demand = nrm_df['kva_i'].max()
        load_factor = round(avg_load / max_demand, 1) if max_demand != 0 else 0

        logger.info(f"MONTHWISE Load Factor calculated: {load_factor}")

        temp_df = nrm_df.copy()
        temp_df['kva_i'] = pd.to_numeric(temp_df['kva_i'], errors='coerce')
        temp_df['load_percent'] = (temp_df['kva_i'] / rating_float) * 100

        bins = [0, 30, 60, 80, float('inf')]
        labels = ['<30%', '30-60%', '60-80%', '>80%']
        temp_df['load_range'] = pd.cut(temp_df['load_percent'], bins=bins, labels=labels, right=False)

        logger.info("Calculating monthwise duration per load category...")

        duration_dict = {label: '0:00 hrs' for label in labels}
        counts = temp_df['load_range'].value_counts()

        for label in labels:
            duration_mins = counts.get(label, 0) * interval_minutes
            if duration_mins > 0:
                hours = duration_mins // 60
                minutes = duration_mins % 60
                duration_dict[label] = f"{int(hours)}:{int(minutes):02d} hrs"

        logger.info(f"MONTHWISE load duration calculations:")
        logger.info(f"  • Load < 30%: {duration_dict['<30%']}")
        logger.info(f"  • Load 30-60%: {duration_dict['30-60%']}")
        logger.info(f"  • Load 60-80%: {duration_dict['60-80%']}")
        logger.info(f"  • Load > 80%: {duration_dict['>80%']}")

        load_table_data = [
            ['Parameter', 'Value'],
            ['Rating (KVA or Ampacity)', rating_value],
            ['Load Factor', load_factor],
            ['Duration Load < 30%', duration_dict['<30%']],
            ['Duration Load 30% - 60%', duration_dict['30-60%']],
            ['Duration Load 60% - 80%', duration_dict['60-80%']],
            ['Duration Load > 80%', duration_dict['>80%']]
        ]

        processed_export_file = f"theoretical_piechart_monthwise_data_{month_safe}_{timestamp}.xlsx"

        with pd.ExcelWriter(processed_export_file, engine="openpyxl") as writer:
            nrm_df.to_excel(writer, sheet_name='NRM_Database_Monthwise', index=False)
            pd.DataFrame(load_table_data[1:], columns=load_table_data[0]).to_excel(writer, sheet_name='Load Table',
                                                                                   index=False)

        logger.info(f"MONTHWISE load calculation and export completed: {processed_export_file}")

        return processed_export_file

    except Exception as e:
        logger.info(f"Error processing monthwise load database comparison: {e}")
        raise


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def complete_comparison_file(chart_file, processed_file, month_info):
    """Enhanced load comparison with color coding"""
    logger.info("Starting load comparison function...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = f"complete_validation_report_piechart_{month_safe}_{timestamp}.xlsx"

        chart_df = pd.read_excel(chart_file, sheet_name="Load Table")
        processed_df = pd.read_excel(processed_file, sheet_name="Load Table")

        logger.info(f"Chart data shape: {chart_df.shape}")
        logger.info(f"Processed data shape: {processed_df.shape}")

        comparison_df = pd.DataFrame()
        comparison_df['Parameter'] = processed_df['Parameter']
        comparison_df['Processed Value'] = processed_df['Value']
        comparison_df['Chart Value'] = chart_df['Value']

        diff_list = []
        match_list = []
        validation_results = {}

        for idx, (param, p_val, c_val) in enumerate(zip(comparison_df['Parameter'],
                                                        comparison_df['Processed Value'],
                                                        comparison_df['Chart Value'])):
            try:
                p_float = float(p_val)
                c_float = float(c_val)
                diff = round(abs(p_float - c_float), 1)

                if diff <= 0.1:
                    diff_list.append(diff)
                    match_list.append("YES")
                    validation_results[param] = {'match': True}
                else:
                    diff_list.append(diff)
                    match_list.append("NO")
                    validation_results[param] = {'match': False}

                logger.info(f"Parameter '{param}': Match = {diff <= 0.1}")

            except:
                if str(p_val).strip() == str(c_val).strip():
                    diff_list.append(0)
                    match_list.append("YES")
                    validation_results[param] = {'match': True}
                else:
                    diff_list.append("NOT A MATCH")
                    match_list.append("NO")
                    validation_results[param] = {'match': False}

                logger.info(f"Parameter '{param}': String match = {str(p_val).strip() == str(c_val).strip()}")

        comparison_df['Difference'] = diff_list
        comparison_df['Match'] = match_list

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name="Load Table Comparison", index=False)

        wb = load_workbook(output_file)
        ws = wb["Load Table Comparison"]

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for row in range(2, ws.max_row + 1):
            diff_cell = ws.cell(row=row, column=4)
            match_cell = ws.cell(row=row, column=5)

            if match_cell.value == "YES":
                diff_cell.fill = green_fill
                match_cell.fill = green_fill
            else:
                diff_cell.fill = red_fill
                match_cell.fill = red_fill

        wb.save(output_file)

        logger.info(f"Load comparison report created: {output_file}")
        return output_file, validation_results

    except Exception as e:
        logger.info(f"Error in load comparison function: {e}")
        return None, None


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_summary_report(config, month_info, chart_file, processed_file, comparison_file, validation_results, nrm_df,
                          meter_name):
    """Create enhanced professional summary report for monthly load validation"""
    logger.info("Creating enhanced monthly load summary report with professional formatting...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_LOAD_MONTHWISE_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_alignment = Alignment(horizontal="center", vertical="center")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # MAIN HEADER
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"LV LOAD/PIE CHART MONTHWISE VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        current_row += 1

        # TEST DETAILS SECTION
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # SYSTEM UNDER TEST
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Load/Pie Chart Monthwise (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # DATA VOLUME ANALYSIS
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        total_chart_points = 0
        try:
            chart_df = pd.read_excel(chart_file, sheet_name='Load Table')
            total_chart_points = len(chart_df)
        except:
            total_chart_points = 6

        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * 96
        data_completeness = (len(nrm_df) / expected_records * 100) if expected_records > 0 else 0

        data_rows = [
            ["Raw Database Records", len(nrm_df), "COMPLETE RECORDS" if len(nrm_df) > 0 else "NO DATA"],
            ["Chart Data Points", total_chart_points, "COMPLETE RECORDS"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"]
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in status or "%" in status:
                if data_completeness >= 90 or "COMPLETE RECORDS" in status:
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = warning_font
                    ws[f'C{current_row}'].fill = warning_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # VALIDATION RESULTS
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        validation_headers = ["Load Parameter", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        overall_passed = 0
        overall_total = 0

        for param, result in validation_results.items():
            match_status = result.get('match', False)

            ws[f'A{current_row}'].value = param
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = 1 if match_status else 0
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = 0 if match_status else 1
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            success_rate = "100%" if match_status else "0%"
            ws[f'D{current_row}'].value = success_rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            status = "PASS" if match_status else "FAIL"
            ws[f'E{current_row}'].value = status
            if match_status:
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

            overall_total += 1
            if match_status:
                overall_passed += 1

        current_row += 1

        # OVERALL ASSESSMENT
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: Monthly load validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor discrepancies found - Review recommended"
            assessment_color = warning_fill
            assessment_font_color = warning_font
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        from openpyxl.utils import get_column_letter
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced monthly load summary report created: {summary_file}")

        logger.info("=" * 60)
        logger.info("MONTHLY LOAD VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"Data: Raw={len(nrm_df)}, Chart={total_chart_points}")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.info(f"Error creating summary report: {e}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_lv_load_piechart_monthwise_automation():
    """Main LV Load/Pie Chart Monthwise automation process"""
    config = None
    driver = None
    output_folder = None
    month_info = None
    chart_file = None
    processed_file = None
    comparison_file = None
    summary_report = None
    raw_df = None
    meter_name = None
    validation_results = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False, None, None, None, None, None, None, None, None, None

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False, config, None, None, None, None, None, None, None, None

        # Apply configuration
        logger.info("Applying LV Load/Pie Chart Monthwise configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False, config, month_info, None, None, None, None, None, None, None

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False, config, month_info, None, None, None, None, None, None, None

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        dt_id, meter_name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False, config, month_info, None, None, None, None, None, None, None

        logger.info(f"Meter found: {meter_name} (ID: {mtr_id})")
        node_id = dt_id

        # Find and click View using search
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False, config, month_info, None, None, None, None, None, meter_name, None

        # Wait for overview page to load
        time.sleep(5)

        # Collect overview data (stays on overview - no detailed view needed)
        logger.info("Collecting monthly load overview data from UI...")
        overview_data = collect_over_data(driver, config['meter_type'])

        # Save overview data
        chart_file = save_chart_data_to_excel(month_info, overview_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get database data for complete month
        raw_df, rating_value = get_database_data_for_load_monthwise(month_info, meter_name, config['meter_serial_no'],
                                                                    node_id, config['meter_type'])

        if raw_df.empty:
            logger.info("No database data found for the month")
            return False, config, month_info, chart_file, None, None, None, None, meter_name, None

        # Process database calculations
        logger.info("Processing database calculations...")
        processed_file = process_load_database_values_monthwise(raw_df, rating_value, month_info)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison report
        logger.info("Creating validation comparison...")
        comparison_file, validation_results = complete_comparison_file(chart_file, processed_file, month_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive summary...")
        summary_report = create_summary_report(
            config, month_info, chart_file, processed_file,
            comparison_file, validation_results, raw_df, meter_name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("LV LOAD/PIE CHART MONTHWISE AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: LV Load/Pie Chart Monthwise (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({meter_name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ LV Load/Pie Chart Monthwise monitoring (fixed)")
        logger.info("   ✓ Complete month data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Tooltip-based load duration extraction")
        logger.info("   ✓ Load distribution across 4 ranges")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True, config, month_info, chart_file, processed_file, comparison_file, summary_report, raw_df, meter_name, validation_results

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"LV Load/Pie Chart Monthwise Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False, config, month_info, chart_file, processed_file, comparison_file, summary_report, raw_df, meter_name, validation_results

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass



# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
# ============================================================================
# HELPER FUNCTIONS FOR MAIN
# ============================================================================
def display_startup_banner():
    """Display startup banner with script information"""
    print("\n" + "=" * 80)
    print("🚀 LV LOAD/PIE CHART MONTHWISE VALIDATION AUTOMATION")
    print("=" * 80)
    print(f"📋 Test Engineer: {TestEngineer.NAME}")
    print(f"🏢 Department: {TestEngineer.DEPARTMENT}")
    print(f"📊 Analysis Type: Complete Month Load Distribution")
    print(f"🗄️ Database Tenant: {DatabaseConfig.TENANT_NAME}")
    print(f"⚙️ Load Ranges: <30%, 30-60%, 60-80%, >80%")
    print("=" * 80 + "\n")


def display_final_summary(success, total_time, output_folder, config, month_info,
                          chart_file, processed_file, comparison_file, summary_report,
                          raw_df, meter_name, validation_results):
    """Display comprehensive final summary"""
    print("\n" + "=" * 80)
    if success:
        print("✅ LV LOAD/PIE CHART MONTHWISE AUTOMATION COMPLETED SUCCESSFULLY!")
    else:
        print("❌ LV LOAD/PIE CHART MONTHWISE AUTOMATION FAILED")
    print("=" * 80)

    # Execution Time
    print(f"\n⏱️  EXECUTION TIME: {total_time:.2f}s ({total_time / 60:.1f} minutes)")

    # Test Details
    print(f"\n📋 TEST DETAILS:")
    print(f"   • Test Engineer: {TestEngineer.NAME}")
    print(f"   • Designation: {TestEngineer.DESIGNATION}")
    print(f"   • Department: {TestEngineer.DEPARTMENT}")
    print(f"   • Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    if success and config:
        # Configuration Used
        print(f"\n🔧 CONFIGURATION USED:")
        print(f"   • Monitoring Type: LV Load/Pie Chart Monthwise")
        print(f"   • Area: {config['area']}")
        print(f"   • Substation: {config['substation']}")
        print(f"   • Feeder: {config['feeder']}")
        print(f"   • Month: {config['target_month_year']}")
        print(f"   • Meter: {config['meter_serial_no']} ({meter_name})")
        print(f"   • Meter Type: {config['meter_type']}")

        # Database Configuration
        print(f"\n🗄️  DATABASE CONFIGURATION:")
        print(f"   • Host: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}")
        print(f"   • Database: {DatabaseConfig.DB1_DATABASE}")
        print(f"   • Tenant: {DatabaseConfig.TENANT_NAME}")

        # Data Analysis
        if raw_df is not None and not raw_df.empty:
            days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
            expected_records = days_in_month * 96
            data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0

            print(f"\n📊 DATA ANALYSIS:")
            print(f"   • Database Records: {len(raw_df):,}")
            print(f"   • Expected Records: {expected_records:,}")
            print(f"   • Data Completeness: {data_completeness:.1f}%")
            print(f"   • Month Coverage: {days_in_month} days")
            print(f"   • Date Range: {month_info['start_date']} to {month_info['end_date']}")

        # Validation Results
        if validation_results:
            passed = sum(1 for r in validation_results.values() if r.get('match', False))
            total = len(validation_results)
            success_rate = (passed / total * 100) if total > 0 else 0

            print(f"\n✅ VALIDATION RESULTS:")
            print(f"   • Total Parameters: {total}")
            print(f"   • Passed: {passed}")
            print(f"   • Failed: {total - passed}")
            print(f"   • Success Rate: {success_rate:.1f}%")

            if success_rate >= 95:
                print(f"   • Assessment: 🌟 EXCELLENT - All validations passed!")
            elif success_rate >= 80:
                print(f"   • Assessment: ✅ GOOD - Minor discrepancies found")
            else:
                print(f"   • Assessment: ⚠️  NEEDS ATTENTION - Review required")

        # Generated Files
        print(f"\n📁 GENERATED FILES:")
        files_generated = []
        if chart_file and os.path.exists(chart_file):
            files_generated.append(f"   1. {os.path.basename(chart_file)}")
        if processed_file and os.path.exists(processed_file):
            files_generated.append(f"   2. {os.path.basename(processed_file)}")
        if comparison_file and os.path.exists(comparison_file):
            files_generated.append(f"   3. {os.path.basename(comparison_file)}")
        if summary_report and os.path.exists(summary_report):
            files_generated.append(f"   4. {os.path.basename(summary_report)}")

        for file_info in files_generated:
            print(file_info)

        print(f"\n📂 Output Folder: {os.path.abspath(output_folder)}")

        # Features Applied
        print(f"\n🎯 KEY FEATURES APPLIED:")
        print(f"   ✓ LV Load/Pie Chart Monthwise monitoring")
        print(f"   ✓ Complete month data processing")
        print(f"   ✓ Search box meter selection")
        print(f"   ✓ Tooltip-based load duration extraction")
        print(f"   ✓ Load distribution: <30%, 30-60%, 60-80%, >80%")
        print(f"   ✓ Rating-based calculations (KVA/Ampacity)")
        print(f"   ✓ Load factor validation")
        print(f"   ✓ Centralized database configuration")
        print(f"   ✓ Test engineer details in reports")
        print(f"   ✓ Enhanced comparison with color coding")
        print(f"   ✓ Professional summary report")

        # Load Insights
        print(f"\n💡 LOAD DISTRIBUTION INSIGHTS:")
        print(f"   • Load <30%: Light loading - Potential for increased capacity")
        print(f"   • Load 30-60%: Moderate loading - Normal operational range")
        print(f"   • Load 60-80%: Heavy loading - Monitor for efficiency")
        print(f"   • Load >80%: Critical loading - Requires immediate attention")

    else:
        print("\n❌ AUTOMATION FAILED")
        print("   • Check configuration file: user_config.xlsx")
        print("   • Review error logs in output folder")
        print("   • Verify database connectivity")
        print("   • Ensure browser driver is up to date")

    print("\n" + "=" * 80)
    print("🏁 LV LOAD/PIE CHART MONTHWISE AUTOMATION FINISHED")
    print("=" * 80 + "\n")


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    # Display startup banner
    display_startup_banner()

    logger.info("=" * 60)
    logger.info("LV LOAD/PIE CHART MONTHWISE AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Department: {TestEngineer.DEPARTMENT}")
    logger.info(f"Monitoring Type: LV Load/Pie Chart Monthwise (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ LV Load/Pie Chart Monthwise monitoring only")
    logger.info("   ✓ Complete month data processing")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Tooltip-based load duration extraction")
    logger.info("   ✓ Load distribution: <30%, 30-60%, 60-80%, >80%")
    logger.info("   ✓ Rating-based load calculations")
    logger.info("   ✓ Load factor validation")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    # Initialize variables for final summary
    config = None
    month_info = None
    chart_file = None
    processed_file = None
    comparison_file = None
    summary_report = None
    raw_df = None
    meter_name = None
    validation_results = None
    output_folder = None

    # Execute automation
    start_time = time.time()

    try:
        print("🔄 Starting automation process...\n")

        # Execute main automation and capture all return values
        result = main_lv_load_piechart_monthwise_automation()

        # Unpack results
        success, config, month_info, chart_file, processed_file, comparison_file, summary_report, raw_df, meter_name, validation_results = result

        # Set output folder
        if os.path.exists('output_files'):
            output_folder = 'output_files'

    except Exception as e:
        logger.error(f"Critical error in main execution: {str(e)}")
        print(f"\n❌ CRITICAL ERROR: {str(e)}")
        success = False
        config = None
        month_info = None
        chart_file = None
        processed_file = None
        comparison_file = None
        summary_report = None
        raw_df = None
        meter_name = None
        validation_results = None

    end_time = time.time()
    total_time = end_time - start_time

    # Log completion
    logger.info("=" * 60)
    if success:
        logger.info("LV LOAD/PIE CHART MONTHWISE AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ LV Load/Pie Chart Monthwise monitoring (fixed)")
        logger.info("   ✓ Complete month processing")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Tooltip extraction for load durations")
        logger.info("   ✓ Load distribution validation")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("LV LOAD/PIE CHART MONTHWISE AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("LV Load/Pie Chart Monthwise Automation Finished")
    logger.info("=" * 60)

    # Display comprehensive final summary
    display_final_summary(
        success=success,
        total_time=total_time,
        output_folder=output_folder if output_folder else 'output_files',
        config=config,
        month_info=month_info,
        chart_file=chart_file,
        processed_file=processed_file,
        comparison_file=comparison_file,
        summary_report=summary_report,
        raw_df=raw_df,
        meter_name=meter_name,
        validation_results=validation_results
    )

    # Exit with appropriate code
    import sys

    sys.exit(0 if success else 1)
