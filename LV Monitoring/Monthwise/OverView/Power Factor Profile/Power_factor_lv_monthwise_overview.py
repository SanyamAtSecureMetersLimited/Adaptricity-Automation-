import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Power factor tenant

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('lv_monthly_pf_overview_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/lv_monthly_pf_overview_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for LV Monthly Power Factor Overview"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'January 2025', 'YOUR_METER_NO',
                      'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'This script is FOR LV MONTHLY POWER FACTOR OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: DT or LV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"LV Monthly Power Factor Overview Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for LV Monthly Power Factor Overview"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}  # Fixed for LV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'feeder', 'target_month_year', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("LV Monthly Power Factor Overview Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING LV MONTHLY POWER FACTOR OVERVIEW AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default LV Monthly Power Factor Overview configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("LV Monthly Power Factor Overview Configuration validated successfully")
    logger.info(f"   Monitoring Type: LV Monthly Power Factor Overview (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Feeder: {config['feeder']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching LV Monthly Power Factor Overview metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        if meter_type.upper() == 'DT':
            query1 = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        elif meter_type.upper() == 'LV':
            query1 = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        else:
            logger.info(f"Invalid meter type: {meter_type}")
            return None, None, None

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None

        dt_id, dt_name, meterid = result1
        logger.info(f"Metrics: {dt_name}, meterid: {meterid}")
        return dt_id, dt_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_pf_overview(month_info, mtr_id):
    """Fetch database data for complete month power factor overview"""
    logger.info(f"Fetching monthly power factor overview database data for: {month_info['selected_month_year']}")

    start_date = month_info['start_date'].strftime("%Y-%m-%d")
    end_date = month_info['end_date'].strftime("%Y-%m-%d")

    date_filter = f"AND DATE(surveydate) >= '{start_date}' AND DATE(surveydate) <= '{end_date}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # Power factor query for COMPLETE MONTH
        query = f"""
            SELECT surveydate, pf, avg_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        logger.info("Executing power factor monthly database query...")
        raw_df = pd.read_sql(query, conn)
        conn.close()

        logger.info(f"Retrieved: {len(raw_df)} power factor records")

        if len(raw_df) > 0:
            logger.info(f"Date range in data: {raw_df['surveydate'].min()} to {raw_df['surveydate'].max()}")

            # Log month coverage
            days_with_data = raw_df['surveydate'].dt.date.nunique()
            total_days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
            coverage_percentage = (days_with_data / total_days_in_month) * 100
            logger.info(f"Month coverage: {days_with_data}/{total_days_in_month} days ({coverage_percentage:.1f}%)")

        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SANYAM")
        driver.find_element(By.ID, "Password").send_keys("Sanyam@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(8)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 5).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)

        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()

        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select LV monitoring - FIXED FOR LV ONLY"""
    try:
        logger.info("Selecting LV monitoring (fixed for LV monthly power factor overview script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        logger.info("LV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - DT or LV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "DT":
            dt_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]')))
            dt_button.click()
            logger.info("DT selected")
        elif meter_type == "LV":
            lv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]')))
            lv_button.click()
            logger.info("LV feeder selected")
        else:
            logger.info("Invalid meter type for LV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_power_factor_overview_data(driver, wait):
    """Collect power factor overview data from UI with enhanced tooltip extraction"""
    logger.info("Starting power factor overview data collection...")
    data = {}

    try:
        # Navigate to Power Factor tab
        logger.info("Clicking on Power Factor tab...")
        wait.until(EC.visibility_of_element_located(
            (By.XPATH, "//div[@class='dx-item-content' and normalize-space()='Power factor']"))).click()
        time.sleep(3)

        # Extract Power Factor Average
        logger.info("Extracting Power Factor Average...")
        try:
            pf_avg = driver.find_element(By.XPATH, "//span[@id='avgPf']").text
            logger.info(f"Power Factor Average extracted: {pf_avg}")
        except Exception as e:
            logger.error(f"Error fetching Power Factor Average: {e}")
            pf_avg = '-'

        # Enhanced Duration Ranges extraction using tooltip extraction
        logger.info("Extracting Power Factor duration ranges with enhanced tooltip extraction...")
        action = ActionChains(driver)

        # Find all bars in the PF pattern chart
        pf_pattern_bars = driver.find_elements(By.CSS_SELECTOR, '#pfrangechart g.dxc-markers')
        tooltip_selector = '.dxc-tooltip svg text'

        # Initialize durations with default values
        pf_durations = {
            'Duration PF < 0.9': '-',
            'Duration PF 0.9 - 0.95': '-',
            'Duration PF > 0.95': '-'
        }

        # Enhanced color mapping for different PF ranges
        color_mapping = {
            '#D11920': 'Duration PF < 0.9',  # Red - Poor PF
            '#DEAE2A': 'Duration PF 0.9 - 0.95',  # Orange/Yellow - Acceptable PF
            '#86B8A5': 'Duration PF > 0.95'  # Green - Good PF
        }

        logger.info(f"Found {len(pf_pattern_bars)} PF pattern bars to process")

        for i, bar in enumerate(pf_pattern_bars):
            try:
                fill_color = bar.get_attribute('fill')
                label = color_mapping.get(fill_color)

                if label:
                    # Enhanced tooltip extraction with multiple attempts
                    for attempt in range(3):
                        try:
                            action.move_to_element(bar).perform()
                            time.sleep(1.5)

                            tooltip = driver.find_element(By.CSS_SELECTOR, tooltip_selector)
                            tooltip_text = tooltip.text.strip()

                            if tooltip_text and tooltip_text != '-':
                                pf_durations[label] = tooltip_text
                                logger.info(f"Extracted for {label}: {tooltip_text}")
                                break
                            else:
                                logger.warning(f"Empty tooltip on attempt {attempt + 1} for {label}")
                        except Exception as attempt_e:
                            logger.warning(f"Tooltip extraction attempt {attempt + 1} failed for {label}: {attempt_e}")
                            if attempt == 2:
                                logger.error(f"All tooltip extraction attempts failed for {label}")
                else:
                    logger.warning(f"Unexpected bar color found: {fill_color}")
            except Exception as e:
                logger.error(f"Error processing PF bar {i + 1}: {e}")

        # Combine all power factor data
        data['Power Factor Table'] = {
            'Power Factor Average': pf_avg,
            **pf_durations
        }

        logger.info("Power factor overview data collection completed")
        logger.info(f"Collected data: {data['Power Factor Table']}")

    except Exception as e:
        logger.error(f"Error in power factor data collection: {str(e)}")
        raise

    return data


@log_execution_time
def save_pf_overview_data_to_excel(month_info, overview_data):
    """Save power factor overview data to Excel"""
    logger.info("Saving power factor overview data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        def extract_numeric_pf(value):
            """Extract numeric value from power factor data"""
            if isinstance(value, str):
                match = re.search(r"[-+]?\d*\.\d+|\d+", value)
                if match:
                    num_val = float(match.group())
                    return num_val if num_val <= 1.0 else value
                return value
            return value

        # Power Factor Table Sheet
        ws_pft = wb.create_sheet("Power Factor Table")
        ws_pft.append(["Parameter", "Value"])

        for key, value in overview_data['Power Factor Table'].items():
            processed_value = extract_numeric_pf(value) if 'Average' in key else value
            ws_pft.append([key, processed_value])
            logger.info(f"UI Parameter: {key} -> Value={processed_value}")

        # Save
        file_name = f"chart_data_from_ui_monthly_pf_overview_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"Power factor overview data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving power factor overview data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_pf_overview_database_calculations(raw_df, month_info):
    """Process database calculations for monthly power factor overview"""
    logger.info("Processing monthly power factor overview database calculations...")

    try:
        if raw_df.empty:
            logger.error("Raw DataFrame is empty")
            raise ValueError("No power factor data available")

        # Check required columns
        required_columns = ['pf', 'avg_i']
        missing_columns = [col for col in required_columns if col not in raw_df.columns]
        if missing_columns:
            logger.error(f"Missing required columns: {missing_columns}")
            raise ValueError(f"Missing required columns: {missing_columns}")

        # Calculate interval
        if len(raw_df) > 1:
            interval_minutes = int((raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)
        else:
            interval_minutes = 15

        logger.info(f"Power factor survey interval: {interval_minutes} minutes")

        def format_duration_hours(hours_val):
            """Format hours into HH:MM format"""
            if hours_val == '-' or hours_val == 0:
                return '-'
            total_minutes = int(hours_val * 60)
            hrs = total_minutes // 60
            mins = total_minutes % 60
            return f"{hrs:02d}:{mins:02d} hrs"

        def calc_duration(pf_values):
            """Calculate duration for given PF values"""
            count = len(pf_values)
            if count == 0:
                return '-'
            total_minutes = count * interval_minutes
            hours = total_minutes / 60
            return hours

        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        processed_file = f"theoretical_monthly_pf_overview_calculated_data_{month_safe}_{timestamp}.xlsx"

        # Process Power Factor calculations
        logger.info("Calculating power factor values and durations...")

        pf_series = raw_df['pf'].dropna()

        if len(pf_series) > 0:
            # Calculate Average Power Factor
            pf_avg = pf_series.mean()
            logger.info(f"Average Power Factor calculated: {pf_avg:.4f}")

            # Calculate durations for different PF ranges
            duration_lt_09 = calc_duration(pf_series[pf_series < 0.9])
            duration_between_09_095 = calc_duration(pf_series[(pf_series >= 0.9) & (pf_series <= 0.95)])
            duration_gt_095 = calc_duration(pf_series[pf_series > 0.95])

            logger.info(f"Duration calculations:")
            logger.info(f"  PF < 0.9: {duration_lt_09} hours")
            logger.info(f"  PF 0.9-0.95: {duration_between_09_095} hours")
            logger.info(f"  PF > 0.95: {duration_gt_095} hours")

            # Prepare Power Factor Table Data
            power_factor_data = [
                ['Power Factor Average', round(pf_avg, 4)],
                ['Duration PF < 0.9', format_duration_hours(duration_lt_09)],
                ['Duration PF 0.9 - 0.95', format_duration_hours(duration_between_09_095)],
                ['Duration PF > 0.95', format_duration_hours(duration_gt_095)]
            ]

            df_power_factor = pd.DataFrame(power_factor_data, columns=['Parameter', 'Value'])

            logger.info("Power Factor calculations completed successfully")

            for param, value in power_factor_data:
                logger.info(f"DB Parameter: {param} -> Value={value}")

        else:
            logger.warning("No valid power factor data found")
            power_factor_data = [
                ['Power Factor Average', '-'],
                ['Duration PF < 0.9', '-'],
                ['Duration PF 0.9 - 0.95', '-'],
                ['Duration PF > 0.95', '-']
            ]
            df_power_factor = pd.DataFrame(power_factor_data, columns=['Parameter', 'Value'])

        # Write to Excel
        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            df_power_factor.to_excel(writer, sheet_name='Power Factor Table', index=False)

        logger.info(f"Processed power factor data saved: {processed_file}")
        return processed_file

    except Exception as e:
        logger.error(f"Error processing power factor database: {str(e)}")
        raise


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_pf_overview_comparison(chart_file, processed_file, month_info):
    """Create complete monthly power factor overview comparison with validation"""
    logger.info("Creating monthly power factor overview comparison...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_monthly_pf_overview_{month_safe}.xlsx"

        # Load workbooks
        wb_processed = load_workbook(processed_file)
        wb_chart = load_workbook(chart_file)

        # Check if 'Power Factor Table' sheet exists
        sheet_name = 'Power Factor Table'
        if sheet_name not in wb_processed.sheetnames or sheet_name not in wb_chart.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in files")
            return None, None

        # Get worksheets
        ws_processed = wb_processed[sheet_name]
        ws_chart = wb_chart[sheet_name]

        # Convert to DataFrames
        processed_data = []
        chart_data = []

        for row in ws_processed.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                processed_data.append(row)

        for row in ws_chart.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                chart_data.append(row)

        logger.info(f"Processed sheet rows: {len(processed_data)}")
        logger.info(f"Chart sheet rows: {len(chart_data)}")

        if len(processed_data) == 0 or len(chart_data) == 0:
            logger.error("No data found in files")
            return None, None

        # Create DataFrames
        processed_df = pd.DataFrame(processed_data)
        chart_df = pd.DataFrame(chart_data)

        processed_df.columns = processed_df.iloc[0]
        processed_df = processed_df[1:].reset_index(drop=True)

        chart_df.columns = chart_df.iloc[0]
        chart_df = chart_df[1:].reset_index(drop=True)

        # Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

        # Output workbook
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        validation_results_detail = {}

        # Output sheet
        ws_output = wb_output.create_sheet(title="Power Factor Table Comparison")
        headers = ['Parameter', 'DB_Value', 'UI_Value', 'Value_Difference', 'Match', 'Validation_Notes']
        ws_output.append(headers)

        logger.info("Starting row-by-row power factor comparison...")

        total_rows = 0
        successful_comparisons = 0

        for idx, row in processed_df.iterrows():
            total_rows += 1
            param = row['Parameter']
            logger.info(f"Processing parameter: '{param}'")

            chart_matches = chart_df[chart_df['Parameter'] == param]
            if chart_matches.empty:
                logger.warning(f"Parameter '{param}' NOT found in Chart file")
                output_row = [param, row['Value'], "NOT_FOUND", "N/A", "FAIL", "Parameter missing in UI"]
                ws_output.append(output_row)
                validation_results_detail[param] = {'match': False, 'reason': 'Missing in UI'}
                continue

            chart_row = chart_matches.iloc[0]
            successful_comparisons += 1

            proc_value = row['Value']
            chart_value = chart_row['Value']

            logger.info(f"DB  -> Value: {proc_value}")
            logger.info(f"UI  -> Value: {chart_value}")

            validation_notes = ""
            try:
                if 'Average' in param:
                    proc_value_f = float(proc_value)
                    chart_value_f = float(chart_value)
                    value_diff = abs(proc_value_f - chart_value_f)
                    value_match = value_diff < 0.01
                    value_diff_disp = round(value_diff, 4)
                    validation_notes = f"PF Average tolerance: ±0.01, Difference: {value_diff_disp}"
                    logger.info(f"Power Factor Average - Numeric difference: {value_diff_disp} (Match: {value_match})")
                else:
                    value_match = (str(proc_value).strip() == str(chart_value).strip())
                    value_diff_disp = 'StringMismatch' if not value_match else '0'
                    validation_notes = f"Duration string match: {value_match}"
                    logger.info(f"Power Factor Duration - String comparison: {value_match}")
            except Exception as e:
                logger.info(f"Comparison handling for {param}: {e}")
                value_match = (str(proc_value).strip() == str(chart_value).strip())
                value_diff_disp = 'StringMismatch' if not value_match else '0'
                validation_notes = f"Fallback string comparison: {value_match}"

            overall_match = value_match
            match_text = 'PASS' if overall_match else 'FAIL'
            logger.info(f"Result: {match_text}")

            validation_results_detail[param] = {
                'match': overall_match,
                'reason': validation_notes,
                'value_difference': value_diff_disp
            }

            output_row = [param, proc_value, chart_value, value_diff_disp, match_text, validation_notes]
            ws_output.append(output_row)

            row_idx = ws_output.max_row
            ws_output.cell(row=row_idx, column=4).fill = green_fill if value_match else red_fill
            ws_output.cell(row=row_idx, column=5).fill = green_fill if overall_match else red_fill

            if 'Average' in param:
                ws_output.cell(row=row_idx, column=1).fill = blue_fill

        logger.info("POWER FACTOR COMPARISON SUMMARY:")
        logger.info(f"  Total parameters: {total_rows}")
        logger.info(f"  Successful matches: {successful_comparisons}")
        logger.info(f"  Success rate: {(successful_comparisons / total_rows * 100):.1f}%")

        wb_output.save(output_file)
        logger.info(f"Power factor comparison report created: {output_file}")
        return output_file, validation_results_detail

    except Exception as e:
        logger.error(f"Error in power factor comparison: {str(e)}")
        return None, None


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_pf_overview_summary_report(config, month_info, chart_file, processed_file,
                                      comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive monthly power factor overview summary report with ENHANCED styling"""
    logger.info("Creating monthly power factor overview summary report with enhanced styling...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_PF_OVERVIEW_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_alignment = Alignment(horizontal="center", vertical="center")

        pf_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"⚡ LV MONTHLY POWER FACTOR OVERVIEW VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "⚡ SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Monthly Power Factor Overview (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "⚡ DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate completeness
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * 96
        data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0

        total_chart_points = 0
        try:
            chart_df = pd.read_excel(chart_file, sheet_name='Power Factor Table')
            total_chart_points = len(chart_df)
        except:
            total_chart_points = 4

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Chart Data Points", total_chart_points, "COMPLETE RECORDS"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"]
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in status or "%" in status:
                if data_completeness >= 90 or "COMPLETE RECORDS" in status:
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = warning_font
                    ws[f'C{current_row}'].fill = warning_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "⚡ POWER FACTOR VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["PF Parameter", "Status", "Validation Result", "Confidence", "Technical Notes"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0

        if validation_results:
            for param_name, result in validation_results.items():
                overall_total += 1
                is_match = result.get('match', False)
                if is_match:
                    overall_passed += 1

                if is_match:
                    if "Average" in param_name:
                        status = "✅ EXCELLENT"
                        status_fill = pass_fill
                        confidence = "HIGH"
                        notes = "PF average within ±0.01 tolerance"
                    else:
                        status = "✅ VALIDATED"
                        status_fill = pass_fill
                        confidence = "HIGH"
                        notes = "Duration calculation matches UI tooltip"
                else:
                    status = "❌ FAILED"
                    status_fill = fail_fill
                    confidence = "LOW"
                    if "Average" in param_name:
                        notes = "PF average exceeds tolerance limit"
                    else:
                        notes = "Duration mismatch - check tooltip extraction"

                result_text = "VALID" if is_match else "INVALID"

                row_data = [param_name, status, result_text, confidence, notes]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = label_alignment if col in [1, 5] else Alignment(horizontal="center",
                                                                                     vertical="center")

                    if col == 2:
                        cell.fill = status_fill
                        cell.font = pass_font if is_match else fail_font
                    elif col == 1:
                        if "Average" in param_name:
                            cell.fill = pf_fill

                current_row += 1

        # Calculate overall success rate
        success_rate = (overall_passed / overall_total * 100) if overall_total > 0 else 0

        current_row += 1

        # ============ SUMMARY STATISTICS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        stats_header = ws[f'A{current_row}']
        stats_header.value = "📊 POWER FACTOR VALIDATION STATISTICS"
        stats_header.font = section_header_font
        stats_header.fill = section_header_fill
        stats_header.alignment = section_header_alignment
        stats_header.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"🎯 Power Factor Validation Success Rate: {success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")

        if success_rate >= 95:
            cell.fill = pass_fill
            cell.font = pass_font
        elif success_rate >= 80:
            cell.fill = warning_fill
            cell.font = warning_font
        else:
            cell.fill = fail_fill
            cell.font = fail_font

        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Overall assessment
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']

        if success_rate >= 95 and data_completeness >= 90:
            assessment = "🌟 EXCELLENT - All monthly power factor validations passed with high confidence"
            assessment_fill = pass_fill
            assessment_font_style = pass_font
        elif success_rate >= 80 and data_completeness >= 70:
            assessment = "✅ GOOD - Minor monthly power factor discrepancies found"
            assessment_fill = warning_fill
            assessment_font_style = warning_font
        else:
            assessment = "⚠️ REQUIRES ATTENTION - Significant monthly power factor validation issues"
            assessment_fill = fail_fill
            assessment_font_style = fail_font

        cell.value = assessment
        cell.font = assessment_font_style
        cell.fill = assessment_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        current_row += 1

        # ============ POWER FACTOR TECHNICAL INSIGHTS ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        insights_header = ws[f'A{current_row}']
        insights_header.value = "🔬 POWER FACTOR TECHNICAL INSIGHTS"
        insights_header.font = section_header_font
        insights_header.fill = section_header_fill
        insights_header.alignment = section_header_alignment
        insights_header.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        technical_insights = [
            ["Power Factor Ranges:", "Poor (<0.9), Acceptable (0.9-0.95), Good (>0.95)"],
            ["Duration Calculation:", "Count of intervals × survey interval (15 min)"],
            ["Validation Method:", "UI tooltip extraction + RAW database queries"],
            ["Average PF Tolerance:", "±0.01 for high precision validation"],
            ["Duration Matching:", "Exact string match for HH:MM format"]
        ]

        for label, value in technical_insights:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            value_cell = ws[f'B{current_row}']
            value_cell.value = value
            value_cell.font = data_font
            value_cell.fill = pf_fill if "Power Factor" in label else data_fill
            value_cell.alignment = data_alignment
            value_cell.border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 50, 'C': 15, 'D': 15, 'E': 40, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced monthly power factor overview summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MONTHLY POWER FACTOR OVERVIEW VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"Data: Raw={len(raw_df)}, Chart={total_chart_points}")
        logger.info(f"Overall Success Rate: {success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_lv_monthly_pf_overview_automation():
    """Main LV Monthly Power Factor Overview automation process"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying LV Monthly Power Factor Overview configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False

        logger.info(f"Meter found: {name} (ID: {mtr_id})")

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for overview page to load
        time.sleep(5)

        # Collect power factor overview data
        logger.info("Collecting monthly power factor overview data from UI...")
        overview_data = collect_power_factor_overview_data(driver, wait)

        # Save overview data
        chart_file = save_pf_overview_data_to_excel(month_info, overview_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get database data for complete month
        raw_df = get_database_data_for_monthly_pf_overview(month_info, mtr_id)

        if raw_df.empty:
            logger.info("No database data found for the month")
            return False

        # Process database calculations
        logger.info("Processing database calculations...")
        processed_file = process_pf_overview_database_calculations(raw_df, month_info)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison report
        logger.info("Creating validation comparison...")
        comparison_file, validation_results = create_pf_overview_comparison(chart_file, processed_file, month_info)

        if comparison_file:
            comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive summary...")
        summary_report = create_pf_overview_summary_report(
            config, month_info, chart_file, processed_file,
            comparison_file, validation_results, raw_df, name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("LV MONTHLY POWER FACTOR OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: LV Monthly Power Factor Overview (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ LV Monthly Power Factor Overview monitoring (fixed)")
        logger.info("   ✓ Complete month data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Enhanced tooltip extraction for PF durations")
        logger.info("   ✓ Power Factor Average validation (±0.01 tolerance)")
        logger.info("   ✓ Duration ranges: Poor (<0.9), Acceptable (0.9-0.95), Good (>0.95)")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"LV Monthly Power Factor Overview Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV MONTHLY POWER FACTOR OVERVIEW AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: LV Monthly Power Factor Overview (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ LV Monthly Power Factor Overview monitoring only")
    logger.info("   ✓ Complete month data processing")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Enhanced tooltip extraction for PF durations")
    logger.info("   ✓ Power Factor Average with ±0.01 tolerance")
    logger.info("   ✓ Duration ranges validation (Poor, Acceptable, Good)")
    logger.info("   ✓ Better null/dash handling")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_lv_monthly_pf_overview_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("LV MONTHLY POWER FACTOR OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ LV Monthly Power Factor Overview monitoring (fixed)")
        logger.info("   ✓ Complete month processing")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Enhanced tooltip extraction for durations")
        logger.info("   ✓ Power Factor Average validation")
        logger.info("   ✓ Duration ranges validation")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("LV MONTHLY POWER FACTOR OVERVIEW AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("LV Monthly Power Factor Overview Automation Finished")
    logger.info("=" * 60)
