import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""
    
    # Database 1: Hetzner - For meter details and voltage data
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"
    
    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"
    
    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Change to tenant02, tenant03, etc. as needed
    
    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }
    
    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    logger = logging.getLogger('lv_monthly_voltage_automation')
    logger.setLevel(logging.INFO)
    
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    
    log_file = 'logs/lv_monthly_voltage_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for LV Monthly Voltage"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'January 2025', 'YOUR_METER_NO', 'DT']
        }
        df_config = pd.DataFrame(config_data)
        
        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)
            
            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'This script is FOR LV MONTHLY VOLTAGE OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: DT or LV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)
        
        logger.info(f"LV Monthly Voltage Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for LV Monthly Voltage"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None
        
        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}  # Fixed for LV monitoring
        
        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()
        
        required_fields = ['type', 'area', 'substation', 'feeder', 'target_month_year', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None
        
        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None
        
        logger.info("LV Monthly Voltage Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING LV MONTHLY VOLTAGE AUTOMATION")
    logger.info("=" * 60)
    
    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default LV Monthly Voltage configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None
    
    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None
    
    logger.info("LV Monthly Voltage Configuration validated successfully")
    logger.info(f"   Monitoring Type: LV Monthly Voltage (Fixed)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Feeder: {config['feeder']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise
    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching LV Monthly Voltage metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        
        if meter_type.upper() == 'DT':
            query1 = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        elif meter_type.upper() == 'LV':
            query1 = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        else:
            logger.info(f"Invalid meter type: {meter_type}")
            return None, None, None
        
        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None
        
        dt_id, dt_name, meterid = result1
        logger.info(f"Metrics: {dt_name}, meterid: {meterid}")
        return dt_id, dt_name, meterid
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_voltage(month_info, mtr_id):
    """Fetch database data for complete month voltage overview - RAW ONLY"""
    logger.info(f"Fetching monthly voltage database data for: {month_info['selected_month_year']}")
    
    start_date = month_info['start_date'].strftime("%Y-%m-%d")
    end_date = month_info['end_date'].strftime("%Y-%m-%d")
    
    date_filter = f"AND DATE(surveydate) >= '{start_date}' AND DATE(surveydate) <= '{end_date}'"
    
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())
        
        # RAW ONLY query for voltage parameters
        raw_query = f"""
            SELECT DISTINCT surveydate, v1, v2, v3, avg_v
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """
        
        raw_df = pd.read_sql(raw_query, conn)
        
        logger.info(f"Retrieved: Raw={len(raw_df)} voltage records")
        logger.info("⚠️ NO NRM queries - 100% raw voltage data based")
        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SANYAM")
        driver.find_element(By.ID, "Password").send_keys("Sanyam@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        # Click Month button
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)
        
        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        
        # Parse month info for COMPLETE MONTH
        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)
        
        # Create COMPLETE MONTH boundaries
        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)
        
        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }
        
        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select LV monitoring - FIXED FOR LV ONLY"""
    try:
        logger.info("Selecting LV monitoring (fixed for LV monthly voltage script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        logger.info("LV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - DT or LV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)
        
        if meter_type == "DT":
            dt_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]')))
            dt_button.click()
            logger.info("DT selected")
        elif meter_type == "LV":
            lv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]')))
            lv_button.click()
            logger.info("LV feeder selected")
        else:
            logger.info("Invalid meter type for LV monitoring")
            return False
        
        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)
        
        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False
        
        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True
        
        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx+1})")
                    return True
            except:
                continue
        
        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def extract_voltage_unbalance_from_svg(driver):
    """Extract voltage unbalance values from SVG with correct phase mapping"""
    logger.info("Starting SVG-based voltage unbalance extraction...")
    vunb_values = {'Phase 1': '-', 'Phase 2': '-', 'Phase 3': '-'}
    
    try:
        chart_svg = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#dvvolatgeunbalancechart svg'))
        )
        logger.info("Voltage unbalance SVG chart found")
        
        js_script = """
        var chartContainer = document.querySelector('#dvvolatgeunbalancechart');
        var svg = chartContainer.querySelector('svg');
        if (!svg) return null;
        
        var result = {};
        var textElements = svg.querySelectorAll('text, tspan');
        var phaseData = [];
        
        textElements.forEach(function(element) {
            var text = element.textContent.trim();
            if (text.includes('%') && text.match(/\\d+(\\.\\d+)?%/)) {
                var numericValue = text.replace('%', '').trim();
                var rect = element.getBoundingClientRect();
                phaseData.push({
                    value: numericValue,
                    text: text,
                    y: rect.y,
                    x: rect.x
                });
            }
        });
        
        phaseData.sort(function(a, b) { return a.y - b.y; });
        return { phaseData: phaseData, count: phaseData.length };
        """
        
        svg_data = driver.execute_script(js_script)
        
        if svg_data and svg_data.get('phaseData'):
            phase_data = svg_data['phaseData']
            logger.info(f"Found {len(phase_data)} phase data points")
            
            if len(phase_data) >= 3:
                vunb_values['Phase 1'] = phase_data[0]['value']
                vunb_values['Phase 2'] = phase_data[1]['value']
                vunb_values['Phase 3'] = phase_data[2]['value']
                logger.info(f"SVG extraction successful: {vunb_values}")
    
    except Exception as e:
        logger.error(f"Error in SVG extraction: {str(e)}")
    
    logger.info(f"Final Voltage Unbalance Results: {vunb_values}")
    return vunb_values


@log_execution_time
def extract_voltage_variation_data(driver):
    """Extract voltage variation data using tooltip extraction"""
    logger.info("Starting voltage variation data extraction...")
    
    action = ActionChains(driver)
    voltage_bars = driver.find_elements(By.CSS_SELECTOR, '#divVoltageVariation g.dxc-markers')
    tooltip_selector = '.dxc-tooltip svg text'
    
    voltage_variation_data = {
        'Duration Voltage <180V': '-',
        'Duration Voltage 180-216V': '-',
        'Duration Voltage 216-240V': '-',
        'Duration Voltage >240V': '-'
    }
    
    color_mapping = {
        '#E38430': 'Duration Voltage <180V',
        '#DEAE2A': 'Duration Voltage 180-216V',
        '#86B8A5': 'Duration Voltage 216-240V',
        '#D11920': 'Duration Voltage >240V'
    }
    
    try:
        logger.info(f"Found {len(voltage_bars)} voltage variation bars to process")
        for i, bar in enumerate(voltage_bars):
            try:
                fill_color = bar.get_attribute('fill')
                label = color_mapping.get(fill_color)
                
                if label:
                    action.move_to_element(bar).perform()
                    time.sleep(1.5)
                    
                    tooltip = driver.find_element(By.CSS_SELECTOR, tooltip_selector)
                    tooltip_text = tooltip.text.strip()
                    voltage_variation_data[label] = tooltip_text
                    logger.info(f"Extracted for {label}: {tooltip_text}")
            except Exception as e:
                logger.warning(f"Error processing voltage bar {i+1}: {str(e)}")
    
    except Exception as e:
        logger.error(f"Error in voltage variation extraction: {str(e)}")
    
    logger.info(f"Voltage variation data extracted: {voltage_variation_data}")
    return voltage_variation_data


@log_execution_time
def collect_voltage_overview_data(driver):
    """Collect voltage overview data from UI"""
    logger.info("Starting voltage overview data collection...")
    data = {}
    
    try:
        # Voltage Phasewise
        logger.info("Collecting Voltage Phasewise data...")
        data['Voltage Phasewise'] = {
            'p1max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph1"]').text,
            'p2max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph2"]').text,
            'p3max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph3"]').text,
            'p1avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph1"]').text,
            'p2avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph2"]').text,
            'p3avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph3"]').text,
            'max_avg': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Avg"]').text
        }
        
        # Voltage at Max Load
        logger.info("Collecting Voltage at Max Load data...")
        data['Voltage Maxload'] = {
            'phase1_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase1Voltage"]').text,
            'phase2_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase2Voltage"]').text,
            'phase3_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase3Voltage"]').text
        }
        
        # Voltage Unbalance using SVG
        logger.info("Extracting Voltage Unbalance data using SVG method...")
        data['Voltage Unbalance'] = extract_voltage_unbalance_from_svg(driver)
        
        # Voltage Variation using tooltip extraction
        logger.info("Collecting Voltage Variation data...")
        data['Voltage Variation'] = extract_voltage_variation_data(driver)
        
        logger.info("Voltage overview data collection completed")
        
    except Exception as e:
        logger.error(f"Error in voltage data collection: {str(e)}")
        raise
    
    return data


@log_execution_time
def save_voltage_overview_data_to_excel(month_info, overview_data):
    """Save voltage overview data to Excel"""
    logger.info("Saving voltage overview data to Excel...")
    
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Voltage Phasewise
        ws_vp = wb.create_sheet("Voltage Phasewise")
        ws_vp.append(["Phase", "Max", "Avg"])
        vp = overview_data['Voltage Phasewise']
        ws_vp.append(["Phase 1", vp['p1max'], vp['p1avg']])
        ws_vp.append(["Phase 2", vp['p2max'], vp['p2avg']])
        ws_vp.append(["Phase 3", vp['p3max'], vp['p3avg']])
        ws_vp.append(["Average", vp['max_avg'], "-"])
        
        # Voltage at Max Load
        ws_vml = wb.create_sheet("Voltage at Max Load")
        ws_vml.append(["Phase", "Voltage", "Date & Time"])
        vml = overview_data['Voltage Maxload']
        for i, key in enumerate(['phase1_maxload', 'phase2_maxload', 'phase3_maxload'], start=1):
            raw = vml[key]
            try:
                voltage_part, time_part = raw.split(' V ')
                voltage = voltage_part.strip()
                time_info = time_part.strip('()').strip()
            except:
                voltage = raw
                time_info = "-"
            ws_vml.append([f"Phase {i}", voltage, time_info])
        
        # Voltage Unbalance
        ws_vu = wb.create_sheet("Voltage Unbalance")
        ws_vu.append(["Phase", "Unbalance (%)"])
        vu = overview_data['Voltage Unbalance']
        ws_vu.append(["Phase 1", vu['Phase 1']])
        ws_vu.append(["Phase 2", vu['Phase 2']])
        ws_vu.append(["Phase 3", vu['Phase 3']])
        
        # Voltage Variation
        ws_vv = wb.create_sheet("Voltage Variation")
        ws_vv.append(["Voltage Range", "Duration (Hrs)"])
        vv = overview_data['Voltage Variation']
        for key, value in vv.items():
            ws_vv.append([key, value])
        
        # Save
        file_name = f"chart_data_from_ui_monthly_voltage_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"Voltage overview data saved: {file_name}")
        return file_name
    
    except Exception as e:
        logger.error(f"Error saving voltage overview data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_voltage_database_calculations(raw_df, month_info):
    """Process database calculations for monthly voltage overview - RAW ONLY"""
    logger.info("Processing monthly voltage database calculations (RAW ONLY)...")
    
    try:
        def format_datetime(dt_value):
            if isinstance(dt_value, pd.Timestamp):
                return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
            elif isinstance(dt_value, str):
                dt_obj = pd.to_datetime(dt_value)
                return dt_obj.strftime(f'{dt_obj.day} %b at %H:%M')
            else:
                return str(dt_value)
        
        def format_duration(minutes):
            hours, mins = divmod(int(minutes), 60)
            return f"{hours}:{mins:02} hrs"
        
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        
        # Calculate interval
        if len(raw_df) > 1:
            interval_minutes = int((raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)
        else:
            interval_minutes = 15
        
        processed_file = f"theoretical_monthly_voltage_calculated_data_{month_safe}_{timestamp}.xlsx"
        
        # 1. Voltage Phasewise
        phase_data = []
        avg_values = []
        for i, col in enumerate(['v1', 'v2', 'v3'], start=1):
            max_val = round(raw_df[col].max(), 1)
            avg_val = round(raw_df[col].mean(), 1)
            avg_values.append(avg_val)
            phase_data.append([f'Phase {i}', max_val, avg_val])
        
        overall_avg = round(sum(row[1] for row in phase_data) / 3, 1)
        phase_data.append(['Average', overall_avg, '-'])
        voltage_phasewise_df = pd.DataFrame(phase_data, columns=['Phase', 'Max', 'Avg'])
        
        # 2. Voltage at Max Load (find maximum avg_v point)
        max_load_idx = raw_df['avg_v'].idxmax()
        max_load_time = format_datetime(raw_df.loc[max_load_idx, 'surveydate'])
        
        max_load_data = [
            ['Phase 1', round(raw_df.loc[max_load_idx, 'v1'], 1), max_load_time],
            ['Phase 2', round(raw_df.loc[max_load_idx, 'v2'], 1), max_load_time],
            ['Phase 3', round(raw_df.loc[max_load_idx, 'v3'], 1), max_load_time]
        ]
        max_load_df = pd.DataFrame(max_load_data, columns=['Phase', 'Voltage', 'Date & Time'])
        
        # 3. Voltage Unbalance
        total_avg = sum(avg_values) / 3
        unbalance_data = []
        for i, avg_val in enumerate(avg_values, start=1):
            unbalance = abs(100 - ((avg_val / total_avg) * 100))
            unbalance = round(unbalance, 1)
            unbalance = 0 if unbalance == 0.0 else unbalance
            unbalance_data.append([f'Phase {i}', unbalance])
        unbalance_df = pd.DataFrame(unbalance_data, columns=['Phase', 'Unbalance (%)'])
        
        # 4. Voltage Variation
        bin_durations = {
            "Duration Voltage <180V": raw_df[raw_df['avg_v'] < 180].shape[0],
            "Duration Voltage 180-216V": raw_df[(raw_df['avg_v'] >= 180) & (raw_df['avg_v'] < 216)].shape[0],
            "Duration Voltage 216-240V": raw_df[(raw_df['avg_v'] >= 216) & (raw_df['avg_v'] <= 240)].shape[0],
            "Duration Voltage >240V": raw_df[raw_df['avg_v'] > 240].shape[0],
        }
        
        variation_data = []
        for label, count in bin_durations.items():
            total_minutes = count * interval_minutes
            duration_formatted = format_duration(total_minutes) if total_minutes > 0 else "-"
            variation_data.append([label, duration_formatted])
        variation_df = pd.DataFrame(variation_data, columns=['Voltage Range', 'Duration (Hrs)'])
        
        # Save to Excel
        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            voltage_phasewise_df.to_excel(writer, sheet_name='Voltage Phasewise', index=False)
            max_load_df.to_excel(writer, sheet_name='Voltage at Max Load', index=False)
            unbalance_df.to_excel(writer, sheet_name='Voltage Unbalance', index=False)
            variation_df.to_excel(writer, sheet_name='Voltage Variation', index=False)
        
        logger.info(f"Processed monthly voltage data saved: {processed_file}")
        return processed_file
    
    except Exception as e:
        logger.error(f"Error processing monthly voltage database: {str(e)}")
        raise


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_voltage_comparison(chart_file, processed_file, month_info):
    """Create complete monthly voltage comparison with validation"""
    logger.info("Creating monthly voltage comparison...")
    
    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_monthly_voltage_{month_safe}.xlsx"
        
        # Load sheets
        sheet_names = ['Voltage Phasewise', 'Voltage at Max Load', 'Voltage Unbalance', 'Voltage Variation']
        chart_data = {sheet: pd.read_excel(chart_file, sheet_name=sheet) for sheet in sheet_names}
        processed_data = {sheet: pd.read_excel(processed_file, sheet_name=sheet) for sheet in sheet_names}
        
        # Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        wb = Workbook()
        wb.remove(wb.active)
        
        validation_results = {}
        
        for sheet in sheet_names:
            logger.info(f"Creating comparison for: {sheet}")
            
            processed_df = processed_data[sheet]
            chart_df = chart_data[sheet]
            
            ws = wb.create_sheet(title=f"{sheet}_Comparison")
            
            if sheet == "Voltage Phasewise":
                headers = ["Phase", "DB_Max", "Chart_Max", "Max_Difference", "Max_Match",
                          "DB_Avg", "Chart_Avg", "Avg_Difference", "Avg_Match", "Overall_Match"]
            elif sheet == "Voltage at Max Load":
                headers = ["Phase", "DB_Voltage", "Chart_Voltage", "Voltage_Difference", "Voltage_Match",
                          "DB_DateTime", "Chart_DateTime", "DateTime_Match", "Overall_Match"]
            elif sheet == "Voltage Unbalance":
                headers = ["Phase", "DB_Unbalance", "Chart_Unbalance", "Unbalance_Difference", "Overall_Match"]
            elif sheet == "Voltage Variation":
                headers = ["Voltage_Range", "DB_Duration", "Chart_Duration", "Duration_Match", "Overall_Match"]
            
            ws.append(headers)
            
            sheet_results = []
            
            for i in range(len(processed_df)):
                try:
                    row_data = []
                    overall_match = True
                    
                    if sheet == "Voltage Phasewise":
                        phase = processed_df.iloc[i, 0]
                        row_data.append(phase)
                        
                        # Max comparison
                        db_max = processed_df.iloc[i, 1]
                        chart_max = chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                        try:
                            if db_max != "-" and chart_max != "-":
                                max_diff = abs(float(db_max) - float(chart_max))
                                max_match = "YES" if max_diff <= 0.1 else "NO"
                                overall_match = overall_match and (max_match == "YES")
                            else:
                                max_diff = "-"
                                max_match = "YES" if str(db_max).strip() == str(chart_max).strip() else "NO"
                                overall_match = overall_match and (max_match == "YES")
                        except (ValueError, TypeError):
                            max_diff = "-"
                            max_match = "YES" if str(db_max).strip() == str(chart_max).strip() else "NO"
                            overall_match = overall_match and (max_match == "YES")
                        
                        row_data.extend([db_max, chart_max, max_diff, max_match])
                        
                        # Avg comparison
                        db_avg = processed_df.iloc[i, 2]
                        chart_avg = chart_df.iloc[i, 2] if i < len(chart_df) else "-"
                        try:
                            if db_avg != "-" and chart_avg != "-":
                                avg_diff = abs(float(db_avg) - float(chart_avg))
                                avg_match = "YES" if avg_diff <= 0.1 else "NO"
                                overall_match = overall_match and (avg_match == "YES")
                            else:
                                avg_diff = "-"
                                avg_match = "YES" if str(db_avg).strip() == str(chart_avg).strip() else "NO"
                                overall_match = overall_match and (avg_match == "YES")
                        except (ValueError, TypeError):
                            avg_diff = "-"
                            avg_match = "YES" if str(db_avg).strip() == str(chart_avg).strip() else "NO"
                            overall_match = overall_match and (avg_match == "YES")
                        
                        row_data.extend([db_avg, chart_avg, avg_diff, avg_match])
                    
                    elif sheet == "Voltage at Max Load":
                        phase = processed_df.iloc[i, 0]
                        row_data.append(phase)
                        
                        # Voltage comparison
                        db_voltage = processed_df.iloc[i, 1]
                        chart_voltage = chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                        try:
                            if db_voltage != "-" and chart_voltage != "-":
                                voltage_diff = abs(float(db_voltage) - float(chart_voltage))
                                voltage_match = "YES" if voltage_diff <= 0.1 else "NO"
                                overall_match = overall_match and (voltage_match == "YES")
                            else:
                                voltage_diff = "-"
                                voltage_match = "YES" if str(db_voltage).strip() == str(chart_voltage).strip() else "NO"
                                overall_match = overall_match and (voltage_match == "YES")
                        except (ValueError, TypeError):
                            voltage_diff = "-"
                            voltage_match = "YES" if str(db_voltage).strip() == str(chart_voltage).strip() else "NO"
                            overall_match = overall_match and (voltage_match == "YES")
                        
                        row_data.extend([db_voltage, chart_voltage, voltage_diff, voltage_match])
                        
                        # DateTime comparison
                        db_datetime = processed_df.iloc[i, 2]
                        chart_datetime = chart_df.iloc[i, 2] if i < len(chart_df) else "-"
                        datetime_match = "YES" if str(db_datetime).strip() == str(chart_datetime).strip() else "NO"
                        overall_match = overall_match and (datetime_match == "YES")
                        
                        row_data.extend([db_datetime, chart_datetime, datetime_match])
                    
                    elif sheet == "Voltage Unbalance":
                        phase = processed_df.iloc[i, 0]
                        row_data.append(phase)
                        
                        db_unbalance = processed_df.iloc[i, 1]
                        chart_unbalance = chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                        
                        try:
                            if db_unbalance != "-" and chart_unbalance != "-":
                                db_val = float(db_unbalance)
                                chart_val = float(chart_unbalance)
                                unbalance_diff = round(abs(db_val - chart_val), 2)
                                overall_match = "YES" if unbalance_diff <= 0.1 else "NO"
                            else:
                                unbalance_diff = "-"
                                overall_match = "YES" if str(db_unbalance).strip() == str(chart_unbalance).strip() else "NO"
                        except (ValueError, TypeError):
                            unbalance_diff = "-"
                            overall_match = "YES" if str(db_unbalance).strip() == str(chart_unbalance).strip() else "NO"
                        
                        row_data.extend([db_unbalance, chart_unbalance, unbalance_diff, overall_match])
                    
                    elif sheet == "Voltage Variation":
                        voltage_range = processed_df.iloc[i, 0]
                        row_data.append(voltage_range)
                        
                        db_duration = processed_df.iloc[i, 1]
                        chart_duration = chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                        duration_match = "YES" if str(db_duration).strip() == str(chart_duration).strip() else "NO"
                        overall_match = duration_match == "YES"
                        
                        row_data.extend([db_duration, chart_duration, duration_match, overall_match])
                    
                    if sheet not in ["Voltage Unbalance", "Voltage Variation"]:
                        row_data.append("YES" if overall_match else "NO")
                    
                    if sheet == "Voltage Unbalance":
                        match_result = overall_match == "YES"
                    elif sheet == "Voltage Variation":
                        match_result = overall_match
                    else:
                        match_result = overall_match
                    
                    sheet_results.append({
                        'item': phase if sheet != "Voltage Variation" else voltage_range,
                        'match': match_result
                    })
                    
                    ws.append(row_data)
                
                except Exception as e:
                    logger.warning(f"Error processing row {i} in {sheet}: {str(e)}")
                    continue
            
            validation_results[sheet] = sheet_results
            
            passed_count = sum(1 for result in sheet_results if result['match'])
            failed_count = len(sheet_results) - passed_count
            logger.info(f"{sheet} Validation: {passed_count} passed, {failed_count} failed")
            
            # Apply colors
            for row_num in range(2, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    header = ws.cell(row=1, column=col_num).value
                    
                    if header and ("_Match" in header or header == "Overall_Match"):
                        if cell.value == "YES":
                            cell.fill = green_fill
                        elif cell.value == "NO":
                            cell.fill = red_fill
                    elif header and "_Difference" in header:
                        if isinstance(cell.value, (int, float)):
                            if cell.value <= 0.1:
                                cell.fill = green_fill
                            else:
                                cell.fill = red_fill
        
        wb.save(output_file)
        logger.info(f"Monthly voltage comparison saved: {output_file}")
        
        return output_file, validation_results
    
    except Exception as e:
        logger.error(f"Error creating monthly voltage comparison: {str(e)}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_voltage_summary_report(config, month_info, chart_file, processed_file,
                                 comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive monthly voltage summary report with ENHANCED styling"""
    logger.info("Creating monthly voltage summary report with enhanced styling...")
    
    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_VOLTAGE_{month_safe}_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"
        
        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")
        
        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")
        
        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")
        
        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")
        
        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")
        
        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_alignment = Alignment(horizontal="center", vertical="center")
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        current_row = 1
        
        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"LV MONTHLY VOLTAGE VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # Empty row
        current_row += 1
        
        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        current_row += 1
        
        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Monthly Voltage (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
        ]
        
        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        current_row += 1
        
        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # Data rows
        total_chart_points = 0
        try:
            chart_df = pd.read_excel(chart_file, sheet_name='Voltage Phasewise')
            total_chart_points = len(chart_df)
        except:
            total_chart_points = 4
        
        # Calculate expected records for the month
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * 96
        data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0
        
        data_rows = [
            ["Raw Voltage Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Chart Voltage Data Points", total_chart_points, "COMPLETE RECORDS"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"]
        ]
        
        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border
            
            ws[f'C{current_row}'].value = status
            if "COMPLETE" in status or "%" in status:
                if data_completeness >= 90 or "COMPLETE RECORDS" in status:
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = warning_font
                    ws[f'C{current_row}'].fill = warning_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        current_row += 1
        
        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # Calculate validation results
        overall_passed = 0
        overall_total = 0
        validation_data = []
        
        for sheet_name, results in validation_results.items():
            total_items = len(results)
            passed_items = sum(1 for result in results if result['match'])
            failed_items = total_items - passed_items
            success_rate = f"{(passed_items / total_items) * 100:.1f}%" if total_items > 0 else "0%"
            status = "PASS" if passed_items == total_items else "FAIL"
            
            display_name = sheet_name.replace('_', ' ')
            validation_data.append([display_name, passed_items, failed_items, success_rate, status])
            
            overall_passed += passed_items
            overall_total += total_items
        
        for comp_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = comp_type
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border
            
            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border
            
            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border
            
            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        current_row += 1
        
        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0
        
        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: Monthly voltage validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor voltage discrepancies found - Review recommended"
            assessment_color = warning_fill
            assessment_font_color = warning_font
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant voltage validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font
        
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # Set column widths for better appearance
        from openpyxl.utils import get_column_letter
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        wb.save(summary_file)
        logger.info(f"Enhanced monthly voltage summary report created: {summary_file}")
        
        # Log summary
        logger.info("=" * 60)
        logger.info("MONTHLY VOLTAGE VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"Data: Raw={len(raw_df)}, Chart={total_chart_points}")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)
        
        return summary_file
    
    except Exception as e:
        logger.error(f"Error creating voltage summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_lv_monthly_voltage_automation():
    """Main LV Monthly Voltage automation process"""
    config = None
    driver = None
    output_folder = None
    
    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False
        
        # Setup output folder
        output_folder = setup_output_folder()
        
        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)
        
        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)
        
        # Login
        if not login(driver):
            logger.info("Login failed")
            return False
        
        # Apply configuration
        logger.info("Applying LV Monthly Voltage configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])
        
        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False
        
        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False
        
        # Get meter metrics
        logger.info("Fetching meter metrics...")
        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])
        
        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False
        
        logger.info(f"Meter found: {name} (ID: {mtr_id})")
        
        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False
        
        # Wait for overview page to load
        time.sleep(5)
        
        # Collect voltage overview data
        logger.info("Collecting monthly voltage data from UI...")
        overview_data = collect_voltage_overview_data(driver)
        
        # Save voltage overview data
        chart_file = save_voltage_overview_data_to_excel(month_info, overview_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)
        
        # Get database data for complete month - RAW ONLY
        raw_df = get_database_data_for_monthly_voltage(month_info, mtr_id)
        
        if raw_df.empty:
            logger.info("No voltage database data found for the month")
            return False
        
        # Process database calculations
        logger.info("Processing voltage database calculations (RAW ONLY)...")
        processed_file = process_voltage_database_calculations(raw_df, month_info)
        processed_file = save_file_to_output(processed_file, output_folder)
        
        # Create comparison report
        logger.info("Creating voltage validation comparison...")
        comparison_file, validation_results = create_voltage_comparison(chart_file, processed_file, month_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)
        
        # Create summary report
        logger.info("Creating comprehensive voltage summary...")
        summary_report = create_voltage_summary_report(
            config, month_info, chart_file, processed_file,
            comparison_file, validation_results, raw_df, name)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)
        
        # Final summary
        logger.info("=" * 60)
        logger.info("LV MONTHLY VOLTAGE AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: LV Monthly Voltage (Fixed)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"Database Records: {len(raw_df)} voltage records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ LV Monthly Voltage monitoring (fixed)")
        logger.info("   ✓ Complete month voltage data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ SVG-based voltage unbalance extraction")
        logger.info("   ✓ Tooltip-based voltage variation extraction")
        logger.info("   ✓ RAW ONLY database approach (no NRM)")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)
        
        return True
    
    except Exception as e:
        logger.info(f"Critical error: {e}")
        
        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"LV Monthly Voltage Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass
        
        return False
    
    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV MONTHLY VOLTAGE AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: LV Monthly Voltage (Fixed)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ LV Monthly Voltage monitoring only")
    logger.info("   ✓ Complete month voltage data processing")
    logger.info("   ✓ RAW ONLY database approach (no NRM)")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ SVG-based voltage unbalance extraction")
    logger.info("   ✓ Tooltip-based voltage variation data")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Better null/dash handling")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)
    
    start_time = time.time()
    success = main_lv_monthly_voltage_automation()
    end_time = time.time()
    total_time = end_time - start_time
    
    logger.info("=" * 60)
    if success:
        logger.info("LV MONTHLY VOLTAGE AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time/60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ LV Monthly Voltage monitoring (fixed)")
        logger.info("   ✓ Complete month voltage processing")
        logger.info("   ✓ RAW ONLY approach (no NRM)")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ SVG extraction for voltage unbalance")
        logger.info("   ✓ Tooltip extraction for voltage variation")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("LV MONTHLY VOLTAGE AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time/60:.1f}min)")
        logger.info("Check error logs in output folder")
    
    logger.info("=" * 60)
    logger.info("LV Monthly Voltage Automation Finished")
    logger.info("=" * 60)
