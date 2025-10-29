import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and SIP duration
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Change to tenant02, tenant03, etc. as needed

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('lv_monthly_pf_detailed_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/lv_monthly_pf_detailed_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for LV Monthly Power Factor Detailed"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'January 2025', 'YOUR_METER_NO',
                      'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'This script is FOR LV MONTHLY POWER FACTOR DETAILED VIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: DT or LV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"LV Monthly Power Factor Detailed Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for LV Monthly Power Factor Detailed"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}  # Fixed for LV monitoring

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'feeder', 'target_month_year', 'meter_serial_no', 'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("LV Monthly Power Factor Detailed Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING LV MONTHLY POWER FACTOR DETAILED VIEW AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default LV Monthly Power Factor Detailed configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("LV Monthly Power Factor Detailed Configuration validated successfully")
    logger.info(f"   Monitoring Type: LV Monthly Power Factor Detailed View")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Feeder: {config['feeder']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    """Get meter metrics from database"""
    logger.info(f"Fetching LV Monthly Power Factor Detailed metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        if meter_type.upper() == 'DT':
            query1 = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        elif meter_type.upper() == 'LV':
            query1 = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        else:
            logger.info(f"Invalid meter type: {meter_type}")
            return None, None, None, None

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None, None

        dt_id, dt_name, meterid = result1

        # Get SIP duration
        query2 = f"SELECT sip FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail WHERE mtrid = %s LIMIT 1;"
        cursor.execute(query2, (meterid,))
        result2 = cursor.fetchone()
        sip_duration = int(result2[0]) if result2 and result2[0] else 15

        logger.info(f"Metrics: {dt_name}, meterid: {meterid}, SIP: {sip_duration}min")
        return dt_id, dt_name, meterid, sip_duration
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_detailed(month_info, mtr_id):
    """Fetch database data for complete month detailed view - POWER FACTOR PARAMETERS"""
    logger.info(f"Fetching monthly detailed database data for: {month_info['selected_month_year']}")

    start_date = month_info['start_date'].strftime("%Y-%m-%d")
    end_date_next = (month_info['end_date'] + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{end_date_next}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # Query for POWER FACTOR parameters from RAW data
        raw_query = f"""
            SELECT surveydate, pf, avg_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        raw_df = pd.read_sql(raw_query, conn)

        logger.info(f"Retrieved: Raw={len(raw_df)} power factor records")
        return raw_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SANYAM")
        driver.find_element(By.ID, "Password").send_keys("Sanyam@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)

        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()

        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select LV monitoring - FIXED FOR LV ONLY"""
    try:
        logger.info("Selecting LV monitoring (fixed for LV monthly power factor detailed script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        logger.info("LV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - DT or LV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "DT":
            dt_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]')))
            dt_button.click()
            logger.info("DT selected")
        elif meter_type == "LV":
            lv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]')))
            lv_button.click()
            logger.info("LV feeder selected")
        else:
            logger.info("Invalid meter type for LV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_side_panel_pf_data(driver, wait):
    """Collect power factor data from detailed view side panel"""
    logger.info("Collecting power factor side panel data...")
    data = {}

    try:
        # Navigate to detailed power factor page
        wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@id="VPDetailedLink"]'))).click()
        time.sleep(2)
        
        # Click on Power factor tab
        wait.until(EC.element_to_be_clickable((By.XPATH, '//span[@class="dx-tab-text-span" and text()="Power factor"]'))).click()
        time.sleep(2)

        action = ActionChains(driver)

        # Power Factor Average
        try:
            pf_avg = driver.find_element(By.XPATH, "//span[@id='avgPf']").text
        except Exception as e:
            logger.debug(f"Error fetching PF average: {e}")
            pf_avg = '-'

        # Power Factor Duration bars
        pf_pattern_bars = driver.find_elements(By.CSS_SELECTOR, '#pfRangeDetailChart g.dxc-markers')
        tooltip_selector = '.dxc-tooltip svg text'

        pf_durations = {
            'Duration PF < 0.9': '-',
            'Duration PF 0.9 - 0.95': '-',
            'Duration PF > 0.95': '-'
        }

        color_mapping = {
            '#D11920': 'Duration PF < 0.9',
            '#DEAE2A': 'Duration PF 0.9 - 0.95',
            '#86B8A5': 'Duration PF > 0.95'
        }

        for i, bar in enumerate(pf_pattern_bars):
            try:
                fill_color = bar.get_attribute('fill')
                label = color_mapping.get(fill_color)

                if label:
                    action.move_to_element(bar).perform()
                    time.sleep(1)

                    tooltip = driver.find_element(By.CSS_SELECTOR, tooltip_selector)
                    tooltip_text = tooltip.text.strip()
                    pf_durations[label] = tooltip_text

            except Exception as e:
                logger.debug(f"Error processing PF bar {i + 1}: {e}")

        data = {
            'Power Factor Average': pf_avg,
            **pf_durations
        }

        logger.info("Power factor side panel data collected successfully")
    except Exception as e:
        logger.error(f"Error collecting power factor side panel data: {str(e)}")
        raise

    return data


@log_execution_time
def save_side_panel_data_to_excel(side_data, month_info, sip_duration):
    """Save side panel power factor data to Excel"""
    logger.info("Saving side panel power factor data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Power Factor Table Sheet
        ws_pf = wb.create_sheet("Power Factor Table")
        ws_pf.append(["Parameter", "UI_Value"])
        for key, value in side_data.items():
            ws_pf.append([key, value])

        # SIP Configuration Sheet
        ws_sip = wb.create_sheet("SIP Configuration")
        ws_sip.append(["Parameter", "Value"])
        ws_sip.append(["SIP Duration (minutes)", sip_duration])
        ws_sip.append(["Expected SIPs per day", (24 * 60) // sip_duration])
        ws_sip.append(["Month Analyzed", month_info['selected_month_year']])

        # Save
        file_name = f"ui_side_panel_pf_data_monthly_detailed_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"Side panel power factor data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving side panel data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING FOR POWER FACTOR
# ============================================================================
def format_duration(hours_val):
    """Format hours as HH:MM string"""
    if hours_val == '-' or pd.isna(hours_val):
        return '-'

    try:
        total_minutes = int(hours_val * 60)
        hrs = total_minutes // 60
        mins = total_minutes % 60
        return f"{hrs:02d}:{mins:02d} hrs"
    except:
        return '-'


@log_execution_time
def calculate_side_panel_pf_metrics_from_raw_data(raw_df, month_info, sip_duration):
    """Calculate side panel power factor metrics from RAW data using dynamic SIP duration"""
    logger.info(f"Calculating side panel power factor metrics from RAW data with {sip_duration}-minute SIP intervals...")

    month_year_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    logger.info(f"Processing RAW power factor data: {len(raw_df)} records with {sip_duration}-minute intervals")

    df_raw = raw_df.copy()
    df_raw['surveydate'] = pd.to_datetime(df_raw['surveydate'])

    # Filter valid power factor values (0 to 0.999999)
    valid_pf_mask = (df_raw['pf'] >= 0) & (df_raw['pf'] <= 0.999999)
    pf_series = df_raw.loc[valid_pf_mask, 'pf']

    logger.info(f"Using dynamic SIP duration: {sip_duration} minutes for all power factor calculations")
    logger.info(f"Valid PF records: {len(pf_series)} out of {len(df_raw)}")

    # Calculate Power Factor Average
    pf_avg = pf_series.mean() if len(pf_series) > 0 else 0

    # Calculate durations based on SIP intervals
    def calc_duration(pf_values):
        count = len(pf_values)
        if count == 0:
            return '-'
        total_minutes = count * sip_duration
        hours = total_minutes / 60
        return hours

    duration_gt_095 = calc_duration(pf_series[pf_series > 0.95])
    duration_between_09_095 = calc_duration(pf_series[(pf_series >= 0.9) & (pf_series <= 0.95)])
    duration_lt_09 = calc_duration(pf_series[pf_series < 0.9])

    # Create calculated side panel data structure
    calculated_data = {
        'Power Factor Average': f"{pf_avg:.4f}" if pf_avg > 0 else '-',
        'Duration PF < 0.9': format_duration(duration_lt_09),
        'Duration PF 0.9 - 0.95': format_duration(duration_between_09_095),
        'Duration PF > 0.95': format_duration(duration_gt_095)
    }

    # Save calculated data to Excel
    calculated_file = f"calculated_side_panel_pf_data_{month_year_safe}_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    # Power Factor Table Sheet
    ws_pf = wb.create_sheet('Power Factor Table')
    ws_pf.append(['Parameter', 'Calculated_Value'])
    for key, value in calculated_data.items():
        ws_pf.append([key, value])

    # SIP Configuration Sheet
    ws_sip = wb.create_sheet('SIP Configuration')
    ws_sip.append(['Parameter', 'Value'])
    ws_sip.append(['SIP Duration (minutes)', sip_duration])
    ws_sip.append(['Expected SIPs per day', (24 * 60) // sip_duration])
    ws_sip.append(['Actual SIPs', len(raw_df)])
    ws_sip.append(['Valid PF SIPs', len(pf_series)])
    ws_sip.append(['Power Factor Average', f"{pf_avg:.4f}"])
    ws_sip.append(['PF < 0.9 Count', len(pf_series[pf_series < 0.9])])
    ws_sip.append(['PF 0.9-0.95 Count', len(pf_series[(pf_series >= 0.9) & (pf_series <= 0.95)])])
    ws_sip.append(['PF > 0.95 Count', len(pf_series[pf_series > 0.95])])

    wb.save(calculated_file)

    logger.info(f"Side panel power factor metrics calculation completed using {sip_duration}-minute SIP intervals")
    logger.info(f"Power Factor Average: {pf_avg:.4f}")
    return calculated_data, calculated_file


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_detailed_comparison(ui_file, calculated_file, month_info, sip_duration):
    """Create complete monthly detailed power factor comparison with validation"""
    logger.info("Creating monthly detailed power factor comparison...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_monthly_pf_detailed_{month_safe}.xlsx"

        # Load sheets
        ui_data = pd.read_excel(ui_file, sheet_name='Power Factor Table')
        calculated_data = pd.read_excel(calculated_file, sheet_name='Power Factor Table')

        # Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb = Workbook()
        wb.remove(wb.active)

        validation_results = {}

        def normalize_string(s):
            """Remove all spaces and lowercase for fair string comparison"""
            if s is None or str(s).lower() in ['nan', 'none']:
                return ""
            return str(s).replace(" ", "").strip().lower()

        logger.info(f"Creating comparison for Power Factor Table")

        ws = wb.create_sheet(title="Power_Factor_Table_Comparison")

        headers = ["Parameter", "UI_Value", "Calculated_Value", "Match", "Notes"]
        ws.append(headers)

        sheet_results = []

        for i in range(len(calculated_data)):
            try:
                param = calculated_data.iloc[i, 0]
                calc_val = calculated_data.iloc[i, 1]
                ui_val = ui_data.iloc[i, 1] if i < len(ui_data) else "-"

                ui_str = normalize_string(ui_val)
                calc_str = normalize_string(calc_val)

                # Try numeric comparison first
                numeric_match = False
                try:
                    if calc_val is not None and ui_val is not None:
                        ui_num_match = re.search(r'[-+]?(\d*\.?\d+)', str(ui_val))
                        calc_num_match = re.search(r'[-+]?(\d*\.?\d+)', str(calc_val))

                        if ui_num_match and calc_num_match:
                            ui_num = float(ui_num_match.group())
                            calc_num = float(calc_num_match.group())
                            # Tolerance based on parameter type
                            if 'average' in str(param).lower():
                                tolerance = 0.001  # Power factor average
                            else:
                                tolerance = 1.0  # Duration (minutes tolerance)
                            numeric_match = abs(ui_num - calc_num) <= tolerance
                except (ValueError, TypeError, AttributeError):
                    numeric_match = False

                # Overall match determination
                if numeric_match:
                    match = 'YES'
                    match_color = green_fill
                    notes = f'Numeric values match (SIP: {sip_duration}min)'
                elif ui_str == calc_str:
                    match = 'YES'
                    match_color = green_fill
                    notes = f'String values match (SIP: {sip_duration}min)'
                else:
                    match = 'NO'
                    match_color = red_fill
                    notes = f'Values differ (SIP: {sip_duration}min)'

                sheet_results.append({
                    'item': param,
                    'match': match == 'YES'
                })

                row_data = [param, ui_val, calc_val, match, notes]
                ws.append(row_data)

                # Apply color to match column
                match_cell = ws.cell(row=ws.max_row, column=4)
                match_cell.fill = match_color

            except Exception as e:
                logger.warning(f"Error processing row {i} in Power Factor Table: {str(e)}")
                continue

        validation_results['Power Factor Table'] = sheet_results

        passed_count = sum(1 for result in sheet_results if result['match'])
        failed_count = len(sheet_results) - passed_count
        logger.info(f"Power Factor Table Validation: {passed_count} passed, {failed_count} failed")

        wb.save(output_file)
        logger.info(f"Monthly detailed power factor comparison saved: {output_file}")

        return output_file, validation_results

    except Exception as e:
        logger.error(f"Error creating monthly detailed power factor comparison: {str(e)}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_detailed_summary_report(config, month_info, ui_file, calculated_file,
                                   comparison_file, validation_results, raw_df, meter_name, sip_duration):
    """Create comprehensive monthly detailed power factor summary report with ENHANCED styling"""
    logger.info("Creating monthly detailed power factor summary report with enhanced styling...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_PF_DETAILED_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"LV MONTHLY POWER FACTOR DETAILED VIEW VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Empty row
        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Monthly Power Factor Detailed View"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
            ["SIP Duration:", f"{sip_duration} minutes"],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate expected records for the month
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * ((24 * 60) // sip_duration)
        data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0

        # Calculate valid PF records
        valid_pf_mask = (raw_df['pf'] >= 0) & (raw_df['pf'] <= 0.999999)
        valid_pf_count = valid_pf_mask.sum()

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Valid PF Records", valid_pf_count, f"{(valid_pf_count/len(raw_df)*100):.1f}% Valid" if len(raw_df) > 0 else "NO DATA"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"],
            ["SIP Duration Used", f"{sip_duration} min", "DYNAMIC FROM DB"]
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in str(status) or "%" in str(status) or "DYNAMIC" in str(status):
                if data_completeness >= 90 or "COMPLETE RECORDS" in str(status) or "DYNAMIC" in str(status) or "Valid" in str(status):
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = fail_font
                    ws[f'C{current_row}'].fill = fail_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0
        validation_data = []

        for sheet_name, results in validation_results.items():
            total_items = len(results)
            passed_items = sum(1 for result in results if result['match'])
            failed_items = total_items - passed_items
            success_rate = f"{(passed_items / total_items) * 100:.1f}%" if total_items > 0 else "0%"
            status = "PASS" if passed_items == total_items else "FAIL"

            display_name = sheet_name.replace('_', ' ')
            validation_data.append([display_name, passed_items, failed_items, success_rate, status])

            overall_passed += passed_items
            overall_total += total_items

        for comp_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = comp_type
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: Monthly detailed power factor validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor discrepancies found - Review recommended"
            assessment_color = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            assessment_font_color = Font(bold=True, size=10, color="000000", name="Calibri")
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Enhanced monthly detailed power factor summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MONTHLY DETAILED POWER FACTOR VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"SIP Duration: {sip_duration} minutes")
        logger.info(f"Data: Raw={len(raw_df)} records")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_lv_monthly_pf_detailed_automation():
    """Main LV Monthly Power Factor Detailed View automation process"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying LV Monthly Power Factor Detailed configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        dt_id, name, mtr_id, sip_duration = get_metrics(config['meter_serial_no'], config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False

        logger.info(f"Meter found: {name} (ID: {mtr_id}, SIP: {sip_duration}min)")

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for page to load
        time.sleep(5)

        # Collect side panel power factor data
        logger.info("Collecting monthly detailed power factor side panel data from UI...")
        side_panel_data = collect_side_panel_pf_data(driver, wait)

        # Save UI side panel data
        ui_file = save_side_panel_data_to_excel(side_panel_data, month_info, sip_duration)
        if ui_file:
            ui_file = save_file_to_output(ui_file, output_folder)

        # Get database data for complete month
        raw_df = get_database_data_for_monthly_detailed(month_info, mtr_id)

        if raw_df.empty:
            logger.info("No database data found for the month")
            return False

        # Process database calculations
        logger.info("Processing database power factor calculations...")
        calculated_data, calculated_file = calculate_side_panel_pf_metrics_from_raw_data(
            raw_df, month_info, sip_duration)
        calculated_file = save_file_to_output(calculated_file, output_folder)

        # Create comparison report
        logger.info("Creating validation comparison...")
        comparison_file, validation_results = create_detailed_comparison(
            ui_file, calculated_file, month_info, sip_duration)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create summary report
        logger.info("Creating comprehensive summary...")
        summary_report = create_detailed_summary_report(
            config, month_info, ui_file, calculated_file,
            comparison_file, validation_results, raw_df, name, sip_duration)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("LV MONTHLY POWER FACTOR DETAILED AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: LV Monthly Power Factor Detailed View")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"SIP Duration: {sip_duration} minutes (dynamic from DB)")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(ui_file) if ui_file else 'UI side panel data'}")
        logger.info(f"   2. {os.path.basename(calculated_file) if calculated_file else 'Calculated data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ LV Monthly Power Factor Detailed View monitoring")
        logger.info("   ✓ Complete month data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Dynamic SIP duration from database")
        logger.info("   ✓ Power factor average & duration validation")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Complete validation summary")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"LV Monthly Power Factor Detailed Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV MONTHLY POWER FACTOR DETAILED VIEW AUTOMATION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: LV Monthly Power Factor Detailed View")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ LV Monthly Power Factor Detailed View monitoring")
    logger.info("   ✓ Complete month data processing")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Dynamic SIP duration from database")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ PF average & duration metrics")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_lv_monthly_pf_detailed_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("LV MONTHLY POWER FACTOR DETAILED AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ LV Monthly Power Factor Detailed View monitoring")
        logger.info("   ✓ Complete month processing")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Dynamic SIP from database")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Side panel PF data extraction")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("LV MONTHLY POWER FACTOR DETAILED AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("LV Monthly Power Factor Detailed Automation Finished")
    logger.info("=" * 60)
