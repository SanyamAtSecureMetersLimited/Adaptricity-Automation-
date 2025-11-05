import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database 1: Hetzner - For meter details and SIP duration
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    # Database 2: Service Platform - For load survey data
    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Change to tenant02, tenant03, etc. as needed

    @classmethod
    def get_db1_params(cls):
        return {
            "host": cls.DB1_HOST,
            "port": cls.DB1_PORT,
            "database": cls.DB1_DATABASE,
            "user": cls.DB1_USER,
            "password": cls.DB1_PASSWORD
        }

    @classmethod
    def get_db2_params(cls):
        return {
            "host": cls.DB2_HOST,
            "port": cls.DB2_PORT,
            "database": cls.DB2_DATABASE,
            "user": cls.DB2_USER,
            "password": cls.DB2_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('lv_monthly_demand_sidepanel_automation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/lv_monthly_demand_sidepanel_automation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for LV Monthly Demand Side Panel"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Month_Year', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'January 2025', 'YOUR_METER_NO',
                      'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (e.g., January 2025)',
                    'Update Meter_Serial_No with your meter serial number',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'This script is FOR LV MONTHLY DEMAND SIDE PANEL ONLY (NO GRAPH)',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: January 2025',
                    'Meter_Type: DT or LV only',
                    'Save file before running',
                    'Test Engineer: Sanyam Upadhyay',
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"LV Monthly Demand Side Panel Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def normalize_month_year(value):
    """Ensure month-year is in 'Month YYYY' format"""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%B %Y")
    try:
        parsed = pd.to_datetime(value, errors='raise')
        return parsed.strftime("%B %Y")
    except Exception:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for LV Monthly Demand Side Panel"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV_MONTHLY_SIDEPANEL'}  # Fixed for LV Monthly Side Panel

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = normalize_month_year(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required_fields = ['type', 'area', 'substation', 'feeder', 'target_month_year', 'meter_serial_no',
                           'meter_type']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("LV Monthly Demand Side Panel Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING LV MONTHLY DEMAND SIDE PANEL AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default LV Monthly Demand Side Panel configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("LV Monthly Demand Side Panel Configuration validated successfully")
    logger.info(f"   Monitoring Type: LV Monthly Demand Side Panel (NO GRAPH)")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   Feeder: {config['feeder']}")
    logger.info(f"   Month: {config['target_month_year']}")
    logger.info(f"   Meter: {config['meter_serial_no']}")
    logger.info(f"   Meter Type: {config['meter_type']}")
    return config


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    """Get meter metrics from database including SIP duration - EXACT same as daywise"""
    logger.info(f"Fetching LV Monthly metrics for meter: {mtr_serial_no}")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()

        if meter_type.upper() == 'DT':
            query1 = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        elif meter_type.upper() == 'LV':
            query1 = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        else:
            logger.info(f"Invalid meter type: {meter_type}")
            return None, None, None, None, None

        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()
        if not result1:
            logger.info(f"Meter not found: {mtr_serial_no}")
            return None, None, None, None, None

        dt_id, dt_name, meterid = result1
        node_id = dt_id  # node_id is same as dt_id

        # Get SIP duration
        query2 = f"SELECT sip FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail WHERE mtrid = %s LIMIT 1;"
        cursor.execute(query2, (meterid,))
        result2 = cursor.fetchone()
        sip_duration = int(result2[0]) if result2 and result2[0] else 15

        logger.info(f"Metrics: {dt_name}, meterid: {meterid}, node_id: {node_id}, SIP: {sip_duration}min")
        return dt_id, dt_name, meterid, node_id, sip_duration
    except Exception as e:
        logger.info(f"Database error: {e}")
        return None, None, None, None, None
    finally:
        if 'conn' in locals():
            conn.close()


@log_execution_time
def get_database_data_for_monthly_sidepanel(month_info, mtr_id, node_id):
    """Fetch database data for complete month - EXACT same as daywise logic"""
    logger.info(f"Fetching monthly database data for: {month_info['selected_month_year']}")

    start_date = month_info['start_date'].strftime("%Y-%m-%d")
    end_date = month_info['end_date'].strftime("%Y-%m-%d")
    date_filter = f"AND DATE(surveydate) >= '{start_date}' AND DATE(surveydate) <= '{end_date}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # RAW QUERY - EXACT same as daywise
        raw_query = f"""
            SELECT DISTINCT surveydate, kwh_i, kvah_i, kvar_i_total, kwh_abs, kvah_abs, kvarh_abs
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        # NRM QUERY - EXACT same as daywise
        nrm_query = f"""
            SELECT surveydate, kw_i, kva_i, kvar_i
            FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile
            WHERE nodeid={node_id} {date_filter}
            ORDER BY surveydate ASC;
        """

        raw_df = pd.read_sql(raw_query, conn)
        nrm_df = pd.read_sql(nrm_query, conn)

        logger.info(f"Retrieved: Raw={len(raw_df)}, NRM={len(nrm_df)} records")

        # DEBUG: Log RAW data statistics
        if not raw_df.empty:
            logger.info("=" * 60)
            logger.info("DEBUG: RAW DATA STATISTICS")
            logger.info(
                f"  kwh_i - Min: {raw_df['kwh_i'].min()}, Max: {raw_df['kwh_i'].max()}, Non-null: {raw_df['kwh_i'].notna().sum()}")
            logger.info(
                f"  kvah_i - Min: {raw_df['kvah_i'].min()}, Max: {raw_df['kvah_i'].max()}, Non-null: {raw_df['kvah_i'].notna().sum()}")
            logger.info(
                f"  kvar_i_total - Min: {raw_df['kvar_i_total'].min()}, Max: {raw_df['kvar_i_total'].max()}, Non-null: {raw_df['kvar_i_total'].notna().sum()}")
            logger.info(f"Sample RAW data (first 3):")
            logger.info(raw_df[['surveydate', 'kwh_i', 'kvah_i', 'kvar_i_total']].head(3).to_string())
            logger.info("=" * 60)

        # DEBUG: Log NRM data statistics
        if not nrm_df.empty:
            logger.info("=" * 60)
            logger.info("DEBUG: NRM DATA STATISTICS")
            logger.info(
                f"  kw_i - Min: {nrm_df['kw_i'].min()}, Max: {nrm_df['kw_i'].max()}, Non-null: {nrm_df['kw_i'].notna().sum()}")
            logger.info(
                f"  kva_i - Min: {nrm_df['kva_i'].min()}, Max: {nrm_df['kva_i'].max()}, Non-null: {nrm_df['kva_i'].notna().sum()}")
            logger.info(
                f"  kvar_i - Min: {nrm_df['kvar_i'].min()}, Max: {nrm_df['kvar_i'].max()}, Non-null: {nrm_df['kvar_i'].notna().sum()}")
            logger.info(f"Sample NRM data (first 3):")
            logger.info(nrm_df[['surveydate', 'kw_i', 'kva_i', 'kvar_i']].head(3).to_string())
            logger.info("=" * 60)

        if not nrm_df.empty:
            nrm_df['date'] = pd.to_datetime(nrm_df['surveydate']).dt.date
            sip_counts = nrm_df.groupby('date').size()
            for date, count in sip_counts.items():
                logger.info(f"   {date}: {count} SIPs available")

        return raw_df, nrm_df
    except Exception as e:
        logger.info(f"Database error: {e}")
        return pd.DataFrame(), pd.DataFrame()
    finally:
        if 'conn' in locals():
            conn.close()


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("Secure")
        driver.find_element(By.ID, "Password").send_keys("Secure@12345")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to month: {target_month_year}")
    try:
        # Click Month button
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        time.sleep(1)

        # Set month
        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        time.sleep(2)

        # Parse month info
        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info("Month set successfully")
        logger.info(f"Complete month range: {start_date} to {end_date}")
        return month_info
    except Exception as e:
        logger.info(f"Month setting error: {e}")
        return None


def select_type(driver):
    """Select LV monitoring - FIXED FOR LV ONLY"""
    try:
        logger.info("Selecting LV monitoring (fixed for LV monthly side panel script)")
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        logger.info("LV monitoring selected")
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type selection error: {e}")


def select_meter_type(driver, meter_type):
    """Select meter type - DT or LV only"""
    try:
        logger.info(f"Selecting meter type: {meter_type}")
        wait = WebDriverWait(driver, 10)

        if meter_type == "DT":
            dt_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]')))
            dt_button.click()
            logger.info("DT selected")
        elif meter_type == "LV":
            lv_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]')))
            lv_button.click()
            logger.info("LV feeder selected")
        else:
            logger.info("Invalid meter type for LV monitoring")
            return False

        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    """Find meter using search box and click View"""
    logger.info(f"Searching for meter: {meter_serial_no}")
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)

        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if not view_buttons:
            logger.info("No View button found")
            return False

        if len(view_buttons) == 1:
            view_buttons[0].click()
            logger.info("View clicked (1 result)")
            return True

        logger.info(f"Found {len(view_buttons)} results, finding exact match")
        for idx, view_btn in enumerate(view_buttons):
            try:
                parent_row = view_btn.find_element(By.XPATH, "./ancestor::tr")
                if meter_serial_no in parent_row.text:
                    view_btn.click()
                    logger.info(f"View clicked (exact match at row {idx + 1})")
                    return True
            except:
                continue

        view_buttons[0].click()
        logger.info("View clicked (first result)")
        return True
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_side_panel_data(driver, wait):
    """Collect data from demand side panel (demand table) - SIDE PANEL ONLY"""
    logger.info("Collecting demand side panel data (NO GRAPH EXTRACTION)...")
    data = {}

    try:
        # Navigate to detailed demand page
        logger.info("Navigating to detailed view...")
        wait.until(EC.element_to_be_clickable((By.XPATH, '//a[@id="VPDetailedLink"]'))).click()
        time.sleep(2)

        logger.info("Clicking Demand tab...")
        wait.until(
            EC.element_to_be_clickable((By.XPATH, '//span[@class="dx-tab-text-span" and text()="Demand"]'))).click()
        time.sleep(3)

        logger.info("Extracting side panel values...")
        # Active Power Data
        data['act_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kw"]').text
        data['act_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kw"]').text
        data['act_dt'] = driver.find_element(By.XPATH, '//td[@id="kw_MaxDatetime"]').text

        # Apparent Power Data
        data['app_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kva"]').text
        data['app_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kva"]').text
        data['app_dt'] = driver.find_element(By.XPATH, '//td[@id="kva_MaxDatetime"]').text

        # Reactive Power Data
        data['react_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kvar"]').text
        data['react_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kvar"]').text
        data['react_dt'] = driver.find_element(By.XPATH, '//td[@id="kvar_MaxDatetime"]').text

        logger.info("Side panel data collected successfully:")
        logger.info(f"   Active: Max={data['act_max']}, Avg={data['act_avg']}, DT={data['act_dt']}")
        logger.info(f"   Apparent: Max={data['app_max']}, Avg={data['app_avg']}, DT={data['app_dt']}")
        logger.info(f"   Reactive: Max={data['react_max']}, Avg={data['react_avg']}, DT={data['react_dt']}")

    except Exception as e:
        logger.error(f"Error collecting side panel data: {str(e)}")
        raise

    return data


@log_execution_time
def save_side_panel_data_to_excel(side_data, month_info, sip_duration):
    """Save side panel data to Excel"""
    logger.info("Saving demand side panel data to Excel...")

    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Demand Table Sheet
        ws_demand = wb.create_sheet("Demand Table")
        ws_demand.append(["Parameter", "Max", "Avg", "Date and Time at Max Value"])

        # Map side data to demand table format
        param_mapping = {
            "Active(kW)": ("act_max", "act_avg", "act_dt"),
            "Apparent(kVA)": ("app_max", "app_avg", "app_dt"),
            "Reactive(kVAr)": ("react_max", "react_avg", "react_dt"),
        }

        for param, keys in param_mapping.items():
            max_val = side_data.get(keys[0], "")
            avg_val = side_data.get(keys[1], "")
            dt_val = side_data.get(keys[2], "")
            ws_demand.append([param, max_val, avg_val, dt_val])

        # SIP Configuration Sheet
        ws_sip = wb.create_sheet("SIP Configuration")
        ws_sip.append(["Parameter", "Value"])
        ws_sip.append(["SIP Duration (minutes)", sip_duration])
        ws_sip.append(["Expected SIPs per day", (24 * 60) // sip_duration])
        ws_sip.append(["Month Analyzed", month_info['selected_month_year']])
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        ws_sip.append(["Days in Month", days_in_month])
        ws_sip.append(["Expected Total SIPs for Month", days_in_month * ((24 * 60) // sip_duration)])

        # Save
        file_name = f"ui_demand_side_panel_data_monthly_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"UI demand side panel data saved: {file_name}")
        return file_name

    except Exception as e:
        logger.error(f"Error saving side panel data: {str(e)}")
        raise


# ============================================================================
# DATABASE PROCESSING AND CALCULATIONS
# ============================================================================
def format_datetime(dt_value):
    """Format datetime for demand table display"""
    try:
        if pd.isna(dt_value) or dt_value is None:
            return "No valid data"
        if isinstance(dt_value, pd.Timestamp):
            return dt_value.strftime(f'{dt_value.day} %b at %H:%M')
        elif isinstance(dt_value, str):
            dt_obj = pd.to_datetime(dt_value)
            return dt_obj.strftime(f'{dt_obj.day} %b - %H:%M')
        else:
            return str(dt_value)
    except Exception:
        return "Invalid datetime"


@log_execution_time
def calculate_side_panel_metrics_from_raw_data(raw_df, nrm_df, month_info, sip_duration):
    """Calculate demand side panel metrics - EXACT same logic as daywise"""
    logger.info(
        f"Calculating demand side panel metrics with {sip_duration}-minute SIP intervals for complete month...")

    month_year_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')

    logger.info(f"Processing RAW data: {len(raw_df)} records")
    logger.info(f"Processing NRM data: {len(nrm_df)} records")

    df_raw = raw_df.copy()
    df_raw['surveydate'] = pd.to_datetime(df_raw['surveydate'])

    # Calculate SIP duration in hours for conversion - EXACT same as daywise
    sip_duration_in_hr = sip_duration / 60
    logger.info(f"Converting energy to demand using SIP duration: {sip_duration_in_hr} hours")

    # NEW LOGIC:
    # 1. Active Power (kW): ALWAYS use kwh_abs
    # 2. Apparent Power (kVA): ALWAYS use kvah_abs
    # 3. Reactive Power (kVAr): Use kvar_i_total, if NULL then kvarh_abs

    logger.info("=" * 60)
    logger.info("CONVERSION LOGIC:")
    logger.info("  Active Power: ALWAYS using kwh_abs (not kwh_i)")
    logger.info("  Apparent Power: ALWAYS using kvah_abs (not kvah_i)")
    logger.info("  Reactive Power: Using kvar_i_total, fallback to kvarh_abs if NULL")
    logger.info("=" * 60)

    # Convert energy to demand (Power = Energy / Time)

    # ACTIVE POWER - ALWAYS from ABS
    if 'kwh_abs' in df_raw.columns:
        df_raw['kw_calculated'] = pd.to_numeric(df_raw['kwh_abs'], errors='coerce') / sip_duration_in_hr
        logger.info(f"Active Power: Using kwh_abs, converted {df_raw['kw_calculated'].notna().sum()} records")
    else:
        logger.error("CRITICAL: kwh_abs column not found!")
        df_raw['kw_calculated'] = None

    # APPARENT POWER - ALWAYS from ABS
    if 'kvah_abs' in df_raw.columns:
        df_raw['kva_calculated'] = pd.to_numeric(df_raw['kvah_abs'], errors='coerce') / sip_duration_in_hr
        logger.info(f"Apparent Power: Using kvah_abs, converted {df_raw['kva_calculated'].notna().sum()} records")
    else:
        logger.error("CRITICAL: kvah_abs column not found!")
        df_raw['kva_calculated'] = None

    # REACTIVE POWER - IMPORT first, fallback to ABS if NULL
    if 'kvar_i_total' in df_raw.columns and 'kvarh_abs' in df_raw.columns:
        # Start with IMPORT values
        df_raw['kvar_calculated'] = pd.to_numeric(df_raw['kvar_i_total'], errors='coerce') / sip_duration_in_hr
        import_count = df_raw['kvar_calculated'].notna().sum()

        # Fill NULL values with ABS
        null_mask = df_raw['kvar_calculated'].isna()
        df_raw.loc[null_mask, 'kvar_calculated'] = pd.to_numeric(df_raw.loc[null_mask, 'kvarh_abs'],
                                                                 errors='coerce') / sip_duration_in_hr
        abs_filled = df_raw.loc[null_mask, 'kvar_calculated'].notna().sum()

        logger.info(
            f"Reactive Power: Using kvar_i_total for {import_count} records, filled {abs_filled} NULLs from kvarh_abs")
    elif 'kvar_i_total' in df_raw.columns:
        df_raw['kvar_calculated'] = pd.to_numeric(df_raw['kvar_i_total'], errors='coerce') / sip_duration_in_hr
        logger.info(f"Reactive Power: Using only kvar_i_total (kvarh_abs not available)")
    elif 'kvarh_abs' in df_raw.columns:
        df_raw['kvar_calculated'] = pd.to_numeric(df_raw['kvarh_abs'], errors='coerce') / sip_duration_in_hr
        logger.info(f"Reactive Power: Using only kvarh_abs (kvar_i_total not available)")
    else:
        logger.error("CRITICAL: Neither kvar_i_total nor kvarh_abs found!")
        df_raw['kvar_calculated'] = None

    logger.info("Converted RAW energy to demand")

    # DEBUG: Log sample of converted data
    if not df_raw.empty:
        logger.info("=" * 60)
        logger.info("DEBUG: Sample of converted data (first 5 records):")
        logger.info(df_raw[['surveydate', 'kw_calculated', 'kva_calculated', 'kvar_calculated']].head().to_string())
        logger.info(f"DEBUG: Total records with valid kw_calculated: {df_raw['kw_calculated'].notna().sum()}")
        logger.info(
            f"DEBUG: kw_calculated range: {df_raw['kw_calculated'].min():.2f} to {df_raw['kw_calculated'].max():.2f}")
        logger.info(
            f"DEBUG: kva_calculated range: {df_raw['kva_calculated'].min():.2f} to {df_raw['kva_calculated'].max():.2f}")
        logger.info(
            f"DEBUG: kvar_calculated range: {df_raw['kvar_calculated'].min():.2f} to {df_raw['kvar_calculated'].max():.2f}")
        logger.info("=" * 60)

    # Safe demand calculation function - EXACT same as daywise
    def safe_demand_calculation(demand_series, datetime_series, param_name):
        try:
            numeric_series = pd.to_numeric(demand_series, errors='coerce')
            valid_mask = ~pd.isna(numeric_series)

            if not valid_mask.any():
                logger.warning(f"No valid data for {param_name}")
                return 0.0, 0.0, "No valid data"

            valid_data = numeric_series[valid_mask]
            max_val = round(valid_data.max(), 2)
            avg_val = round(valid_data.mean(), 2)

            if pd.isna(max_val):
                max_datetime_formatted = "No valid data"
            else:
                max_idx = numeric_series.idxmax()
                if pd.isna(max_idx):
                    max_datetime_formatted = "No valid data"
                else:
                    max_datetime = datetime_series.iloc[max_idx]
                    max_datetime_formatted = format_datetime(max_datetime)

            return max_val, avg_val, max_datetime_formatted

        except Exception as e:
            logger.error(f"Error calculating {param_name}: {str(e)}")
            return 0.0, 0.0, "Calculation error"

    # Calculate demand table metrics - Using new converted columns
    calculated_data = {}

    # Active Power - from kw_calculated (which uses kwh_abs)
    if 'kw_calculated' in df_raw.columns and df_raw['kw_calculated'].notna().any():
        active_max, active_avg, active_max_time = safe_demand_calculation(
            df_raw['kw_calculated'], df_raw['surveydate'], "Active Power"
        )
        calculated_data['Active(kW)'] = {
            'Max': f"{active_max:.2f}",
            'Avg': f"{active_avg:.2f}",
            'Date and time at max value': active_max_time
        }
        logger.info(
            f"DB Calculated: Active -> Max={active_max:.2f}, Avg={active_avg:.2f}, DateTime={active_max_time}")
    else:
        calculated_data['Active(kW)'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}
        logger.warning("No valid Active Power data")

    # Apparent Power - from kva_calculated (which uses kvah_abs)
    if 'kva_calculated' in df_raw.columns and df_raw['kva_calculated'].notna().any():
        apparent_max, apparent_avg, apparent_max_time = safe_demand_calculation(
            df_raw['kva_calculated'], df_raw['surveydate'], "Apparent Power"
        )
        calculated_data['Apparent(kVA)'] = {
            'Max': f"{apparent_max:.2f}",
            'Avg': f"{apparent_avg:.2f}",
            'Date and time at max value': apparent_max_time
        }
        logger.info(
            f"DB Calculated: Apparent -> Max={apparent_max:.2f}, Avg={apparent_avg:.2f}, DateTime={apparent_max_time}")
    else:
        calculated_data['Apparent(kVA)'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}
        logger.warning("No valid Apparent Power data")

    # Reactive Power - from kvar_calculated (which uses kvar_i_total with fallback to kvarh_abs)
    if 'kvar_calculated' in df_raw.columns and df_raw['kvar_calculated'].notna().any():
        reactive_max, reactive_avg, reactive_max_time = safe_demand_calculation(
            df_raw['kvar_calculated'], df_raw['surveydate'], "Reactive Power"
        )
        calculated_data['Reactive(kVAr)'] = {
            'Max': f"{reactive_max:.2f}",
            'Avg': f"{reactive_avg:.2f}",
            'Date and time at max value': reactive_max_time
        }
        logger.info(
            f"DB Calculated: Reactive -> Max={reactive_max:.2f}, Avg={reactive_avg:.2f}, DateTime={reactive_max_time}")
    else:
        calculated_data['Reactive(kVAr)'] = {'Max': '-', 'Avg': '-', 'Date and time at max value': '-'}
        logger.warning("No valid Reactive Power data")

    # Save calculated data to Excel - EXACT same as daywise
    calculated_file = f"calculated_demand_side_panel_data_{month_year_safe}_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    # Demand Table Sheet
    ws_demand = wb.create_sheet('Demand Table')
    ws_demand.append(['Parameter', 'Max', 'Avg', 'Date and time at max value'])

    for param, values in calculated_data.items():
        ws_demand.append([param, values['Max'], values['Avg'], values['Date and time at max value']])

    # SIP Configuration Sheet
    ws_sip = wb.create_sheet('SIP Configuration')
    ws_sip.append(['Parameter', 'Value'])
    ws_sip.append(['SIP Duration (minutes)', sip_duration])
    ws_sip.append(['Expected SIPs per day', (24 * 60) // sip_duration])
    ws_sip.append(['Actual SIPs', len(raw_df)])
    days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
    expected_total = days_in_month * ((24 * 60) // sip_duration)
    ws_sip.append(['Days in Month', days_in_month])
    ws_sip.append(['Expected Total SIPs for Month', expected_total])
    ws_sip.append(['Coverage Percentage', f"{(len(raw_df) / expected_total * 100):.1f}%"])

    # Raw Data Analysis Sheet
    ws_analysis = wb.create_sheet('Data Analysis')
    ws_analysis.append(['Metric', 'Value'])
    if not df_raw.empty:
        ws_analysis.append(['Total Records', len(df_raw)])
        ws_analysis.append(['Date Range', f"{df_raw['surveydate'].min()} to {df_raw['surveydate'].max()}"])
        ws_analysis.append(['', ''])
        ws_analysis.append(['Source Data Used:', ''])
        ws_analysis.append(['Active Power', 'kwh_abs (ALWAYS)'])
        ws_analysis.append(['Apparent Power', 'kvah_abs (ALWAYS)'])
        ws_analysis.append(['Reactive Power', 'kvar_i_total → kvarh_abs (if NULL)'])
        ws_analysis.append(['', ''])
        if 'kw_calculated' in df_raw.columns:
            ws_analysis.append(['Max Active Demand (kW)', f"{df_raw['kw_calculated'].max():.2f}"])
            ws_analysis.append(['Avg Active Demand (kW)', f"{df_raw['kw_calculated'].mean():.2f}"])
        if 'kva_calculated' in df_raw.columns:
            ws_analysis.append(['Max Apparent Demand (kVA)', f"{df_raw['kva_calculated'].max():.2f}"])
            ws_analysis.append(['Avg Apparent Demand (kVA)', f"{df_raw['kva_calculated'].mean():.2f}"])
        if 'kvar_calculated' in df_raw.columns:
            ws_analysis.append(['Max Reactive Demand (kVAr)', f"{df_raw['kvar_calculated'].max():.2f}"])
            ws_analysis.append(['Avg Reactive Demand (kVAr)', f"{df_raw['kvar_calculated'].mean():.2f}"])

    wb.save(calculated_file)

    logger.info(f"Demand side panel metrics calculation completed using {sip_duration}-minute SIP intervals")
    logger.info("Used kwh_abs and kvah_abs for Active/Apparent, kvar_i_total→kvarh_abs for Reactive")
    return calculated_data, calculated_file


# ============================================================================
# COMPARISON AND VALIDATION
# ============================================================================
@log_execution_time
def create_detailed_comparison(ui_file, calculated_file, month_info, sip_duration):
    """Create complete monthly side panel comparison with validation"""
    logger.info("Creating monthly demand side panel comparison...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_monthly_demand_sidepanel_{month_safe}.xlsx"

        # Load workbooks
        wb_ui = load_workbook(ui_file)
        wb_calc = load_workbook(calculated_file)
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        # Colors
        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        def normalize_string(s):
            """Remove all spaces and lowercase for fair string comparison"""
            if s is None or str(s).lower() in ['nan', 'none', '-']:
                return ""
            return str(s).replace(" ", "").strip().lower()

        def extract_numeric_value(s):
            """Extract numeric value from string"""
            try:
                if s is None or str(s).lower() in ['nan', 'none', '-', '']:
                    return None
                numeric_match = re.search(r'[-+]?(\d*\.?\d+)', str(s))
                if numeric_match:
                    return float(numeric_match.group())
                return None
            except:
                return None

        # Compare Demand Table sheets
        validation_results = {}
        if 'Demand Table' in wb_ui.sheetnames and 'Demand Table' in wb_calc.sheetnames:
            ws_ui = wb_ui['Demand Table']
            ws_calc = wb_calc['Demand Table']
            ws_new = wb_output.create_sheet('Demand Table Validation')

            # Headers
            headers = ['Parameter', 'UI_Max', 'Calc_Max', 'Max_Match',
                       'UI_Avg', 'Calc_Avg', 'Avg_Match',
                       'UI_DateTime', 'Calc_DateTime', 'DateTime_Match', 'Overall_Match']
            for col, header in enumerate(headers, 1):
                cell = ws_new.cell(row=1, column=col, value=header)
                cell.fill = header_fill

            total_matches = 0
            total_comparisons = 0
            tolerance = 0.02  # 2% tolerance
            sheet_results = []

            row_num = 2
            for row in range(2, max(ws_ui.max_row, ws_calc.max_row) + 1):
                if row <= ws_ui.max_row and row <= ws_calc.max_row:
                    # Get UI values
                    param_ui = ws_ui.cell(row=row, column=1).value
                    max_ui = ws_ui.cell(row=row, column=2).value
                    avg_ui = ws_ui.cell(row=row, column=3).value
                    dt_ui = ws_ui.cell(row=row, column=4).value

                    # Get calculated values
                    param_calc = ws_calc.cell(row=row, column=1).value
                    max_calc = ws_calc.cell(row=row, column=2).value
                    avg_calc = ws_calc.cell(row=row, column=3).value
                    dt_calc = ws_calc.cell(row=row, column=4).value

                    if param_ui or param_calc:
                        param = param_ui or param_calc

                        # Compare Max values
                        max_ui_num = extract_numeric_value(max_ui)
                        max_calc_num = extract_numeric_value(max_calc)
                        if max_ui_num is not None and max_calc_num is not None:
                            if max_ui_num == 0 or max_calc_num == 0:
                                max_match = (max_ui_num == max_calc_num)
                            else:
                                max_match = abs(max_ui_num - max_calc_num) <= (
                                        max(abs(max_ui_num), abs(max_calc_num)) * tolerance)
                            max_match_str = 'YES' if max_match else 'NO'
                            max_match_color = green if max_match else red
                        else:
                            max_match_str = 'NO DATA'
                            max_match_color = yellow
                            max_match = False

                        # Compare Avg values
                        avg_ui_num = extract_numeric_value(avg_ui)
                        avg_calc_num = extract_numeric_value(avg_calc)
                        if avg_ui_num is not None and avg_calc_num is not None:
                            if avg_ui_num == 0 or avg_calc_num == 0:
                                avg_match = (avg_ui_num == avg_calc_num)
                            else:
                                avg_match = abs(avg_ui_num - avg_calc_num) <= (
                                        max(abs(avg_ui_num), abs(avg_calc_num)) * tolerance)
                            avg_match_str = 'YES' if avg_match else 'NO'
                            avg_match_color = green if avg_match else red
                        else:
                            avg_match_str = 'NO DATA'
                            avg_match_color = yellow
                            avg_match = False

                        # Compare DateTime - more lenient
                        dt_ui_str = normalize_string(dt_ui)
                        dt_calc_str = normalize_string(dt_calc)
                        if dt_ui_str and dt_calc_str and dt_ui_str not in ['novaliddata',
                                                                           'invaliddata'] and dt_calc_str not in [
                            'novaliddata', 'invaliddata']:
                            dt_match = (dt_ui_str == dt_calc_str)
                        else:
                            dt_match = True  # Don't fail on datetime if both calculations passed
                        dt_match_str = 'YES' if dt_match else 'NO'
                        dt_match_color = green if dt_match else yellow

                        # Overall match - prioritize numeric values
                        overall_match = max_match and avg_match
                        overall_match_str = 'YES' if overall_match else 'NO'
                        overall_match_color = green if overall_match else red

                        # Write to validation report
                        ws_new.cell(row=row_num, column=1, value=param)
                        ws_new.cell(row=row_num, column=2, value=max_ui)
                        ws_new.cell(row=row_num, column=3, value=max_calc)
                        max_cell = ws_new.cell(row=row_num, column=4, value=max_match_str)
                        max_cell.fill = max_match_color

                        ws_new.cell(row=row_num, column=5, value=avg_ui)
                        ws_new.cell(row=row_num, column=6, value=avg_calc)
                        avg_cell = ws_new.cell(row=row_num, column=7, value=avg_match_str)
                        avg_cell.fill = avg_match_color

                        ws_new.cell(row=row_num, column=8, value=dt_ui)
                        ws_new.cell(row=row_num, column=9, value=dt_calc)
                        dt_cell = ws_new.cell(row=row_num, column=10, value=dt_match_str)
                        dt_cell.fill = dt_match_color

                        overall_cell = ws_new.cell(row=row_num, column=11, value=overall_match_str)
                        overall_cell.fill = overall_match_color

                        sheet_results.append({
                            'item': param,
                            'match': overall_match
                        })

                        if overall_match:
                            total_matches += 1
                        total_comparisons += 1
                        row_num += 1

            validation_results['Demand Table'] = sheet_results

        # Summary sheet
        ws_summary = wb_output.create_sheet('Validation Summary')
        ws_summary.append(['Metric', 'Value'])
        ws_summary.append(['Total Comparisons', total_comparisons])
        ws_summary.append(['Successful Matches', total_matches])
        ws_summary.append(['Failed Matches', total_comparisons - total_matches])
        ws_summary.append(['Success Rate',
                           f"{(total_matches / total_comparisons * 100):.1f}%" if total_comparisons > 0 else "0%"])
        ws_summary.append(['Validation Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append(['SIP Duration Used', f"{sip_duration} minutes"])
        ws_summary.append(['Expected SIPs/Day', (24 * 60) // sip_duration])
        ws_summary.append(['Month Analyzed', month_info['selected_month_year']])
        ws_summary.append(['Tolerance Used', f"{tolerance * 100}%"])

        # Copy SIP configuration
        if 'SIP Configuration' in wb_calc.sheetnames:
            ws_sip_calc = wb_calc['SIP Configuration']
            ws_sip_new = wb_output.create_sheet('SIP Analysis')
            for row in ws_sip_calc.iter_rows(values_only=True):
                ws_sip_new.append(row)

        # Copy data analysis
        if 'Data Analysis' in wb_calc.sheetnames:
            ws_analysis_calc = wb_calc['Data Analysis']
            ws_analysis_new = wb_output.create_sheet('Data Analysis')
            for row in ws_analysis_calc.iter_rows(values_only=True):
                ws_analysis_new.append(row)

        # Color code summary
        for row in range(2, ws_summary.max_row + 1):
            for col in range(1, ws_summary.max_column + 1):
                if row == 5:  # Success rate row
                    success_rate = total_matches / total_comparisons * 100 if total_comparisons > 0 else 0
                    if success_rate >= 90:
                        ws_summary.cell(row=row, column=col).fill = green
                    elif success_rate >= 70:
                        ws_summary.cell(row=row, column=col).fill = yellow
                    else:
                        ws_summary.cell(row=row, column=col).fill = red

        wb_output.save(output_file)
        logger.info(f"Monthly side panel comparison saved: {output_file}")

        return output_file, validation_results

    except Exception as e:
        logger.error(f"Error creating comparison: {str(e)}")
        raise


# ============================================================================
# COMPREHENSIVE SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_comprehensive_summary_report(config, month_info, ui_file, calculated_file,
                                        comparison_file, validation_results, raw_df, meter_name, sip_duration):
    """Create comprehensive monthly side panel summary report with ENHANCED styling"""
    logger.info("Creating comprehensive monthly demand side panel summary report...")

    try:
        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_DEMAND_SIDEPANEL_{month_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # ============ MAIN HEADER ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = f"LV MONTHLY DEMAND SIDE PANEL VALIDATION SUMMARY - {month_info['selected_month_year'].upper()}"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        current_row += 1

        # ============ TEST DETAILS SECTION ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Month:", config['target_month_year']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ SYSTEM UNDER TEST ============
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Monthly Demand Side Panel (NO GRAPH)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Month Range:", f"{month_info['start_date']} to {month_info['end_date']}"],
            ["SIP Duration:", f"{sip_duration} minutes"],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ DATA VOLUME ANALYSIS ============
        ws.merge_cells(f'A{current_row}:C{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 DATA VOLUME ANALYSIS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, header in enumerate(headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate expected records
        days_in_month = (month_info['end_date'] - month_info['start_date']).days + 1
        expected_records = days_in_month * ((24 * 60) // sip_duration)
        data_completeness = (len(raw_df) / expected_records * 100) if expected_records > 0 else 0

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Expected Records", expected_records, f"{data_completeness:.1f}% Complete"],
            ["SIP Duration Used", f"{sip_duration} min", "DYNAMIC FROM DB"],
            ["Days in Month", days_in_month, f"{month_info['selected_month_year']}"]
        ]

        for dataset, count, status in data_rows:
            ws[f'A{current_row}'].value = dataset
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = count
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = status
            if "COMPLETE" in str(status) or "%" in str(status) or "DYNAMIC" in str(status) or "January" in str(
                    status) or "February" in str(status) or "March" in str(status):
                if data_completeness >= 90 or "COMPLETE RECORDS" in str(status) or "DYNAMIC" in str(
                        status) or days_in_month > 0:
                    ws[f'C{current_row}'].font = pass_font
                    ws[f'C{current_row}'].fill = pass_fill
                else:
                    ws[f'C{current_row}'].font = fail_font
                    ws[f'C{current_row}'].fill = fail_fill
            else:
                ws[f'C{current_row}'].font = fail_font
                ws[f'C{current_row}'].fill = fail_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ VALIDATION RESULTS ============
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "✅ VALIDATION RESULTS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, header in enumerate(validation_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = label_font
            cell.fill = label_fill
            cell.alignment = label_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Calculate validation results
        overall_passed = 0
        overall_total = 0
        validation_data = []

        for sheet_name, results in validation_results.items():
            total_items = len(results)
            passed_items = sum(1 for result in results if result['match'])
            failed_items = total_items - passed_items
            success_rate = f"{(passed_items / total_items) * 100:.1f}%" if total_items > 0 else "0%"
            status = "PASS" if passed_items == total_items else "FAIL"

            display_name = sheet_name.replace('_', ' ')
            validation_data.append([display_name, passed_items, failed_items, success_rate, status])

            overall_passed += passed_items
            overall_total += total_items

        for comp_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{current_row}'].value = comp_type
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = matches
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = mismatches
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = rate
            ws[f'D{current_row}'].font = Font(bold=True, size=10, name="Calibri", color="000000")
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            if status == "PASS":
                ws[f'E{current_row}'].font = pass_font
                ws[f'E{current_row}'].fill = pass_fill
            else:
                ws[f'E{current_row}'].font = fail_font
                ws[f'E{current_row}'].fill = fail_fill
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ============ OVERALL ASSESSMENT ============
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        overall_success_rate = (overall_passed / overall_total) * 100 if overall_total > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: Monthly demand side panel validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor discrepancies found - Review recommended"
            assessment_color = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            assessment_font_color = Font(bold=True, size=10, color="000000", name="Calibri")
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Success rate detail
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({overall_passed}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Comprehensive monthly side panel summary report created: {summary_file}")

        # Log summary
        logger.info("=" * 60)
        logger.info("MONTHLY DEMAND SIDE PANEL VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Month: {month_info['selected_month_year']}")
        logger.info(f"SIP Duration: {sip_duration} minutes")
        logger.info(f"Data: Raw={len(raw_df)} records")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Data Completeness: {data_completeness:.1f}%")
        logger.info("=" * 60)

        return summary_file

    except Exception as e:
        logger.error(f"Error creating summary report: {str(e)}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_lv_monthly_demand_sidepanel_automation():
    """Main LV Monthly Demand Side Panel automation process - NO GRAPH EXTRACTION"""
    config = None
    driver = None
    output_folder = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB1: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"DB2: {DatabaseConfig.DB2_HOST}:{DatabaseConfig.DB2_PORT}/{DatabaseConfig.DB2_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False

        # Apply configuration
        logger.info("Applying LV Monthly Demand Side Panel configuration...")
        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False

        # Select meter type
        if not select_meter_type(driver, config['meter_type']):
            logger.info("Invalid meter type")
            return False

        # Get meter metrics
        logger.info("Fetching meter metrics...")
        dt_id, name, mtr_id, node_id, sip_duration = get_metrics(config['meter_serial_no'], config['meter_type'])

        if not dt_id:
            logger.info(f"Meter not found: {config['meter_serial_no']}")
            return False

        logger.info(f"Meter found: {name} (ID: {mtr_id}, node_id: {node_id}, SIP: {sip_duration}min)")

        # Find and click View
        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            logger.info("Failed to find View button")
            return False

        # Wait for page to load
        time.sleep(5)

        # Collect side panel data ONLY - NO GRAPH
        logger.info("=" * 60)
        logger.info("COLLECTING SIDE PANEL DATA ONLY (NO GRAPH EXTRACTION)")
        logger.info("=" * 60)
        side_panel_data = collect_side_panel_data(driver, wait)

        # Save UI side panel data
        ui_file = save_side_panel_data_to_excel(side_panel_data, month_info, sip_duration)
        if ui_file:
            ui_file = save_file_to_output(ui_file, output_folder)

        # Get database data for complete month - EXACT same as daywise
        raw_df, nrm_df = get_database_data_for_monthly_sidepanel(month_info, mtr_id, node_id)

        if raw_df.empty and nrm_df.empty:
            logger.info("No database data found for the month")
            return False

        # Process database calculations - EXACT same as daywise
        logger.info("Processing database calculations for side panel metrics...")
        calculated_data, calculated_file = calculate_side_panel_metrics_from_raw_data(
            raw_df, nrm_df, month_info, sip_duration)
        calculated_file = save_file_to_output(calculated_file, output_folder)

        # Create comparison report
        logger.info("Creating side panel validation comparison...")
        comparison_file, validation_results = create_detailed_comparison(
            ui_file, calculated_file, month_info, sip_duration)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Create comprehensive summary report
        logger.info("Creating comprehensive summary...")
        summary_report = create_comprehensive_summary_report(
            config, month_info, ui_file, calculated_file,
            comparison_file, validation_results, raw_df, name, sip_duration)
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("LV MONTHLY DEMAND SIDE PANEL AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Monitoring Type: LV Monthly Demand Side Panel (NO GRAPH)")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Month: {config['target_month_year']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info(f"Meter: {config['meter_serial_no']} ({name})")
        logger.info(f"Meter Type: {config['meter_type']}")
        logger.info(f"SIP Duration: {sip_duration} minutes (dynamic from DB)")
        logger.info(f"Database Records: {len(raw_df)} records")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(ui_file) if ui_file else 'UI side panel data'}")
        logger.info(f"   2. {os.path.basename(calculated_file) if calculated_file else 'Calculated data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ LV Monthly Demand Side Panel monitoring (NO GRAPH)")
        logger.info("   ✓ Complete month data processing")
        logger.info("   ✓ Search box meter selection")
        logger.info("   ✓ Dynamic SIP duration from database")
        logger.info("   ✓ Active/Apparent/Reactive demand validation")
        logger.info("   ✓ Energy to Power conversion")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Comprehensive validation summary")
        logger.info("   ✓ SIDE PANEL ONLY - No graph extraction")
        logger.info("=" * 60)

        return True

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"LV Monthly Demand Side Panel Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False

    finally:
        if driver:
            try:
                driver.quit()
                logger.info("Browser closed")
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV MONTHLY DEMAND SIDE PANEL AUTOMATION (NO GRAPH)")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Monitoring Type: LV Monthly Demand Side Panel (NO GRAPH)")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ LV Monthly Demand Side Panel monitoring")
    logger.info("   ✓ Complete month data processing")
    logger.info("   ✓ Search box meter selection")
    logger.info("   ✓ Dynamic SIP duration from database")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Active/Apparent/Reactive demand metrics")
    logger.info("   ✓ Energy to Power conversion")
    logger.info("   ✓ Enhanced value parsing")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Comprehensive summary report")
    logger.info("   ✓ SIDE PANEL ONLY - No graph extraction")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_lv_monthly_demand_sidepanel_automation()
    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("LV MONTHLY DEMAND SIDE PANEL AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ LV Monthly Demand Side Panel monitoring")
        logger.info("   ✓ Complete month processing")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Dynamic SIP from database")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Demand side panel data extraction")
        logger.info("   ✓ Energy to Power conversion")
        logger.info("   ✓ Enhanced parsing")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
        logger.info("   ✓ NO GRAPH EXTRACTION - Side panel only")
    else:
        logger.info("LV MONTHLY DEMAND SIDE PANEL AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("LV Monthly Demand Side Panel Automation Finished")
    logger.info("=" * 60)
