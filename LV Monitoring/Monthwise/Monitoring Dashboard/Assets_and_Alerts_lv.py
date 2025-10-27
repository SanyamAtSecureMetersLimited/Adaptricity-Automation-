import os
import time
import shutil
import logging
import functools
import pandas as pd
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION - EASY TO MODIFY
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""

    # Database Configuration
    DB_HOST = "10.11.16.146"
    DB_PORT = "5434"
    DB_DATABASE = "Prod_LVMV_Test"
    DB_USER = "postgres"
    DB_PASSWORD = "postgres"

    # Tenant Configuration - Change this if needed
    TENANT_NAME = "tenant01"  # Change to tenant02, tenant03, etc. as needed

    @classmethod
    def get_db_params(cls):
        return {
            "host": cls.DB_HOST,
            "port": cls.DB_PORT,
            "database": cls.DB_DATABASE,
            "user": cls.DB_USER,
            "password": cls.DB_PASSWORD
        }


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    """Setup simple logging system"""
    if not os.path.exists('logs'):
        os.makedirs('logs')

    logger = logging.getLogger('monthly_dashboard_validation')
    logger.setLevel(logging.INFO)

    for handler in logger.handlers[:]:
        logger.removeHandler(handler)

    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

    log_file = 'logs/monthly_dashboard_validation.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    """Create output folder and clean previous run files"""
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    """Move generated file to output folder"""
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    """Create default configuration Excel file for Monthly Dashboard Validation"""
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'MV_Feeder', 'Target_Month_Year'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'April 2025']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)

            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7', '8'],
                'Instructions': [
                    'Open the "User_Configuration" sheet',
                    'Replace "YOUR_AREA_HERE" with your actual area name',
                    'Replace "YOUR_SUBSTATION_HERE" with your actual substation name',
                    'Replace "YOUR_FEEDER_HERE" with your actual feeder name',
                    'Update Target_Month_Year with desired month (Format: Month YYYY)',
                    'Values must match EXACTLY with dropdown options',
                    'Month format: January 2025, February 2024, etc.',
                    'Save file before running'
                ],
                'Important_Notes': [
                    'This script is FOR MONTHLY DASHBOARD VALIDATION',
                    'Values are case-sensitive',
                    'No extra spaces before/after values',
                    'Month format: Month YYYY (e.g., April 2025)',
                    'Validates Assets (DTs, LV Feeders)',
                    'Validates Alerts (OV, VUB, PPF, HC)',
                    'Test Engineer: Sanyam Upadhyay',
                    'Analysis period: Entire month'
                ]
            }
            df_instructions = pd.DataFrame(instructions)
            df_instructions.to_excel(writer, sheet_name='Setup_Instructions', index=False)

        logger.info(f"Monthly Dashboard Validation Configuration template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config file: {e}")
        return False


def read_user_configuration(config_file="user_config.xlsx"):
    """Read user configuration from Excel file for Monthly Dashboard Validation"""
    try:
        if not os.path.exists(config_file):
            logger.info(f"Configuration file not found: {config_file}")
            return None

        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {}

        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'MV_Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Month_Year':
                config['target_month_year'] = str(value).strip()

        required_fields = ['area', 'substation', 'feeder', 'target_month_year']
        missing_fields = [f for f in required_fields if f not in config or not config[f]]
        if missing_fields:
            logger.info(f"Missing required configuration: {missing_fields}")
            return None

        placeholders = ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE']
        for key, value in config.items():
            if value in placeholders:
                logger.info(f"Placeholder value found: {key} = {value}")
                return None

        logger.info("Monthly Dashboard Validation Configuration loaded successfully")
        return config
    except Exception as e:
        logger.info(f"Error reading configuration file: {e}")
        return None


def validate_config_at_startup():
    """Validate configuration before starting browser"""
    logger.info("=" * 60)
    logger.info("STARTING MONTHLY DASHBOARD VALIDATION AUTOMATION")
    logger.info("=" * 60)

    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info(f"Configuration file not found: {config_file}")
        logger.info("Creating default Monthly Dashboard Validation configuration template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the configuration file and restart")
        return None

    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None

    logger.info("Monthly Dashboard Validation Configuration validated successfully")
    logger.info(f"   Area: {config['area']}")
    logger.info(f"   Substation: {config['substation']}")
    logger.info(f"   MV Feeder: {config['feeder']}")
    logger.info(f"   Target Month: {config['target_month_year']}")
    return config


# ============================================================================
# DECORATOR FOR EXECUTION TIME
# ============================================================================
def log_execution_time(func):
    """Decorator to log function execution time"""

    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start_time:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# WEB AUTOMATION FUNCTIONS
# ============================================================================
def login(driver):
    """Login to web application"""
    try:
        logger.info("Logging in...")
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("SANYAM")
        driver.find_element(By.ID, "Password").send_keys("Sanyam@1234")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(7)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    """Select dropdown option"""
    try:
        logger.info(f"Selecting {option_name} in {dropdown_id}")
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        options = driver.find_elements(By.CSS_SELECTOR, ".dx-list-item")
        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        logger.info(f"Option not found: {option_name}")
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_month(driver, target_month_year):
    """Set calendar to target month and return month info"""
    logger.info(f"Setting calendar to: {target_month_year}")

    try:
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Month']").click()
        month_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")

        month_input.clear()
        month_input.send_keys(target_month_year)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()

        # Parse month info
        month_name, year = target_month_year.split()
        month_num = datetime.strptime(month_name, "%B").month
        year = int(year)

        # Create month boundaries
        start_date = datetime(year, month_num, 1).date()
        if month_num == 12:
            end_date = datetime(year + 1, 1, 1).date() - pd.Timedelta(days=1)
        else:
            end_date = datetime(year, month_num + 1, 1).date() - pd.Timedelta(days=1)

        month_info = {
            'selected_month_year': target_month_year,
            'month_num': month_num,
            'year': year,
            'start_date': start_date,
            'end_date': end_date
        }

        logger.info(f"Calendar set successfully: {target_month_year}")
        logger.info(f"Analysis period: {start_date} to {end_date}")
        return month_info

    except Exception as e:
        logger.info(f"Error setting calendar month: {e}")
        return None


@log_execution_time
def collect_data(driver):
    """Collect dashboard data"""
    logger.info("Collecting data from dashboard...")
    data = {}
    data['Assets'] = {
        'DTs': driver.find_element(By.XPATH, '//span[@id="dtmonitoringdtcount"]').text,
        'LV Feeders': driver.find_element(By.XPATH, '//span[@id="dtmonitoringlvfeedercount"]').text
    }
    data['Alerts'] = {
        'Over Voltage': driver.find_element(By.XPATH, '//span[@id="dtmonitoringovervoltagecount"]').text,
        'Voltage Unbalance': driver.find_element(By.XPATH, '//span[@id="dtmonitoringvoltageunbalancecount"]').text,
        'Poor Power Factor': driver.find_element(By.XPATH, '//span[@id="dtmonitoringpoorpfcount"]').text,
        "High Current": driver.find_element(By.XPATH, '//span[@id="dtmonitoringhighcurrentcount"]').text
    }
    logger.info(f"Dashboard data collected successfully")
    return data


@log_execution_time
def save_date_to_excel(month_info, data):
    """Save dashboard UI data to Excel"""
    logger.info("Saving dashboard data to Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    # Assets Sheet
    ws_assets = wb.create_sheet("Assets")
    ws_assets.append(["Parameter", "Value"])
    for key, value in data.get("Assets", {}).items():
        ws_assets.append([key, value])

    # Alerts Sheet
    ws_alerts = wb.create_sheet("Alerts")
    ws_alerts.append(["Parameter", "Value"])
    for key, value in data.get("Alerts", {}).items():
        ws_alerts.append([key, value])

    filename = f"chart_data_from_ui_dashboard_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    logger.info(f"Chart data saved: {filename}")
    return filename


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_dts():
    """Fetch active DT information from database"""
    logger.info("Fetching DT information from database...")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db_params())
        cursor = conn.cursor()
        cursor.execute(f"SELECT meter_serial_no FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE isactive = 1;")
        rows = cursor.fetchall()
        dts = [row[0] for row in rows]
        cursor.close()
        conn.close()
        logger.info(f"Found {len(dts)} active DTs")
        return dts
    except Exception as e:
        logger.info(f"Error fetching DTs: {e}")
        return []


@log_execution_time
def get_lvfs():
    """Fetch active LV Feeder information from database"""
    logger.info("Fetching LV Feeder information from database...")
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db_params())
        cursor = conn.cursor()
        cursor.execute(f"SELECT meter_serial_no FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE isactive = 1;")
        rows = cursor.fetchall()
        lvs = [row[0] for row in rows]
        cursor.close()
        conn.close()
        logger.info(f"Found {len(lvs)} active LV Feeders")
        return lvs
    except Exception as e:
        logger.info(f"Error fetching LV Feeders: {e}")
        return []


def get_metrics_dt(mtr_serial_no, nodetypeid=153):
    """Get DT metrics with meter name"""
    conn = psycopg2.connect(**DatabaseConfig.get_db_params())
    cursor = conn.cursor()

    try:
        query1 = f"""
            SELECT meterid, dt_name
            FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt
            WHERE meter_serial_no = %s
            LIMIT 1;
        """
        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()

        if not result1:
            return (None,) * 9

        meterid, meter_name = result1

        query2 = f"""
            SELECT voltagerating, currentrating
            FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail
            WHERE mtrid = %s
            LIMIT 1;
        """
        cursor.execute(query2, (meterid,))
        result2 = cursor.fetchone()

        if not result2:
            return meterid, meter_name, None, None, None, None, None, None, None

        voltagerating, currentrating = result2

        cursor.execute("""
            SELECT overvoltage, voltageunbalance
            FROM servicemeta.tb_voltage_threshold_configuration
            WHERE nodetypeid = %s AND voltagerating = %s
            LIMIT 1;
        """, (nodetypeid, voltagerating))
        result3 = cursor.fetchone()
        overvoltage, voltageunbalance = result3 if result3 else (None, None)

        cursor.execute("""
            SELECT overload, currentunbalance
            FROM servicemeta.tb_current_threshold_configuration
            WHERE nodetypeid = %s AND currentrating = %s
            LIMIT 1;
        """, (nodetypeid, currentrating))
        result4 = cursor.fetchone()
        overload, currentunbalance = result4 if result4 else (None, None)

        cursor.execute("""
            SELECT powerfactorthreshold
            FROM servicemeta.tb_powerfactor_threshold_configuration
            WHERE nodetypeid = %s
            LIMIT 1;
        """, (nodetypeid,))
        result5 = cursor.fetchone()
        powerfactorthreshold = result5[0] if result5 else None

        return meterid, meter_name, voltagerating, overvoltage, voltageunbalance, currentrating, overload, currentunbalance, powerfactorthreshold

    except Exception as e:
        logger.info(f"Error getting metrics for DT {mtr_serial_no}: {e}")
        return (None,) * 9
    finally:
        cursor.close()
        conn.close()


def get_metrics_lvf(mtr_serial_no, nodetypeid=157):
    """Get LV Feeder metrics with meter name"""
    conn = psycopg2.connect(**DatabaseConfig.get_db_params())
    cursor = conn.cursor()

    try:
        query1 = f"""
            SELECT lvfeeder_id, meterid, lvfeeder_name
            FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder
            WHERE meter_serial_no = %s
            LIMIT 1;
        """
        cursor.execute(query1, (mtr_serial_no,))
        result1 = cursor.fetchone()

        if not result1:
            return (None,) * 10

        lvfeeder_id, meterid, meter_name = result1

        query2 = f"""
            SELECT voltagerating, currentrating
            FROM {DatabaseConfig.TENANT_NAME}.tb_metermasterdetail
            WHERE mtrid = %s
            LIMIT 1;
        """
        cursor.execute(query2, (meterid,))
        result2 = cursor.fetchone()

        if not result2:
            return lvfeeder_id, meterid, meter_name, None, None, None, None, None, None, None

        voltagerating, currentrating = result2

        cursor.execute("""
            SELECT overvoltage, voltageunbalance
            FROM servicemeta.tb_voltage_threshold_configuration
            WHERE nodetypeid = %s AND voltagerating = %s
            LIMIT 1;
        """, (nodetypeid, voltagerating))
        result3 = cursor.fetchone()
        overvoltage, voltageunbalance = result3 if result3 else (None, None)

        cursor.execute("""
            SELECT overload, currentunbalance
            FROM servicemeta.tb_current_threshold_configuration
            WHERE nodetypeid = %s AND currentrating = %s
            LIMIT 1;
        """, (nodetypeid, currentrating))
        result4 = cursor.fetchone()
        overload, currentunbalance = result4 if result4 else (None, None)

        cursor.execute("""
            SELECT powerfactorthreshold
            FROM servicemeta.tb_powerfactor_threshold_configuration
            WHERE nodetypeid = %s
            LIMIT 1;
        """, (nodetypeid,))
        result5 = cursor.fetchone()
        powerfactorthreshold = result5[0] if result5 else None

        return lvfeeder_id, meterid, meter_name, voltagerating, overvoltage, voltageunbalance, currentrating, overload, currentunbalance, powerfactorthreshold

    except Exception as e:
        logger.info(f"Error getting metrics for LVF {mtr_serial_no}: {e}")
        return (None,) * 10
    finally:
        cursor.close()
        conn.close()


def get_raw_database_data_dt(month_info, mtr_id):
    """Get raw database data for DT for entire month"""
    start_date = month_info["start_date"].strftime("%Y-%m-%d")
    next_day = (month_info["end_date"] + timedelta(days=1)).strftime("%Y-%m-%d")

    query = f"""
        SELECT DISTINCT surveydate, v1, v2, v3, avg_v, i1_line, i2_line, i3_line, avg_i, pf
        FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata
        WHERE mtrid = %s AND surveydate >= %s AND surveydate < %s
        ORDER BY surveydate ASC;
    """

    try:
        with psycopg2.connect(**DatabaseConfig.get_db_params()) as conn:
            raw_df = pd.read_sql(query, conn, params=(mtr_id, start_date, next_day))
        return raw_df
    except Exception as e:
        logger.info(f"Error fetching raw data: {e}")
        return pd.DataFrame()


def get_raw_database_data_lvf(month_info, mtrid):
    """Get raw database data for LV Feeder for entire month"""
    start_date = month_info["start_date"].strftime("%Y-%m-%d")
    next_day = (month_info["end_date"] + timedelta(days=1)).strftime("%Y-%m-%d")

    query = f"""
        SELECT DISTINCT surveydate, v1, v2, v3, avg_v, i1_line, i2_line, i3_line, avg_i, pf
        FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata
        WHERE mtrid = %s AND surveydate >= %s AND surveydate < %s
        ORDER BY surveydate ASC;
    """

    try:
        with psycopg2.connect(**DatabaseConfig.get_db_params()) as conn:
            raw_df = pd.read_sql(query, conn, params=(mtrid, start_date, next_day))
        return raw_df
    except Exception as e:
        logger.info(f"Error fetching raw data: {e}")
        return pd.DataFrame()


# ============================================================================
# CALCULATION FUNCTIONS
# ============================================================================
def calculate_over_voltage_events(raw_df, voltagerating, overvoltage):
    """Calculate over voltage events"""
    if raw_df.empty or voltagerating is None or overvoltage is None:
        return 0

    try:
        threshold = voltagerating + (overvoltage / 100) * voltagerating
        over_mask = (raw_df['v1'] > threshold) | (raw_df['v2'] > threshold) | (raw_df['v3'] > threshold)
        return 1 if over_mask.sum() > 0 else 0
    except Exception as e:
        logger.info(f"Error calculating over voltage events: {e}")
        return 0


def calculate_voltage_unbalance_events(raw_df, voltageunbalance):
    """Calculate voltage unbalance events"""
    if raw_df.empty or 'avg_v' not in raw_df.columns or voltageunbalance is None:
        return 0

    try:
        df = raw_df.copy()
        df['v1_dev'] = abs(df['v1'] - df['avg_v'])
        df['v2_dev'] = abs(df['v2'] - df['avg_v'])
        df['v3_dev'] = abs(df['v3'] - df['avg_v'])
        df['max_dev'] = df[['v1_dev', 'v2_dev', 'v3_dev']].max(axis=1)
        df['unbalance_pct'] = (df['max_dev'] / df['avg_v']) * 100
        unbalance_mask = df['unbalance_pct'] > float(voltageunbalance)
        return 1 if unbalance_mask.sum() > 0 else 0
    except Exception as e:
        logger.info(f"Error calculating voltage unbalance events: {e}")
        return 0


def calculate_high_current_events(raw_df, currentrating, overload):
    """Calculate high current events"""
    if raw_df.empty or currentrating is None or overload is None:
        return 0

    if not all(col in raw_df.columns for col in ['i1_line', 'i2_line', 'i3_line']):
        return 0

    try:
        high_current_threshold = currentrating + (overload / 100) * currentrating
        over_mask = (raw_df['i1_line'] > high_current_threshold) | \
                    (raw_df['i2_line'] > high_current_threshold) | \
                    (raw_df['i3_line'] > high_current_threshold)
        return 1 if over_mask.sum() > 0 else 0
    except Exception as e:
        logger.info(f"Error calculating high current events: {e}")
        return 0


def calculate_low_power_factor_events(raw_df, powerfactor_threshold):
    """Calculate low power factor events"""
    if raw_df.empty or 'pf' not in raw_df.columns or powerfactor_threshold is None:
        return 0

    try:
        low_pf_mask = raw_df['pf'] < float(powerfactor_threshold)
        return 1 if low_pf_mask.sum() > 0 else 0
    except Exception as e:
        logger.info(f"Error calculating low power factor events: {e}")
        return 0


# ============================================================================
# SAVE PROCESSED DATA
# ============================================================================
@log_execution_time
def save_calculated_data(month_info, ov_count, vub_count, hc_count, ppf_count, dt_count, lvf_count, alert_contributors):
    """Save calculated data with meter names"""
    logger.info("Saving calculated data to Excel with alert contributors...")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Assets
    ws_assets = wb.create_sheet("Assets")
    ws_assets.append(["Parameter", "Value"])
    ws_assets.append(["DTs", dt_count])
    ws_assets.append(["LV Feeders", lvf_count])

    # Sheet 2: Alerts
    ws_alerts = wb.create_sheet("Alerts")
    ws_alerts.append(["Parameter", "Value"])
    ws_alerts.append(["Over Voltage", ov_count])
    ws_alerts.append(["Voltage Unbalance", vub_count])
    ws_alerts.append(["Poor Power Factor", ppf_count])
    ws_alerts.append(["High Current", hc_count])

    # Sheet 3: Alert Contributors with Meter Names
    ws_contributors = wb.create_sheet("Alert Contributors")

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    ws_contributors['A1'] = "ALERT CONTRIBUTORS - WITH METER NAMES (MONTHLY)"
    ws_contributors['A1'].fill = header_fill
    ws_contributors['A1'].font = header_font
    ws_contributors.merge_cells('A1:D1')

    current_row = 3

    alert_types = [
        ('Over Voltage', alert_contributors['over_voltage']),
        ('Voltage Unbalance', alert_contributors['voltage_unbalance']),
        ('Poor Power Factor', alert_contributors['poor_power_factor']),
        ('High Current', alert_contributors['high_current'])
    ]

    for alert_type, contributors in alert_types:
        ws_contributors[f'A{current_row}'] = f"{alert_type} Contributors"
        ws_contributors[f'A{current_row}'].fill = subheader_fill
        ws_contributors[f'A{current_row}'].font = subheader_font
        ws_contributors.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 1

        ws_contributors[f'A{current_row}'] = "S.No."
        ws_contributors[f'B{current_row}'] = "Meter Serial Number"
        ws_contributors[f'C{current_row}'] = "Meter Name"
        ws_contributors[f'D{current_row}'] = "Device Type"

        for col in ['A', 'B', 'C', 'D']:
            cell = ws_contributors[f'{col}{current_row}']
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.border = border
        current_row += 1

        if contributors:
            for idx, (meter_serial, meter_name, device_type) in enumerate(contributors, 1):
                ws_contributors[f'A{current_row}'] = idx
                ws_contributors[f'B{current_row}'] = meter_serial
                ws_contributors[f'C{current_row}'] = meter_name
                ws_contributors[f'D{current_row}'] = device_type

                for col in ['A', 'B', 'C', 'D']:
                    ws_contributors[f'{col}{current_row}'].border = border
                current_row += 1
        else:
            ws_contributors[f'A{current_row}'] = "-"
            ws_contributors[f'B{current_row}'] = "No contributors found"
            ws_contributors[f'C{current_row}'] = "-"
            ws_contributors[f'D{current_row}'] = "-"

            for col in ['A', 'B', 'C', 'D']:
                ws_contributors[f'{col}{current_row}'].border = border
            current_row += 1

        current_row += 1

        # Summary
    ws_contributors[f'A{current_row}'] = "SUMMARY"
    ws_contributors[f'A{current_row}'].fill = subheader_fill
    ws_contributors[f'A{current_row}'].font = subheader_font
    ws_contributors.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1

    for alert_type, contributors in alert_types:
        ws_contributors[f'A{current_row}'] = f"Total {alert_type}:"
        ws_contributors[f'B{current_row}'] = len(contributors)
        ws_contributors[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

    ws_contributors.column_dimensions['A'].width = 8
    ws_contributors.column_dimensions['B'].width = 25
    ws_contributors.column_dimensions['C'].width = 30
    ws_contributors.column_dimensions['D'].width = 15

    filename = f"theoretical_dashboard_data_{month_info['selected_month_year'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    logger.info(f"Calculated data saved: {filename}")
    return filename


# ============================================================================
# COMPARISON FUNCTIONS
# ============================================================================
@log_execution_time
def complete_data_comparison_dashboard(month_info, chart_file, processed_file):
    """Create comparison dashboard"""
    logger.info("Creating comparison dashboard...")

    chart_wb = load_workbook(chart_file)
    processed_wb = load_workbook(processed_file)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    comparison_wb = Workbook()
    comparison_wb.remove(comparison_wb.active)

    def compare_sheets(sheet_name, output_sheet_name):
        chart_ws = chart_wb[sheet_name]
        processed_ws = processed_wb[sheet_name]
        comp_ws = comparison_wb.create_sheet(output_sheet_name)

        comp_ws.append(["Parameter", "Chart Value", "Processed Value", "Difference", "Match"])

        output_row = 2
        for i in range(2, chart_ws.max_row + 1):
            param = chart_ws.cell(row=i, column=1).value
            chart_val = chart_ws.cell(row=i, column=2).value
            proc_val = processed_ws.cell(row=i, column=2).value

            try:
                chart_val = float(chart_val)
                proc_val = float(proc_val)
                diff = abs(chart_val - proc_val)
            except:
                diff = "N/A"

            match = "YES" if diff == 0 else "NO" if diff != "N/A" else "N/A"
            row = [param, chart_val, proc_val, diff, match]
            comp_ws.append(row)

            diff_cell = comp_ws.cell(row=output_row, column=4)
            match_cell = comp_ws.cell(row=output_row, column=5)

            if match == "YES":
                diff_cell.fill = green_fill
                match_cell.fill = green_fill
            elif match == "NO":
                diff_cell.fill = red_fill
                match_cell.fill = red_fill

            output_row += 1

    compare_sheets("Assets", "Assets Comparison")
    compare_sheets("Alerts", "Alerts Comparison")

    month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
    filename = f"Complete_Dashboard_Comparison_{month_safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    comparison_wb.save(filename)
    logger.info(f"Comparison saved: {filename}")
    return filename


# ============================================================================
# ENHANCED SUMMARY REPORT
# ============================================================================
@log_execution_time
def generate_summary_report(month_info, config, dashboard_data, dt_count, lv_count,
                            total_ov, total_vub, total_hc, total_ppf,
                            chart_file, processed_file, comparison_file):
    """Generate professional summary report with enhanced styling"""
    logger.info("Creating enhanced summary report...")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly_Validation_Summary"

        # Enhanced Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        main_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_header_alignment = Alignment(horizontal="left", vertical="center")

        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_alignment = Alignment(horizontal="left", vertical="center")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        label_alignment = Alignment(horizontal="left", vertical="center")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        pass_alignment = Alignment(horizontal="center", vertical="center")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        fail_alignment = Alignment(horizontal="center", vertical="center")

        warning_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        current_row = 1

        # MAIN HEADER
        ws.merge_cells(f'A{current_row}:H{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = "MONTHLY DASHBOARD VALIDATION SUMMARY REPORT"
        header_cell.font = main_header_font
        header_cell.fill = main_header_fill
        header_cell.alignment = main_header_alignment
        header_cell.border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Timestamp
        ws.merge_cells(f'A{current_row}:H{current_row}')
        timestamp_cell = ws[f'A{current_row}']
        timestamp_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(size=10, italic=True, color="666666", name="Calibri")
        timestamp_cell.alignment = Alignment(horizontal="center", vertical="center")
        timestamp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_cell.border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        current_row += 1

        # TEST DETAILS SECTION
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📋 TEST DETAILS"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Target Month:", config.get('target_month_year', 'Unknown')],
            ["Analysis Period:", f"{month_info['start_date']} to {month_info['end_date']}"],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        for label, value in test_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # SYSTEM UNDER TEST
        ws.merge_cells(f'A{current_row}:B{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🔧 SYSTEM UNDER TEST"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Validation Type:", "Monthly Dashboard (Assets & Alerts)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME],
            ["Target Month:", config['target_month_year']],
            ["Start Date:", str(month_info['start_date'])],
            ["End Date:", str(month_info['end_date'])],
        ]

        for label, value in system_details:
            ws[f'A{current_row}'].value = label
            ws[f'A{current_row}'].font = label_font
            ws[f'A{current_row}'].fill = label_fill
            ws[f'A{current_row}'].alignment = label_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = value
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = data_alignment
            ws[f'B{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ASSET VALIDATION
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "📊 ASSET VALIDATION"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Asset headers
        asset_headers = ["Asset Type", "UI Count", "DB Count", "Difference", "Status"]
        for i, header in enumerate(asset_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Extract UI values
        try:
            ui_dt_count = int(dashboard_data.get('Assets', {}).get('DTs', 0))
            ui_lv_count = int(dashboard_data.get('Assets', {}).get('LV Feeders', 0))
        except (KeyError, ValueError, TypeError):
            ui_dt_count = 0
            ui_lv_count = 0

        asset_data = [
            ("Distribution Transformers", ui_dt_count, dt_count),
            ("LV Feeders", ui_lv_count, lv_count)
        ]

        asset_matches = 0
        for asset_name, ui_val, db_val in asset_data:
            diff = abs(ui_val - db_val)
            status = "✓ MATCH" if diff == 0 else "✗ MISMATCH"
            fill_color = pass_fill if diff == 0 else fail_fill
            font_color = pass_font if diff == 0 else fail_font

            if diff == 0:
                asset_matches += 1

            ws[f'A{current_row}'].value = asset_name
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = ui_val
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = db_val
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = diff
            ws[f'D{current_row}'].font = data_font
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            ws[f'E{current_row}'].font = font_color
            ws[f'E{current_row}'].fill = fill_color
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # ALERT VALIDATION
        ws.merge_cells(f'A{current_row}:E{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "⚠️ ALERT VALIDATION"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Alert headers
        alert_headers = ["Alert Type", "UI Count", "Calculated Count", "Difference", "Status"]
        for i, header in enumerate(alert_headers, start=1):
            col_letter = chr(64 + i)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.alignment = subsection_alignment
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Extract UI alert values
        try:
            ui_ov = int(dashboard_data.get('Alerts', {}).get('Over Voltage', 0))
            ui_vub = int(dashboard_data.get('Alerts', {}).get('Voltage Unbalance', 0))
            ui_ppf = int(dashboard_data.get('Alerts', {}).get('Poor Power Factor', 0))
            ui_hc = int(dashboard_data.get('Alerts', {}).get('High Current', 0))
        except (KeyError, ValueError, TypeError):
            ui_ov = ui_vub = ui_ppf = ui_hc = 0

        alert_data = [
            ("Over Voltage", ui_ov, total_ov),
            ("Voltage Unbalance", ui_vub, total_vub),
            ("Poor Power Factor", ui_ppf, total_ppf),
            ("High Current", ui_hc, total_hc)
        ]

        alert_matches = 0
        for alert_name, ui_val, calc_val in alert_data:
            diff = abs(ui_val - calc_val)
            status = "✓ MATCH" if diff == 0 else "✗ MISMATCH"
            fill_color = pass_fill if diff == 0 else fail_fill
            font_color = pass_font if diff == 0 else fail_font

            if diff == 0:
                alert_matches += 1

            ws[f'A{current_row}'].value = alert_name
            ws[f'A{current_row}'].font = data_font
            ws[f'A{current_row}'].fill = data_fill
            ws[f'A{current_row}'].alignment = data_alignment
            ws[f'A{current_row}'].border = thin_border

            ws[f'B{current_row}'].value = ui_val
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].fill = data_fill
            ws[f'B{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'B{current_row}'].border = thin_border

            ws[f'C{current_row}'].value = calc_val
            ws[f'C{current_row}'].font = data_font
            ws[f'C{current_row}'].fill = data_fill
            ws[f'C{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'C{current_row}'].border = thin_border

            ws[f'D{current_row}'].value = diff
            ws[f'D{current_row}'].font = data_font
            ws[f'D{current_row}'].fill = data_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'D{current_row}'].border = thin_border

            ws[f'E{current_row}'].value = status
            ws[f'E{current_row}'].font = font_color
            ws[f'E{current_row}'].fill = fill_color
            ws[f'E{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            ws[f'E{current_row}'].border = thin_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        current_row += 1

        # OVERALL ASSESSMENT
        ws.merge_cells(f'A{current_row}:H{current_row}')
        section_cell = ws[f'A{current_row}']
        section_cell.value = "🏆 OVERALL ASSESSMENT"
        section_cell.font = section_header_font
        section_cell.fill = section_header_fill
        section_cell.alignment = section_header_alignment
        section_cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        total_validations = len(asset_data) + len(alert_data)
        total_passed = asset_matches + alert_matches
        overall_success_rate = (total_passed / total_validations) * 100 if total_validations > 0 else 0

        if overall_success_rate >= 95:
            assessment = "✓ EXCELLENT: Monthly dashboard validation passed with high confidence"
            assessment_color = pass_fill
            assessment_font_color = pass_font
        elif overall_success_rate >= 80:
            assessment = "⚠ GOOD: Minor discrepancies found - Review recommended"
            assessment_color = warning_fill
            assessment_font_color = warning_font
        else:
            assessment = "❌ REQUIRES ATTENTION: Significant validation failures detected"
            assessment_color = fail_fill
            assessment_font_color = fail_font

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = assessment
        cell.font = assessment_font_color
        cell.fill = assessment_color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Overall Success Rate: {overall_success_rate:.1f}% ({total_passed}/{total_validations} validations passed)"
        cell.font = Font(bold=True, size=11, name="Calibri", color="000000")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 25, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        month_safe = month_info['selected_month_year'].replace(' ', '_').replace('/', '_')
        filename = f"COMPLETE_VALIDATION_SUMMARY_MONTHLY_{month_safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        logger.info(f"Enhanced summary report created: {filename}")

        overall_status = "PASS" if overall_success_rate >= 95 else "FAIL"

        logger.info("=" * 60)
        logger.info("MONTHLY DASHBOARD VALIDATION SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Target Month: {config['target_month_year']}")
        logger.info(f"Analysis Period: {month_info['start_date']} to {month_info['end_date']}")
        logger.info(f"Overall Success Rate: {overall_success_rate:.1f}%")
        logger.info(f"Status: {overall_status}")
        logger.info("=" * 60)

        return filename, overall_status, overall_success_rate

    except Exception as e:
        logger.info(f"Error creating summary report: {e}")
        raise


# ============================================================================
# MAIN AUTOMATION FUNCTION
# ============================================================================
@log_execution_time
def main_monthly_dashboard_validation():
    """Main Monthly Dashboard Validation automation process"""
    config = None
    driver = None
    output_folder = None
    month_info = None
    chart_file = None
    processed_file = None
    comparison_file = None
    summary_report = None
    dashboard_data = None

    try:
        # Validate config
        config = validate_config_at_startup()
        if not config:
            logger.info("Cannot proceed without valid configuration")
            return False, None, None, None, None, None, None, None

        # Setup output folder
        output_folder = setup_output_folder()

        # Display database config
        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB: {DatabaseConfig.DB_HOST}:{DatabaseConfig.DB_PORT}/{DatabaseConfig.DB_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        # Start browser
        logger.info("Starting browser...")
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        # Login
        if not login(driver):
            logger.info("Login failed")
            return False, config, None, None, None, None, None, None

        time.sleep(2)

        # Apply configuration
        logger.info("Applying Monthly Dashboard Validation configuration...")
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        # Set month
        month_info = set_calendar_month(driver, config['target_month_year'])
        if not month_info:
            logger.info("Failed to set month")
            return False, config, month_info, None, None, None, None, None

        time.sleep(3)

        # Collect dashboard data
        dashboard_data = collect_data(driver)

        # Save dashboard data
        chart_file = save_date_to_excel(month_info, dashboard_data)
        if chart_file:
            chart_file = save_file_to_output(chart_file, output_folder)

        # Get assets from database
        dts = get_dts()
        dt_count = len(dts)
        lvfs = get_lvfs()
        lv_count = len(lvfs)

        # Initialize alert tracking
        total_ov = 0
        total_vub = 0
        total_hc = 0
        total_ppf = 0

        alert_contributors = {
            'over_voltage': [],
            'voltage_unbalance': [],
            'poor_power_factor': [],
            'high_current': []
        }

        # Process DTs
        logger.info(f"Processing {dt_count} DTs for month: {config['target_month_year']}...")
        for i, mtr_serial_no in enumerate(dts, start=1):
            if i <= 5 or i % 50 == 0 or i == dt_count:
                logger.info(f"Processing DT #{i}/{dt_count}: {mtr_serial_no}")

            (meterid, meter_name, voltagerating, overvoltage, voltageunbalance,
             currentrating, overload, currentunbalance, powerfactorthreshold) = get_metrics_dt(mtr_serial_no,
                                                                                               nodetypeid=153)

            if meterid is None:
                continue

            raw_df = get_raw_database_data_dt(month_info, meterid)

            ov_dt = calculate_over_voltage_events(raw_df, voltagerating, overvoltage)
            vub_dt = calculate_voltage_unbalance_events(raw_df, voltageunbalance)
            ppf_dt = calculate_low_power_factor_events(raw_df, powerfactorthreshold)
            hc_dt = calculate_high_current_events(raw_df, currentrating, overload)

            if ov_dt == 1:
                alert_contributors['over_voltage'].append((mtr_serial_no, meter_name or 'N/A', 'DT'))
            if vub_dt == 1:
                alert_contributors['voltage_unbalance'].append((mtr_serial_no, meter_name or 'N/A', 'DT'))
            if ppf_dt == 1:
                alert_contributors['poor_power_factor'].append((mtr_serial_no, meter_name or 'N/A', 'DT'))
            if hc_dt == 1:
                alert_contributors['high_current'].append((mtr_serial_no, meter_name or 'N/A', 'DT'))

            total_ov += ov_dt
            total_vub += vub_dt
            total_ppf += ppf_dt
            total_hc += hc_dt

            # Process LV Feeders


        logger.info(f"Processing {lv_count} LV Feeders for month: {config['target_month_year']}...")
        for i, mtr_serial_no in enumerate(lvfs, start=1):
            if i <= 5 or i % 50 == 0 or i == lv_count:
                logger.info(f"Processing LV Feeder #{i}/{lv_count}: {mtr_serial_no}")

            (lvfeeder_id, meterid, meter_name, voltagerating, overvoltage, voltageunbalance,
             currentrating, overload, currentunbalance, powerfactorthreshold) = get_metrics_lvf(mtr_serial_no,
                                                                                                nodetypeid=157)

            if meterid is None:
                continue

            raw_df = get_raw_database_data_lvf(month_info, meterid)

            ov_lvf = calculate_over_voltage_events(raw_df, voltagerating, overvoltage)
            vub_lvf = calculate_voltage_unbalance_events(raw_df, voltageunbalance)
            ppf_lvf = calculate_low_power_factor_events(raw_df, powerfactorthreshold)
            hc_lvf = calculate_high_current_events(raw_df, currentrating, overload)

            if ov_lvf == 1:
                alert_contributors['over_voltage'].append((mtr_serial_no, meter_name or 'N/A', 'LV Feeder'))
            if vub_lvf == 1:
                alert_contributors['voltage_unbalance'].append((mtr_serial_no, meter_name or 'N/A', 'LV Feeder'))
            if ppf_lvf == 1:
                alert_contributors['poor_power_factor'].append((mtr_serial_no, meter_name or 'N/A', 'LV Feeder'))
            if hc_lvf == 1:
                alert_contributors['high_current'].append((mtr_serial_no, meter_name or 'N/A', 'LV Feeder'))

            total_ov += ov_lvf
            total_vub += vub_lvf
            total_ppf += ppf_lvf
            total_hc += hc_lvf

        # Log alert summary
        logger.info("=" * 60)
        logger.info("ALERT CONTRIBUTORS SUMMARY (MONTHLY):")
        logger.info("=" * 60)
        logger.info(f"Over Voltage Contributors: {len(alert_contributors['over_voltage'])}")
        logger.info(f"Voltage Unbalance Contributors: {len(alert_contributors['voltage_unbalance'])}")
        logger.info(f"Poor Power Factor Contributors: {len(alert_contributors['poor_power_factor'])}")
        logger.info(f"High Current Contributors: {len(alert_contributors['high_current'])}")
        logger.info("=" * 60)

        # Save calculated data
        processed_file = save_calculated_data(month_info, total_ov, total_vub, total_hc, total_ppf,
                                              dt_count, lv_count, alert_contributors)
        processed_file = save_file_to_output(processed_file, output_folder)

        # Create comparison
        comparison_file = complete_data_comparison_dashboard(month_info, chart_file, processed_file)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        # Generate summary report
        summary_report, overall_status, accuracy_percentage = generate_summary_report(
            month_info, config, dashboard_data, dt_count, lv_count,
            total_ov, total_vub, total_hc, total_ppf,
            chart_file, processed_file, comparison_file
        )
        if summary_report:
            summary_report = save_file_to_output(summary_report, output_folder)

        # Final summary
        logger.info("=" * 60)
        logger.info("MONTHLY DASHBOARD VALIDATION AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Test Engineer: {TestEngineer.NAME}")
        logger.info(f"Validation Type: Monthly Dashboard Assets & Alerts")
        logger.info(f"Output Folder: {output_folder}")
        logger.info(f"Target Month: {config['target_month_year']}")
        logger.info(f"Analysis Period: {month_info['start_date']} to {month_info['end_date']}")
        logger.info(f"Area: {config['area']}")
        logger.info(f"Substation: {config['substation']}")
        logger.info(f"Feeder: {config['feeder']}")
        logger.info("")
        logger.info("Generated Files (4 total):")
        logger.info(f"   1. {os.path.basename(chart_file) if chart_file else 'Chart data'}")
        logger.info(f"   2. {os.path.basename(processed_file) if processed_file else 'Processed data'}")
        logger.info(f"   3. {os.path.basename(comparison_file) if comparison_file else 'Comparison report'}")
        logger.info(f"   4. {os.path.basename(summary_report) if summary_report else 'Summary report'}")
        logger.info("")
        logger.info("KEY FEATURES APPLIED:")
        logger.info("   ✓ Monthly dashboard validation (Assets & Alerts)")
        logger.info("   ✓ Centralized DB configuration")
        logger.info("   ✓ Test engineer details included")
        logger.info("   ✓ Alert contributors with meter names")
        logger.info("   ✓ Enhanced comparison with color coding")
        logger.info("   ✓ Professional summary report")
        logger.info("   ✓ Complete month analysis")
        logger.info("=" * 60)

        return True, config, month_info, chart_file, processed_file, comparison_file, summary_report, dashboard_data

    except Exception as e:
        logger.info(f"Critical error: {e}")

        if output_folder and os.path.exists(output_folder):
            try:
                error_file = os.path.join(output_folder, f"error_log_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
                with open(error_file, 'w') as f:
                    f.write(f"Monthly Dashboard Validation Automation Error\n")
                    f.write(f"Time: {datetime.now()}\n")
                    f.write(f"Error: {str(e)}\n")
                    f.write(f"Config: {config}\n")
                    f.write(f"Engineer: {TestEngineer.NAME}\n")
                logger.info(f"Error log saved: {os.path.basename(error_file)}")
            except:
                pass

        return False, config, month_info, chart_file, processed_file, comparison_file, summary_report, dashboard_data

    finally:
            if driver:
                try:
                    driver.quit()
                    logger.info("Browser closed")
                except:
                    pass


# ============================================================================
# HELPER FUNCTIONS FOR MAIN
# ============================================================================
def display_startup_banner():
    """Display startup banner with script information"""
    print("\n" + "=" * 80)
    print("🚀 MONTHLY DASHBOARD VALIDATION AUTOMATION")
    print("=" * 80)
    print(f"📋 Test Engineer: {TestEngineer.NAME}")
    print(f"🏢 Department: {TestEngineer.DEPARTMENT}")
    print(f"📊 Validation Type: Monthly Dashboard Assets & Alerts")
    print(f"🗄️ Database Tenant: {DatabaseConfig.TENANT_NAME}")
    print(f"⚙️ Validates: DTs, LV Feeders, OV, VUB, PPF, HC (Monthly)")
    print(f"📅 Analysis: Complete Month Period")
    print("=" * 80 + "\n")


def display_final_summary(success, total_time, output_folder, config, month_info,
                          chart_file, processed_file, comparison_file, summary_report, dashboard_data):
    """Display comprehensive final summary"""
    print("\n" + "=" * 80)
    if success:
        print("✅ MONTHLY DASHBOARD VALIDATION AUTOMATION COMPLETED SUCCESSFULLY!")
    else:
        print("❌ MONTHLY DASHBOARD VALIDATION AUTOMATION FAILED")
    print("=" * 80)

    print(f"\n⏱️  EXECUTION TIME: {total_time:.2f}s ({total_time / 60:.1f} minutes)")

    print(f"\n📋 TEST DETAILS:")
    print(f"   • Test Engineer: {TestEngineer.NAME}")
    print(f"   • Designation: {TestEngineer.DESIGNATION}")
    print(f"   • Department: {TestEngineer.DEPARTMENT}")
    print(f"   • Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    if success and config:
        print(f"\n🔧 CONFIGURATION USED:")
        print(f"   • Validation Type: Monthly Dashboard Assets & Alerts")
        print(f"   • Area: {config['area']}")
        print(f"   • Substation: {config['substation']}")
        print(f"   • Feeder: {config['feeder']}")
        print(f"   • Target Month: {config['target_month_year']}")

        if month_info:
            print(f"\n📅 ANALYSIS PERIOD:")
            print(f"   • Start Date: {month_info['start_date']}")
            print(f"   • End Date: {month_info['end_date']}")
            print(f"   • Month: {month_info['selected_month_year']}")
            print(f"   • Days Analyzed: {(month_info['end_date'] - month_info['start_date']).days + 1}")

        print(f"\n🗄️  DATABASE CONFIGURATION:")
        print(f"   • Host: {DatabaseConfig.DB_HOST}:{DatabaseConfig.DB_PORT}")
        print(f"   • Database: {DatabaseConfig.DB_DATABASE}")
        print(f"   • Tenant: {DatabaseConfig.TENANT_NAME}")

        if dashboard_data:
            print(f"\n📊 VALIDATION RESULTS:")
            print(f"   • Assets Validated: DTs, LV Feeders")
            print(f"   • Alerts Validated: OV, VUB, PPF, HC")
            print(f"   • UI Data Collected: ✓")
            print(f"   • Database Calculations: ✓ (Monthly)")

        print(f"\n📁 GENERATED FILES:")
        files_generated = []
        if chart_file and os.path.exists(chart_file):
            files_generated.append(f"   1. {os.path.basename(chart_file)}")
        if processed_file and os.path.exists(processed_file):
            files_generated.append(f"   2. {os.path.basename(processed_file)}")
        if comparison_file and os.path.exists(comparison_file):
            files_generated.append(f"   3. {os.path.basename(comparison_file)}")
        if summary_report and os.path.exists(summary_report):
            files_generated.append(f"   4. {os.path.basename(summary_report)}")

        for file_info in files_generated:
            print(file_info)

        print(f"\n📂 Output Folder: {os.path.abspath(output_folder)}")

        print(f"\n🎯 KEY FEATURES APPLIED:")
        print(f"   ✓ Monthly dashboard validation (Assets & Alerts)")
        print(f"   ✓ Centralized database configuration")
        print(f"   ✓ Test engineer details in reports")
        print(f"   ✓ Alert contributors with meter names")
        print(f"   ✓ Enhanced comparison with color coding")
        print(f"   ✓ Professional summary report")
        print(f"   ✓ Complete month period analysis")

    else:
        print("\n❌ AUTOMATION FAILED")
        print("   • Check configuration file: user_config.xlsx")
        print("   • Review error logs in output folder")
        print("   • Verify database connectivity")
        print("   • Ensure browser driver is up to date")
        print("   • Confirm month format: 'Month YYYY' (e.g., April 2025)")

    print("\n" + "=" * 80)
    print("🏁 MONTHLY DASHBOARD VALIDATION AUTOMATION FINISHED")
    print("=" * 80 + "\n")


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    display_startup_banner()

    logger.info("=" * 60)
    logger.info("MONTHLY DASHBOARD VALIDATION AUTOMATION - COMPLETE VERSION")
    logger.info("=" * 60)
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Department: {TestEngineer.DEPARTMENT}")
    logger.info(f"Validation Type: Monthly Dashboard Assets & Alerts")
    logger.info(f"Database Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("")
    logger.info("FEATURES:")
    logger.info("   ✓ Monthly dashboard validation automation")
    logger.info("   ✓ Assets: DTs, LV Feeders")
    logger.info("   ✓ Alerts: OV, VUB, PPF, HC")
    logger.info("   ✓ Centralized database configuration")
    logger.info("   ✓ Alert contributors with meter names")
    logger.info("   ✓ Test engineer details in reports")
    logger.info("   ✓ Professional summary report")
    logger.info("   ✓ Complete month period analysis")
    logger.info("=" * 60)

    config = None
    month_info = None
    chart_file = None
    processed_file = None
    comparison_file = None
    summary_report = None
    dashboard_data = None
    output_folder = None

    start_time = time.time()

    try:
        print("🔄 Starting monthly automation process...\n")

        result = main_monthly_dashboard_validation()
        success, config, month_info, chart_file, processed_file, comparison_file, summary_report, dashboard_data = result

        if os.path.exists('output_files'):
            output_folder = 'output_files'

    except Exception as e:
        logger.error(f"Critical error in main execution: {str(e)}")
        print(f"\n❌ CRITICAL ERROR: {str(e)}")
        success = False

    end_time = time.time()
    total_time = end_time - start_time

    logger.info("=" * 60)
    if success:
        logger.info("MONTHLY DASHBOARD VALIDATION AUTOMATION COMPLETED SUCCESSFULLY ✓")
        logger.info(f"Total Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("All optimizations verified:")
        logger.info("   ✓ Monthly dashboard validation")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Alert contributors with meter names")
        logger.info("   ✓ Professional summary")
        logger.info("   ✓ All 4 output files generated")
        logger.info("   ✓ Complete month period processed")
    else:
        logger.info("MONTHLY DASHBOARD VALIDATION AUTOMATION FAILED ✗")
        logger.info(f"Failed after: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Check error logs in output folder")

    logger.info("=" * 60)
    logger.info("Monthly Dashboard Validation Automation Finished")
    logger.info("=" * 60)

    display_final_summary(
        success=success,
        total_time=total_time,
        output_folder=output_folder if output_folder else 'output_files',
        config=config,
        month_info=month_info,
        chart_file=chart_file,
        processed_file=processed_file,
        comparison_file=comparison_file,
        summary_report=summary_report,
        dashboard_data=dashboard_data
    )

    import sys

    sys.exit(0 if success else 1)

