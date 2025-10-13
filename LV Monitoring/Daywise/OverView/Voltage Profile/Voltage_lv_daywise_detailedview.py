import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    TENANT_NAME = "tenant01"

    @classmethod
    def get_db1_params(cls):
        return {"host": cls.DB1_HOST, "port": cls.DB1_PORT, "database": cls.DB1_DATABASE,
                "user": cls.DB1_USER, "password": cls.DB1_PASSWORD}

    @classmethod
    def get_db2_params(cls):
        return {"host": cls.DB2_HOST, "port": cls.DB2_PORT, "database": cls.DB2_DATABASE,
                "user": cls.DB2_USER, "password": cls.DB2_PASSWORD}


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    logger = logging.getLogger('lv_voltage_overview')
    logger.setLevel(logging.INFO)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    log_file = 'logs/lv_voltage_overview.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setFormatter(formatter)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file: {e}")
        return file_path


# ============================================================================
# CONFIGURATION
# ============================================================================
def create_default_config_file(config_file):
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Date', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA', 'YOUR_SUBSTATION', 'YOUR_FEEDER', 'DD/MM/YYYY', 'YOUR_METER', 'DT']
        }
        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            pd.DataFrame(config_data).to_excel(writer, sheet_name='User_Configuration', index=False)
        logger.info(f"Config created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(value, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    try:
        if not os.path.exists(config_file):
            return None
        df = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}
        for _, row in df.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()
        return config
    except Exception as e:
        logger.info(f"Error: {e}")
        return None


def validate_config_at_startup():
    logger.info("=" * 60)
    logger.info("LV VOLTAGE OVERVIEW AUTOMATION")
    logger.info("=" * 60)
    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        create_default_config_file(config_file)
        logger.info("Please edit config and restart")
        return None
    config = read_user_configuration(config_file)
    if config:
        logger.info("Configuration validated")
    return config


# ============================================================================
# DECORATOR
# ============================================================================
def log_execution_time(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        logger.info(f"{func.__name__} completed in {time.time() - start:.2f}s")
        return result

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        if meter_type.upper() == 'DT':
            query = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        else:
            query = f"SELECT dt_id, lvfeeder_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        conn.close()
        return result if result else (None, None, None)
    except Exception as e:
        logger.info(f"DB Error: {e}")
        return None, None, None


@log_execution_time
def get_database_data_for_voltage_overview(target_date, mtr_id, node_id):
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{next_day}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())
        raw_query = f"SELECT DISTINCT surveydate, v1, v2, v3, avg_v FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata WHERE mtrid={mtr_id} {date_filter} ORDER BY surveydate;"
        nrm_query = f"SELECT surveydate, kva_i FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile WHERE nodeid={node_id} {date_filter} ORDER BY surveydate;"
        raw_df = pd.read_sql(raw_query, conn)
        nrm_df = pd.read_sql(nrm_query, conn)
        conn.close()
        logger.info(f"Retrieved: Raw={len(raw_df)}, NRM={len(nrm_df)}")
        return raw_df, nrm_df
    except Exception as e:
        logger.info(f"DB Error: {e}")
        return pd.DataFrame(), pd.DataFrame()


# ============================================================================
# WEB AUTOMATION
# ============================================================================
def login(driver):
    try:
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("Secure")
        driver.find_element(By.ID, "Password").send_keys("Secure@12345")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    try:
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        for option in driver.find_elements(By.CSS_SELECTOR, ".dx-list-item"):
            if option.text.strip().lower() == option_name.lower():
                option.click()
                return True
        return False
    except:
        return False


def set_calendar_date(driver, target_date):
    try:
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        return {'selected_date': target_dt.strftime("%B %Y"), 'start_date': target_dt.strftime("%Y-%m-%d")}
    except:
        return None


def select_type(driver):
    try:
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        time.sleep(3)
    except Exception as e:
        logger.info(f"Error: {e}")


def select_meter_type(driver, meter_type):
    try:
        wait = WebDriverWait(driver, 10)
        if meter_type == "DT":
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]'))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]'))).click()
        time.sleep(3)
        return True
    except:
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)
        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if view_buttons:
            view_buttons[0].click()
            return True
        return False
    except:
        return False


@log_execution_time
def extract_voltage_unbalance_from_svg(driver):
    vunb_values = {'Phase 1': '-', 'Phase 2': '-', 'Phase 3': '-'}
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#dvvolatgeunbalancechart svg')))
        js_script = """
        var svg = document.querySelector('#dvvolatgeunbalancechart svg');
        if (!svg) return null;
        var phaseData = [];
        svg.querySelectorAll('text, tspan').forEach(function(el) {
            var text = el.textContent.trim();
            if (text.includes('%') && text.match(/\\d+(\\.\\d+)?%/)) {
                phaseData.push({value: text.replace('%', '').trim(), y: el.getBoundingClientRect().y});
            }
        });
        phaseData.sort(function(a, b) { return a.y - b.y; });
        return {phaseData: phaseData};
        """
        svg_data = driver.execute_script(js_script)
        if svg_data and svg_data.get('phaseData') and len(svg_data['phaseData']) >= 3:
            vunb_values['Phase 1'] = svg_data['phaseData'][0]['value']
            vunb_values['Phase 2'] = svg_data['phaseData'][1]['value']
            vunb_values['Phase 3'] = svg_data['phaseData'][2]['value']
    except Exception as e:
        logger.error(f"SVG error: {e}")
    return vunb_values


@log_execution_time
def collect_voltage_overview_data(driver):
    data = {}
    try:
        # Voltage Phasewise
        data['Voltage Phasewise'] = {
            'p1max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph1"]').text,
            'p2max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph2"]').text,
            'p3max': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Ph3"]').text,
            'p1avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph1"]').text,
            'p2avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph2"]').text,
            'p3avg': driver.find_element(By.XPATH, '//*[@id="avgVoltage_Ph3"]').text,
            'max_avg': driver.find_element(By.XPATH, '//*[@id="maxVoltage_Avg"]').text
        }

        # Voltage at Max Load
        data['Voltage Maxload'] = {
            'phase1_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase1Voltage"]').text,
            'phase2_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase2Voltage"]').text,
            'phase3_maxload': driver.find_element(By.XPATH, '//*[@id="dtPeakPhase3Voltage"]').text
        }

        # Voltage Unbalance
        data['Voltage Unbalance'] = extract_voltage_unbalance_from_svg(driver)

        # Voltage Variation
        action = ActionChains(driver)
        voltage_bars = driver.find_elements(By.CSS_SELECTOR, '#divVoltageVariation g.dxc-markers')
        voltage_variation_data = {
            'Duration Voltage <180V': '-',
            'Duration Voltage 180-216V': '-',
            'Duration Voltage 216-240V': '-',
            'Duration Voltage >240V': '-'
        }
        color_mapping = {
            '#E38430': 'Duration Voltage <180V',
            '#DEAE2A': 'Duration Voltage 180-216V',
            '#86B8A5': 'Duration Voltage 216-240V',
            '#D11920': 'Duration Voltage >240V'
        }
        for bar in voltage_bars:
            try:
                fill_color = bar.get_attribute('fill')
                label = color_mapping.get(fill_color)
                if label:
                    action.move_to_element(bar).perform()
                    time.sleep(1)
                    tooltip = driver.find_element(By.CSS_SELECTOR, '.dxc-tooltip svg text')
                    voltage_variation_data[label] = tooltip.text.strip()
            except:
                pass
        data['Voltage Variation'] = voltage_variation_data
    except Exception as e:
        logger.error(f"Collection error: {e}")
    return data


@log_execution_time
def save_voltage_overview_data_to_excel(date_info, overview_data):
    try:
        wb = Workbook()
        wb.remove(wb.active)

        # Voltage Phasewise
        ws = wb.create_sheet("Voltage Phasewise")
        ws.append(["Phase", "Max", "Avg"])
        vp = overview_data['Voltage Phasewise']
        ws.append(["Phase 1", vp['p1max'], vp['p1avg']])
        ws.append(["Phase 2", vp['p2max'], vp['p2avg']])
        ws.append(["Phase 3", vp['p3max'], vp['p3avg']])
        ws.append(["Average", vp['max_avg'], "-"])

        # Voltage at Max Load
        ws = wb.create_sheet("Voltage at Max Load")
        ws.append(["Phase", "Voltage", "Date & Time"])
        vml = overview_data['Voltage Maxload']
        for i, key in enumerate(['phase1_maxload', 'phase2_maxload', 'phase3_maxload'], 1):
            raw = vml[key]
            try:
                v, t = raw.split(' V ')
                ws.append([f"Phase {i}", v.strip(), t.strip('()').strip()])
            except:
                ws.append([f"Phase {i}", raw, "-"])

        # Voltage Unbalance
        ws = wb.create_sheet("Voltage Unbalance")
        ws.append(["Phase", "Unbalance (%)"])
        vu = overview_data['Voltage Unbalance']
        ws.append(["Phase 1", vu['Phase 1']])
        ws.append(["Phase 2", vu['Phase 2']])
        ws.append(["Phase 3", vu['Phase 3']])

        # Voltage Variation
        ws = wb.create_sheet("Voltage Variation")
        ws.append(["Voltage Range", "Duration (Hrs)"])
        for k, v in overview_data['Voltage Variation'].items():
            ws.append([k, v])

        filename = f"chart_data_voltage_overview_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Save error: {e}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_voltage_overview_database_calculations(raw_df, nrm_df, date_info):
    try:
        def format_duration(minutes):
            h, m = divmod(int(minutes), 60)
            return f"{h}:{m:02} hrs"

        date_safe = date_info['selected_date'].replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        interval_minutes = 15 if len(raw_df) <= 1 else int(
            (raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)

        # Voltage Phasewise
        phase_data = []
        avg_values = []
        for i, col in enumerate(['v1', 'v2', 'v3'], 1):
            max_val = round(raw_df[col].max(), 1)
            avg_val = round(raw_df[col].mean(), 1)
            avg_values.append(avg_val)
            phase_data.append([f'Phase {i}', max_val, avg_val])
        phase_data.append(['Average', round(sum(r[1] for r in phase_data) / 3, 1), '-'])
        voltage_df = pd.DataFrame(phase_data, columns=['Phase', 'Max', 'Avg'])

        # Voltage at Max Load
        if not nrm_df.empty:
            max_kva_row = nrm_df.loc[nrm_df['kva_i'].idxmax()]
            max_time = max_kva_row['surveydate']
            matching = raw_df[raw_df['surveydate'] == max_time]
            if not matching.empty:
                v1, v2, v3 = round(matching['v1'].iloc[0], 1), round(matching['v2'].iloc[0], 1), round(
                    matching['v3'].iloc[0], 1)
                time_str = max_time.strftime('%#d %b - %H:%M')
            else:
                v1, v2, v3 = round(raw_df['v1'].max(), 1), round(raw_df['v2'].max(), 1), round(raw_df['v3'].max(), 1)
                time_str = raw_df.loc[raw_df['v1'].idxmax(), 'surveydate'].strftime('%#d %b - %H:%M')
        else:
            v1, v2, v3 = round(raw_df['v1'].max(), 1), round(raw_df['v2'].max(), 1), round(raw_df['v3'].max(), 1)
            time_str = raw_df.loc[raw_df['v1'].idxmax(), 'surveydate'].strftime('%#d %b - %H:%M')

        maxload_df = pd.DataFrame([['Phase 1', v1, time_str], ['Phase 2', v2, time_str], ['Phase 3', v3, time_str]],
                                  columns=['Phase', 'Voltage', 'Date & Time'])

        # Voltage Unbalance
        total_avg = sum(avg_values) / 3
        unbalance_data = []
        for i, avg_val in enumerate(avg_values, 1):
            unb = round(abs(100 - (avg_val / total_avg * 100)), 1)
            unbalance_data.append([f'Phase {i}', 0 if unb == 0.0 else unb])
        unbalance_df = pd.DataFrame(unbalance_data, columns=['Phase', 'Unbalance (%)'])

        # Voltage Variation
        bins = {
            "Duration Voltage <180V": raw_df[raw_df['avg_v'] < 180].shape[0],
            "Duration Voltage 180-216V": raw_df[(raw_df['avg_v'] >= 180) & (raw_df['avg_v'] < 216)].shape[0],
            "Duration Voltage 216-240V": raw_df[(raw_df['avg_v'] >= 216) & (raw_df['avg_v'] <= 240)].shape[0],
            "Duration Voltage >240V": raw_df[raw_df['avg_v'] > 240].shape[0]
        }
        variation_data = [[label, format_duration(count * interval_minutes) if count > 0 else "-"] for label, count in
                          bins.items()]
        variation_df = pd.DataFrame(variation_data, columns=['Voltage Range', 'Duration (Hrs)'])

        # Save
        filename = f"theoretical_voltage_overview_{date_safe}_{timestamp}.xlsx"
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            voltage_df.to_excel(writer, sheet_name='Voltage Phasewise', index=False)
            maxload_df.to_excel(writer, sheet_name='Voltage at Max Load', index=False)
            unbalance_df.to_excel(writer, sheet_name='Voltage Unbalance', index=False)
            variation_df.to_excel(writer, sheet_name='Voltage Variation', index=False)

        return filename
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise


# ============================================================================
# COMPARISON
# ============================================================================
@log_execution_time
def create_voltage_overview_comparison(chart_file, processed_file, date_info):
    try:
        output_file = f"validation_report_voltage_overview_{date_info['selected_date'].replace(' ', '_')}.xlsx"
        sheet_names = ['Voltage Phasewise', 'Voltage at Max Load', 'Voltage Unbalance', 'Voltage Variation']
        chart_data = {s: pd.read_excel(chart_file, sheet_name=s) for s in sheet_names}
        processed_data = {s: pd.read_excel(processed_file, sheet_name=s) for s in sheet_names}

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        wb = Workbook()
        wb.remove(wb.active)
        validation_results = {}

        for sheet in sheet_names:
            processed_df = processed_data[sheet]
            chart_df = chart_data[sheet]
            ws = wb.create_sheet(title=f"{sheet}_Comparison")

            if sheet == "Voltage Phasewise":
                headers = ["Phase", "DB_Max", "Chart_Max", "Max_Diff", "Max_Match", "DB_Avg", "Chart_Avg", "Avg_Diff",
                           "Avg_Match", "Overall"]
            elif sheet == "Voltage at Max Load":
                headers = ["Phase", "DB_V", "Chart_V", "V_Diff", "V_Match", "DB_Time", "Chart_Time", "Time_Match",
                           "Overall"]
            elif sheet == "Voltage Unbalance":
                headers = ["Phase", "DB_Unb", "Chart_Unb", "Unb_Diff", "Overall"]
            else:
                headers = ["Range", "DB_Dur", "Chart_Dur", "Dur_Match", "Overall"]

            ws.append(headers)
            sheet_results = []

            for i in range(len(processed_df)):
                row_data = []
                overall_match = True

                if sheet == "Voltage Phasewise":
                    phase = processed_df.iloc[i, 0]
                    row_data.append(phase)
                    db_max, chart_max = processed_df.iloc[i, 1], chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                    try:
                        max_diff = abs(float(db_max) - float(chart_max)) if db_max != "-" and chart_max != "-" else "-"
                        max_match = "YES" if (max_diff != "-" and max_diff <= 0.1) or str(db_max).strip() == str(
                            chart_max).strip() else "NO"
                    except:
                        max_diff, max_match = "-", "YES" if str(db_max).strip() == str(chart_max).strip() else "NO"
                    overall_match = overall_match and (max_match == "YES")
                    row_data.extend([db_max, chart_max, max_diff, max_match])

                    db_avg, chart_avg = processed_df.iloc[i, 2], chart_df.iloc[i, 2] if i < len(chart_df) else "-"
                    try:
                        avg_diff = abs(float(db_avg) - float(chart_avg)) if db_avg != "-" and chart_avg != "-" else "-"
                        avg_match = "YES" if (avg_diff != "-" and avg_diff <= 0.1) or str(db_avg).strip() == str(
                            chart_avg).strip() else "NO"
                    except:
                        avg_diff, avg_match = "-", "YES" if str(db_avg).strip() == str(chart_avg).strip() else "NO"
                    overall_match = overall_match and (avg_match == "YES")
                    row_data.extend([db_avg, chart_avg, avg_diff, avg_match, "YES" if overall_match else "NO"])

                elif sheet == "Voltage at Max Load":
                    phase = processed_df.iloc[i, 0]
                    row_data.append(phase)
                    db_v, chart_v = processed_df.iloc[i, 1], chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                    try:
                        v_diff = abs(float(db_v) - float(chart_v)) if db_v != "-" and chart_v != "-" else "-"
                        v_match = "YES" if (v_diff != "-" and v_diff <= 0.1) or str(db_v).strip() == str(
                            chart_v).strip() else "NO"
                    except:
                        v_diff, v_match = "-", "YES" if str(db_v).strip() == str(chart_v).strip() else "NO"
                    overall_match = overall_match and (v_match == "YES")
                    row_data.extend([db_v, chart_v, v_diff, v_match])

                    db_time, chart_time = processed_df.iloc[i, 2], chart_df.iloc[i, 2] if i < len(chart_df) else "-"
                    time_match = "YES" if str(db_time).strip() == str(chart_time).strip() else "NO"
                    overall_match = overall_match and (time_match == "YES")
                    row_data.extend([db_time, chart_time, time_match, "YES" if overall_match else "NO"])

                elif sheet == "Voltage Unbalance":
                    phase = processed_df.iloc[i, 0]
                    row_data.append(phase)
                    db_unb, chart_unb = processed_df.iloc[i, 1], chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                    try:
                        unb_diff = round(abs(float(db_unb) - float(chart_unb)),
                                         2) if db_unb != "-" and chart_unb != "-" else "-"
                        overall_match = "YES" if (unb_diff != "-" and unb_diff <= 0.1) or str(db_unb).strip() == str(
                            chart_unb).strip() else "NO"
                    except:
                        unb_diff, overall_match = "-", "YES" if str(db_unb).strip() == str(chart_unb).strip() else "NO"
                    row_data.extend([db_unb, chart_unb, unb_diff, overall_match])

                else:  # Voltage Variation
                    v_range = processed_df.iloc[i, 0]
                    row_data.append(v_range)
                    db_dur, chart_dur = processed_df.iloc[i, 1], chart_df.iloc[i, 1] if i < len(chart_df) else "-"
                    dur_match = "YES" if str(db_dur).strip() == str(chart_dur).strip() else "NO"
                    overall_match = dur_match == "YES"
                    row_data.extend([db_dur, chart_dur, dur_match, overall_match])

                sheet_results.append({'item': phase if sheet != "Voltage Variation" else v_range,
                                      'match': overall_match == "YES" if isinstance(overall_match,
                                                                                    str) else overall_match})
                ws.append(row_data)

            validation_results[sheet] = sheet_results

            # Apply colors
            for row_num in range(2, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    header = ws.cell(row=1, column=col_num).value
                    if header and ("Match" in header or header == "Overall"):
                        if cell.value == "YES":
                            cell.fill = green_fill
                        elif cell.value == "NO":
                            cell.fill = red_fill
                    elif header and "Diff" in header:
                        if isinstance(cell.value, (int, float)) and cell.value <= 0.1:
                            cell.fill = green_fill
                        elif isinstance(cell.value, (int, float)):
                            cell.fill = red_fill

        wb.save(output_file)
        logger.info(f"Comparison saved: {output_file}")
        return output_file, validation_results
    except Exception as e:
        logger.error(f"Comparison error: {e}")
        raise


# ============================================================================
# SUMMARY REPORT - COMPLETE VERSION
# ============================================================================
@log_execution_time
def create_voltage_overview_summary_report(config, date_info, chart_file, processed_file, comparison_file,
                                           validation_results, raw_df, meter_name):
    """Create comprehensive voltage overview summary report"""
    try:
        date_safe = date_info['selected_date'].replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"SUMMARY_VOLTAGE_OVERVIEW_{date_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary_Report"

        # Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subsection_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        label_font = Font(bold=True, size=10, name="Calibri")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_font = Font(size=10, name="Calibri")
        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                              bottom=Side(style='medium'))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        row = 1

        # Main Header
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"LV VOLTAGE OVERVIEW VALIDATION - {date_info['selected_date'].upper()}"
        cell.font = main_header_font
        cell.fill = main_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        ws.row_dimensions[row].height = 30
        row += 1

        # Timestamp
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(size=10, italic=True, color="666666")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        row += 2

        # Test Details
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "📋 TEST DETAILS"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        ws[f'B{row}'].border = thick_border
        row += 1

        for label, value in [["Engineer:", TestEngineer.NAME], ["Designation:", TestEngineer.DESIGNATION],
                             ["Date:", config['target_date']], ["Department:", TestEngineer.DEPARTMENT]]:
            ws[f'A{row}'].value = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = value
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            row += 1
        row += 1

        # System Under Test
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "🔧 SYSTEM UNDER TEST"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        ws[f'B{row}'].border = thick_border
        row += 1

        for label, value in [["Area:", config['area']], ["Substation:", config['substation']],
                             ["Feeder:", config['feeder']], ["Meter:", config['meter_serial_no']],
                             ["Name:", meter_name], ["Type:", config['meter_type']],
                             ["Monitoring:", "LV Voltage Overview"], ["Tenant:", DatabaseConfig.TENANT_NAME]]:
            ws[f'A{row}'].value = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = value
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            row += 1
        row += 1

        # Data Volume
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "📊 DATA VOLUME"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        for i, header in enumerate(["Dataset", "Count", "Status"], 1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
        row += 1

        chart_points = 4
        for dataset, count, status in [["Raw DB Records", len(raw_df), "COMPLETE" if len(raw_df) > 0 else "NO DATA"],
                                       ["Chart Points", chart_points, "COMPLETE"]]:
            ws[f'A{row}'].value = dataset
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = count
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].value = status
            ws[f'C{row}'].font = pass_font if "COMPLETE" in status else fail_font
            ws[f'C{row}'].fill = pass_fill if "COMPLETE" in status else fail_fill
            ws[f'C{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].border = thin_border
            row += 1
        row += 1

        # Validation Results
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "✅ VALIDATION RESULTS"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        for i, header in enumerate(["Type", "Pass", "Fail", "Rate", "Status"], 1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
        row += 1

        overall_pass = overall_total = 0
        for sheet_name, results in validation_results.items():
            total = len(results)
            passed = sum(1 for r in results if r['match'])
            failed = total - passed
            rate = f"{(passed / total * 100):.1f}%" if total > 0 else "0%"
            status = "PASS" if passed == total else "FAIL"
            overall_pass += passed
            overall_total += total

            ws[f'A{row}'].value = sheet_name
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = passed
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].value = failed
            ws[f'C{row}'].font = data_font
            ws[f'C{row}'].border = thin_border
            ws[f'C{row}'].alignment = Alignment(horizontal="center")
            ws[f'D{row}'].value = rate
            ws[f'D{row}'].font = Font(bold=True, size=10)
            ws[f'D{row}'].border = thin_border
            ws[f'D{row}'].alignment = Alignment(horizontal="center")
            ws[f'E{row}'].value = status
            ws[f'E{row}'].font = pass_font if status == "PASS" else fail_font
            ws[f'E{row}'].fill = pass_fill if status == "PASS" else fail_fill
            ws[f'E{row}'].alignment = Alignment(horizontal="center")
            ws[f'E{row}'].border = thin_border
            row += 1
        row += 1

        # Overall Assessment
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "🏆 OVERALL ASSESSMENT"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        success_rate = (overall_pass / overall_total * 100) if overall_total > 0 else 0
        if success_rate >= 95:
            assessment = "✓ EXCELLENT: Validation passed"
            color = pass_fill
            font_color = pass_font
        elif success_rate >= 80:
            assessment = "⚠ GOOD: Minor issues found"
            color = warning_fill
            font_color = Font(bold=True, size=10, color="000000")
        else:
            assessment = "❌ ATTENTION: Significant failures"
            color = fail_fill
            font_color = fail_font

        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = assessment
        cell.font = font_color
        cell.fill = color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thick_border
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Success Rate: {success_rate:.1f}% ({overall_pass}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        wb.save(summary_file)
        logger.info(f"Summary created: {summary_file}")
        return summary_file
    except Exception as e:
        logger.error(f"Summary error: {e}")
        raise


# ============================================================================
# MAIN FUNCTION
# ============================================================================
@log_execution_time
def main_lv_voltage_overview_automation():
    config = driver = output_folder = None
    try:
        config = validate_config_at_startup()
        if not config:
            return False

        output_folder = setup_output_folder()

        logger.info("=" * 60)
        logger.info(f"DB: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        if not login(driver):
            return False

        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info:
            return False

        if not select_meter_type(driver, config['meter_type']):
            return False

        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])
        if not dt_id:
            logger.info(f"Meter not found")
            return False

        logger.info(f"Meter: {name}")
        time.sleep(3)

        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            return False

        time.sleep(5)

        overview_data = collect_voltage_overview_data(driver)
        chart_file = save_voltage_overview_data_to_excel(date_info, overview_data)
        chart_file = save_file_to_output(chart_file, output_folder)

        raw_df, nrm_df = get_database_data_for_voltage_overview(config['target_date'], mtr_id, dt_id)
        if raw_df.empty:
            logger.info("No database data")
            return False

        processed_file = process_voltage_overview_database_calculations(raw_df, nrm_df, date_info)
        processed_file = save_file_to_output(processed_file, output_folder)

        comparison_file, validation_results = create_voltage_overview_comparison(chart_file, processed_file, date_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        summary_report = create_voltage_overview_summary_report(config, date_info, chart_file, processed_file,
                                                                comparison_file, validation_results, raw_df, name)
        summary_report = save_file_to_output(summary_report, output_folder)

        logger.info("=" * 60)
        logger.info("LV VOLTAGE OVERVIEW AUTOMATION COMPLETED!")
        logger.info("=" * 60)
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info(f"Output: {output_folder}")
        logger.info(f"Files Generated:")
        logger.info(f"  1. {os.path.basename(chart_file)}")
        logger.info(f"  2. {os.path.basename(processed_file)}")
        logger.info(f"  3. {os.path.basename(comparison_file)}")
        logger.info(f"  4. {os.path.basename(summary_report)}")
        logger.info("=" * 60)

        return True
    except Exception as e:
        logger.info(f"Error: {e}")
        return False
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV VOLTAGE OVERVIEW AUTOMATION")
    logger.info(f"Engineer: {TestEngineer.NAME}")
    logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_lv_voltage_overview_automation()
    total_time = time.time() - start_time

    logger.info("=" * 60)
    if success:
        logger.info("✓ COMPLETED SUCCESSFULLY")
        logger.info(f"Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
    else:
        logger.info("✗ FAILED")
        logger.info(f"Failed after: {total_time:.2f}s")
    logger.info("=" * 60)
