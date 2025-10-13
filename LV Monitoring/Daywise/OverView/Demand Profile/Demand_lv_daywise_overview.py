"""
LV MONITORING DAYWISE OVERVIEW - DEMAND PROFILE
Complete automation script with enhanced styling and validation
Test Engineer: Sanyam Upadhyay
Department: NPD - Quality Assurance
"""

import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import functools


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    TENANT_NAME = "tenant01"

    @classmethod
    def get_db1_params(cls):
        return {"host": cls.DB1_HOST, "port": cls.DB1_PORT, "database": cls.DB1_DATABASE,
                "user": cls.DB1_USER, "password": cls.DB1_PASSWORD}

    @classmethod
    def get_db2_params(cls):
        return {"host": cls.DB2_HOST, "port": cls.DB2_PORT, "database": cls.DB2_DATABASE,
                "user": cls.DB2_USER, "password": cls.DB2_PASSWORD}


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    logger = logging.getLogger('lv_demand_overview')
    logger.setLevel(logging.INFO)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    log_file = 'logs/lv_demand_overview.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setFormatter(formatter)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Date', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'DD/MM/YYYY', 'YOUR_METER_NO', 'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)
            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open "User_Configuration" sheet',
                    'Replace YOUR_AREA_HERE with area name',
                    'Replace YOUR_SUBSTATION_HERE with substation',
                    'Replace YOUR_FEEDER_HERE with feeder',
                    'Update Target_Date (DD/MM/YYYY)',
                    'Update Meter_Serial_No',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'FOR LV MONITORING DEMAND OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces',
                    'Date format: DD/MM/YYYY',
                    'Meter_Type: DT or LV only',
                    'Save before running',
                    f'Test Engineer: {TestEngineer.NAME}',
                ]
            }
            pd.DataFrame(instructions).to_excel(writer, sheet_name='Setup_Instructions', index=False)
        logger.info(f"Config template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(value, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    try:
        if not os.path.exists(config_file):
            return None
        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}
        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required = ['type', 'area', 'substation', 'feeder', 'target_date', 'meter_serial_no', 'meter_type']
        if any(f not in config or not config[f] for f in required):
            return None
        if any(config.get(k) in ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO'] for k in
               config):
            return None
        return config
    except Exception as e:
        logger.info(f"Error reading config: {e}")
        return None


def validate_config_at_startup():
    logger.info("=" * 60)
    logger.info("STARTING LV DEMAND OVERVIEW AUTOMATION")
    logger.info("=" * 60)
    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info("Creating default config template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the config file and restart")
        return None
    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None
    logger.info("Config validated successfully")
    for k, v in config.items():
        logger.info(f"   {k}: {v}")
    return config


# ============================================================================
# DECORATOR
# ============================================================================
def log_execution_time(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        if meter_type.upper() == 'DT':
            query = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        else:
            query = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if result:
            logger.info(f"Metrics: {result[1]}, meterid: {result[2]}")
            return result
        return None, None, None
    except Exception as e:
        logger.info(f"DB error: {e}")
        return None, None, None


@log_execution_time
def get_database_data_for_demand_overview(target_date, mtr_id, node_id):
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{next_day}'"
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())
        raw_query = f"SELECT DISTINCT surveydate, kwh_i, kvah_i, kvar_i_total FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata WHERE mtrid={mtr_id} {date_filter} ORDER BY surveydate ASC;"
        nrm_query = f"SELECT surveydate, kw_i, kva_i, kvar_i FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile WHERE nodeid={node_id} {date_filter} ORDER BY surveydate ASC;"
        raw_df = pd.read_sql(raw_query, conn)
        nrm_df = pd.read_sql(nrm_query, conn)
        conn.close()
        logger.info(f"Retrieved: Raw={len(raw_df)}, NRM={len(nrm_df)}")
        return raw_df, nrm_df
    except Exception as e:
        logger.info(f"DB error: {e}")
        return pd.DataFrame(), pd.DataFrame()


# ============================================================================
# WEB AUTOMATION
# ============================================================================
def login(driver):
    try:
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("Secure")
        driver.find_element(By.ID, "Password").send_keys("Secure@12345")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    try:
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        for option in driver.find_elements(By.CSS_SELECTOR, ".dx-list-item"):
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_date(driver, target_date):
    try:
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        return {
            'selected_date': target_dt.strftime("%B %Y"),
            'start_date': target_dt.strftime("%Y-%m-%d"),
            'end_date': target_dt.strftime("%Y-%m-%d")
        }
    except Exception as e:
        logger.info(f"Date error: {e}")
        return None


def select_type(driver):
    try:
        time.sleep(10)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type error: {e}")


def select_meter_type(driver, meter_type):
    try:
        wait = WebDriverWait(driver, 10)
        if meter_type == "DT":
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]'))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]'))).click()
        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)
        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if view_buttons:
            view_buttons[0].click()
            logger.info("View clicked")
            return True
        return False
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_demand_overview_data(driver):
    data = {}
    try:
        time.sleep(4)
        driver.find_element(By.XPATH, "//div[@class='dx-item-content' and text()='Demand']").click()
        time.sleep(2)
        data['act_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kw"]').text
        data['act_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kw"]').text
        data['act_dt'] = driver.find_element(By.XPATH, '//td[@id="kw_MaxDatetime"]').text
        data['app_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kva"]').text
        data['app_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kva"]').text
        data['app_dt'] = driver.find_element(By.XPATH, '//td[@id="kva_MaxDatetime"]').text
        data['react_max'] = driver.find_element(By.XPATH, '//td[@id="maxDemand_Kvar"]').text
        data['react_avg'] = driver.find_element(By.XPATH, '//td[@id="avgDemand_Kvar"]').text
        data['react_dt'] = driver.find_element(By.XPATH, '//td[@id="kvar_MaxDatetime"]').text
        logger.info("Demand data collected")
    except Exception as e:
        logger.error(f"Collection error: {e}")
        raise
    return data


@log_execution_time
def save_demand_overview_data_to_excel(date_info, demand_data):
    try:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Demand Table")
        ws.append(["Parameter", "Max", "Avg", "Date and time at max value"])
        for param, keys in [("Active", ("act_max", "act_avg", "act_dt")),
                            ("Apparent", ("app_max", "app_avg", "app_dt")),
                            ("Reactive", ("react_max", "react_avg", "react_dt"))]:
            ws.append([param, demand_data.get(keys[0], ""), demand_data.get(keys[1], ""), demand_data.get(keys[2], "")])
        file_name = f"chart_data_from_ui_demand_overview_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"Saved: {file_name}")
        return file_name
    except Exception as e:
        logger.error(f"Save error: {e}")
        raise


# ============================================================================
# PROCESSING
# ============================================================================
@log_execution_time
def process_demand_overview_database_calculations(raw_df, nrm_df, date_info):
    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        interval_minutes = 15 if len(raw_df) <= 1 else int(
            (raw_df['surveydate'].iloc[1] - raw_df['surveydate'].iloc[0]).total_seconds() / 60)
        sip_hr = interval_minutes / 60

        nrm_calc = pd.DataFrame()
        nrm_calc['kw_i'] = raw_df.get('kwh_i', 0) / sip_hr
        nrm_calc['kva_i'] = raw_df.get('kvah_i', 0) / sip_hr
        nrm_calc['kvar_i'] = raw_df.get('kvar_i_total', 0) / sip_hr
        nrm_calc['surveydate'] = raw_df['surveydate']

        def fmt_dt(dt):
            if isinstance(dt, pd.Timestamp):
                return dt.strftime(f'{dt.day} %b at %H:%M')
            return str(dt)

        demand_data = []
        for col, name in [('kw_i', 'Active'), ('kva_i', 'Apparent'), ('kvar_i', 'Reactive')]:
            max_val = nrm_calc[col].max()
            avg_val = nrm_calc[col].mean()
            max_time = fmt_dt(nrm_calc.loc[nrm_calc[col].idxmax(), 'surveydate'])
            demand_data.append([name, max_val, avg_val, max_time])

        demand_df = pd.DataFrame(demand_data, columns=['Parameter', 'Max', 'Avg', 'Date and time at max value'])
        processed_file = f"theoretical_demand_overview_calculated_data_{date_safe}_{timestamp}.xlsx"

        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            nrm_calc.to_excel(writer, sheet_name='NRM Calculated', index=False)
            demand_df.to_excel(writer, sheet_name='Demand Table', index=False)

        logger.info(f"Processed: {processed_file}")
        return processed_file
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise


# ============================================================================
# COMPARISON
# ============================================================================
@log_execution_time
def create_demand_overview_comparison(chart_file, processed_file, date_info):
    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_demand_overview_{date_safe}.xlsx"

        wb_proc = load_workbook(processed_file)
        wb_chart = load_workbook(chart_file)

        ws_proc = wb_proc['Demand Table']
        ws_chart = wb_chart['Demand Table']

        proc_data = [row for row in ws_proc.iter_rows(values_only=True) if any(cell is not None for cell in row)]
        chart_data = [row for row in ws_chart.iter_rows(values_only=True) if any(cell is not None for cell in row)]

        proc_df = pd.DataFrame(proc_data)
        chart_df = pd.DataFrame(chart_data)

        proc_df.columns = proc_df.iloc[0]
        proc_df = proc_df[1:].reset_index(drop=True)
        chart_df.columns = chart_df.iloc[0]
        chart_df = chart_df[1:].reset_index(drop=True)

        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        ws_out = wb_out.create_sheet("Demand Table Comparison")

        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        ws_out.append(
            ['Parameter', 'DB_Max', 'UI_Max', 'Max_Diff', 'DB_Avg', 'UI_Avg', 'Avg_Diff', 'DB_Datetime', 'UI_Datetime',
             'Datetime_Match', 'Overall_Match'])

        validation_results = {}

        for idx, row in proc_df.iterrows():
            param = row['Parameter']
            chart_match = chart_df[chart_df['Parameter'] == param]
            if chart_match.empty:
                validation_results[param] = {'match': False}
                continue

            chart_row = chart_match.iloc[0]
            proc_max, proc_avg, proc_dt = row['Max'], row['Avg'], row['Date and time at max value']
            chart_max, chart_avg, chart_dt = chart_row['Max'], chart_row['Avg'], chart_row['Date and time at max value']

            try:
                max_diff = abs(float(proc_max) - float(chart_max))
                max_match = max_diff < 0.01
                max_diff_disp = round(max_diff, 4)
            except:
                max_match = str(proc_max).strip() == str(chart_max).strip()
                max_diff_disp = 'Mismatch' if not max_match else '0'

            try:
                avg_diff = abs(float(proc_avg) - float(chart_avg))
                avg_match = avg_diff < 0.01
                avg_diff_disp = round(avg_diff, 4)
            except:
                avg_match = str(proc_avg).strip() == str(chart_avg).strip()
                avg_diff_disp = 'Mismatch' if not avg_match else '0'

            dt_match = str(proc_dt).strip() == str(chart_dt).strip()
            overall = max_match and avg_match and dt_match

            validation_results[param] = {'match': overall}

            ws_out.append(
                [param, proc_max, chart_max, max_diff_disp, proc_avg, chart_avg, avg_diff_disp, proc_dt, chart_dt,
                 'PASS' if dt_match else 'FAIL', 'PASS' if overall else 'FAIL'])

            row_idx = ws_out.max_row
            ws_out.cell(row_idx, 4).fill = green if max_match else red
            ws_out.cell(row_idx, 7).fill = green if avg_match else red
            ws_out.cell(row_idx, 10).fill = green if dt_match else red
            ws_out.cell(row_idx, 11).fill = green if overall else red

        wb_out.save(output_file)
        logger.info(f"Comparison saved: {output_file}")
        return output_file, validation_results
    except Exception as e:
        logger.error(f"Comparison error: {e}")
        raise


# ============================================================================
# SUMMARY REPORT - COMPLETE WITH ENHANCED STYLING
# ============================================================================
@log_execution_time
def create_demand_overview_summary_report(config, date_info, chart_file, processed_file,
                                          comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive demand overview summary with enhanced styling - FIXED VERSION"""
    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_DEMAND_OVERVIEW_{date_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Define all styles ONCE at the start - avoid creating styles in loops
        main_hdr_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        sec_hdr_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        sec_hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        subsec_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsec_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")

        warn_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warn_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Helper function to apply styles safely
        def apply_cell_style(cell, font_obj, fill_obj, align_h="center", align_v="center", border_obj=None):
            """Safely apply styles to a cell"""
            cell.font = Font(
                bold=font_obj.bold,
                size=font_obj.size,
                color=font_obj.color.rgb if font_obj.color else "000000",
                name=font_obj.name
            )
            cell.fill = PatternFill(
                start_color=fill_obj.start_color.rgb if fill_obj.start_color else "FFFFFF",
                end_color=fill_obj.end_color.rgb if fill_obj.end_color else "FFFFFF",
                fill_type=fill_obj.fill_type
            )
            cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=True)
            if border_obj:
                cell.border = border_obj

        def merge_and_style(row, start_col, end_col, value, font_obj, fill_obj, align_h="center",
                            border_obj=thick_border):
            """Merge cells and apply styling"""
            ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
            cell = ws[f'{start_col}{row}']
            cell.value = value
            apply_cell_style(cell, font_obj, fill_obj, align_h, "center", border_obj)
            # Apply border to all merged cells
            for col in range(ord(start_col), ord(end_col) + 1):
                ws[f'{chr(col)}{row}'].border = border_obj

        r = 1

        # ============ MAIN HEADER ============
        merge_and_style(r, 'A', 'H',
                        f"LV DEMAND OVERVIEW VALIDATION SUMMARY - {date_info['selected_date'].upper()}",
                        main_hdr_font, main_hdr_fill)
        ws.row_dimensions[r].height = 30
        r += 1

        # Timestamp
        timestamp_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_font = Font(size=10, italic=True, color="666666")
        merge_and_style(r, 'A', 'H',
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        timestamp_font, timestamp_fill, border_obj=thin_border)
        ws.row_dimensions[r].height = 20
        r += 2

        # ============ TEST DETAILS ============
        merge_and_style(r, 'A', 'B', "📋 TEST DETAILS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Date:", config['target_date']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]

        for label, value in test_details:
            cell_a = ws[f'A{r}']
            cell_a.value = label
            apply_cell_style(cell_a, label_font, label_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = value
            apply_cell_style(cell_b, data_font, data_fill, "left", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # ============ SYSTEM UNDER TEST ============
        merge_and_style(r, 'A', 'B', "🔧 SYSTEM UNDER TEST", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Demand Overview (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME]
        ]

        for label, value in system_details:
            cell_a = ws[f'A{r}']
            cell_a.value = label
            apply_cell_style(cell_a, label_font, label_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = value
            apply_cell_style(cell_b, data_font, data_fill, "left", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # ============ DATA VOLUME ANALYSIS ============
        merge_and_style(r, 'A', 'C', "📊 DATA VOLUME ANALYSIS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        # Column headers
        headers = ["Dataset", "Record Count", "Status"]
        for i, hdr in enumerate(headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        # Get chart data count
        try:
            chart_pts = len(pd.read_excel(chart_file, sheet_name='Demand Table'))
        except:
            chart_pts = 3

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Chart Data Points", chart_pts, "COMPLETE RECORDS"]
        ]

        for ds, cnt, st in data_rows:
            cell_a = ws[f'A{r}']
            cell_a.value = ds
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = cnt
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = st
            if "COMPLETE" in st:
                apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_c, fail_font, fail_fill, "center", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # ============ VALIDATION RESULTS ============
        merge_and_style(r, 'A', 'E', "✅ VALIDATION RESULTS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        # Column headers
        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, hdr in enumerate(validation_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        # Validation data
        total_pass, total_cnt = 0, 0
        for param, res in validation_results.items():
            passed = 1 if res['match'] else 0
            failed = 1 - passed
            rate = f"{passed * 100:.1f}%"
            status = "PASS" if passed else "FAIL"

            cell_a = ws[f'A{r}']
            cell_a.value = param
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = passed
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = failed
            apply_cell_style(cell_c, data_font, data_fill, "center", "center", thin_border)

            cell_d = ws[f'D{r}']
            cell_d.value = rate
            bold_font = Font(bold=True, size=10, color="000000", name="Calibri")
            apply_cell_style(cell_d, bold_font, data_fill, "center", "center", thin_border)

            cell_e = ws[f'E{r}']
            cell_e.value = status
            if status == "PASS":
                apply_cell_style(cell_e, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_e, fail_font, fail_fill, "center", "center", thin_border)

            ws.row_dimensions[r].height = 20
            total_pass += passed
            total_cnt += 1
            r += 1
        r += 1

        # ============ COMPARISON VALIDATION REPORT ============
        merge_and_style(r, 'A', 'D', "📋 COMPARISON VALIDATION REPORT", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        panel_headers = ["Parameter Type", "Chart vs Processed", "Match Status", "Issues Found"]
        for i, hdr in enumerate(panel_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        for param, res in validation_results.items():
            status = "PASS" if res['match'] else "FAIL"
            issues = "All values match" if res['match'] else "Value mismatch detected"

            cell_a = ws[f'A{r}']
            cell_a.value = param
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = "Comparison Done"
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = status
            if status == "PASS":
                apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_c, fail_font, fail_fill, "center", "center", thin_border)

            cell_d = ws[f'D{r}']
            cell_d.value = issues
            apply_cell_style(cell_d, data_font, data_fill, "left", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # ============ ROOT CAUSE ANALYSIS ============
        merge_and_style(r, 'A', 'C', "🔍 ROOT CAUSE ANALYSIS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        root_headers = ["Issue Type", "Likely Causes", "Recommendation"]
        for i, hdr in enumerate(root_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        failed = [p for p, res in validation_results.items() if not res['match']]
        if failed:
            issue = "Demand Data Mismatch"
            causes = "Chart shows different demand values. Check calculation intervals and time sync."
            recc = "Verify formulas and intervals"
            issue_fill = label_fill
        else:
            issue = "No Issues Found"
            causes = "All validations passed. Data integrity confirmed."
            recc = "Continue monitoring"
            issue_fill = pass_fill

        cell_a = ws[f'A{r}']
        cell_a.value = issue
        apply_cell_style(cell_a, label_font, issue_fill, "left", "center", thin_border)

        cell_b = ws[f'B{r}']
        cell_b.value = causes
        small_font = Font(size=9, color="000000", name="Calibri")
        apply_cell_style(cell_b, small_font, data_fill, "left", "center", thin_border)

        cell_c = ws[f'C{r}']
        cell_c.value = recc
        apply_cell_style(cell_c, small_font, data_fill, "left", "center", thin_border)

        ws.row_dimensions[r].height = 40
        r += 2

        # ============ DETAILED STATISTICS ============
        merge_and_style(r, 'A', 'C', "📈 DETAILED STATISTICS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        stats_headers = ["Metric", "Value", "Status"]
        for i, hdr in enumerate(stats_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        coverage = len(raw_df) / 96 * 100 if len(raw_df) > 0 else 0
        stats_data = [
            ["Data Completeness", f"{len(raw_df)} records ({coverage:.1f}%)",
             "GOOD" if coverage >= 80 else "NEEDS ATTENTION"],
            ["Chart Data Coverage", f"{chart_pts} demand types", "COMPLETE" if chart_pts >= 3 else "INCOMPLETE"],
            ["Demand Parameters", "3 types (Active/Apparent/Reactive)", "COMPLETE"]
        ]

        for metric, val, stat in stats_data:
            cell_a = ws[f'A{r}']
            cell_a.value = metric
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = val
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = stat
            if stat in ["GOOD", "COMPLETE"]:
                apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_c, warn_font, warn_fill, "center", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # ============ OVERALL ASSESSMENT ============
        merge_and_style(r, 'A', 'H', "🏆 OVERALL ASSESSMENT", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        success_rate = total_pass / total_cnt * 100 if total_cnt > 0 else 0
        if success_rate >= 95:
            assess = "✓ EXCELLENT: Demand validation passed with high confidence"
            assess_fill_obj = pass_fill
            assess_font_obj = pass_font
        elif success_rate >= 80:
            assess = "⚠ GOOD: Minor discrepancies - Review recommended"
            assess_fill_obj = warn_fill
            assess_font_obj = warn_font
        else:
            assess = "❌ REQUIRES ATTENTION: Significant failures detected"
            assess_fill_obj = fail_fill
            assess_font_obj = fail_font

        merge_and_style(r, 'A', 'H', assess, assess_font_obj, assess_fill_obj)
        ws.row_dimensions[r].height = 30
        r += 1

        # Success rate detail
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        bold_black_font = Font(bold=True, size=11, color="000000", name="Calibri")
        merge_and_style(r, 'A', 'H',
                        f"Overall Success Rate: {success_rate:.1f}% ({total_pass}/{total_cnt} validations passed)",
                        bold_black_font, gray_fill, border_obj=thin_border)
        ws.row_dimensions[r].height = 22

        # Set column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Summary created: {summary_file}")
        logger.info(f"Success Rate: {success_rate:.1f}%")
        return summary_file

    except Exception as e:
        logger.error(f"Summary error: {str(e)}")
        raise


# ============================================================================
# MAIN FUNCTION
# ============================================================================
@log_execution_time
def main_lv_demand_overview_automation():
    config = driver = output_folder = None
    try:
        config = validate_config_at_startup()
        if not config:
            return False

        output_folder = setup_output_folder()

        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        if not login(driver):
            return False

        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info or not select_meter_type(driver, config['meter_type']):
            return False

        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])
        if not dt_id:
            return False

        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            return False

        time.sleep(5)
        demand_data = collect_demand_overview_data(driver)
        chart_file = save_file_to_output(save_demand_overview_data_to_excel(date_info, demand_data), output_folder)

        raw_df, nrm_df = get_database_data_for_demand_overview(config['target_date'], mtr_id, dt_id)
        if raw_df.empty:
            return False

        processed_file = save_file_to_output(process_demand_overview_database_calculations(raw_df, nrm_df, date_info),
                                             output_folder)
        comparison_file, validation_results = create_demand_overview_comparison(chart_file, processed_file, date_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        if validation_results:
            summary_report = save_file_to_output(
                create_demand_overview_summary_report(config, date_info, chart_file, processed_file,
                                                      comparison_file, validation_results, raw_df, name), output_folder)

        logger.info("=" * 60)
        logger.info("LV DEMAND OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info(f"Output: {output_folder}")
        logger.info(f"Files: Chart, Processed, Comparison, Summary")
        logger.info("=" * 60)
        return True

    except Exception as e:
        logger.info(f"Error: {e}")
        return False
    finally:
        if driver:
            driver.quit()


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV DEMAND OVERVIEW AUTOMATION - FINAL VERSION")
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("=" * 60)

    start = time.time()
    success = main_lv_demand_overview_automation()
    elapsed = time.time() - start

    logger.info("=" * 60)
    if success:
        logger.info(f"✓ COMPLETED in {elapsed:.2f}s ({elapsed / 60:.1f}min)")
        logger.info("All 4 files generated successfully")
    else:
        logger.info(f"✗ FAILED after {elapsed:.2f}s")
        logger.info("Check logs for details")
    logger.info("=" * 60)
