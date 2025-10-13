import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
import re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import functools


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    TENANT_NAME = "tenant01"

    @classmethod
    def get_db1_params(cls):
        return {"host": cls.DB1_HOST, "port": cls.DB1_PORT, "database": cls.DB1_DATABASE,
                "user": cls.DB1_USER, "password": cls.DB1_PASSWORD}

    @classmethod
    def get_db2_params(cls):
        return {"host": cls.DB2_HOST, "port": cls.DB2_PORT, "database": cls.DB2_DATABASE,
                "user": cls.DB2_USER, "password": cls.DB2_PASSWORD}


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    logger = logging.getLogger('lv_piechart_overview')
    logger.setLevel(logging.INFO)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    log_file = 'logs/lv_piechart_overview.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setFormatter(formatter)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file: {e}")
        return file_path


# ============================================================================
# CONFIGURATION
# ============================================================================
def create_default_config_file(config_file):
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Date', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA', 'YOUR_SUBSTATION', 'YOUR_FEEDER', 'DD/MM/YYYY', 'YOUR_METER', 'DT']
        }
        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            pd.DataFrame(config_data).to_excel(writer, sheet_name='User_Configuration', index=False)
        logger.info(f"Config created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(value, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    try:
        if not os.path.exists(config_file):
            return None
        df = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}
        for _, row in df.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()
        return config
    except Exception as e:
        logger.info(f"Error: {e}")
        return None


def validate_config_at_startup():
    logger.info("=" * 60)
    logger.info("LV PIE CHART OVERVIEW AUTOMATION")
    logger.info("=" * 60)
    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        create_default_config_file(config_file)
        logger.info("Please edit config and restart")
        return None
    config = read_user_configuration(config_file)
    if config:
        logger.info("Configuration validated")
    return config


# ============================================================================
# DECORATOR
# ============================================================================
def log_execution_time(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        logger.info(f"{func.__name__} completed in {time.time() - start:.2f}s")
        return result

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        if meter_type.upper() == 'DT':
            query = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        else:
            query = f"SELECT dt_id, lvfeeder_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        conn.close()
        return result if result else (None, None, None)
    except Exception as e:
        logger.info(f"DB Error: {e}")
        return None, None, None


@log_execution_time
def get_database_data_for_piechart(target_date, dt_name, mtr_serial_no, node_id, meter_type):
    logger.info(f"Fetching pie chart database data for date: {target_date}")
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{next_day}'"

    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())

        # NRM query for KVA data
        nrm_query = f"SELECT surveydate, kva_i FROM {DatabaseConfig.TENANT_NAME}.tb_nrm_loadsurveyprofile WHERE nodeid={node_id} {date_filter} ORDER BY surveydate;"

        # Rating query based on meter type
        if meter_type.upper() == "DT":
            rating_query = f"SELECT kva_rating FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE dt_name = '{dt_name}' AND meter_serial_no = '{mtr_serial_no}' LIMIT 1;"
        else:
            rating_query = f"SELECT conductor_ampacity FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE lvfeeder_name = '{dt_name}' AND meter_serial_no = '{mtr_serial_no}' LIMIT 1;"

        nrm_df = pd.read_sql(nrm_query, conn)

        cursor = conn.cursor()
        cursor.execute(rating_query)
        rating_result = cursor.fetchone()
        cursor.close()
        conn.close()

        rating_value = rating_result[0] if rating_result else None
        logger.info(f"Retrieved: NRM={len(nrm_df)}, Rating={rating_value}")
        return nrm_df, rating_value
    except Exception as e:
        logger.info(f"DB Error: {e}")
        return pd.DataFrame(), None


# ============================================================================
# WEB AUTOMATION
# ============================================================================
def login(driver):
    try:
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("Secure")
        driver.find_element(By.ID, "Password").send_keys("Secure@12345")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    try:
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        for option in driver.find_elements(By.CSS_SELECTOR, ".dx-list-item"):
            if option.text.strip().lower() == option_name.lower():
                option.click()
                return True
        return False
    except:
        return False


def set_calendar_date(driver, target_date):
    try:
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        return {'selected_date': target_dt.strftime("%B %Y"), 'start_date': target_dt.strftime("%Y-%m-%d")}
    except:
        return None


def select_type(driver):
    try:
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        time.sleep(3)
    except Exception as e:
        logger.info(f"Error: {e}")


def select_meter_type(driver, meter_type):
    try:
        wait = WebDriverWait(driver, 10)
        if meter_type == "DT":
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]'))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]'))).click()
        time.sleep(3)
        return True
    except:
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)
        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if view_buttons:
            view_buttons[0].click()
            return True
        return False
    except:
        return False


@log_execution_time
def collect_piechart_overview_data(driver, meter_type):
    """Collect pie chart load data from overview using tooltip extraction"""
    logger.info("Starting pie chart overview data collection...")
    data = {}

    try:
        action = ActionChains(driver)
        wait = WebDriverWait(driver, 5)

        logger.info("Deactivating all chart legends...")
        # Deactivate all legends first
        driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#257E94"]').click()
        driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#86B8A5"]').click()
        driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#DEAE2A"]').click()
        driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#E38430"]').click()
        time.sleep(1)

        # Duration Load <30%
        rect1 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#257E94"]')
        rect1.click()
        time.sleep(1)
        path1 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxc-markers path[fill="#257E94"]')
        action.move_to_element(path1).perform()
        time.sleep(1)
        tooltip = driver.find_element(By.CSS_SELECTOR, '.dxc-tooltip svg text')
        p1_text = tooltip.text
        rect1 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#257E94"]')
        rect1.click()

        # Duration Load 30%-60%
        rect2 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#86B8A5"]')
        rect2.click()
        time.sleep(1)
        path2 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxc-markers path[fill="#86B8A5"]')
        action.move_to_element(path2).perform()
        time.sleep(1)
        tooltip = driver.find_element(By.CSS_SELECTOR, '.dxc-tooltip svg text')
        p2_text = tooltip.text
        rect2 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#86B8A5"]')
        rect2.click()

        # Duration Load 60%-80%
        rect3 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#DEAE2A"]')
        rect3.click()
        time.sleep(1)
        path3 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxc-markers path[fill="#DEAE2A"]')
        action.move_to_element(path3).perform()
        time.sleep(1)
        tooltip = driver.find_element(By.CSS_SELECTOR, '.dxc-tooltip svg text')
        p3_text = tooltip.text
        rect3 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#DEAE2A"]')
        rect3.click()

        # Duration Load >80%
        rect4 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#E38430"]')
        rect4.click()
        time.sleep(1)
        path4 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxc-markers path[fill="#E38430"]')
        action.move_to_element(path4).perform()
        time.sleep(1)
        tooltip = driver.find_element(By.CSS_SELECTOR, '.dxc-tooltip svg text')
        p4_text = tooltip.text
        rect4 = driver.find_element(By.CSS_SELECTOR, '#dvLoadingTrend g.dxl-marker rect[fill="#E38430"]')
        rect4.click()

        # Extract rating and load factor
        if meter_type.upper() == "DT":
            rating_label = "KVA Rating"
            raw_text = driver.find_element(By.XPATH, '//label[@id="lblKvaRating"]').text
        else:
            rating_label = "Conductor Ampacity"
            raw_text = driver.find_element(By.XPATH, '//label[@id="lblConductorAmpacity"]').text

        load_fac = driver.find_element(By.XPATH, '//label[@id="lblLoadFactor"]').text
        rating_value = re.findall(r"[\d.]+", raw_text)[0] if raw_text else "-"

        data['Load Table'] = {
            rating_label: rating_value,
            'Load Factor': load_fac,
            'Duration Load < 30%': p1_text,
            'Duration Load 30% - 60%': p2_text,
            'Duration Load 60% - 80%': p3_text,
            'Duration Load > 80%': p4_text
        }

        logger.info("Pie chart overview data collection completed")
    except Exception as e:
        logger.error(f"Error in pie chart data collection: {e}")
        raise

    return data


@log_execution_time
def save_piechart_overview_data_to_excel(date_info, overview_data):
    try:
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Load Table")
        ws.append(['Parameter', 'Value'])
        for key, value in overview_data['Load Table'].items():
            ws.append([key, value])

        filename = f"chart_data_piechart_overview_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        logger.error(f"Save error: {e}")
        raise


# ============================================================================
# DATABASE PROCESSING
# ============================================================================
@log_execution_time
def process_piechart_overview_database_calculations(nrm_df, rating_value, date_info, meter_type):
    try:
        def format_duration(minutes):
            h, m = divmod(int(minutes), 60)
            return f"{h}:{m:02} hrs"

        date_safe = date_info['selected_date'].replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        interval_minutes = 15 if len(nrm_df) <= 1 else int(
            (nrm_df['surveydate'].iloc[1] - nrm_df['surveydate'].iloc[0]).total_seconds() / 60)

        rating_float = float(rating_value) if rating_value else 1.0
        if rating_float == 0:
            rating_float = 1.0

        # Calculate Load Factor
        avg_load = nrm_df['kva_i'].mean()
        max_demand = nrm_df['kva_i'].max()
        load_factor = round(avg_load / max_demand, 1) if max_demand != 0 else 0

        # Calculate load percentages
        temp_df = nrm_df.copy()
        temp_df['kva_i'] = pd.to_numeric(temp_df['kva_i'], errors='coerce')
        temp_df['load_percent'] = (temp_df['kva_i'] / rating_float) * 100

        # Categorize into bins
        bins = [0, 30, 60, 80, float('inf')]
        labels = ['<30%', '30-60%', '60-80%', '>80%']
        temp_df['load_range'] = pd.cut(temp_df['load_percent'], bins=bins, labels=labels, right=False)

        # Calculate durations
        duration_dict = {label: '0:00 hrs' for label in labels}
        counts = temp_df['load_range'].value_counts()
        for label in labels:
            duration_mins = counts.get(label, 0) * interval_minutes
            if duration_mins > 0:
                duration_dict[label] = format_duration(duration_mins)

        # Prepare Load Table
        load_table_data = [
            ['Parameter', 'Value'],
            [f'{"KVA Rating" if meter_type == "DT" else "Conductor Ampacity"}', rating_value],
            ['Load Factor', load_factor],
            ['Duration Load < 30%', duration_dict['<30%']],
            ['Duration Load 30% - 60%', duration_dict['30-60%']],
            ['Duration Load 60% - 80%', duration_dict['60-80%']],
            ['Duration Load > 80%', duration_dict['>80%']]
        ]

        filename = f"theoretical_piechart_overview_{date_safe}_{timestamp}.xlsx"
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            nrm_df.to_excel(writer, sheet_name='NRM_Database', index=False)
            pd.DataFrame(load_table_data[1:], columns=load_table_data[0]).to_excel(writer, sheet_name='Load Table',
                                                                                   index=False)

        return filename
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise


# ============================================================================
# COMPARISON
# ============================================================================
@log_execution_time
def create_piechart_overview_comparison(chart_file, processed_file, date_info):
    try:
        output_file = f"validation_report_piechart_overview_{date_info['selected_date'].replace(' ', '_')}.xlsx"

        chart_df = pd.read_excel(chart_file, sheet_name="Load Table")
        processed_df = pd.read_excel(processed_file, sheet_name="Load Table")

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        comparison_df = pd.DataFrame()
        comparison_df['Parameter'] = processed_df['Parameter']
        comparison_df['Processed Value'] = processed_df['Value']
        comparison_df['Chart Value'] = chart_df['Value']

        diff_list = []
        match_list = []
        validation_results = {}

        for param, p_val, c_val in zip(comparison_df['Parameter'], comparison_df['Processed Value'],
                                       comparison_df['Chart Value']):
            try:
                p_float = float(p_val)
                c_float = float(c_val)
                diff = round(abs(p_float - c_float), 1)
                if diff <= 0.1:
                    diff_list.append(diff)
                    match_list.append("YES")
                    validation_results[param] = {'match': True}
                else:
                    diff_list.append(diff)
                    match_list.append("NO")
                    validation_results[param] = {'match': False}
            except:
                if str(p_val).strip() == str(c_val).strip():
                    diff_list.append(0)
                    match_list.append("YES")
                    validation_results[param] = {'match': True}
                else:
                    diff_list.append("NOT A MATCH")
                    match_list.append("NO")
                    validation_results[param] = {'match': False}

        comparison_df['Difference'] = diff_list
        comparison_df['Match'] = match_list

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name="Load Table Comparison", index=False)

        # Apply colors
        wb = load_workbook(output_file)
        ws = wb["Load Table Comparison"]
        for row in range(2, ws.max_row + 1):
            diff_cell = ws.cell(row=row, column=4)
            match_cell = ws.cell(row=row, column=5)
            if match_cell.value == "YES":
                diff_cell.fill = green_fill
                match_cell.fill = green_fill
            else:
                diff_cell.fill = red_fill
                match_cell.fill = red_fill
        wb.save(output_file)

        logger.info(f"Comparison saved: {output_file}")
        return output_file, validation_results
    except Exception as e:
        logger.error(f"Comparison error: {e}")
        raise


# ============================================================================
# SUMMARY REPORT
# ============================================================================
@log_execution_time
def create_piechart_overview_summary_report(config, date_info, chart_file, processed_file, comparison_file,
                                            validation_results, nrm_df, meter_name):
    """Create comprehensive pie chart overview summary report"""
    try:
        date_safe = date_info['selected_date'].replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"SUMMARY_PIECHART_OVERVIEW_{date_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary_Report"

        # Styles
        main_header_font = Font(bold=True, size=14, color="FFFFFF")
        main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_header_font = Font(bold=True, size=11, color="FFFFFF")
        section_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subsection_font = Font(bold=True, size=10)
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        label_font = Font(bold=True, size=10)
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_font = Font(size=10)
        pass_font = Font(bold=True, size=10, color="FFFFFF")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        fail_font = Font(bold=True, size=10, color="FFFFFF")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                              bottom=Side(style='medium'))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        row = 1

        # Main Header
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"LV PIE CHART OVERVIEW VALIDATION - {date_info['selected_date'].upper()}"
        cell.font = main_header_font
        cell.fill = main_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        ws.row_dimensions[row].height = 30
        row += 1

        # Timestamp
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        cell.font = Font(size=10, italic=True, color="666666")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        row += 2

        # Test Details
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "📋 TEST DETAILS"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        ws[f'B{row}'].border = thick_border
        row += 1

        for label, value in [["Engineer:", TestEngineer.NAME], ["Designation:", TestEngineer.DESIGNATION],
                             ["Date:", config['target_date']], ["Department:", TestEngineer.DEPARTMENT]]:
            ws[f'A{row}'].value = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = value
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            row += 1
        row += 1

        # System Under Test
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "🔧 SYSTEM UNDER TEST"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        ws[f'B{row}'].border = thick_border
        row += 1

        for label, value in [["Area:", config['area']], ["Substation:", config['substation']],
                             ["Feeder:", config['feeder']], ["Meter:", config['meter_serial_no']],
                             ["Name:", meter_name], ["Type:", config['meter_type']],
                             ["Monitoring:", "LV Pie Chart Overview"], ["Tenant:", DatabaseConfig.TENANT_NAME]]:
            ws[f'A{row}'].value = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = value
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            row += 1
        row += 1

        # Data Volume
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "📊 DATA VOLUME"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        for i, header in enumerate(["Dataset", "Count", "Status"], 1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
        row += 1

        chart_points = 6
        for dataset, count, status in [["NRM DB Records", len(nrm_df), "COMPLETE" if len(nrm_df) > 0 else "NO DATA"],
                                       ["Chart Points", chart_points, "COMPLETE"]]:
            ws[f'A{row}'].value = dataset
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = count
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].value = status
            ws[f'C{row}'].font = pass_font if "COMPLETE" in status else fail_font
            ws[f'C{row}'].fill = pass_fill if "COMPLETE" in status else fail_fill
            ws[f'C{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].border = thin_border
            row += 1
        row += 1

        # Validation Results
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "✅ VALIDATION RESULTS"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        for i, header in enumerate(["Type", "Pass", "Fail", "Rate", "Status"], 1):
            cell = ws.cell(row=row, column=i)
            cell.value = header
            cell.font = subsection_font
            cell.fill = subsection_fill
            cell.border = thin_border
        row += 1

        overall_pass = overall_total = 0
        rating_match = \
        validation_results.get('KVA Rating', validation_results.get('Conductor Ampacity', {'match': False}))['match']
        load_factor_match = validation_results.get('Load Factor', {'match': False})['match']
        duration_params = ['Duration Load < 30%', 'Duration Load 30% - 60%', 'Duration Load 60% - 80%',
                           'Duration Load > 80%']
        duration_matches = sum(
            1 for param in duration_params if validation_results.get(param, {'match': False})['match'])
        duration_total = len(duration_params)

        validation_data = [
            ['Rating/Ampacity', 1 if rating_match else 0, 0 if rating_match else 1,
             "100.0%" if rating_match else "0.0%", "PASS" if rating_match else "FAIL"],
            ['Load Factor', 1 if load_factor_match else 0, 0 if load_factor_match else 1,
             "100.0%" if load_factor_match else "0.0%", "PASS" if load_factor_match else "FAIL"],
            ['Duration Parameters', duration_matches, duration_total - duration_matches,
             f"{(duration_matches / duration_total * 100):.1f}%",
             "PASS" if duration_matches == duration_total else "FAIL"]
        ]

        overall_pass = (1 if rating_match else 0) + (1 if load_factor_match else 0) + duration_matches
        overall_total = 2 + duration_total

        for param_type, matches, mismatches, rate, status in validation_data:
            ws[f'A{row}'].value = param_type
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].value = matches
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = thin_border
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'C{row}'].value = mismatches
            ws[f'C{row}'].font = data_font
            ws[f'C{row}'].border = thin_border
            ws[f'C{row}'].alignment = Alignment(horizontal="center")
            ws[f'D{row}'].value = rate
            ws[f'D{row}'].font = Font(bold=True, size=10)
            ws[f'D{row}'].border = thin_border
            ws[f'D{row}'].alignment = Alignment(horizontal="center")
            ws[f'E{row}'].value = status
            ws[f'E{row}'].font = pass_font if status == "PASS" else fail_font
            ws[f'E{row}'].fill = pass_fill if status == "PASS" else fail_fill
            ws[f'E{row}'].alignment = Alignment(horizontal="center")
            ws[f'E{row}'].border = thin_border
            row += 1
        row += 1

        # Overall Assessment
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = "🏆 OVERALL ASSESSMENT"
        cell.font = section_header_font
        cell.fill = section_header_fill
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thick_border
        row += 1

        success_rate = (overall_pass / overall_total * 100) if overall_total > 0 else 0
        if success_rate >= 95:
            assessment = "✓ EXCELLENT: Validation passed"
            color = pass_fill
            font_color = pass_font
        elif success_rate >= 80:
            assessment = "⚠ GOOD: Minor issues found"
            color = warning_fill
            font_color = Font(bold=True, size=10)
        else:
            assessment = "❌ ATTENTION: Significant failures"
            color = fail_fill
            font_color = fail_font

        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = assessment
        cell.font = font_color
        cell.fill = color
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thick_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thick_border
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Success Rate: {success_rate:.1f}% ({overall_pass}/{overall_total} validations passed)"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{row}'].border = thin_border

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

        wb.save(summary_file)
        logger.info(f"Summary created: {summary_file}")
        return summary_file
    except Exception as e:
        logger.error(f"Summary error: {e}")
        raise


# ============================================================================
# MAIN FUNCTION
# ============================================================================
@log_execution_time
def main_lv_piechart_overview_automation():
    config = driver = output_folder = None
    try:
        config = validate_config_at_startup()
        if not config:
            return False

        output_folder = setup_output_folder()

        logger.info("=" * 60)
        logger.info(f"DB: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        if not login(driver):
            return False

        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info:
            return False

        if not select_meter_type(driver, config['meter_type']):
            return False

        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])
        if not dt_id:
            logger.info(f"Meter not found")
            return False

        logger.info(f"Meter: {name}")
        time.sleep(3)

        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            return False

        time.sleep(5)

        overview_data = collect_piechart_overview_data(driver, config['meter_type'])
        chart_file = save_piechart_overview_data_to_excel(date_info, overview_data)
        chart_file = save_file_to_output(chart_file, output_folder)

        nrm_df, rating_value = get_database_data_for_piechart(config['target_date'], name, config['meter_serial_no'],
                                                              dt_id, config['meter_type'])
        if nrm_df.empty:
            logger.info("No database data")
            return False

        processed_file = process_piechart_overview_database_calculations(nrm_df, rating_value, date_info,
                                                                         config['meter_type'])
        processed_file = save_file_to_output(processed_file, output_folder)

        comparison_file, validation_results = create_piechart_overview_comparison(chart_file, processed_file, date_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        summary_report = create_piechart_overview_summary_report(config, date_info, chart_file, processed_file,
                                                                 comparison_file, validation_results, nrm_df, name)
        summary_report = save_file_to_output(summary_report, output_folder)

        logger.info("=" * 60)
        logger.info("LV PIE CHART OVERVIEW AUTOMATION COMPLETED!")
        logger.info("=" * 60)
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info(f"Output: {output_folder}")
        logger.info(f"Files Generated:")
        logger.info(f"  1. {os.path.basename(chart_file)}")
        logger.info(f"  2. {os.path.basename(processed_file)}")
        logger.info(f"  3. {os.path.basename(comparison_file)}")
        logger.info(f"  4. {os.path.basename(summary_report)}")
        logger.info("=" * 60)

        return True
    except Exception as e:
        logger.info(f"Error: {e}")
        return False
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV PIE CHART OVERVIEW AUTOMATION")
    logger.info(f"Engineer: {TestEngineer.NAME}")
    logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("=" * 60)

    start_time = time.time()
    success = main_lv_piechart_overview_automation()
    total_time = time.time() - start_time

    logger.info("=" * 60)
    if success:
        logger.info("✓ COMPLETED SUCCESSFULLY")
        logger.info(f"Time: {total_time:.2f}s ({total_time / 60:.1f}min)")
        logger.info("Features Applied:")
        logger.info("   ✓ LV Pie Chart Overview monitoring")
        logger.info("   ✓ Search box selection")
        logger.info("   ✓ Centralized DB config")
        logger.info("   ✓ Tooltip extraction for load durations")
        logger.info("   ✓ Load factor calculation")
        logger.info("   ✓ Test engineer details")
        logger.info("   ✓ All 4 output files generated")
    else:
        logger.info("✗ FAILED")
        logger.info(f"Failed after: {total_time:.2f}s")
    logger.info("=" * 60)
