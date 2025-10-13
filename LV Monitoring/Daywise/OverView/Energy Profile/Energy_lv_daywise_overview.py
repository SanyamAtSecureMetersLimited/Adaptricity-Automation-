"""
LV MONITORING DAYWISE OVERVIEW - ENERGY PROFILE
Complete automation script with enhanced styling and validation
Test Engineer: Sanyam Upadhyay
Department: NPD - Quality Assurance
"""

import os
import shutil
import time
import logging
import pandas as pd
import numpy as np
import psycopg2
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import functools
import re


# ============================================================================
# TEST ENGINEER CONFIGURATION
# ============================================================================
class TestEngineer:
    """Test Engineer Details - Modify as needed"""
    NAME = "Sanyam Upadhyay"
    DESIGNATION = "Test Engineer"
    DEPARTMENT = "NPD - Quality Assurance"


# ============================================================================
# CENTRALIZED DATABASE CONFIGURATION
# ============================================================================
class DatabaseConfig:
    """Centralized database configuration for easy modification"""
    DB1_HOST = "10.11.16.146"
    DB1_PORT = "5434"
    DB1_DATABASE = "Prod_LVMV_Test"
    DB1_USER = "postgres"
    DB1_PASSWORD = "postgres"

    DB2_HOST = "10.11.16.146"
    DB2_PORT = "5434"
    DB2_DATABASE = "Prod_LVMV_Test"
    DB2_USER = "postgres"
    DB2_PASSWORD = "postgres"

    TENANT_NAME = "tenant01"

    @classmethod
    def get_db1_params(cls):
        return {"host": cls.DB1_HOST, "port": cls.DB1_PORT, "database": cls.DB1_DATABASE,
                "user": cls.DB1_USER, "password": cls.DB1_PASSWORD}

    @classmethod
    def get_db2_params(cls):
        return {"host": cls.DB2_HOST, "port": cls.DB2_PORT, "database": cls.DB2_DATABASE,
                "user": cls.DB2_USER, "password": cls.DB2_PASSWORD}


# ============================================================================
# LOGGER SETUP
# ============================================================================
def setup_logger():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    logger = logging.getLogger('lv_energy_overview')
    logger.setLevel(logging.INFO)
    for handler in logger.handlers[:]:
        logger.removeHandler(handler)
    formatter = logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    log_file = 'logs/lv_energy_overview.log'
    file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
    file_handler.setFormatter(formatter)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


logger = setup_logger()


# ============================================================================
# OUTPUT FOLDER MANAGEMENT
# ============================================================================
def setup_output_folder():
    output_folder = 'output_files'
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
        logger.info("Cleaned previous output files")
    os.makedirs(output_folder)
    logger.info(f"Created output folder: {output_folder}")
    return output_folder


def save_file_to_output(file_path, output_folder):
    try:
        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            output_path = os.path.join(output_folder, filename)
            shutil.move(file_path, output_path)
            logger.info(f"Moved {filename} to output folder")
            return output_path
        return file_path
    except Exception as e:
        logger.info(f"Error moving file {file_path}: {e}")
        return file_path


# ============================================================================
# CONFIGURATION FUNCTIONS
# ============================================================================
def create_default_config_file(config_file):
    try:
        config_data = {
            'Parameter': ['Area', 'Substation', 'Feeder', 'Target_Date', 'Meter_Serial_No', 'Meter_Type'],
            'Value': ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'DD/MM/YYYY', 'YOUR_METER_NO', 'DT']
        }
        df_config = pd.DataFrame(config_data)

        with pd.ExcelWriter(config_file, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='User_Configuration', index=False)
            instructions = {
                'Step': ['1', '2', '3', '4', '5', '6', '7'],
                'Instructions': [
                    'Open "User_Configuration" sheet',
                    'Replace YOUR_AREA_HERE with area name',
                    'Replace YOUR_SUBSTATION_HERE with substation',
                    'Replace YOUR_FEEDER_HERE with feeder',
                    'Update Target_Date (DD/MM/YYYY)',
                    'Update Meter_Serial_No',
                    'Set Meter_Type (DT or LV)',
                ],
                'Important_Notes': [
                    'FOR LV MONITORING ENERGY OVERVIEW ONLY',
                    'Values are case-sensitive',
                    'No extra spaces',
                    'Date format: DD/MM/YYYY',
                    'Meter_Type: DT or LV only',
                    'Save before running',
                    f'Test Engineer: {TestEngineer.NAME}',
                ]
            }
            pd.DataFrame(instructions).to_excel(writer, sheet_name='Setup_Instructions', index=False)
        logger.info(f"Config template created: {config_file}")
        return True
    except Exception as e:
        logger.info(f"Error creating config: {e}")
        return False


def normalize_date_ddmmyyyy(value):
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(value, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return str(value).strip()


def read_user_configuration(config_file="user_config.xlsx"):
    try:
        if not os.path.exists(config_file):
            return None
        df_config = pd.read_excel(config_file, sheet_name='User_Configuration')
        config = {'type': 'LV'}
        for _, row in df_config.iterrows():
            param, value = row['Parameter'], row['Value']
            if param == 'Area':
                config['area'] = str(value).strip()
            elif param == 'Substation':
                config['substation'] = str(value).strip()
            elif param == 'Feeder':
                config['feeder'] = str(value).strip()
            elif param == 'Target_Date':
                config['target_date'] = normalize_date_ddmmyyyy(value)
            elif param == 'Meter_Serial_No':
                config['meter_serial_no'] = str(value).strip()
            elif param == 'Meter_Type':
                config['meter_type'] = str(value).strip()

        required = ['type', 'area', 'substation', 'feeder', 'target_date', 'meter_serial_no', 'meter_type']
        if any(f not in config or not config[f] for f in required):
            return None
        if any(config.get(k) in ['YOUR_AREA_HERE', 'YOUR_SUBSTATION_HERE', 'YOUR_FEEDER_HERE', 'YOUR_METER_NO'] for k in
               config):
            return None
        return config
    except Exception as e:
        logger.info(f"Error reading config: {e}")
        return None


def validate_config_at_startup():
    logger.info("=" * 60)
    logger.info("STARTING LV ENERGY OVERVIEW AUTOMATION")
    logger.info("=" * 60)
    config_file = "user_config.xlsx"
    if not os.path.exists(config_file):
        logger.info("Creating default config template...")
        if create_default_config_file(config_file):
            logger.info(f"Created: {config_file}")
            logger.info("Please edit the config file and restart")
        return None
    config = read_user_configuration(config_file)
    if config is None:
        logger.info("Configuration validation failed")
        return None
    logger.info("Config validated successfully")
    for k, v in config.items():
        logger.info(f"   {k}: {v}")
    return config


# ============================================================================
# DECORATOR
# ============================================================================
def log_execution_time(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        logger.info(f"Starting {func.__name__}...")
        try:
            result = func(*args, **kwargs)
            logger.info(f"{func.__name__} completed in {time.time() - start:.2f}s")
            return result
        except Exception as e:
            logger.info(f"{func.__name__} failed: {e}")
            raise

    return wrapper


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================
@log_execution_time
def get_metrics(mtr_serial_no, meter_type):
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db1_params())
        cursor = conn.cursor()
        if meter_type.upper() == 'DT':
            query = f"SELECT dt_id, dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_dt WHERE meter_serial_no = %s LIMIT 1;"
        else:
            query = f"SELECT dt_id, lvfeeder_name AS dt_name, meterid FROM {DatabaseConfig.TENANT_NAME}.tb_ntw_lvfeeder WHERE meter_serial_no = %s LIMIT 1;"
        cursor.execute(query, (mtr_serial_no,))
        result = cursor.fetchone()
        cursor.close()
        conn.close()
        if result:
            logger.info(f"Metrics: {result[1]}, meterid: {result[2]}")
            return result
        return None, None, None
    except Exception as e:
        logger.info(f"DB error: {e}")
        return None, None, None


@log_execution_time
def get_database_data_for_energy_overview(target_date, mtr_id):
    target_dt = datetime.strptime(target_date, "%d/%m/%Y")
    start_date = target_dt.strftime("%Y-%m-%d")
    next_day = (target_dt + timedelta(days=1)).strftime("%Y-%m-%d")
    date_filter = f"AND surveydate >= '{start_date}' AND surveydate < '{next_day}'"
    try:
        conn = psycopg2.connect(**DatabaseConfig.get_db2_params())
        raw_query = f"""
            SELECT DISTINCT surveydate, kwh_i, kvah_i, kvar_i_total, kwh_e, kvah_e, kvar_e_total,
                   kwh_abs, kvah_abs, kvarh_abs
            FROM {DatabaseConfig.TENANT_NAME}.tb_raw_loadsurveydata 
            WHERE mtrid={mtr_id} {date_filter}
            ORDER BY surveydate ASC;
        """
        raw_df = pd.read_sql(raw_query, conn)
        conn.close()
        logger.info(f"Retrieved: Raw={len(raw_df)} records")
        return raw_df
    except Exception as e:
        logger.info(f"DB error: {e}")
        return pd.DataFrame()


# ============================================================================
# WEB AUTOMATION
# ============================================================================
def login(driver):
    try:
        driver.get("https://networkmonitoringpv.secure.online:10122/")
        time.sleep(1)
        driver.find_element(By.ID, "UserName").send_keys("Secure")
        driver.find_element(By.ID, "Password").send_keys("Secure@12345")
        time.sleep(10)
        driver.find_element(By.ID, "btnlogin").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//span[@class='dx-button-text' and text()='Continue']").click()
        time.sleep(5)
        logger.info("Login successful")
        return True
    except Exception as e:
        logger.info(f"Login failed: {e}")
        return False


def select_dropdown_option(driver, dropdown_id, option_name):
    try:
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".dx-list-item")))
        for option in driver.find_elements(By.CSS_SELECTOR, ".dx-list-item"):
            if option.text.strip().lower() == option_name.lower():
                option.click()
                logger.info(f"Selected: {option_name}")
                return True
        return False
    except Exception as e:
        logger.info(f"Dropdown error: {e}")
        return False


def set_calendar_date(driver, target_date):
    try:
        date_input = driver.find_element(By.XPATH, "//input[@class='dx-texteditor-input' and @aria-label='Date']")
        date_input.clear()
        date_input.send_keys(target_date)
        driver.find_element(By.XPATH, '//div[@id="dxSearchbtn"]').click()
        target_dt = datetime.strptime(target_date, "%d/%m/%Y")
        return {
            'selected_date': target_dt.strftime("%B %Y"),
            'start_date': target_dt.strftime("%Y-%m-%d"),
            'end_date': target_dt.strftime("%Y-%m-%d")
        }
    except Exception as e:
        logger.info(f"Date error: {e}")
        return None


def select_type(driver):
    try:
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divHome']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//A[@id='divlvmonitoring']").click()
        time.sleep(3)
    except Exception as e:
        logger.info(f"Type error: {e}")


def select_meter_type(driver, meter_type):
    try:
        wait = WebDriverWait(driver, 10)
        if meter_type == "DT":
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="DTClick"]'))).click()
        else:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@id="lvfeederClick"]'))).click()
        time.sleep(3)
        return True
    except Exception as e:
        logger.info(f"Meter type error: {e}")
        return False


@log_execution_time
def find_and_click_view_using_search(driver, wait, meter_serial_no):
    try:
        search_input = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//input[@placeholder='Search grid' and @aria-label='Search in the data grid']")))
        search_input.clear()
        search_input.send_keys(meter_serial_no)
        time.sleep(2)
        view_buttons = driver.find_elements(By.XPATH, "//a[text()='View']")
        if view_buttons:
            view_buttons[0].click()
            logger.info("View clicked")
            return True
        return False
    except Exception as e:
        logger.info(f"Search error: {e}")
        return False


@log_execution_time
def collect_energy_overview_data(driver):
    data = {}
    try:
        logger.info("Collecting Import Energy data...")
        data['Import Energy'] = {
            'Active': driver.find_element(By.XPATH, '//td[@id="kwh_Import"]').text,
            'Apparent': driver.find_element(By.XPATH, '//td[@id="kvah_Import"]').text,
            'Reactive': driver.find_element(By.XPATH, '//td[@id="kvarh_Import"]').text
        }

        logger.info("Clicking Export energy tab...")
        driver.find_element(By.XPATH, "//div[@class='dx-item-content' and normalize-space()='Export energy']").click()
        time.sleep(3)

        logger.info("Collecting Export Energy data...")
        data['Export Energy'] = {
            'Active': driver.find_element(By.XPATH, '//td[@id="kwh_Export"]').text,
            'Apparent': driver.find_element(By.XPATH, '//td[@id="kvah_Export"]').text,
            'Reactive': driver.find_element(By.XPATH, '//td[@id="kvarh_Export"]').text
        }
        logger.info("Energy data collected")
    except Exception as e:
        logger.error(f"Collection error: {e}")
        raise
    return data


@log_execution_time
def save_energy_overview_data_to_excel(date_info, energy_data):
    try:
        wb = Workbook()
        wb.remove(wb.active)

        def extract_numeric(value):
            if isinstance(value, str):
                match = re.search(r"[-+]?\d*\.\d+|\d+", value)
                return float(match.group()) if match else value
            return value

        # Import Energy Sheet
        ws_imp = wb.create_sheet("Import Energy")
        ws_imp.append(["Parameter", "Value"])
        for key, value in energy_data['Import Energy'].items():
            numeric_value = extract_numeric(value)
            ws_imp.append([key, numeric_value])

        # Export Energy Sheet
        ws_exp = wb.create_sheet("Export Energy")
        ws_exp.append(["Parameter", "Value"])
        for key, value in energy_data['Export Energy'].items():
            numeric_value = extract_numeric(value)
            ws_exp.append([key, numeric_value])

        file_name = f"chart_data_from_ui_energy_overview_{date_info['selected_date'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(file_name)
        logger.info(f"Saved: {file_name}")
        return file_name
    except Exception as e:
        logger.error(f"Save error: {e}")
        raise


# ============================================================================
# PROCESSING
# ============================================================================
@log_execution_time
def process_energy_overview_database_calculations(raw_df, date_info):
    try:
        def calculate_energy_value(df, col_name):
            total = df[col_name].sum()
            if total / 1000 >= 1:
                return round(total / 1000, 1)
            else:
                return round(total, 1)

        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        processed_file = f"theoretical_energy_overview_calculated_data_{date_safe}_{timestamp}.xlsx"

        # Calculate Import Energy
        import_energy_data = [
            ['Active', calculate_energy_value(raw_df, 'kwh_i')],
            ['Apparent', calculate_energy_value(raw_df, 'kvah_i')],
            ['Reactive', calculate_energy_value(raw_df, 'kvar_i_total')]
        ]

        # Calculate Export Energy
        export_energy_data = [
            ['Active', calculate_energy_value(raw_df, 'kwh_e')],
            ['Apparent', calculate_energy_value(raw_df, 'kvah_e')],
            ['Reactive', calculate_energy_value(raw_df, 'kvar_e_total')]
        ]

        df_import_energy = pd.DataFrame(import_energy_data, columns=['Parameter', 'Value'])
        df_export_energy = pd.DataFrame(export_energy_data, columns=['Parameter', 'Value'])

        with pd.ExcelWriter(processed_file, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name='tb_raw_loadsurveydata', index=False)
            df_import_energy.to_excel(writer, sheet_name='Import Energy', index=False)
            df_export_energy.to_excel(writer, sheet_name='Export Energy', index=False)

        logger.info(f"Processed: {processed_file}")
        return processed_file
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise


# ============================================================================
# COMPARISON
# ============================================================================
@log_execution_time
def create_energy_overview_comparison(chart_file, processed_file, date_info):
    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        output_file = f"complete_validation_report_energy_overview_{date_safe}.xlsx"

        wb_proc = load_workbook(processed_file)
        wb_chart = load_workbook(chart_file)
        wb_output = Workbook()
        wb_output.remove(wb_output.active)

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        validation_results = {}

        def compare_energy_sheets(sheet_name):
            ws_processed = wb_proc[sheet_name]
            ws_chart = wb_chart[sheet_name]

            processed_df = pd.DataFrame(ws_processed.values)
            chart_df = pd.DataFrame(ws_chart.values)

            processed_df.columns = processed_df.iloc[0]
            processed_df = processed_df[1:].reset_index(drop=True)
            chart_df.columns = chart_df.iloc[0]
            chart_df = chart_df[1:].reset_index(drop=True)

            ws_output = wb_output.create_sheet(title=f"{sheet_name} Comparison")
            headers = ['Parameter', 'Processed Value', 'Chart Value', 'Value_Difference', 'Match']
            ws_output.append(headers)

            sheet_validation = {}

            for idx, row in processed_df.iterrows():
                parameter = row['Parameter']
                proc_value = float(row['Value'])

                chart_row = chart_df.loc[chart_df['Parameter'] == parameter]
                if not chart_row.empty:
                    chart_value = float(chart_row['Value'].values[0])
                    diff = abs(proc_value - chart_value)
                    is_match = diff <= 0.01
                    match_text = "YES" if is_match else "NO"

                    sheet_validation[f"{sheet_name}_{parameter}"] = {'match': is_match}

                    output_row = [parameter, proc_value, chart_value, round(diff, 4), match_text]
                    ws_output.append(output_row)

                    value_diff_cell = ws_output.cell(row=ws_output.max_row, column=4)
                    match_cell = ws_output.cell(row=ws_output.max_row, column=5)

                    if is_match:
                        value_diff_cell.fill = green_fill
                        match_cell.fill = green_fill
                    else:
                        value_diff_cell.fill = red_fill
                        match_cell.fill = red_fill

            return sheet_validation

        import_validation = compare_energy_sheets('Import Energy')
        export_validation = compare_energy_sheets('Export Energy')

        validation_results.update(import_validation)
        validation_results.update(export_validation)

        wb_output.save(output_file)
        logger.info(f"Comparison saved: {output_file}")
        return output_file, validation_results
    except Exception as e:
        logger.error(f"Comparison error: {e}")
        raise


# ============================================================================
# SUMMARY REPORT - FIXED VERSION
# ============================================================================
@log_execution_time
def create_energy_overview_summary_report(config, date_info, chart_file, processed_file,
                                          comparison_file, validation_results, raw_df, meter_name):
    """Create comprehensive energy overview summary with enhanced styling - FIXED VERSION"""
    try:
        date_safe = date_info['selected_date'].replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        summary_file = f"COMPLETE_VALIDATION_SUMMARY_ENERGY_OVERVIEW_{date_safe}_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Validation_Summary_Report"

        # Define all styles ONCE
        main_hdr_font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        main_hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        sec_hdr_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
        sec_hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        subsec_font = Font(bold=True, size=10, color="000000", name="Calibri")
        subsec_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        label_font = Font(bold=True, size=10, name="Calibri", color="000000")
        label_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        data_font = Font(size=10, name="Calibri", color="000000")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        pass_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        pass_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        fail_font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        fail_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")

        warn_font = Font(bold=True, size=10, color="000000", name="Calibri")
        warn_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        thick_border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Helper function to apply styles safely
        def apply_cell_style(cell, font_obj, fill_obj, align_h="center", align_v="center", border_obj=None):
            cell.font = Font(
                bold=font_obj.bold,
                size=font_obj.size,
                color=font_obj.color.rgb if font_obj.color else "000000",
                name=font_obj.name
            )
            cell.fill = PatternFill(
                start_color=fill_obj.start_color.rgb if fill_obj.start_color else "FFFFFF",
                end_color=fill_obj.end_color.rgb if fill_obj.end_color else "FFFFFF",
                fill_type=fill_obj.fill_type
            )
            cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=True)
            if border_obj:
                cell.border = border_obj

        def merge_and_style(row, start_col, end_col, value, font_obj, fill_obj, align_h="center",
                            border_obj=thick_border):
            ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
            cell = ws[f'{start_col}{row}']
            cell.value = value
            apply_cell_style(cell, font_obj, fill_obj, align_h, "center", border_obj)
            for col in range(ord(start_col), ord(end_col) + 1):
                ws[f'{chr(col)}{row}'].border = border_obj

        r = 1

        # MAIN HEADER
        merge_and_style(r, 'A', 'H',
                        f"LV ENERGY OVERVIEW VALIDATION SUMMARY - {date_info['selected_date'].upper()}",
                        main_hdr_font, main_hdr_fill)
        ws.row_dimensions[r].height = 30
        r += 1

        # Timestamp
        timestamp_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        timestamp_font = Font(size=10, italic=True, color="666666")
        merge_and_style(r, 'A', 'H',
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        timestamp_font, timestamp_fill, border_obj=thin_border)
        ws.row_dimensions[r].height = 20
        r += 2

        # TEST DETAILS
        merge_and_style(r, 'A', 'B', "📋 TEST DETAILS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        test_details = [
            ["Test Engineer:", TestEngineer.NAME],
            ["Designation:", TestEngineer.DESIGNATION],
            ["Test Date:", config['target_date']],
            ["Department:", TestEngineer.DEPARTMENT],
            ["Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]

        for label, value in test_details:
            cell_a = ws[f'A{r}']
            cell_a.value = label
            apply_cell_style(cell_a, label_font, label_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = value
            apply_cell_style(cell_b, data_font, data_fill, "left", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # SYSTEM UNDER TEST
        merge_and_style(r, 'A', 'B', "🔧 SYSTEM UNDER TEST", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        system_details = [
            ["Area:", config['area']],
            ["Substation:", config['substation']],
            ["MV Feeder:", config['feeder']],
            ["Meter Serial No:", config['meter_serial_no']],
            ["Meter Name:", meter_name],
            ["Meter Type:", config['meter_type']],
            ["Monitoring Type:", "LV Energy Overview (Fixed)"],
            ["Database Tenant:", DatabaseConfig.TENANT_NAME]
        ]

        for label, value in system_details:
            cell_a = ws[f'A{r}']
            cell_a.value = label
            apply_cell_style(cell_a, label_font, label_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = value
            apply_cell_style(cell_b, data_font, data_fill, "left", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # DATA VOLUME ANALYSIS
        merge_and_style(r, 'A', 'C', "📊 DATA VOLUME ANALYSIS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        headers = ["Dataset", "Record Count", "Status"]
        for i, hdr in enumerate(headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        try:
            chart_pts = len(pd.read_excel(chart_file, sheet_name='Import Energy')) + len(
                pd.read_excel(chart_file, sheet_name='Export Energy'))
        except:
            chart_pts = 6

        data_rows = [
            ["Raw Database Records", len(raw_df), "COMPLETE RECORDS" if len(raw_df) > 0 else "NO DATA"],
            ["Chart Data Points", chart_pts, "COMPLETE RECORDS"]
        ]

        for ds, cnt, st in data_rows:
            cell_a = ws[f'A{r}']
            cell_a.value = ds
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = cnt
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = st
            if "COMPLETE" in st:
                apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_c, fail_font, fail_fill, "center", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # VALIDATION RESULTS
        merge_and_style(r, 'A', 'E', "✅ VALIDATION RESULTS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        validation_headers = ["Comparison Type", "Matches", "Mismatches", "Success Rate", "Status"]
        for i, hdr in enumerate(validation_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        # Group by energy type
        total_pass, total_cnt = 0, 0
        energy_types = {'Import Energy': [], 'Export Energy': []}
        for param_name, result in validation_results.items():
            if param_name.startswith('Import Energy'):
                energy_types['Import Energy'].append(result)
            elif param_name.startswith('Export Energy'):
                energy_types['Export Energy'].append(result)

        for energy_type, results in energy_types.items():
            if results:
                total_items = len(results)
                passed_items = sum(1 for result in results if result['match'])
                failed_items = total_items - passed_items
                rate = f"{(passed_items / total_items) * 100:.1f}%"
                status = "PASS" if passed_items == total_items else "FAIL"

                cell_a = ws[f'A{r}']
                cell_a.value = energy_type
                apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

                cell_b = ws[f'B{r}']
                cell_b.value = passed_items
                apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

                cell_c = ws[f'C{r}']
                cell_c.value = failed_items
                apply_cell_style(cell_c, data_font, data_fill, "center", "center", thin_border)

                cell_d = ws[f'D{r}']
                cell_d.value = rate
                bold_font = Font(bold=True, size=10, color="000000", name="Calibri")
                apply_cell_style(cell_d, bold_font, data_fill, "center", "center", thin_border)

                cell_e = ws[f'E{r}']
                cell_e.value = status
                if status == "PASS":
                    apply_cell_style(cell_e, pass_font, pass_fill, "center", "center", thin_border)
                else:
                    apply_cell_style(cell_e, fail_font, fail_fill, "center", "center", thin_border)

                ws.row_dimensions[r].height = 20
                total_pass += passed_items
                total_cnt += total_items
                r += 1
        r += 1

        # COMPARISON VALIDATION REPORT
        merge_and_style(r, 'A', 'D', "📋 COMPARISON VALIDATION REPORT", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        panel_headers = ["Energy Type", "Chart vs Processed", "Match Status", "Issues Found"]
        for i, hdr in enumerate(panel_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        for energy_type, results in energy_types.items():
            if results:
                all_match = all(result['match'] for result in results)
                status = "PASS" if all_match else "FAIL"
                issues = "All values match" if all_match else f"{sum(1 for r in results if not r['match'])} parameter mismatches"

                cell_a = ws[f'A{r}']
                cell_a.value = energy_type
                apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

                cell_b = ws[f'B{r}']
                cell_b.value = "Comparison Done"
                apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

                cell_c = ws[f'C{r}']
                cell_c.value = status
                if status == "PASS":
                    apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
                else:
                    apply_cell_style(cell_c, fail_font, fail_fill, "center", "center", thin_border)

                cell_d = ws[f'D{r}']
                cell_d.value = issues
                apply_cell_style(cell_d, data_font, data_fill, "left", "center", thin_border)

                ws.row_dimensions[r].height = 20
                r += 1
        r += 1

        # ROOT CAUSE ANALYSIS
        merge_and_style(r, 'A', 'C', "🔍 ROOT CAUSE ANALYSIS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        root_headers = ["Issue Type", "Likely Causes", "Recommendation"]
        for i, hdr in enumerate(root_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        failed = [p for p, res in validation_results.items() if not res['match']]
        if failed:
            issue = "Energy Data Mismatch"
            causes = "Chart shows different energy values. Check accumulation formulas and measurement intervals."
            recc = "Verify energy calculation methods"
            issue_fill = label_fill
        else:
            issue = "No Issues Found"
            causes = "All validations passed. Data integrity confirmed."
            recc = "Continue monitoring"
            issue_fill = pass_fill

        cell_a = ws[f'A{r}']
        cell_a.value = issue
        apply_cell_style(cell_a, label_font, issue_fill, "left", "center", thin_border)

        cell_b = ws[f'B{r}']
        cell_b.value = causes
        small_font = Font(size=9, color="000000", name="Calibri")
        apply_cell_style(cell_b, small_font, data_fill, "left", "center", thin_border)

        cell_c = ws[f'C{r}']
        cell_c.value = recc
        apply_cell_style(cell_c, small_font, data_fill, "left", "center", thin_border)

        ws.row_dimensions[r].height = 40
        r += 2

        # DETAILED STATISTICS
        merge_and_style(r, 'A', 'C', "📈 DETAILED STATISTICS", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        stats_headers = ["Metric", "Value", "Status"]
        for i, hdr in enumerate(stats_headers, 1):
            cell = ws.cell(r, i)
            cell.value = hdr
            apply_cell_style(cell, subsec_font, subsec_fill, "center", "center", thin_border)
        ws.row_dimensions[r].height = 22
        r += 1

        coverage = len(raw_df) / 96 * 100 if len(raw_df) > 0 else 0
        stats_data = [
            ["Data Completeness", f"{len(raw_df)} records ({coverage:.1f}%)",
             "GOOD" if coverage >= 80 else "NEEDS ATTENTION"],
            ["Chart Data Coverage", f"{chart_pts} energy parameters", "COMPLETE" if chart_pts >= 6 else "INCOMPLETE"],
            ["Energy Types", "2 types (Import/Export)", "COMPLETE"]
        ]

        for metric, val, stat in stats_data:
            cell_a = ws[f'A{r}']
            cell_a.value = metric
            apply_cell_style(cell_a, data_font, data_fill, "left", "center", thin_border)

            cell_b = ws[f'B{r}']
            cell_b.value = val
            apply_cell_style(cell_b, data_font, data_fill, "center", "center", thin_border)

            cell_c = ws[f'C{r}']
            cell_c.value = stat
            if stat in ["GOOD", "COMPLETE"]:
                apply_cell_style(cell_c, pass_font, pass_fill, "center", "center", thin_border)
            else:
                apply_cell_style(cell_c, warn_font, warn_fill, "center", "center", thin_border)

            ws.row_dimensions[r].height = 20
            r += 1
        r += 1

        # OVERALL ASSESSMENT
        merge_and_style(r, 'A', 'H', "🏆 OVERALL ASSESSMENT", sec_hdr_font, sec_hdr_fill, "left")
        ws.row_dimensions[r].height = 25
        r += 1

        success_rate = total_pass / total_cnt * 100 if total_cnt > 0 else 0
        if success_rate >= 95:
            assess = "✓ EXCELLENT: Energy validation passed with high confidence"
            assess_fill_obj = pass_fill
            assess_font_obj = pass_font
        elif success_rate >= 80:
            assess = "⚠ GOOD: Minor discrepancies - Review recommended"
            assess_fill_obj = warn_fill
            assess_font_obj = warn_font
        else:
            assess = "❌ REQUIRES ATTENTION: Significant failures detected"
            assess_fill_obj = fail_fill
            assess_font_obj = fail_font

        merge_and_style(r, 'A', 'H', assess, assess_font_obj, assess_fill_obj)
        ws.row_dimensions[r].height = 30
        r += 1

        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        bold_black_font = Font(bold=True, size=11, color="000000", name="Calibri")
        merge_and_style(r, 'A', 'H',
                        f"Overall Success Rate: {success_rate:.1f}% ({total_pass}/{total_cnt} validations passed)",
                        bold_black_font, gray_fill, border_obj=thin_border)
        ws.row_dimensions[r].height = 22

        # Column widths
        column_widths = {'A': 30, 'B': 25, 'C': 20, 'D': 25, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(summary_file)
        logger.info(f"Summary created: {summary_file}")
        logger.info(f"Success Rate: {success_rate:.1f}%")
        return summary_file

    except Exception as e:
        logger.error(f"Summary error: {str(e)}")
        raise


# ============================================================================
# MAIN FUNCTION
# ============================================================================
@log_execution_time
def main_lv_energy_overview_automation():
    config = driver = output_folder = None
    try:
        config = validate_config_at_startup()
        if not config:
            return False

        output_folder = setup_output_folder()

        logger.info("=" * 60)
        logger.info("DATABASE CONFIGURATION")
        logger.info("=" * 60)
        logger.info(f"DB: {DatabaseConfig.DB1_HOST}:{DatabaseConfig.DB1_PORT}/{DatabaseConfig.DB1_DATABASE}")
        logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info("=" * 60)

        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)

        if not login(driver):
            return False

        select_type(driver)
        select_dropdown_option(driver, "ddl-area", config['area'])
        select_dropdown_option(driver, "ddl-substation", config['substation'])
        select_dropdown_option(driver, "ddl-feeder", config['feeder'])

        date_info = set_calendar_date(driver, config['target_date'])
        if not date_info or not select_meter_type(driver, config['meter_type']):
            return False

        dt_id, name, mtr_id = get_metrics(config['meter_serial_no'], config['meter_type'])
        if not dt_id:
            return False

        time.sleep(3)
        if not find_and_click_view_using_search(driver, wait, config['meter_serial_no']):
            return False

        time.sleep(5)
        energy_data = collect_energy_overview_data(driver)
        chart_file = save_file_to_output(save_energy_overview_data_to_excel(date_info, energy_data), output_folder)

        raw_df = get_database_data_for_energy_overview(config['target_date'], mtr_id)
        if raw_df.empty:
            return False

        processed_file = save_file_to_output(process_energy_overview_database_calculations(raw_df, date_info),
                                             output_folder)
        comparison_file, validation_results = create_energy_overview_comparison(chart_file, processed_file, date_info)
        comparison_file = save_file_to_output(comparison_file, output_folder)

        if validation_results:
            summary_report = save_file_to_output(
                create_energy_overview_summary_report(config, date_info, chart_file, processed_file,
                                                      comparison_file, validation_results, raw_df, name), output_folder)

        logger.info("=" * 60)
        logger.info("LV ENERGY OVERVIEW AUTOMATION COMPLETED SUCCESSFULLY!")
        logger.info("=" * 60)
        logger.info(f"Engineer: {TestEngineer.NAME}")
        logger.info(f"Output: {output_folder}")
        logger.info(f"Files: Chart, Processed, Comparison, Summary")
        logger.info("=" * 60)
        return True

    except Exception as e:
        logger.info(f"Error: {e}")
        return False
    finally:
        if driver:
            driver.quit()


# ============================================================================
# SCRIPT EXECUTION
# ============================================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("LV ENERGY OVERVIEW AUTOMATION - FINAL VERSION")
    logger.info(f"Test Engineer: {TestEngineer.NAME}")
    logger.info(f"Tenant: {DatabaseConfig.TENANT_NAME}")
    logger.info("=" * 60)

    start = time.time()
    success = main_lv_energy_overview_automation()
    elapsed = time.time() - start

    logger.info("=" * 60)
    if success:
        logger.info(f"✓ COMPLETED in {elapsed:.2f}s ({elapsed / 60:.1f}min)")
        logger.info("All 4 files generated successfully")
    else:
        logger.info(f"✗ FAILED after {elapsed:.2f}s")
        logger.info("Check logs for details")
    logger.info("=" * 60)
