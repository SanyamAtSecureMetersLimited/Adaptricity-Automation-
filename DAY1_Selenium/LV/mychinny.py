import io

import os
import cv2
import re
import numpy as np
import openpyxl
import pandas as pd
import psycopg2
import selenium
from PIL import Image
import pytesseract
from openpyxl.styles import PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
from openpyxl.styles import PatternFill, Border, Side
from xlsxwriter import Workbook

# Define chart type, month, and year - modify these as needed
CHART_TYPE = "Demand"  # Options: "Voltage", "Current", "Energy", "Demand"
MONTH = 6  # Month number (1-12)
YEAR = 2024  # Year (e.g., 2023)

# Parameter configurations by chart type
PARAMETER_CONFIGS = {
    "Voltage": {
        "chart_params": ["Date", "Phase 1", "Phase 2", "Phase 3", "Avg"],
        "db_columns": ["dateandtime", "avgvoltage_ph1", "avgvoltage_ph2", "avgvoltage_ph3", "avgvoltage_avg"],
        "tab_selector": "//span[@class='dx-tab-text-span' and text()='Voltage']"
    },
    "Current": {
        "chart_params": ["Date", "Line 1", "Line 2", "Line 3", "Avg", "Neutral"],
        "db_columns": ["dateandtime", "avgcurrent_ph1", "avgcurrent_ph2", "avgcurrent_ph3", "avgcurrent_avg",
                      "avgneutralcurrent"],
        "tab_selector": "//span[@class='dx-tab-text-span' and text()='Current']"
    },
    "Energy": {
        "chart_params": ["Date", "Active", "Apparent", "Reactive"],
        "db_columns": ["dateandtime", "activepowersum", "apparentpowersum", "reactivepowersum"],
        "tab_selector": "//span[@class='dx-tab-text-span' and text()='Energy']"
    },
    "Demand": {
        "chart_params": ["Date", "Active", "Apparent", "Reactive"],
        "db_columns": ["dateandtime", "avgdemand_kw", "avgdemand_kva", "avgdemand_kvar"],
        "tab_selector": "//span[@class='dx-tab-text-span' and text()='Demand']"
    }
}
# Get the configuration for the selected chart type
# Get the configuration for the selected chart type
if CHART_TYPE in PARAMETER_CONFIGS:
    config = PARAMETER_CONFIGS[CHART_TYPE]
    CHART_PARAMETERS = config["chart_params"]
    DB_COLUMNS = config["db_columns"]
    TAB_SELECTOR = config["tab_selector"]
else:
    # Default if chart type not recognized
    print(f"⚠️ Warning: Chart type '{CHART_TYPE}' not recognized, using Voltage parameters")
    CHART_PARAMETERS = PARAMETER_CONFIGS["Demand"]["chart_params"]
    DB_COLUMNS = PARAMETER_CONFIGS["Demand"]["db_columns"]
    TAB_SELECTOR = PARAMETER_CONFIGS["Demand"]["tab_selector"]
# **Step 1: Set Up WebDriver**
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

driver = webdriver.Chrome(options=options)
driver.get("https://networkmonitoring.secure.online:10122/Account/Login")
time.sleep(3)  # Allow page to load
driver.maximize_window()

# **Step 2: Enter Login Credentials**
username_input = WebDriverWait(driver, 5).until(
    EC.presence_of_element_located((By.XPATH, "//input[@name='UserName']"))
)
username_input.clear()
username_input.send_keys("SanyamU")

password_input = driver.find_element(By.XPATH, "//input[@name='Password']")
password_input.clear()
password_input.send_keys("Secure@12345")

login_button = driver.find_element(By.ID, 'btnlogin')
login_button.click()


# it will select tenanant name
# driver.find_element(By.XPATH,"//button[@id='btnGo']").click()
# **Step 3: Function to Process CAPTCHA Image**
def preprocess_captcha(image_bytes):
    """Preprocess CAPTCHA image for better OCR accuracy."""
    img = Image.open(io.BytesIO(image_bytes))
    img = img.convert("L")  # Convert to grayscale
    img = np.array(img)

    # Apply adaptive thresholding for better clarity
    img = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)

    # Morphological operations to clean up noise
    kernel = np.ones((2, 2), np.uint8)
    img = cv2.erode(img, kernel, iterations=1)
    img = cv2.dilate(img, kernel, iterations=1)

    return Image.fromarray(img)


# **Step 4: Function to Extract Text from CAPTCHA**
def extract_captcha_text(image):
    """Extract text from processed CAPTCHA using OCR with regex filtering."""
    pytesseract.pytesseract.tesseract_cmd = r"C:\\Users\\110573\\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract.exe"
    custom_config = r"--psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"

    for _ in range(3):  # Try multiple times to get a valid CAPTCHA
        raw_text = pytesseract.image_to_string(image, config=custom_config).strip()
        match = re.search(r"(\d{3}[A-Z]{3})", raw_text)  # Match 'NNNLLL' pattern
        if match:
            return match.group(0)  # Return only the valid part

    return ""  # Return empty if no valid match found


# **Step 5: Function to Refresh CAPTCHA**
def refresh_captcha():
    """Refresh CAPTCHA if login fails."""
    try:
        refresh_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//i[@class='float-right fas fa-sync-alt']"))
        )
        refresh_button.click()
        time.sleep(2)
        print("🔄 CAPTCHA Refreshed")
    except (NoSuchElementException, TimeoutException):
        print("⚠️ CAPTCHA refresh button not found!")


# **Step 6: Function to Attempt Login**
def attempt_login():
    """Attempts login and ensures CAPTCHA is always filled."""
    try:
        # Locate CAPTCHA Image and Capture Screenshot
        captcha_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//canvas[@id='canvas']"))
        )
        captcha_png = captcha_element.screenshot_as_png

        # Process and Extract CAPTCHA Text
        processed_captcha = preprocess_captcha(captcha_png)
        processed_captcha.save("debug_captcha.png")  # Save for debugging
        captcha_text = extract_captcha_text(processed_captcha)

        if not captcha_text:
            print("❌ Failed to extract valid CAPTCHA. Refreshing...")
            return False

        print(f"🔹 Attempting login with CAPTCHA: {captcha_text}")

        # Locate and fill CAPTCHA input field
        captcha_input = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Enter Captcha here..']"))
        )
        captcha_input.clear()
        captcha_input.send_keys(captcha_text)
        time.sleep(1)  # Allow field to register input

        # Ensure CAPTCHA is filled
        filled_captcha = captcha_input.get_attribute("value")
        print(f"🔍 CAPTCHA field value: '{filled_captcha}'")

        if not filled_captcha:
            print("⚠️ CAPTCHA field is empty! Retrying input...")
            captcha_input.send_keys(captcha_text)
            time.sleep(1)

        # Click Login Button using JavaScript (More Reliable)

        driver.execute_script("arguments[0].click();", login_button)

        # Wait for response
        time.sleep(5)

        # Check if login was successful
        current_url = driver.current_url
        print(f"🌍 Current URL after login: {current_url}")

        if "dashboard" in current_url.lower():
            print("✅ Login Successful!")
            return True
        else:
            print("❌ Login Failed. Trying again...")
            return False

    except NoSuchElementException as e:
        print(f"⚠️ Element not found: {e}")
    except TimeoutException:
        print("⏳ Page took too long to load!")

    return False  # Return False if login is not successful


# **Step 7: Execute Login Attempts**
MAX_ATTEMPTS = 4
for attempt in range(MAX_ATTEMPTS):
    login_successful = attempt_login()

    if login_successful:
        break  # Stop immediately if login succeeds
    else:
        refresh_captcha()  # Refresh only if login fails

# **Final Handling**
if not login_successful:
    print("❌ All attempts failed. Please try manually.")

time.sleep(1)


def select_dropdown_option(dropdown_id, option_name):
    """Selects an option from a dropdown dynamically."""
    try:
        # 🔹 Click dropdown to open it
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()

        # 🔹 Wait for dropdown options to be visible
        options_list_css = ".dx-list-item"
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, options_list_css)))

        # 🔹 Find all dropdown options
        options = driver.find_elements(By.CSS_SELECTOR, options_list_css)

        # Debugging: Print all available options
        print(f"🔍 Available options for {dropdown_id}: {[opt.text.strip() for opt in options]}")

        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                print(f"✅ Selected: {option_name}")
                return

        print(f"❌ Option '{option_name}' not found in {dropdown_id}!")

    except Exception as e:
        print(f"⚠️ Error selecting '{option_name}' in {dropdown_id}: {e}")


def select_month():
    """Selects the 'Month' button from the button group."""
    try:
        # 🔹 Find and click the 'Month' button
        month_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'dx-buttongroup-item') and .//span[text()='Month']]")))
        month_button.click()
        print("✅ Selected: Month")

    except Exception as e:
        print(f"⚠️ Error selecting Month: {e}")


# def select_day():
#     """Selects the 'Month' button from the button group."""
#     try:
#         # 🔹 Find and click the 'Month' button
#         day_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
#             (By.XPATH, "//div[contains(@class, 'dx-buttongroup-item') and .//span[text()='Day']]")))
#         day_button.click()
#         print("✅ Selected: Day")
#
#     except Exception as e:
#         print(f"⚠️ Error selecting Month: {e}")


def fill_date_input(date_value):
    """Fills the date input field with the given value (e.g., 'November 2024')."""
    try:
        # 🔹 Locate the date input field
        date_input = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@aria-label='Date']")))

        # 🔹 Clear existing value and enter new date
        date_input.clear()
        date_input.send_keys(date_value)
        ##date_input.send_keys(Keys.TAB)  # 🔹 Ensures the input is registered

        print(f"✅ Entered Date: {date_value}")

    except Exception as e:
        print(f"⚠️ Error entering date '{date_value}': ")


# 🔹 Select Area
select_dropdown_option("ddl-area", "PIA   ")  # Replace with the actual Area option

# 🔹 Select Substation
select_dropdown_option("ddl-substation", "SS_001")  # Replace with the actual Substation option

# 🔹 Select MV Feeder
select_dropdown_option("ddl-feeder", "SS_001")  # Replace with the actual MV Feeder option

# 🔹 Select Month/day
select_month()
# select_day()
# 🔹 Fill Date Input
fill_date_input("June 2024")

# 🔹 Click Right Arrow Button
try:
    arrow_button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "i.dx-icon.dx-icon-arrowright"))
    )
    arrow_button.click()
    print("✅ Clicked Right Arrow Button")
except Exception as e:
    print(f"⚠️ Error clicking right arrow button: {e}")

time.sleep(8)

# Find the row containing both the specific meter number and MV feeder
meter_number = "Q0792230_1"
mv_feeder = "MV_001"

row_xpath = f"//tr[td[contains(text(), '{meter_number}')] and td[contains(text(), '{mv_feeder}')]]"

try:
    # Locate the row
    row_element = driver.find_element(By.XPATH, row_xpath)

    # Find the "View" link within the same row
    view_link = row_element.find_element(By.XPATH, ".//a[contains(text(), 'View')]")

    # Click the "View" link
    view_link.click()
    print("Successfully clicked on View.")
    time.sleep(4)

except Exception as e:
    print(f"Error: {e}")
time.sleep(4)
# click the "Detailed view for Voltage profile "
driver.find_element(By.XPATH, "//a[@id='VPDetailedLink']").click()

time.sleep(4)

# Click the appropriate chart tab based on chart type
try:
    chart_tab = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, TAB_SELECTOR))
    )
    chart_tab.click()
    print(f"✅ Selected {CHART_TYPE} tab")
except Exception as e:
    print(f"⚠️ Error selecting {CHART_TYPE} tab: {e}")

time.sleep(5)
import os
import re
import time
import calendar
import numpy as np
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

# Global variables
last_successful_position = None
# Define parameters centrally - modify this list if parameters change
# CHART_PARAMETERS = ["Date", "Line 1", "Line 2", "Line 3", "Avg", "Neutral"]
CHART_PARAMETERS = ["Date", "Active","Apparent","Reactive"]
# CHART_PARAMETERS = ["Date", "Phase 1","Phase 2","Phase 3","Avg"]
# CHART_PARAMETERS = ["Date", "Active","Apparent","Reactive","Active","Apparent","Reactive"]
class ChartDataExtractor:
    def __init__(self, driver, month, year):
        self.driver = driver
        self.month = month
        self.year = year
        self.month_name = calendar.month_name[month]
        self.month_abbr = calendar.month_abbr[month]
        self.num_days = calendar.monthrange(year, month)[1]
        self.chart_container = None
        self.chart_x = None
        self.chart_y = None
        self.chart_width = None
        self.chart_height = None
        self.collected_data = []
        self.target_parameters = None

        # Create a directory for screenshots if it doesn't exist
        os.makedirs("chart_screenshots", exist_ok=True)

    def find_and_setup_chart(self):
        """Find the chart container and set up initial chart properties"""
        try:
            self.chart_container = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "(//*[local-name()='svg' and @class='dxc dxc-chart'])[1]"))
            )
            print("Chart container found successfully")

            chart_rect = self.driver.find_element(By.XPATH,
                                                  "(//*[local-name()='svg' and @class='dxc dxc-chart'])[1]//*[name()='rect']")
            self.chart_width = chart_rect.size['width']
            self.chart_height = chart_rect.size['height']
            self.chart_x = chart_rect.location['x']
            self.chart_y = chart_rect.location['y']

            # Scroll chart into view
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", self.chart_container)
            time.sleep(2)  # Allow time for scrolling and rendering
            return True
        except Exception as e:
            print(f"Chart container not found: {str(e)}")
            return False

    def extract_tooltip_data(self, target_params=None):
        """Extract data from tooltip with improved detection for any parameter names"""
        if target_params is None:
            target_params = CHART_PARAMETERS

        # First, look for tooltip elements with specific class names
        tooltip_selectors = [
            ".dxc-tooltip", ".chart-tooltip", "[role='tooltip']",
            ".highcharts-tooltip", ".tooltip", ".c3-tooltip",
            ".nvtooltip", ".chartjs-tooltip"
        ]

        tooltip_text = None
        for selector in tooltip_selectors:
            try:
                tooltip_elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                if tooltip_elements and tooltip_elements[0].is_displayed():
                    tooltip_text = tooltip_elements[0].text
                    break
            except:
                pass

        if not tooltip_text:
            try:
                # If no tooltip found by selector, try to find any element that might be a tooltip
                month_abbr_lower = self.month_abbr.lower()
                month_name_lower = self.month_name.lower()

                tooltip_elements = self.driver.execute_script(f"""
                    return Array.from(document.querySelectorAll('*')).filter(el => {{
                        const style = window.getComputedStyle(el);
                        return style.display !== 'none' &&
                               style.visibility !== 'hidden' &&
                               style.opacity !== '0' &&
                               el.innerText &&
                               (el.innerText.toLowerCase().includes('{month_name_lower}') ||
                                el.innerText.toLowerCase().includes('{month_abbr_lower}') ||
                                el.innerText.includes('{self.month}/') ||
                                el.innerText.match(/\\d+[\\s]*\\w+/));
                    }}).map(el => el.innerText);
                """)

                if tooltip_elements and len(tooltip_elements) > 0:
                    tooltip_text = " | ".join(tooltip_elements)
            except:
                pass

        extracted_data = {}
        if tooltip_text:
            # First, extract the explicitly specified parameters
            for param in target_params:
                match = re.search(rf"{param}:\s*([^\n]+)", tooltip_text, re.IGNORECASE)
                if match:
                    extracted_data[param] = match.group(1).strip()
                else:
                    extracted_data[param] = None  # Parameter not found

            # Now, try to extract any other parameters that might be present
            # Look for patterns like "Parameter: Value" or "Parameter - Value"
            additional_params = re.findall(r"([A-Za-z0-9\s]+)(?:\:|\-)\s*([^\n]+)", tooltip_text)
            for param, value in additional_params:
                param_key = param.strip()
                if param_key not in extracted_data and param_key.lower() not in [p.lower() for p in
                                                                                 extracted_data.keys()]:
                    extracted_data[param_key] = value.strip()

        return extracted_data, tooltip_text

    def detect_available_parameters(self):
        """Try to detect what parameters are available in the chart by scanning tooltips"""
        print("Detecting available chart parameters...")

        # Move to the middle of the chart to get a sample tooltip
        middle_x = self.chart_x + (self.chart_width / 2)
        middle_y = self.chart_y + (self.chart_height / 2)

        try:
            js_script = f"""
            var evt = new MouseEvent('mousemove', {{
                'view': window,
                'bubbles': true,
                'cancelable': true,
                'clientX': {int(middle_x)},
                'clientY': {int(middle_y)}
            }});
            document.elementFromPoint({int(middle_x)}, {int(middle_y)}).dispatchEvent(evt);
            """
            self.driver.execute_script(js_script)
            time.sleep(1)

            # Try to get the tooltip text
            tooltip_text = None
            tooltip_selectors = [
                ".dxc-tooltip", ".chart-tooltip", "[role='tooltip']",
                ".highcharts-tooltip", ".tooltip", ".c3-tooltip",
                ".nvtooltip", ".chartjs-tooltip"
            ]

            for selector in tooltip_selectors:
                try:
                    tooltip_elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if tooltip_elements and tooltip_elements[0].is_displayed():
                        tooltip_text = tooltip_elements[0].text
                        break
                except:
                    pass

            if not tooltip_text:
                # If no tooltip found by selector, try to find any element that might be a tooltip
                tooltip_elements = self.driver.execute_script("""
                    return Array.from(document.querySelectorAll('*')).filter(el => {
                        const style = window.getComputedStyle(el);
                        return style.display !== 'none' &&
                               style.visibility !== 'hidden' &&
                               style.opacity !== '0' &&
                               el.innerText &&
                               (el.innerText.includes(':') || el.innerText.match(/\\d+[\\s]*\\w+/));
                    }).map(el => el.innerText);
                """)

                if tooltip_elements and len(tooltip_elements) > 0:
                    tooltip_text = " | ".join(tooltip_elements)

            # Try to extract parameter names from the tooltip text
            detected_params = ["Date"]  # Always include Date

            if tooltip_text:
                # Look for patterns like "Parameter: Value" or "Parameter - Value"
                pattern_matches = re.findall(r"([A-Za-z0-9\s]+)(?:\:|\-)\s*([^\n]+)", tooltip_text)
                for param, _ in pattern_matches:
                    param_key = param.strip()
                    if param_key.lower() != "date" and param_key not in detected_params:
                        detected_params.append(param_key)

            if len(detected_params) > 1:
                print(f"Detected {len(detected_params)} parameters: {', '.join(detected_params)}")
                return detected_params
        except Exception as e:
            print(f"Error detecting parameters: {e}")

        # If we couldn't detect parameters, return the default ones
        print(f"Using default parameters: {', '.join(CHART_PARAMETERS)}")
        return CHART_PARAMETERS

    def scan_available_days(self):
        """Scan the chart to discover what days are actually present"""
        print("Scanning chart to discover available days...")

        # Set parameters for scanning
        left_buffer = self.chart_x + (self.chart_width * 0.02)  # 2% buffer from left edge
        right_edge = self.chart_x + self.chart_width - (self.chart_width * 0.02)
        middle_y = self.chart_y + (self.chart_height / 2)

        # Create a higher resolution scan for more precise results
        positions = np.linspace(left_buffer, right_edge, 100)  # Scan 100 positions

        found_days = {}  # Dictionary to store {day: (x_position, extracted_info, tooltip_text)}

        # Reset mouse position before starting
        actions = ActionChains(self.driver)
        actions.move_to_element(self.chart_container).move_by_offset(-100, -100).perform()
        time.sleep(1)

        for x_pos in positions:
            try:
                # Move to the position
                js_script = f"""
                var evt = new MouseEvent('mousemove', {{
                    'view': window,
                    'bubbles': true,
                    'cancelable': true,
                    'clientX': {int(x_pos)},
                    'clientY': {int(middle_y)}
                }});
                document.elementFromPoint({int(x_pos)}, {int(middle_y)}).dispatchEvent(evt);
                """
                self.driver.execute_script(js_script)
                time.sleep(0.5)

                # Extract tooltip data
                extracted_info, full_tooltip_text = self.extract_tooltip_data(self.target_parameters)

                if extracted_info:
                    # Try to extract the day from the tooltip
                    found_day = None
                    date_match = re.search(r"Date:\s*(\d+)\s*(?:-|of)\s*\w+", full_tooltip_text, re.IGNORECASE)
                    if date_match:
                        try:
                            found_day = int(date_match.group(1).strip())
                        except ValueError:
                            pass
                    elif "Date" in extracted_info and extracted_info["Date"]:
                        day_match_simple = re.search(r"(\d+)", extracted_info["Date"])
                        if day_match_simple:
                            try:
                                found_day = int(day_match_simple.group(1).strip())
                            except ValueError:
                                pass

                    if found_day is not None and 1 <= found_day <= self.num_days:
                        # Only add this position if it's the first time we've seen this day
                        # or if it's closer to a previously found position for the same day
                        if found_day not in found_days:
                            found_days[found_day] = (x_pos, extracted_info, full_tooltip_text)
                            print(f"Found day {found_day} at position {x_pos}")

                        # Take a quick screenshot for debugging
                        screenshot_path = f"chart_screenshots/scan_day_{found_day}.png"
                        self.driver.save_screenshot(screenshot_path)
            except Exception as e:
                print(f"Error at position {x_pos}: {str(e)}")

        print(f"Scan complete. Found {len(found_days)} unique days: {sorted(found_days.keys())}")
        return found_days

    def extract_chart_data(self):
        """Modified main function to extract only available data points"""
        print(f"Extracting data for {self.month_name} {self.year}...")

        # Find and set up the chart
        if not self.find_and_setup_chart():
            return None

        # Detect available parameters or use default ones
        self.target_parameters = self.detect_available_parameters()

        # Reset mouse position before starting
        actions = ActionChains(self.driver)
        actions.move_to_element(self.chart_container).move_by_offset(-100, -100).perform()
        time.sleep(1)

        # Scan the chart to discover available days
        available_days = self.scan_available_days()

        if not available_days:
            print("No data points were found in the chart.")
            return None

        print(f"Found {len(available_days)} unique days: {sorted(available_days.keys())}")

        # Clear previously collected data to ensure fresh start
        self.collected_data = []

        # Extract data for each available day
        for day, (x_pos, extracted_info, full_tooltip_text) in available_days.items():
            # Move to the position to ensure tooltip is visible
            js_script = f"""
            var evt = new MouseEvent('mousemove', {{
                'view': window,
                'bubbles': true,
                'cancelable': true,
                'clientX': {int(x_pos)},
                'clientY': {int(self.chart_y + (self.chart_height / 2))}
            }});
            document.elementFromPoint({int(x_pos)}, {int(self.chart_y + (self.chart_height / 2))}).dispatchEvent(evt);
            """
            self.driver.execute_script(js_script)
            time.sleep(0.5)

            # Take a screenshot for verification
            screenshot_path = f"chart_screenshots/data_day_{day}.png"
            self.driver.save_screenshot(screenshot_path)

            # Initialize day data
            day_data = {
                'Date': f"{day} {self.month_name}"
            }

            # Add data for each parameter (excluding Date which we formatted above)
            for param in self.target_parameters:
                if param != "Date":
                    day_data[param] = self.clean_numeric_value(extracted_info.get(param, ''))

            self.collected_data.append(day_data)

            # Log the results
            param_values = ", ".join(
                [f"{param}={day_data.get(param, '')}" for param in self.target_parameters if param != "Date"])
            print(f"Day {day}: Extracted data: {param_values}")

        # Save to Excel
        return self.save_to_excel()

    def clean_numeric_value(self, value):
        """Extract numeric value from string like '29.88 kW'"""
        if value and isinstance(value, str):
            numeric_match = re.search(r'([\d.]+)', value)
            if numeric_match:
                return numeric_match.group(1)
        return value

    def save_to_excel(self):
            """Save collected data to Excel file"""
            print("Saving data to Excel...")

            if not self.collected_data:
                print("No data to save.")
                return None

            # Create a pandas DataFrame
            df = pd.DataFrame(self.collected_data)

            # Ensure we have the 'Date' column
            if 'Date' not in df.columns:
                print("Warning: 'Date' column not found in collected data.")
                return None

            # Convert numeric columns
            for col in df.columns:
                if col != 'Date':
                    # Use safer approach instead of deprecated errors='ignore'
                    try:
                        df[col] = pd.to_numeric(df[col])
                    except (ValueError, TypeError):
                        pass  # Keep as is if conversion fails

            # Sort by the day number in the date
            def extract_day(date_str):
                try:
                    return int(re.search(r'^(\d+)', date_str).group(1))
                except:
                    return 0

            df['_day_sort'] = df['Date'].apply(extract_day)
            df = df.sort_values('_day_sort')
            df = df.drop('_day_sort', axis=1)

            # Save to Excel
            filename = f"chart_{CHART_TYPE}_data_{self.month_name}_{self.year}.xlsx"
            df.to_excel(filename, index=False)
            print(f"Data saved to {filename}")

            return filename

# ADD THE NEW FUNCTIONS BELOW THIS LINE
def extract_from_database(year, month, meter_id=None):
    """Extracts data from PostgreSQL database for the specified month/year and meter ID."""
    print(f"Extracting data from database for {calendar.month_name[month]} {year}...")

    try:
        # PostgreSQL connection details
        conn = psycopg2.connect(
            host="10.11.16.102",
            database="serviceplatformdb11003_v1",
            user="postgres",
            password="postgres",
            port="5432"
        )

        # Construct date range for the query
        start_date = f"{year}-{month:02d}-01"
        _, last_day = calendar.monthrange(year, month)
        end_date = f"{year}-{month:02d}-{last_day}"

        # Create column selection string for SQL query from DB_COLUMNS
        db_cols = DB_COLUMNS.copy()
        date_col = db_cols.pop(0)  # Remove date column from mapping

        # Build column list for SQL query with aliases that match chart parameters
        column_selection = f"{date_col}"
        for i, col in enumerate(db_cols):
            param_name = CHART_PARAMETERS[i + 1]  # Skip Date in chart parameters
            column_selection += f', {col} as "{param_name}"'

        # SQL query with appropriate filtering
        query = f"""
        SELECT {column_selection}
        FROM tenant01.tb_rpt_dtmis_daily
        WHERE version_no = 2
          AND {date_col} >= '{start_date}'
          AND {date_col} <= '{end_date}'
        ORDER BY {date_col};
        """

        # Fetch the data into a pandas DataFrame
        df = pd.read_sql(query, conn)

        # Transform dateandtime to match the chart format (e.g., "15 July")
        df['Date'] = df[date_col].dt.strftime('%d %B')

        # Reorder columns to match chart data format
        columns = ['Date'] + [col for col in df.columns if col not in ['Date', date_col]]
        df = df[columns]

        # Close the connection
        conn.close()

        # Write to Excel
        filename = f"{CHART_TYPE}_FROM_db_data_{calendar.month_name[month]}_{year}.xlsx"
        df.to_excel(filename, index=False)
        print(f"Database data written to '{filename}'")

        return df, filename

    except Exception as e:
        print(f"Error extracting data from database: {e}")
        return None, None

def compare_chart_and_db_data(chart_file, db_file, output_file=None):
        """Compares chart data with database data and generates a highlighted comparison file."""
        print("Comparing chart data with database data...")

        try:
            # Check if files exist
            if not os.path.exists(chart_file) or not os.path.exists(db_file):
                print(f"Chart file exists: {os.path.exists(chart_file)}")
                print(f"DB file exists: {os.path.exists(db_file)}")
                return None

            # Load Excel files
            df_chart = pd.read_excel(chart_file)
            df_db = pd.read_excel(db_file)

            print(f"Chart data shape: {df_chart.shape}")
            print(f"DB data shape: {df_db.shape}")
            print(f"Chart data columns: {df_chart.columns.tolist()}")
            print(f"DB data columns: {df_db.columns.tolist()}")

            # Debug first few rows of each DataFrame
            print("\nChart data preview:")
            print(df_chart.head(2))
            print("\nDB data preview:")
            print(df_db.head(2))

            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Comparison"

            # Add headers
            headers = ["Day", "Parameter", "Chart Value", "DB Value", "Difference", "Match"]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
                ws.cell(row=1, column=col_idx).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Define styles
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            # Direct cell-by-cell comparison (ignoring column names)
            # This approach works even if column naming is different

            # First, extract days from both datasets
            if 'Date' in df_chart.columns:
                df_chart['Day'] = df_chart['Date'].str.extract(r'^(\d+)').astype(int)
            else:
                print("Warning: No 'Date' column in chart data")
                return None

            if 'dateandtime' in df_db.columns:
                df_db['Day'] = pd.to_datetime(df_db['dateandtime']).dt.day
            elif 'Date' in df_db.columns:
                df_db['Day'] = df_db['Date'].str.extract(r'^(\d+)').astype(int)
            else:
                print("Warning: No 'Date' or 'dateandtime' column in DB data")
                return None

            # Get all parameters to compare (excluding date columns)
            chart_params = [col for col in df_chart.columns if col != 'Date' and col != 'Day']
            db_params = [col for col in df_db.columns if col not in ['Date', 'dateandtime', 'Day']]

            # Force specific parameter mapping if needed
            param_mapping = {}
            # Try auto-matching on exact names first
            for cp in chart_params:
                for dp in db_params:
                    if cp == dp:
                        param_mapping[cp] = dp
                        break

            # If auto-matching didn't find all mappings, try with case-insensitive matching
            for cp in chart_params:
                if cp not in param_mapping:
                    for dp in db_params:
                        if cp.lower() == dp.lower():
                            param_mapping[cp] = dp
                            break

            # If we still don't have mappings for all chart parameters, try substring matching
            for cp in chart_params:
                if cp not in param_mapping:
                    for dp in db_params:
                        # Check if chart param is contained in db param or vice versa
                        if cp.lower() in dp.lower() or dp.lower() in cp.lower():
                            param_mapping[cp] = dp
                            break

            # If still missing mappings, assign in order
            remaining_chart_params = [cp for cp in chart_params if cp not in param_mapping]
            remaining_db_params = [dp for dp in db_params if dp not in param_mapping.values()]

            for i, cp in enumerate(remaining_chart_params):
                if i < len(remaining_db_params):
                    param_mapping[cp] = remaining_db_params[i]

            print(f"Parameter mapping: {param_mapping}")

            # Perform comparison
            row_idx = 2
            all_days = sorted(set(df_chart['Day'].tolist() + df_db['Day'].tolist()))
            print(f"All days to compare: {all_days}")

            for day in all_days:
                chart_rows = df_chart[df_chart['Day'] == day]
                db_rows = df_db[df_db['Day'] == day]

                if not chart_rows.empty and not db_rows.empty:
                    chart_row = chart_rows.iloc[0]
                    db_row = db_rows.iloc[0]

                    for chart_param, db_param in param_mapping.items():
                        # Get values (handle missing columns)
                        chart_val = chart_row.get(chart_param, "N/A")
                        db_val = db_row.get(db_param, "N/A")

                        # Handle comparison
                        try:
                            # Try numeric comparison with tolerance
                            chart_num = float(chart_val)
                            db_num = float(db_val)
                            diff = chart_num - db_num
                            match = abs(diff) < 0.01
                        except (ValueError, TypeError):
                            # Fall back to string comparison
                            str_chart = str(chart_val).strip()
                            str_db = str(db_val).strip()
                            match = str_chart == str_db
                            diff = "N/A"

                        # Write data to worksheet
                        ws.cell(row=row_idx, column=1, value=day)
                        ws.cell(row=row_idx, column=2, value=chart_param)
                        ws.cell(row=row_idx, column=3, value=chart_val)
                        ws.cell(row=row_idx, column=4, value=db_val)
                        ws.cell(row=row_idx, column=5, value=diff)
                        ws.cell(row=row_idx, column=6, value="YES" if match else "NO")

                        # Apply formatting
                        for col in range(1, 7):
                            cell = ws.cell(row=row_idx, column=col)
                            cell.border = thin_border
                            if col == 6:  # Match column
                                cell.fill = green_fill if match else red_fill

                        row_idx += 1
                else:
                    # Handle days that exist in one dataset but not the other
                    missing_in = "chart" if chart_rows.empty else "database"

                    # Write a summary row for the missing day
                    ws.cell(row=row_idx, column=1, value=day)
                    ws.cell(row=row_idx, column=2, value=f"DATA MISSING IN {missing_in.upper()}")

                    for col in range(1, 7):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.border = thin_border
                        cell.fill = red_fill

                    row_idx += 1

            # If we didn't add any comparison rows, add an error message to the sheet
            if row_idx == 2:
                for col_idx in range(1, 7):
                    if col_idx == 1:
                        ws.cell(row=2, column=col_idx, value="NO MATCHING DATA FOUND")
                    else:
                        ws.cell(row=2, column=col_idx, value="")
                    ws.cell(row=2, column=col_idx).border = thin_border
                    ws.cell(row=2, column=col_idx).fill = red_fill

            # Save the workbook
            if output_file is None:
                output_file = f"comparison_report_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"

            wb.save(output_file)
            print(f"Comparison report saved to {output_file} with {row_idx - 2} comparison rows")
            return output_file

        except Exception as e:
            import traceback
            print(f"Error comparing data: {e}")
            print(f"Detailed traceback: {traceback.format_exc()}")
            return None
    # Main execution
if __name__ == "__main__":


    # Get month and year
    manual_month = 6  # July
    manual_year = 2024

    # Optional: Specify meter ID for database filtering
    meter_id = "Q0792230_1"  # Use None to get all meters

    # Validate the month and year
    if 1 <= manual_month <= 12 and manual_year > 0:
        # First extract from the web chart
        extractor = ChartDataExtractor(driver, manual_month, manual_year)
        chart_file = extractor.extract_chart_data()

        # Then extract from database for the same period
        db_data, db_file = extract_from_database(manual_year, manual_month, meter_id)

        # Compare the data if both extractions were successful
        if chart_file and db_file:
            comparison_file = compare_chart_and_db_data(
                chart_file,
                db_file,
                f"comparison_{calendar.month_name[manual_month]}_{manual_year}.xlsx"
            )

            print("Process completed successfully!")
            print(f"Chart data: {chart_file}")
            print(f"Database data: {db_file}")
            print(f"Comparison report: {comparison_file}")
    else:
        print("Invalid month or year specified")

    # Cleanup
        driver.quit()

