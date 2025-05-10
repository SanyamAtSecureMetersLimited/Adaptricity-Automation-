import io

import os

import cv2
import re
import numpy as np
import pandas as pd
import selenium
from PIL import Image
import pytesseract
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time


# --- setup driver as you already have it ---
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.get("https://networkmonitoring.secure.online:10122/Account/Login")
# … do your login steps …
username_input = WebDriverWait(driver, 5).until(
    EC.presence_of_element_located((By.XPATH, "//input[@name='UserName']"))
)
username_input.clear()
username_input.send_keys("SanyamU")

password_input = driver.find_element(By.XPATH, "//input[@name='Password']")
password_input.clear()
password_input.send_keys("Secure@12345")

login_button = driver.find_element(By.ID, 'btnlogin')
login_button.click()

# WebDriverWait(driver, 10).until(
#     EC.url_to_be("https://networkmonitoring.secure.online:10097/LVMV/DTMonitoring")
# )
# WebDriverWait(driver, 20).until(
#     EC.visibility_of_element_located((By.ID, "assetalertbox"))
# )


def select_dropdown_option(dropdown_id, option_name):
    """Selects an option from a dropdown dynamically."""
    try:
        # 🔹 Click dropdown to open it
        dropdown = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, dropdown_id)))
        dropdown.click()

        # 🔹 Wait for dropdown options to be visible
        options_list_css = ".dx-list-item"
        WebDriverWait(driver, 3).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, options_list_css)))

        # 🔹 Find all dropdown options
        options = driver.find_elements(By.CSS_SELECTOR, options_list_css)

        # Debugging: Print all available options
        print(f"🔍 Available options for {dropdown_id}: {[opt.text.strip() for opt in options]}")

        for option in options:
            if option.text.strip().lower() == option_name.lower():
                option.click()
                print(f"✅ Selected: {option_name}")
                return

        print(f"❌ Option '{option_name}' not found in {dropdown_id}!")

    except Exception as e:
        print(f"⚠️ Error selecting '{option_name}' in {dropdown_id}: {e}")


def select_month():
    """Selects the 'Month' button from the button group."""
    try:
        # 🔹 Find and click the 'Month' button
        month_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'dx-buttongroup-item') and .//span[text()='Month']]")))
        month_button.click()
        print("✅ Selected: Month")

    except Exception as e:
        print(f"⚠️ Error selecting Month: {e}")
# def select_day():
#     """Selects the 'Month' button from the button group."""
#     try:
#         # 🔹 Find and click the 'Month' button
#         day_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable(
#             (By.XPATH, "//div[contains(@class, 'dx-buttongroup-item') and .//span[text()='Day']]")))
#         day_button.click()
#         print("✅ Selected: Day")
#
#     except Exception as e:
#         print(f"⚠️ Error selecting Month: {e}")


def fill_date_input(date_value):
    """Fills the date input field with the given value (e.g., 'November 2024')."""
    try:
        # 🔹 Locate the date input field
        date_input = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@aria-label='Date']")))

        # 🔹 Clear existing value and enter new date
        date_input.clear()
        date_input.send_keys(date_value)
        ##date_input.send_keys(Keys.TAB)  # 🔹 Ensures the input is registered

        print(f"✅ Entered Date: {date_value}")

    except Exception as e:
        print(f"⚠️ Error entering date '{date_value}': ")


# 🔹 Select Area
select_dropdown_option("ddl-area", "PIA   ")  # Replace with the actual Area option

# 🔹 Select Substation
select_dropdown_option("ddl-substation", "SS_001")  # Replace with the actual Substation option

# 🔹 Select MV Feeder
select_dropdown_option("ddl-feeder", "SS_001")  # Replace with the actual MV Feeder option

# 🔹 Select Month/day
select_month()
# select_day()
# 🔹 Fill Date Input
fill_date_input("June 2024")
# 🔹 Click Right Arrow Button
try:
    arrow_button = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "i.dx-icon.dx-icon-arrowright"))
    )
    arrow_button.click()
    print("✅ Clicked Right Arrow Button")
except Exception as e:
    print(f"⚠️ Error clicking right arrow button: {e}")

time.sleep(10)
# ASSETS
assets = []
asset_cards = driver.find_elements(
    By.XPATH,
    "//div[@id='LVMonitoring']//div[contains(@class,'containerbox')]/div"
)
for card in asset_cards:
    # look for the big number
    value = card.find_element(
        By.XPATH,
        ".//span[contains(@class,'lbl_header_label2') and contains(@id,'count')]"
    ).text
    # look for the small label
    label = card.find_element(
        By.XPATH,
        ".//span[contains(@class,'lvmv-fs-7')]"
    ).text
    assets.append((label, value))

# ALERTS
alerts = []
alert_cards = driver.find_elements(
    By.XPATH,
    "//div[@id='divDTMonitoringCount']//div[contains(@class,'containerbox')]/div"
)
for card in alert_cards:
    value = card.find_element(
        By.XPATH,
        ".//span[contains(@class,'lbl_header_label2') and contains(@id,'count')]"
    ).text
    label = card.find_element(
        By.XPATH,
        ".//span[contains(@class,'lvmv-fs-7')]"
    ).text
    alerts.append((label, value))

print("Assets:", assets)
print("Alerts:", alerts)

import psycopg2
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# PostgreSQL connection details
conn = psycopg2.connect(
    host="10.11.16.102",
    database="serviceplatformdb11003_v1",
    user="postgres",
    password="postgres",
    port="5432"
)

# Queries
query = """
SELECT DISTINCT 
    vtc.nodetypeid, 
    vtc.voltagerating, 
    vtc.overvoltage, 
    vtc.voltageunbalance, 
    nm.nodetype, 
    nm.isactive,
    pfc.powerfactorthreshold
FROM servicemeta.tb_voltage_threshold_configuration vtc
JOIN servicemeta.tb_ntw_nodemaster nm 
    ON vtc.nodetypeid = nm.nodetypeid
LEFT JOIN servicemeta.tb_powerfactor_threshold_configuration pfc 
    ON vtc.nodetypeid = pfc.nodetypeid
WHERE vtc.nodetypeid = 153 
  AND vtc.voltagerating = 230;
"""

query2 = """
SELECT DISTINCT nodetype, 
    surveydate,
    avg_v,
    pf 
FROM tenant01.tb_nrm_loadsurveyprofile 
WHERE surveydate >= '2024-06-01' 
  AND surveydate <= '2024-06-30' 
  AND nodetype = 153 
ORDER BY surveydate ASC;
"""

# Fetch data
df_config = pd.read_sql(query, conn)
df_survey = pd.read_sql(query2, conn)
conn.close()

# Extract configuration values
voltagerating = df_config.iloc[0]["voltagerating"]
overvoltage = df_config.iloc[0]["overvoltage"]
voltageunbalance = df_config.iloc[0]["voltageunbalance"]
powerfactorthreshold = df_config.iloc[0]["powerfactorthreshold"]

# Calculate thresholds
overvoltage_condition = (voltagerating * overvoltage) / 100
voltageunbalance_condition = (voltagerating * voltageunbalance) / 100

# Add calculated columns
df_survey["overvoltagecondition"] = (230 - df_survey["avg_v"]).abs()
df_survey["voltageunbalancecondition"] = (230 - df_survey["avg_v"]).abs()  # This might need adjustment if not voltage-based
# You could alternatively use: (1 - df_survey["pf"]).abs()

# Write to Excel
file_name = "Combined_MeterData_Report.xlsx"
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df_config.to_excel(writer, sheet_name="meter config", index=False)
    df_survey.to_excel(writer, sheet_name="loadsurveydata", index=False)

# Load workbook and sheet
wb = load_workbook(file_name)
ws = wb["loadsurveydata"]

# Define styles
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # Overvoltage
orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Voltage unbalance
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Poor PF

# Column headers
header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
ovc_col = header.get("overvoltagecondition")
vuc_col = header.get("voltageunbalancecondition")
pf_col = header.get("pf")

# Flags
overvoltage_occurred_flag = False
voltageunbalance_occurred_flag = False
poor_powerfactor_occurred_flag = False

# Check each row
for row in range(2, ws.max_row + 1):
    # Overvoltage
    ovc_cell = ws.cell(row=row, column=ovc_col)
    if ovc_cell.value is not None and ovc_cell.value >= overvoltage_condition:
        ovc_cell.fill = red_fill
        overvoltage_occurred_flag = True

    # Voltage Unbalance
    vuc_cell = ws.cell(row=row, column=vuc_col)
    if vuc_cell.value is not None and vuc_cell.value >= voltageunbalance_condition:
        vuc_cell.fill = orange_fill
        voltageunbalance_occurred_flag = True

    # Poor Power Factor
    pf_cell = ws.cell(row=row, column=pf_col)
    if pf_cell.value is not None and pf_cell.value <= powerfactorthreshold:
        pf_cell.fill = yellow_fill
        poor_powerfactor_occurred_flag = True

# Save workbook
wb.save(file_name)

# Final status variables
overvoltage_occurred = "Overvoltage occurred" if overvoltage_occurred_flag else "No overvoltage"
voltageunbalance_occurred = "Voltage unbalance occurred" if voltageunbalance_occurred_flag else "No voltage unbalance"
poor_pf_occurred = "Poor power factor occurred" if poor_powerfactor_occurred_flag else "Power factor is within threshold"

# Output statuses
print(f"Excel file updated with conditional formatting, Check {file_name}" )
print("Status:", overvoltage_occurred)
print("Status:", voltageunbalance_occurred)
print("Status:", poor_pf_occurred)

time.sleep(5)
driver.quit()




